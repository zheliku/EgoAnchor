"""公式求值器与缓存注入的独立自检。"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, r"p:\VSCode-Project\EgoAnchor\2026-EgoAnchor\material")
import sim_opus5_v3_formula as FM

OUT = Path(__file__).with_name("formula_selftest.xlsx")


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for idx, value in enumerate([3, 5, 5, 6, 7, 2, 4, 4], start=1):
        ws.cell(row=idx, column=1, value=value)
    other = wb.create_sheet("Calc")
    other["A1"] = "=AVERAGE(Data!A1:A8)"
    other["A2"] = "=MEDIAN(Data!A1:A8)"
    other["A3"] = "=QUARTILE.INC(Data!A1:A8,1)"
    other["A4"] = "=QUARTILE.INC(Data!A1:A8,3)"
    other["A5"] = "=STDEV.S(Data!A1:A8)"
    other["A6"] = "=COUNTIF(Data!A1:A8,\">=5\")"
    other["A7"] = "=IF(A1>4,\"high\",\"low\")"
    other["A8"] = "=ROUND(A1,2)"
    other["A9"] = "=(A1-A2)/A5"
    other["A10"] = "=ABS(A2-A1)"
    other["A11"] = "=SUM(Data!A1:A8)/COUNT(Data!A1:A8)"
    other["A12"] = "=MIN(Data!A1:A8)&\"/\"&MAX(Data!A1:A8)"
    other["A13"] = "=6-Data!A6"
    wb.save(OUT)

    wb2 = openpyxl.load_workbook(OUT)
    ctx = FM.EvalContext(wb2)
    values = ctx.evaluate_all()
    names = [ws.title for ws in wb2.worksheets]
    got = {f"{s}!{FM.index_to_col(c)}{r}": v for (s, r, c), v in values.items()}
    import numpy as np
    data = np.array([3, 5, 5, 6, 7, 2, 4, 4], dtype=float)
    expect = {
        "Calc!A1": data.mean(),
        "Calc!A2": float(np.median(data)),
        "Calc!A3": float(np.percentile(data, 25)),
        "Calc!A4": float(np.percentile(data, 75)),
        "Calc!A5": float(np.std(data, ddof=1)),
        "Calc!A6": 4.0,
        "Calc!A7": "high",
        "Calc!A8": round(data.mean(), 2),
        "Calc!A9": (data.mean() - float(np.median(data))) / float(np.std(data, ddof=1)),
        "Calc!A10": abs(float(np.median(data)) - data.mean()),
        "Calc!A11": data.mean(),
        "Calc!A12": "2/7",
        "Calc!A13": 4.0,
    }
    ok = True
    for key, want in expect.items():
        have = got[key]
        if isinstance(want, str):
            good = have == want
        else:
            good = abs(float(have) - float(want)) < 1e-9
        print(f"{'OK ' if good else 'BAD'} {key}: got={have!r} want={want!r}")
        ok &= good

    n = FM.inject_cached_values(OUT, values, names)
    print(f"injected {n} cached values")
    wb3 = openpyxl.load_workbook(OUT, data_only=True)
    for key, want in expect.items():
        sheet, ref = key.split("!")
        have = wb3[sheet][ref].value
        if isinstance(want, str):
            good = have == want
        else:
            good = have is not None and abs(float(have) - float(want)) < 1e-9
        print(f"{'OK ' if good else 'BAD'} cached {key}: {have!r}")
        ok &= good

    # 未支持语法必须抛错，不得静默降级
    for bad in ["=VLOOKUP(A1,Data!A1:A8,1)", "=SUMPRODUCT(A1:A2,A1:A2)", "=A1+{1,2}"]:
        try:
            FM.Parser(FM.tokenize(bad[1:]), bad[1:]).parse().eval(ctx, "Calc")
            print(f"BAD 未支持语法未抛错: {bad}")
            ok = False
        except FM.FormulaError as exc:
            print(f"OK  拒绝 {bad}: {str(exc)[:52]}")
    print("RESULT:", "PASS" if ok else "FAIL")
    OUT.unlink(missing_ok=True)


if __name__ == "__main__":
    main()
