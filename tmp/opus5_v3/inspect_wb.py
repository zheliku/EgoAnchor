"""检查产出工作簿：表结构、公式数量、缓存值可读性与关键单元格内容。"""
from __future__ import annotations

import sys
import zipfile
from pathlib import Path

import openpyxl

XLSX = Path(r"p:\VSCode-Project\EgoAnchor\2026-EgoAnchor\material"
            r"\EgoAnchor_Experiment3_Simulated_Claude-Opus-5_v3_24P.xlsx")


def main() -> int:
    print(f"file size = {XLSX.stat().st_size:,} bytes")
    with zipfile.ZipFile(XLSX) as zf:
        bad = zf.testzip()
        print(f"zip integrity: {'OK' if bad is None else 'BAD ' + str(bad)}")
        book = zf.read("xl/workbook.xml").decode()
        print(f"fullCalcOnLoad present: {'fullCalcOnLoad=\"1\"' in book}")

    wf = openpyxl.load_workbook(XLSX, data_only=False)
    wv = openpyxl.load_workbook(XLSX, data_only=True)
    print("\nsheet | rows x cols | formulas | cached-empty")
    total_f = total_empty = 0
    for ws in wf.worksheets:
        vs = wv[ws.title]
        n_f = n_empty = 0
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    n_f += 1
                    if vs.cell(row=cell.row, column=cell.column).value is None:
                        n_empty += 1
        total_f += n_f
        total_empty += n_empty
        print(f"{ws.title:<19} {ws.max_row:>4} x {ws.max_column:<3} {n_f:>6} {n_empty:>8}")
    print(f"{'TOTAL':<19} {'':>4}   {'':<3} {total_f:>6} {total_empty:>8}")
    if total_empty:
        print("!! 存在公式单元格缓存值为空 —— 离线查看器会显示空白")

    print("\n== Analysis A 段（主证实家族，公式实时计算列） ==")
    av = wv["Analysis"]
    print("row | item  short  OE Mdn[IQR]        EA Mdn[IQR]        dmed   dmean  dSD    dz     signs")
    for r in range(5, 12):
        vals = [av.cell(row=r, column=c).value for c in (1, 2, 3, 4, 5, 6, 7, 8, 15)]
        print(f"{r:>3} | " + "  ".join(str(v) for v in vals))

    print("\n== Analysis B 段（已发表量表家族） ==")
    for r in range(14, 19):
        vals = [av.cell(row=r, column=c).value for c in (1, 2, 3, 4, 5, 6, 13, 14, 15, 16)]
        print(f"{r:>3} | " + "  ".join(str(v) for v in vals))

    print("\n== Derived L1 前 3 行 ==")
    dv = wv["Derived"]
    for r in range(5, 8):
        vals = [dv.cell(row=r, column=c).value for c in range(1, 22)]
        print(f"{r:>3} | " + " ".join(str(v) for v in vals))

    print("\n== Derived L3 首行（TiA 反向计分与量表分） ==")
    hdr_row = None
    for r in range(1, dv.max_row + 1):
        if dv.cell(row=r, column=1).value == "Participant_ID" and \
           dv.cell(row=r, column=3).value == "Records行":
            hdr_row = r
    if hdr_row:
        hdrs = [dv.cell(row=hdr_row, column=c).value for c in range(1, 21)]
        vals = [dv.cell(row=hdr_row + 1, column=c).value for c in range(1, 21)]
        for h, v in zip(hdrs, vals):
            if h is not None:
                print(f"   {str(h):<14} = {v}")

    print("\n== Records A 段首行 ==")
    rv = wv["Records"]
    hdrs = [rv.cell(row=4, column=c).value for c in range(1, rv.max_column + 1)]
    vals = [rv.cell(row=5, column=c).value for c in range(1, rv.max_column + 1)]
    for h, v in zip(hdrs, vals):
        print(f"   {str(h):<22} = {v}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
