"""结构一致性抽查：Participants 的物体顺序/标签映射是否与 Records A 段逐行一致。"""
from __future__ import annotations

from pathlib import Path

import openpyxl

XLSX = Path(r"p:\VSCode-Project\EgoAnchor\2026-EgoAnchor\material"
            r"\EgoAnchor_Experiment3_Simulated_Claude-Opus-5_v3_24P.xlsx")

LABEL_SEQ = {"S1": (("A", "B"), ("B", "A"), ("A", "B")),
             "S2": (("B", "A"), ("A", "B"), ("B", "A"))}


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ps, rec = wb["Participants"], wb["Records"]

    # Participants：pid -> (三个物体, 序列, A=?, B=?, 先行方法)
    info: dict[str, tuple[list[str], str, str, str, str]] = {}
    for r in range(3, 27):
        pid = ps.cell(row=r, column=1).value
        info[str(pid)] = (
            [str(ps.cell(row=r, column=c).value) for c in (5, 6, 7)],
            str(ps.cell(row=r, column=4).value),
            str(ps.cell(row=r, column=9).value),
            str(ps.cell(row=r, column=10).value),
            str(ps.cell(row=r, column=11).value),
        )

    errors: list[str] = []
    per_pid: dict[str, list[tuple]] = {}
    for r in range(5, 149):
        pid = str(rec.cell(row=r, column=1).value)
        per_pid.setdefault(pid, []).append((
            int(rec.cell(row=r, column=2).value),      # block index
            int(rec.cell(row=r, column=4).value),      # object position
            str(rec.cell(row=r, column=5).value),      # object label
            str(rec.cell(row=r, column=7).value),      # shown label
            str(rec.cell(row=r, column=8).value),      # condition
            int(rec.cell(row=r, column=9).value),      # within-object order
            int(rec.cell(row=r, column=10).value),     # method repeat
        ))

    for pid, rows in per_pid.items():
        objects, seq, method_a, method_b, leading = info[pid]
        if len(rows) != 6:
            errors.append(f"{pid} 区块数 {len(rows)} != 6")
            continue
        rows.sort(key=lambda t: t[0])
        if [t[0] for t in rows] != [1, 2, 3, 4, 5, 6]:
            errors.append(f"{pid} block_index 不连续")
        # 物体顺序：每个物体连续两个区块，顺序须与 Participants 一致
        got_objects = [rows[0][2], rows[2][2], rows[4][2]]
        if got_objects != objects:
            errors.append(f"{pid} 物体顺序 Records={got_objects} != Participants={objects}")
        if rows[0][2] != rows[1][2] or rows[2][2] != rows[3][2] or rows[4][2] != rows[5][2]:
            errors.append(f"{pid} 同物体两区块未相邻")
        # 标签序列
        got_labels = tuple((rows[i][3][-1], rows[i + 1][3][-1]) for i in (0, 2, 4))
        if got_labels != LABEL_SEQ[seq]:
            errors.append(f"{pid} 标签序列 {got_labels} != {seq} 的 {LABEL_SEQ[seq]}")
        # 标签→方法映射在参与者内必须全程稳定
        for t in rows:
            expect = method_a if t[3].endswith("A") else method_b
            if t[4] != expect:
                errors.append(f"{pid} block{t[0]} 标签{t[3]} 映射为 {t[4]}，应为 {expect}")
        # 先行方法
        if rows[0][4] != leading:
            errors.append(f"{pid} 先行方法 Records={rows[0][4]} != Participants={leading}")
        # 每方法各 3 个区块，method_repeat 为 1,2,3
        for method in (method_a, method_b):
            reps = sorted(t[6] for t in rows if t[4] == method)
            if reps != [1, 2, 3]:
                errors.append(f"{pid} {method} 的 method_repeat={reps}")
        # 物体内先后须为 1,2
        for i in (0, 2, 4):
            if (rows[i][5], rows[i + 1][5]) != (1, 2):
                errors.append(f"{pid} 物体内先后异常 {rows[i][5]},{rows[i+1][5]}")

    # 平衡性核对
    leads = [info[p][4] for p in info]
    ea_labels = ["A" if info[p][2] == "EgoAnchor" else "B" for p in info]
    perms = [tuple(info[p][0]) for p in info]
    print(f"参与者数 = {len(info)}")
    print(f"先行方法：EgoAnchor {leads.count('EgoAnchor')} / One-Euro {leads.count('One-Euro')}")
    print(f"标签映射：A=EgoAnchor {ea_labels.count('A')} / B=EgoAnchor {ea_labels.count('B')}")
    print(f"物体排列种类 = {len(set(perms))}，每种人数 = "
          f"{sorted({p: perms.count(p) for p in set(perms)}.values())}")
    seqs = [info[p][1] for p in info]
    print(f"序列：S1 {seqs.count('S1')} / S2 {seqs.count('S2')}")

    print(f"\n结构错误数 = {len(errors)}")
    for e in errors[:15]:
        print("  " + e)

    # 抽查 P001 的六行
    print("\nP001 六个区块：")
    for t in sorted(per_pid["P001"], key=lambda x: x[0]):
        print(f"  block{t[0]} pos{t[1]} {t[2]:<8} {t[3]} = {t[4]:<10} within={t[5]} rep={t[6]}")
    print(f"P001 Participants: objects={info['P001'][0]} seq={info['P001'][1]} "
          f"A={info['P001'][2]} B={info['P001'][3]} leading={info['P001'][4]}")


if __name__ == "__main__":
    main()
