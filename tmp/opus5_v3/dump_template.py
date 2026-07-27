"""转储实验三正式采集工作簿结构，供 v3 模拟工作簿对齐使用。"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

SRC = Path(
    r"p:\VSCode-Project\EgoAnchor\2026-EgoAnchor\material"
    r"\EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx"
)


def cell_repr(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", "\\n")
    return text


def main() -> int:
    wb = openpyxl.load_workbook(SRC, data_only=False)
    out: dict[str, object] = {}
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            rows.append([cell_repr(c.value) for c in row])
        out[ws.title] = {
            "dims": [ws.max_row, ws.max_column],
            "merged": [str(r) for r in ws.merged_cells.ranges],
            "col_widths": {k: v.width for k, v in ws.column_dimensions.items()},
            "rows": rows,
        }
    dest = Path(__file__).with_name("template_dump.json")
    dest.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    for title, info in out.items():
        print(f"{title}: {info['dims'][0]} rows x {info['dims'][1]} cols, "
              f"{len(info['merged'])} merged")
    return 0


if __name__ == "__main__":
    sys.exit(main())
