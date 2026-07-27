"""独立转储正式采集工作簿结构，只读官方模板，不接触任何已有模拟产物。"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

TPL = Path(
    r"p:\VSCode-Project\EgoAnchor\2026-EgoAnchor\material"
    r"\EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx"
)
OUT = Path(__file__).with_name("template_dump.txt")


def main() -> int:
    wb = openpyxl.load_workbook(TPL, data_only=False)
    lines: list[str] = []
    meta: dict[str, dict] = {}
    for ws in wb.worksheets:
        lines.append("=" * 100)
        lines.append(f"SHEET {ws.title}  dims={ws.dimensions}  max_row={ws.max_row} max_col={ws.max_column}")
        lines.append(f"  merged={[str(r) for r in ws.merged_cells.ranges]}")
        lines.append(f"  col_widths={{ {', '.join(f'{k}:{v.width}' for k, v in ws.column_dimensions.items())} }}")
        lines.append("=" * 100)
        meta[ws.title] = {
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged": [str(r) for r in ws.merged_cells.ranges],
        }
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            cells = []
            for c in row:
                if c.value is None:
                    continue
                v = str(c.value).replace("\n", "\\n")
                if len(v) > 220:
                    v = v[:220] + "...<TRUNC>"
                cells.append(f"{c.coordinate}={v}")
            if cells:
                lines.append(f"r{row[0].row}: " + " | ".join(cells))
        lines.append("")
    OUT.write_text("\n".join(lines), encoding="utf-8")
    Path(__file__).with_name("template_meta.json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")
    print("sheets:", [ws.title for ws in wb.worksheets])
    return 0


if __name__ == "__main__":
    sys.exit(main())
