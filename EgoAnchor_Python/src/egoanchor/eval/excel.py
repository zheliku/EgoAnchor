"""评估分析的多 sheet Excel 导出共享工具。

论文图只呈现结论，Excel 面向作者和审稿人保留完整可复核数据：逐帧原始误差
（可自绘时间图）、按 trial/场景的统计汇总，以及便于横向对比的透视矩阵。写出
使用 openpyxl，并对表头加粗、冻结首行、按内容估算列宽，保证可读性。
"""

from __future__ import annotations

from pathlib import Path
from typing import Mapping

import pandas as pd


_MAX_COLUMN_WIDTH = 48
"""列宽上限，避免长字符串把整表撑开。"""

_MAX_SHEET_ROWS = 1_048_576
"""Excel 单 sheet 行上限；超出的原始帧会被截断并在末列标注。"""


def write_workbook(sheets: Mapping[str, pd.DataFrame], path: str | Path) -> Path:
    """把若干命名表写入一个 ``.xlsx``，套用统一表头样式与列宽。

    传入顺序即 sheet 顺序。空表也会写出仅含表头的 sheet，保持产物结构稳定，
    方便下游脚本按固定 sheet 名读取。
    """

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3B5B7A")
    header_align = Alignment(vertical="center")

    with pd.ExcelWriter(destination, engine="openpyxl") as writer:
        for raw_name, frame in sheets.items():
            sheet_name = _safe_sheet_name(raw_name)
            table = frame.copy()
            truncated = False
            if len(table) > _MAX_SHEET_ROWS - 1:
                table = table.iloc[: _MAX_SHEET_ROWS - 1].copy()
                truncated = True
            table.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            for column_index, column in enumerate(table.columns, start=1):
                cell = worksheet.cell(row=1, column=column_index)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                worksheet.column_dimensions[get_column_letter(column_index)].width = (
                    _estimate_width(table[column], str(column))
                )
            if truncated:
                worksheet.cell(
                    row=1,
                    column=len(table.columns) + 1,
                    value=f"TRUNCATED to {_MAX_SHEET_ROWS - 1} rows",
                )
    return destination


def _safe_sheet_name(name: str) -> str:
    """把表名规范为 Excel 允许的 sheet 名（<=31 字符，去非法字符）。"""

    cleaned = "".join(character for character in name if character not in set("[]:*?/\\"))
    return cleaned[:31] or "sheet"


def _estimate_width(series: pd.Series, header: str) -> float:
    """按表头与样本值估算列宽，限制在合理上限内。"""

    longest = 0
    for value in series.head(200):
        text = str(value)
        if text not in ("nan", "None", "<NA>"):
            longest = max(longest, len(text))
    return float(min(max(longest, len(header)) + 2, _MAX_COLUMN_WIDTH))


def flatten_pose_columns(frame: pd.DataFrame, columns: tuple[str, ...]) -> pd.DataFrame:
    """把 3/4 维 pose 列展开为标量子列，使原始帧表可直接在 Excel 中绘图。"""

    result = frame.copy()
    axis_names = {3: ("x", "y", "z"), 4: ("x", "y", "z", "w")}
    for column in columns:
        if column not in result.columns:
            continue
        sample = result[column].dropna()
        length = len(sample.iloc[0]) if not sample.empty and hasattr(sample.iloc[0], "__len__") else 0
        names = axis_names.get(length)
        if names is None:
            continue
        for axis_index, axis in enumerate(names):
            result[f"{column}_{axis}"] = result[column].map(
                lambda value, axis_index=axis_index: (
                    float(value[axis_index])
                    if isinstance(value, (list, tuple)) and len(value) > axis_index
                    else None
                )
            )
        result = result.drop(columns=[column])
    return result


__all__ = ["flatten_pose_columns", "write_workbook"]
