"""实验一/二本地表格、绘图数据和手工引入用 TeX 片段物化。"""

from __future__ import annotations

import csv
import json
from collections.abc import Callable, Mapping
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from .figures import summarize_risk_coverage
from .metrics import (
    LINEAR_SLERP_VARIANT,
    SMOOTHED_EXTRAPOLATION_VARIANT,
    FULL_VARIANT,
    METHODS,
    NO_STATIC_LOCK,
    PaperResults,
    TEMPORAL_STRATEGY_VARIANTS,
    paired_metric_matrix,
    segment_identity,
)


# 关闭 StaticLock 后平滑外推与 Hermite 插值的片段级配对输出契约。
_STRATEGY_COMPARISON_METRICS = (
    ("static", "centered_p95_mm", "mm"),
    ("static", "frame_increment_p95_mm", "mm"),
    ("static", "centered_rotation_p95_deg", "deg"),
    ("static", "frame_rotation_increment_p95_deg", "deg"),
    ("translation", "effective_lag_ms", "ms"),
    ("translation", "aligned_rmse_mm", "mm"),
    ("translation", "current_time_rmse_mm", "mm"),
    ("translation", "aligned_residual_increment_p95_mm", "mm"),
    ("rotation", "effective_lag_ms", "ms"),
    ("rotation", "aligned_rmse_deg", "deg"),
    ("rotation", "current_time_rmse_deg", "deg"),
    ("rotation", "aligned_residual_increment_p95_deg", "deg"),
    ("occlusion", "translation_p95_mm", "mm"),
    ("occlusion", "translation_max_mm", "mm"),
    ("occlusion", "rotation_p95_deg", "deg"),
    ("occlusion", "rotation_max_deg", "deg"),
    ("occlusion", "catastrophic_gt40", "episode"),
    ("transition", "response_ms", "ms"),
    ("stop", "forward_overshoot_mm", "mm"),
    ("stop", "reverse_return_mm", "mm"),
    ("stop", "settling_time_ms", "ms"),
    ("correction", "position_step_p95_mm", "mm"),
    ("correction", "rotation_step_p95_deg", "deg"),
)

# 表格短名称与图例保持一致，但留在发布层，避免显示文本使指标缓存失效。
_METHOD_LABELS = {
    "Arrival-Hold": "Arrival",
    "Capture-Hold": "Capture",
    "One-Euro Anchor": "One-Euro",
    FULL_VARIANT: "EgoAnchor",
}


def _fmt(value: float) -> str:
    """按论文表格精度固定格式化有限数值。"""

    if not np.isfinite(value):
        return "--"
    return f"{value:.2f}"


def _fmt_ms(value: float) -> str:
    """按毫秒量级的实际分辨率格式化时延。

    有效时延在 5~ms 的时延网格上取极小值，起动转换由帧间隔决定，两者均无
    0.01~ms 级的分辨率；固定一位小数既保留配对中位数可能出现的半格，又不
    暗示不存在的精度。
    """

    if not np.isfinite(value):
        return "--"
    return f"{value:.1f}"


def _summary(rows: tuple[Mapping[str, Any], ...], key: str) -> tuple[float, float, float]:
    """返回片段值的 median、Q1 和 Q3。"""

    values = np.asarray([float(row[key]) for row in rows], dtype=float)
    values = values[np.isfinite(values)]
    if values.size == 0:
        raise ValueError(f"论文表缺少指标：{key}")
    return tuple(float(item) for item in np.quantile(values, (0.5, 0.25, 0.75)))  # type: ignore[return-value]


def _cell(summary: tuple[float, float, float]) -> str:
    """写出 ``median [Q1, Q3]`` 读者表格单元格。"""

    median, q1, q3 = summary
    return f"{_fmt(median)} [{_fmt(q1)}, {_fmt(q3)}]"


def _best_median_cells(
    rows: Mapping[str, tuple[Mapping[str, Any], ...]],
    key: str,
    formatter: Callable[[float], str] = _fmt,
) -> Mapping[str, str]:
    """生成只含中位数的四方法单元格，并加粗较低的最优值。"""

    medians = {
        method: _summary(rows[method], key)[0]
        for method in METHODS
    }
    best = min(medians.values())
    return {
        method: (
            rf"\textbf{{{formatter(value)}}}"
            if np.isclose(value, best)
            else formatter(value)
        )
        for method, value in medians.items()
    }


def _paired_channel_cells(
    channels: tuple[tuple[Mapping[str, tuple[Mapping[str, Any], ...]], str], ...],
) -> Mapping[str, str]:
    """把平移/旋转两个通道并入同一单元格，各通道独立加粗最优中位数。"""

    per_channel = [_best_median_cells(rows, key) for rows, key in channels]
    return {
        method: " / ".join(channel[method] for channel in per_channel)
        for method in METHODS
    }


_CHANNEL_SEPARATOR = "/"
"""单元格内平移与旋转两个通道之间的分隔。

裸斜杠而非窄空包围的斜杠：双通道单元格是三张表里最宽的成分，去掉两侧窄空
使承载指标最多的动态跟随表净宽从 229.50~pt 降到 220.49~pt，列距因而不必压
到视觉上过挤的档位。
"""


def _channel_cells(
    translation: tuple[Mapping[str, tuple[Mapping[str, Any], ...]], str],
    rotation: tuple[Mapping[str, tuple[Mapping[str, Any], ...]], str],
    formatter: Callable[[float], str] = _fmt,
) -> Mapping[str, str]:
    """把平移、旋转两个通道并入一个单元格，两通道各自独立加粗最优中位数。

    单位统一写在表头，单元格只留裸数值；通道并入单元格后每个方法只占一行，
    使按评价方面拆分的三张表都能在单栏宽度内保持模板字号与列距。
    """

    left = _best_median_cells(*translation, formatter)
    right = _best_median_cells(*rotation, formatter)
    return {
        method: f"{left[method]}{_CHANNEL_SEPARATOR}{right[method]}"
        for method in METHODS
    }


def _method_rows(columns: tuple[Mapping[str, str], ...]) -> list[str]:
    """每个方法排成一行，各列按给定顺序取该方法的单元格。"""

    return [
        f"{_METHOD_LABELS[method]} & "
        + " & ".join(column[method] for column in columns)
        + r" \\"
        for method in METHODS
    ]


def _metric_header(name: str, unit: str) -> str:
    r"""排出「指标名 $\downarrow$ / 单位」两行表头。

    箭头紧随指标名，标示整列的优劣方向；统计量写在题注而非表头，使按评价
    方面拆分后的各表表头压到两行。
    """

    return r"\shortstack{" + rf"{name}\,$\downarrow$" + r"\\" + unit + "}"


_CHANNEL_UNIT = r"mm/$^\circ$"
"""双通道指标的单位，与单元格内的平移/旋转书写顺序一致。"""


_DEFAULT_TABCOLSEP_PT = 6.0
"""``vgtc`` 模板的 ``\\tabcolsep`` 默认值，仅作为是否需要收紧列距的判据。"""


def _behavior_table(
    caption: str,
    label: str,
    columns: tuple[tuple[str, str, Mapping[str, str]], ...],
    tabcolsep_pt: float = _DEFAULT_TABCOLSEP_PT,
) -> str:
    """排出一张单栏的实验一评价方面表。

    每列给出指标名、单位与四个方法的单元格；方法占一行，通道并入单元格。表格
    按自然宽度排版（``tabular`` 而非 ``tabular*``），不横向撑满单栏，也不改动
    模板字号；宽度只由 ``tabcolsep_pt`` 调节，且仅在需要收紧时才写出
    ``\\setlength{\\tabcolsep}``——该赋值局限在 ``table`` 环境内，不外溢。

    实测（``\\columnwidth`` 为 240.94~pt）：静态配准三指标在模板列距下为
    203.49~pt，转换响应两指标为 150.49~pt，均无需干预；动态跟随四指标为
    277.49~pt，超出 36.55~pt，收到 2~pt 列距后为 236.49~pt。瓶颈是数值而非表头，
    把单位从表头移入题注对宽度无影响（同为 220.49~pt），故单位保留在表头。
    """

    lines = [
        r"\begin{table}[t]",
        r"\centering",
        rf"\caption{{{caption}}}",
        rf"\label{{{label}}}",
        *(
            [rf"\setlength{{\tabcolsep}}{{{tabcolsep_pt:g}pt}}"]
            if tabcolsep_pt != _DEFAULT_TABCOLSEP_PT
            else []
        ),
        r"\begin{tabular}{@{}l" + "c" * len(columns) + r"@{}}",
        r"\toprule",
        "方法 & "
        + " & ".join(_metric_header(name, unit) for name, unit, _ in columns)
        + r" \\",
        r"\midrule",
    ]
    lines.extend(_method_rows(tuple(cells for _, _, cells in columns)))
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{table}"])
    return "\n".join(lines) + "\n"


def build_exp1_static_table(results: PaperResults) -> str:
    """生成实验一静态配准表：头动泄漏、绝对配准与静止抖动。"""

    static = results.static_segments
    return _behavior_table(
        r"实验一的静态配准。各单元格按平移/旋转顺序给出重复测量的中位数，"
        r"三项指标均取P95；箭头标记优劣方向，粗体为各指标在该通道上的最优"
        r"中位数。分布见图~\ref{fig:exp1-final}(a,\,b)。",
        "tab:exp1-static",
        (
            (
                "头动泄漏",
                _CHANNEL_UNIT,
                _channel_cells(
                    (static, "centered_p95_mm"),
                    (static, "centered_rotation_p95_deg"),
                ),
            ),
            (
                "绝对配准",
                _CHANNEL_UNIT,
                _channel_cells(
                    (static, "absolute_p95_mm"),
                    (static, "absolute_rotation_p95_deg"),
                ),
            ),
            (
                "静止抖动",
                _CHANNEL_UNIT,
                _channel_cells(
                    (static, "frame_increment_p95_mm"),
                    (static, "frame_rotation_increment_p95_deg"),
                ),
            ),
        ),
    )


def build_exp1_dynamic_table(results: PaperResults) -> str:
    """生成实验一动态跟随表：有效时延、两个RMSE与残余抖动。"""

    translation = results.translation_segments
    rotation = results.rotation_segments
    return _behavior_table(
        r"实验一的动态跟随。各单元格按平移/旋转顺序给出重复测量的中位数；"
        r"LA-RMSE为时延对齐均方根误差，CT-RMSE为当前时刻均方根误差，"
        r"残余抖动取P95。箭头标记优劣方向，粗体为各指标在该通道上的最优"
        r"中位数。分布见图~\ref{fig:exp1-final}(c,\,d)。",
        "tab:exp1-dynamic",
        (
            (
                "有效时延",
                "ms",
                _channel_cells(
                    (translation, "effective_lag_ms"),
                    (rotation, "effective_lag_ms"),
                    _fmt_ms,
                ),
            ),
            (
                "LA-RMSE",
                _CHANNEL_UNIT,
                _channel_cells(
                    (translation, "aligned_rmse_mm"),
                    (rotation, "aligned_rmse_deg"),
                ),
            ),
            (
                "CT-RMSE",
                _CHANNEL_UNIT,
                _channel_cells(
                    (translation, "current_time_rmse_mm"),
                    (rotation, "current_time_rmse_deg"),
                ),
            ),
            (
                "残余抖动",
                _CHANNEL_UNIT,
                _channel_cells(
                    (translation, "aligned_residual_increment_p95_mm"),
                    (rotation, "aligned_residual_increment_p95_deg"),
                ),
            ),
        ),
        # 四指标×两通道是三张表中最宽的一张，模板列距下超出单栏 36.55~pt；
        # 2~pt 列距使其为 236.49~pt，留 4.45~pt 余量以容纳数据更新带来的位数变化。
        tabcolsep_pt=2.0,
    )


def build_exp1_transition_table(results: PaperResults) -> str:
    """生成实验一转换响应表：遮挡误差与起动转换。"""

    occlusion = results.occlusion_episodes
    return _behavior_table(
        r"实验一的转换响应。遮挡误差按平移/旋转顺序给出重复测量的中位数并取"
        r"P95，起动转换为跨通道标量。箭头标记优劣方向，粗体为最优中位数。",
        "tab:exp1-transition",
        (
            (
                "遮挡误差",
                _CHANNEL_UNIT,
                _channel_cells(
                    (occlusion, "translation_p95_mm"),
                    (occlusion, "rotation_p95_deg"),
                ),
            ),
            (
                "起动转换",
                "ms",
                _best_median_cells(
                    results.transition_segments, "response_ms", _fmt_ms
                ),
            ),
        ),
    )



def _paired_values(
    rows: Mapping[str, tuple[Mapping[str, Any], ...]],
    alternative_variant: str,
    key: str,
    reference_variant: str = FULL_VARIANT,
) -> tuple[np.ndarray, np.ndarray]:
    """返回按片段身份严格配对的参照值和对照值。"""

    matrix = paired_metric_matrix(
        rows,
        (reference_variant, alternative_variant),
        (key,),
    )[:, :, 0]
    return matrix[:, 0], matrix[:, 1]


def _array_cell(values: np.ndarray) -> str:
    """把有限数组写成 median [Q1, Q3] 单元格。"""

    finite = values[np.isfinite(values)]
    if finite.size == 0:
        raise ValueError("论文表缺少有限配对指标")
    median, q1, q3 = (float(item) for item in np.quantile(finite, (0.5, 0.25, 0.75)))
    return _cell((median, q1, q3))


def _effect_summary(reference: np.ndarray, alternative: np.ndarray) -> tuple[str, str]:
    """按配对样本汇总对照/参照倍率及对照更差的一致性。"""

    valid = np.isfinite(reference) & np.isfinite(alternative) & (reference > 0.0)
    if not bool(valid.any()):
        raise ValueError("论文表效应倍率缺少正的有限参照值")
    reference = reference[valid]
    alternative = alternative[valid]
    ratios = alternative / reference
    return (
        f"{_array_cell(ratios)}$\\times$",
        f"{int(np.sum(alternative > reference))}/{len(reference)}",
    )


def build_exp2_attribution_table(results: PaperResults) -> str:
    """生成实验二四项设计归因与配对时序策略表。"""

    capture = np.asarray([float(row["capture_p95_mm"]) for row in results.capture_alignment])
    arrival = np.asarray([float(row["arrival_p95_mm"]) for row in results.capture_alignment])
    if capture.size == 0 or capture.size != arrival.size:
        raise ValueError("采集时刻对齐表缺少完整配对片段")
    capture_effect, capture_consistency = _effect_summary(capture, arrival)

    full_static, disabled_static = _paired_values(
        results.static_segments,
        NO_STATIC_LOCK,
        "centered_p95_mm",
    )
    static_effect, static_consistency = _effect_summary(full_static, disabled_static)

    linear_translation, extrapolation_translation = _paired_values(
        results.translation_segments,
        SMOOTHED_EXTRAPOLATION_VARIANT,
        "aligned_rmse_mm",
        LINEAR_SLERP_VARIANT,
    )
    linear_rotation, extrapolation_rotation = _paired_values(
        results.rotation_segments,
        SMOOTHED_EXTRAPOLATION_VARIANT,
        "aligned_rmse_deg",
        LINEAR_SLERP_VARIANT,
    )
    translation_effect, translation_consistency = _effect_summary(
        linear_translation,
        extrapolation_translation,
    )
    rotation_effect, rotation_consistency = _effect_summary(
        linear_rotation,
        extrapolation_rotation,
    )

    vcd_aurc = np.asarray([float(row["aurc_mm"]) for row in results.vcd_aurc_segments])
    vcd_full_risk = np.asarray(
        [float(row["full_coverage_risk_mm"]) for row in results.vcd_aurc_segments]
    )
    vcd_effect, vcd_consistency = _effect_summary(vcd_aurc, vcd_full_risk)

    lines = [
        r"\begin{table*}[t]",
        r"\centering",
        r"\caption{实验二的设计归因。效应倍率先在同一次重复测量或同一遮挡过程内计算组件关闭与启用之比，再报告 median [Q1, Q3]；所有指标越低越好，倍率大于 1 表示启用设计更优。VCD 行的启用与关闭分别表示分数排序的 AURC 与忽略排序的全覆盖风险，并非冻结阈值接纳消融；时序策略行比较均关闭 StaticLock 的历史轨迹取值（Linear/SLERP）与平滑卡尔曼外推。}",
        r"\label{tab:exp2-final}",
        r"\small",
        r"\setlength{\tabcolsep}{3.0pt}",
        r"\resizebox{\textwidth}{!}{%",
        r"\begin{tabular}{llllll}",
        r"\toprule",
        r"受评设计 & 参照指标 & 启用 & 关闭 & 效应倍率 & 一致性 \\",
        r"\midrule",
        f"采集时刻对齐 & 位姿泄漏 P95 & 采集时刻 {_array_cell(capture)}~mm & 到达时刻 {_array_cell(arrival)}~mm & {capture_effect} & {capture_consistency} " + r"\\",
        f"StaticLock & 静止平移波动 P95 & {_array_cell(full_static)}~mm & {_array_cell(disabled_static)}~mm & {static_effect} & {static_consistency} " + r"\\",
        f"VCD 判别性 & 遮挡过程 AURC & {_array_cell(vcd_aurc)}~mm & 全覆盖 {_array_cell(vcd_full_risk)}~mm & {vcd_effect} & {vcd_consistency} " + r"\\",
        f"时序策略 & 平移 / 旋转 aligned RMSE & \\shortstack{{历史取值 (StaticLock off)\\\\{_array_cell(linear_translation)}~mm / {_array_cell(linear_rotation)}$^\\circ$}} & \\shortstack{{平滑外推 (StaticLock off)\\\\{_array_cell(extrapolation_translation)}~mm / {_array_cell(extrapolation_rotation)}$^\\circ$}} & \\shortstack{{T: {translation_effect}\\\\R: {rotation_effect}}} & \\shortstack{{T: {translation_consistency}\\\\R: {rotation_consistency}}} " + r"\\",
        r"\bottomrule",
        r"\end{tabular}%",
        r"}",
        r"\end{table*}",
    ]
    return "\n".join(lines) + "\n"


def _write_strategy_candidate_data(
    results: PaperResults,
    data_root: Path,
) -> tuple[Path, Path]:
    """写出关闭 StaticLock 后外推、Linear/SLERP 与 Hermite 的配对数据和汇总。"""

    rows_by_family = {
        "static": results.static_segments,
        "translation": results.translation_segments,
        "rotation": results.rotation_segments,
        "occlusion": results.occlusion_episodes,
        "transition": results.transition_segments,
        "stop": results.stop_segments,
        "correction": results.correction_segments,
    }
    metrics_path = data_root / "strategy_comparison_segments.csv"
    summary_path = data_root / "strategy_comparison_summary.csv"
    metric_fields = (
        "family",
        "metric",
        "unit",
        "session_id",
        "trial_id",
        "segment_id",
        "smoothed_kf_extrapolation",
        "linear_slerp_interpolation",
        "hermite_interpolation",
        "extrapolation_minus_linear",
        "hermite_minus_linear",
    )
    summary_fields = (
        "family",
        "metric",
        "unit",
        "n",
        "smoothed_kf_extrapolation_median",
        "smoothed_kf_extrapolation_q1",
        "smoothed_kf_extrapolation_q3",
        "linear_slerp_interpolation_median",
        "linear_slerp_interpolation_q1",
        "linear_slerp_interpolation_q3",
        "hermite_interpolation_median",
        "hermite_interpolation_q1",
        "hermite_interpolation_q3",
        "extrapolation_minus_linear_median",
        "extrapolation_minus_linear_q1",
        "extrapolation_minus_linear_q3",
        "hermite_minus_linear_median",
        "hermite_minus_linear_q1",
        "hermite_minus_linear_q3",
    )
    summaries: list[dict[str, Any]] = []
    with metrics_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=metric_fields)
        writer.writeheader()
        for family, metric, unit in _STRATEGY_COMPARISON_METRICS:
            rows = rows_by_family[family]
            matrix = paired_metric_matrix(
                rows,
                TEMPORAL_STRATEGY_VARIANTS,
                (metric,),
            )[:, :, 0]
            identities = sorted(
                segment_identity(row)
                for row in rows[SMOOTHED_EXTRAPOLATION_VARIANT]
                if np.isfinite(float(row[metric]))
            )
            if len(identities) != matrix.shape[0]:
                raise ValueError(f"时序策略身份与配对矩阵不一致：{family}/{metric}")
            extrapolation_minus_linear = matrix[:, 0] - matrix[:, 1]
            hermite_minus_linear = matrix[:, 2] - matrix[:, 1]
            for identity, values, extrapolation_delta, hermite_delta in zip(
                identities,
                matrix,
                extrapolation_minus_linear,
                hermite_minus_linear,
                strict=True,
            ):
                writer.writerow(
                    {
                        "family": family,
                        "metric": metric,
                        "unit": unit,
                        "session_id": identity[0],
                        "trial_id": identity[1],
                        "segment_id": identity[2],
                        "smoothed_kf_extrapolation": values[0],
                        "linear_slerp_interpolation": values[1],
                        "hermite_interpolation": values[2],
                        "extrapolation_minus_linear": extrapolation_delta,
                        "hermite_minus_linear": hermite_delta,
                    }
                )
            extrapolation_quantiles = np.quantile(matrix[:, 0], (0.5, 0.25, 0.75))
            linear_quantiles = np.quantile(matrix[:, 1], (0.5, 0.25, 0.75))
            hermite_quantiles = np.quantile(matrix[:, 2], (0.5, 0.25, 0.75))
            extrapolation_delta_quantiles = np.quantile(extrapolation_minus_linear, (0.5, 0.25, 0.75))
            hermite_delta_quantiles = np.quantile(hermite_minus_linear, (0.5, 0.25, 0.75))
            summaries.append(
                {
                    "family": family,
                    "metric": metric,
                    "unit": unit,
                    "n": matrix.shape[0],
                    "smoothed_kf_extrapolation_median": extrapolation_quantiles[0],
                    "smoothed_kf_extrapolation_q1": extrapolation_quantiles[1],
                    "smoothed_kf_extrapolation_q3": extrapolation_quantiles[2],
                    "linear_slerp_interpolation_median": linear_quantiles[0],
                    "linear_slerp_interpolation_q1": linear_quantiles[1],
                    "linear_slerp_interpolation_q3": linear_quantiles[2],
                    "hermite_interpolation_median": hermite_quantiles[0],
                    "hermite_interpolation_q1": hermite_quantiles[1],
                    "hermite_interpolation_q3": hermite_quantiles[2],
                    "extrapolation_minus_linear_median": extrapolation_delta_quantiles[0],
                    "extrapolation_minus_linear_q1": extrapolation_delta_quantiles[1],
                    "extrapolation_minus_linear_q3": extrapolation_delta_quantiles[2],
                    "hermite_minus_linear_median": hermite_delta_quantiles[0],
                    "hermite_minus_linear_q1": hermite_delta_quantiles[1],
                    "hermite_minus_linear_q3": hermite_delta_quantiles[2],
                }
            )
    with summary_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=summary_fields)
        writer.writeheader()
        writer.writerows(summaries)
    return metrics_path, summary_path


def _write_plot_sheet(
    worksheet: Any,
    fields: tuple[str, ...],
    rows: list[dict[str, Any]],
) -> None:
    """写入一张可筛选、可直接核对的绘图长表。"""

    worksheet.append(fields)
    for row in rows:
        worksheet.append(tuple(row.get(field) for field in fields))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_index, field in enumerate(fields, start=1):
        values = [field, *(row.get(field, "") for row in rows)]
        width = min(34, max(11, max(len(str(value)) for value in values) + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
        if field in {"x_value", "y_value"}:
            for cell in worksheet.iter_cols(
                min_col=column_index,
                max_col=column_index,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = "0.000000"


def _workbooks_equivalent(first: Path, second: Path) -> bool:
    """比较两本绘图工作簿的可见数据与版式契约。"""

    left = load_workbook(first, read_only=False, data_only=False)
    right = load_workbook(second, read_only=False, data_only=False)
    try:
        if left.sheetnames != right.sheetnames:
            return False
        for sheet_name in left.sheetnames:
            left_sheet = left[sheet_name]
            right_sheet = right[sheet_name]
            if (
                left_sheet.max_row != right_sheet.max_row
                or left_sheet.max_column != right_sheet.max_column
                or left_sheet.freeze_panes != right_sheet.freeze_panes
                or left_sheet.auto_filter.ref != right_sheet.auto_filter.ref
            ):
                return False
            for row in range(1, left_sheet.max_row + 1):
                for column in range(1, left_sheet.max_column + 1):
                    left_cell = left_sheet.cell(row=row, column=column)
                    right_cell = right_sheet.cell(row=row, column=column)
                    if (
                        left_cell.value != right_cell.value
                        or left_cell.data_type != right_cell.data_type
                        or left_cell.number_format != right_cell.number_format
                        or left_cell.style_id != right_cell.style_id
                    ):
                        return False
            for column in range(1, left_sheet.max_column + 1):
                letter = get_column_letter(column)
                if left_sheet.column_dimensions[letter].width != right_sheet.column_dimensions[letter].width:
                    return False
        return True
    finally:
        left.close()
        right.close()


def _publish_plot_workbook(temporary: Path, destination: Path) -> None:
    """原子发布绘图工作簿；Windows 占用时仅复用完全等价的正式文件。"""

    try:
        temporary.replace(destination)
    except PermissionError:
        if not destination.exists() or not _workbooks_equivalent(temporary, destination):
            raise
        temporary.unlink()


def _write_figure_source_data(
    results: PaperResults,
    plot_root: Path,
) -> Path:
    """把图二和图三的全部可见数据点写入专用 XLSX。"""

    fields = (
        "figure",
        "panel",
        "series",
        "variant_id",
        "session_id",
        "trial_id",
        "segment_id",
        "x_metric",
        "x_value",
        "y_metric",
        "y_value",
    )

    def append_metric_rows(
        destination: list[dict[str, Any]],
        *,
        figure: str,
        panel: str,
        rows: Mapping[str, tuple[Mapping[str, Any], ...]],
        variants: tuple[str, ...],
        x_key: str | None,
        y_key: str,
        metric_role: str | None = None,
    ) -> None:
        """把一个面板的各配置片段追加到统一长表。"""

        for index, variant in enumerate(variants):
            for row in sorted(rows[variant], key=segment_identity):
                y_value = float(row[y_key])
                x_value = float(row[x_key]) if x_key is not None else float(index)
                if not np.isfinite((x_value, y_value)).all():
                    continue
                identity = segment_identity(row)
                destination.append(
                    {
                        "figure": figure,
                        "panel": panel,
                        "series": f"{variant}: {metric_role}" if metric_role else variant,
                        "variant_id": variant,
                        "session_id": identity[0],
                        "trial_id": identity[1],
                        "segment_id": identity[2],
                        "x_metric": x_key or "category_index",
                        "x_value": x_value,
                        "y_metric": y_key,
                        "y_value": y_value,
                    }
                )

    figure2_rows: list[dict[str, Any]] = []
    append_metric_rows(
        figure2_rows,
        figure="Figure 2",
        panel="(a) Static translation",
        rows=results.static_segments,
        variants=METHODS,
        x_key=None,
        y_key="centered_p95_mm",
        metric_role="Centered P95",
    )
    append_metric_rows(
        figure2_rows,
        figure="Figure 2",
        panel="(a) Static translation",
        rows=results.static_segments,
        variants=METHODS,
        x_key=None,
        y_key="frame_increment_p95_mm",
        metric_role="Frame-increment P95",
    )
    append_metric_rows(
        figure2_rows,
        figure="Figure 2",
        panel="(b) Static rotation",
        rows=results.static_segments,
        variants=METHODS,
        x_key=None,
        y_key="centered_rotation_p95_deg",
        metric_role="Centered P95",
    )
    append_metric_rows(
        figure2_rows,
        figure="Figure 2",
        panel="(b) Static rotation",
        rows=results.static_segments,
        variants=METHODS,
        x_key=None,
        y_key="frame_rotation_increment_p95_deg",
        metric_role="Frame-increment P95",
    )
    append_metric_rows(
        figure2_rows,
        figure="Figure 2",
        panel="(c) Dynamic translation",
        rows=results.translation_segments,
        variants=METHODS,
        x_key=None,
        y_key="aligned_rmse_mm",
        metric_role="Lag-aligned RMSE",
    )
    append_metric_rows(
        figure2_rows,
        figure="Figure 2",
        panel="(c) Dynamic translation",
        rows=results.translation_segments,
        variants=METHODS,
        x_key=None,
        y_key="aligned_residual_increment_p95_mm",
        metric_role="Residual jitter P95",
    )
    append_metric_rows(
        figure2_rows,
        figure="Figure 2",
        panel="(d) Dynamic rotation",
        rows=results.rotation_segments,
        variants=METHODS,
        x_key=None,
        y_key="aligned_rmse_deg",
        metric_role="Lag-aligned RMSE",
    )
    append_metric_rows(
        figure2_rows,
        figure="Figure 2",
        panel="(d) Dynamic rotation",
        rows=results.rotation_segments,
        variants=METHODS,
        x_key=None,
        y_key="aligned_residual_increment_p95_deg",
        metric_role="Residual jitter P95",
    )

    figure3_rows: list[dict[str, Any]] = []
    for row in sorted(results.capture_alignment, key=segment_identity):
        identity = segment_identity(row)
        for index, (series, key) in enumerate(
            (("Capture time", "capture_p95_mm"), ("Arrival time", "arrival_p95_mm"))
        ):
            figure3_rows.append(
                {
                    "figure": "Figure 3",
                    "panel": "(a) Capture-time alignment",
                    "series": series,
                    "variant_id": "",
                    "session_id": identity[0],
                    "trial_id": identity[1],
                    "segment_id": identity[2],
                    "x_metric": "condition_index",
                    "x_value": index,
                    "y_metric": key,
                    "y_value": float(row[key]),
                }
            )
    append_metric_rows(
        figure3_rows,
        figure="Figure 3",
        panel="(b) StaticLock",
        rows=results.static_segments,
        variants=(FULL_VARIANT, NO_STATIC_LOCK),
        x_key=None,
        y_key="centered_p95_mm",
    )
    for row in sorted(
        results.vcd_risk_coverage,
        key=lambda item: (*segment_identity(item), float(item["coverage"])),
    ):
        identity = segment_identity(row)
        figure3_rows.append(
            {
                "figure": "Figure 3",
                "panel": "(c) VCD score risk-coverage",
                "series": "Event curve",
                "variant_id": FULL_VARIANT,
                "session_id": identity[0],
                "trial_id": identity[1],
                "segment_id": identity[2],
                "x_metric": "coverage",
                "x_value": float(row["coverage"]),
                "y_metric": "selective_risk_mm",
                "y_value": float(row["selective_risk_mm"]),
            }
        )
    for row in summarize_risk_coverage(results.vcd_risk_coverage):
        for series, key in (
            ("Median", "selective_risk_median_mm"),
            ("IQR lower", "selective_risk_q1_mm"),
            ("IQR upper", "selective_risk_q3_mm"),
        ):
            figure3_rows.append(
                {
                    "figure": "Figure 3",
                    "panel": "(c) VCD score risk-coverage",
                    "series": series,
                    "variant_id": FULL_VARIANT,
                    "session_id": "",
                    "trial_id": "",
                    "segment_id": "",
                    "x_metric": "coverage",
                    "x_value": float(row["coverage"]),
                    "y_metric": key,
                    "y_value": float(row[key]),
                }
            )
    append_metric_rows(
        figure3_rows,
        figure="Figure 3",
        panel="(d) Runtime temporal strategies",
        rows=results.translation_segments,
        variants=TEMPORAL_STRATEGY_VARIANTS,
        x_key="effective_lag_ms",
        y_key="aligned_rmse_mm",
    )

    plot_root.mkdir(parents=True, exist_ok=True)
    destination = plot_root / "figure_plot_data.xlsx"
    temporary = plot_root / "figure_plot_data.tmp.xlsx"
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    _write_plot_sheet(
        readme,
        ("项目", "说明"),
        [
            {"项目": "用途", "说明": "图 2 和图 3 的逐点绘图数据；每行是一条实际显示的片段/episode 记录。"},
            {"项目": "数据来源", "说明": "五本只读 Stage 1 工作簿，由论文分析重新计算，不回读 raw JSONL。"},
            {"项目": "配对语义", "说明": "session_id、trial_id、segment_id 相同的记录属于严格配对。"},
            {"项目": "图 2", "说明": "四个面板均在同一线性纵轴上记录主指标与残差抖动；动态抖动是 lag 对齐残差的帧间增量 P95。"},
            {"项目": "图 3(c)", "说明": "event 曲线不拆分同分候选；固定 coverage 汇总取第一个不小于目标 coverage 的完整同分组。"},
            {"项目": "数值精度", "说明": "XLSX 保留计算得到的浮点值；论文表格另行格式化。"},
        ],
    )
    figure2 = workbook.create_sheet("Figure2")
    figure3 = workbook.create_sheet("Figure3")
    _write_plot_sheet(figure2, fields, figure2_rows)
    _write_plot_sheet(figure3, fields, figure3_rows)
    if temporary.exists():
        temporary.unlink()
    workbook.save(temporary)
    _publish_plot_workbook(temporary, destination)

    verification = load_workbook(destination, read_only=True, data_only=False)
    try:
        if verification.sheetnames != ["README", "Figure2", "Figure3"]:
            raise ValueError("绘图工作簿 sheet 集合不正确")
        if verification["Figure2"].max_row != len(figure2_rows) + 1:
            raise ValueError("绘图工作簿 Figure2 行数不正确")
        if verification["Figure3"].max_row != len(figure3_rows) + 1:
            raise ValueError("绘图工作簿 Figure3 行数不正确")
    finally:
        verification.close()
    return destination


def write_analysis_artifacts(
    results: PaperResults,
    output_root: Path,
) -> Mapping[str, Path]:
    """只在活动批次写出指标、绘图 XLSX 和论文表格。"""

    metrics_root = output_root / "metrics"
    plot_root = output_root / "plots"
    tex_root = output_root / "tex"
    table_root = tex_root / "tables"
    metrics_root.mkdir(parents=True, exist_ok=True)
    table_root.mkdir(parents=True, exist_ok=True)
    summary_path = metrics_root / "experiment1_summary.csv"
    fields = (
        "method",
        "variant_id",
        "head_motion_leakage_p95_mm",
        "absolute_registration_p95_mm",
        "stationary_frame_increment_p95_mm",
        "stationary_rotation_leakage_p95_deg",
        "absolute_rotation_registration_p95_deg",
        "stationary_rotation_frame_increment_p95_deg",
        "translation_lag_ms",
        "translation_aligned_rmse_mm",
        "translation_current_time_rmse_mm",
        "translation_aligned_residual_increment_p95_mm",
        "rotation_lag_ms",
        "rotation_aligned_rmse_deg",
        "rotation_current_time_rmse_deg",
        "rotation_aligned_residual_increment_p95_deg",
        "occlusion_p95_mm",
        "occlusion_max_mm",
        "occlusion_rotation_p95_deg",
        "occlusion_rotation_max_deg",
        "catastrophic_failures_gt40",
        "occlusion_episodes",
        "start_transition_response_ms",
        "stop_forward_overshoot_mm",
        "stop_reverse_return_mm",
        "stop_settling_time_ms",
        "correction_position_step_p95_mm",
        "correction_rotation_step_p95_deg",
    )
    with summary_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for method in METHODS:
            writer.writerow(
                {
                    "method": method,
                    "variant_id": method,
                    "head_motion_leakage_p95_mm": _summary(results.static_segments[method], "centered_p95_mm")[0],
                    "absolute_registration_p95_mm": _summary(results.static_segments[method], "absolute_p95_mm")[0],
                    "stationary_frame_increment_p95_mm": _summary(results.static_segments[method], "frame_increment_p95_mm")[0],
                    "stationary_rotation_leakage_p95_deg": _summary(results.static_segments[method], "centered_rotation_p95_deg")[0],
                    "absolute_rotation_registration_p95_deg": _summary(results.static_segments[method], "absolute_rotation_p95_deg")[0],
                    "stationary_rotation_frame_increment_p95_deg": _summary(results.static_segments[method], "frame_rotation_increment_p95_deg")[0],
                    "translation_lag_ms": _summary(results.translation_segments[method], "effective_lag_ms")[0],
                    "translation_aligned_rmse_mm": _summary(results.translation_segments[method], "aligned_rmse_mm")[0],
                    "translation_current_time_rmse_mm": _summary(results.translation_segments[method], "current_time_rmse_mm")[0],
                    "translation_aligned_residual_increment_p95_mm": _summary(results.translation_segments[method], "aligned_residual_increment_p95_mm")[0],
                    "rotation_lag_ms": _summary(results.rotation_segments[method], "effective_lag_ms")[0],
                    "rotation_aligned_rmse_deg": _summary(results.rotation_segments[method], "aligned_rmse_deg")[0],
                    "rotation_current_time_rmse_deg": _summary(results.rotation_segments[method], "current_time_rmse_deg")[0],
                    "rotation_aligned_residual_increment_p95_deg": _summary(results.rotation_segments[method], "aligned_residual_increment_p95_deg")[0],
                    "occlusion_p95_mm": _summary(results.occlusion_episodes[method], "translation_p95_mm")[0],
                    "occlusion_max_mm": _summary(results.occlusion_episodes[method], "translation_max_mm")[0],
                    "occlusion_rotation_p95_deg": _summary(results.occlusion_episodes[method], "rotation_p95_deg")[0],
                    "occlusion_rotation_max_deg": _summary(results.occlusion_episodes[method], "rotation_max_deg")[0],
                    "catastrophic_failures_gt40": sum(bool(row["catastrophic_gt40"]) for row in results.occlusion_episodes[method]),
                    "occlusion_episodes": len(results.occlusion_episodes[method]),
                    "start_transition_response_ms": _summary(results.transition_segments[method], "response_ms")[0],
                    "stop_forward_overshoot_mm": _summary(results.stop_segments[method], "forward_overshoot_mm")[0],
                    "stop_reverse_return_mm": _summary(results.stop_segments[method], "reverse_return_mm")[0],
                    "stop_settling_time_ms": _summary(results.stop_segments[method], "settling_time_ms")[0],
                    "correction_position_step_p95_mm": _summary(results.correction_segments[method], "position_step_p95_mm")[0],
                    "correction_rotation_step_p95_deg": _summary(results.correction_segments[method], "rotation_step_p95_deg")[0],
                }
            )
    capture_path = metrics_root / "capture_alignment.csv"
    with capture_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=("session_id", "trial_id", "segment_id", "capture_p95_mm", "arrival_p95_mm", "paired_reduction_mm", "n_candidates"))
        writer.writeheader()
        writer.writerows(results.capture_alignment)
    vcd_curve_path = metrics_root / "vcd_risk_coverage.csv"
    with vcd_curve_path.open("w", newline="", encoding="utf-8") as handle:
        vcd_curve_fields = (
            "session_id",
            "trial_id",
            "segment_id",
            "score_threshold",
            "score_tie_count",
            "retained_candidates",
            "evaluable_candidates",
            "coverage",
            "selective_risk_mm",
        )
        writer = csv.DictWriter(handle, fieldnames=vcd_curve_fields)
        writer.writeheader()
        writer.writerows(results.vcd_risk_coverage)
    vcd_aurc_path = metrics_root / "vcd_aurc_segments.csv"
    with vcd_aurc_path.open("w", newline="", encoding="utf-8") as handle:
        vcd_aurc_fields = (
            "session_id",
            "trial_id",
            "segment_id",
            "candidate_rows",
            "evaluable_candidates",
            "excluded_candidates",
            "score_levels",
            "full_coverage_risk_mm",
            "aurc_mm",
            "risk_gain_mm",
        )
        writer = csv.DictWriter(handle, fieldnames=vcd_aurc_fields)
        writer.writeheader()
        writer.writerows(results.vcd_aurc_segments)
    performance_path = metrics_root / "runtime_performance.json"
    performance_path.write_text(json.dumps(results.performance, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    strategy_metrics_path, strategy_summary_path = _write_strategy_candidate_data(results, metrics_root)
    plot_data_path = _write_figure_source_data(results, plot_root)
    exp1_tables = {
        "exp1_static_table": (
            "experiment1_static_fidelity.tex",
            build_exp1_static_table(results),
        ),
        "exp1_dynamic_table": (
            "experiment1_dynamic_fidelity.tex",
            build_exp1_dynamic_table(results),
        ),
        "exp1_transition_table": (
            "experiment1_transition_response.tex",
            build_exp1_transition_table(results),
        ),
    }
    exp1_table_paths = {}
    for key, (name, content) in exp1_tables.items():
        path = table_root / name
        path.write_text(content, encoding="utf-8")
        exp1_table_paths[key] = path
    exp2_table = build_exp2_attribution_table(results)
    exp2_table_path = table_root / "experiment2_design_attribution.tex"
    exp2_table_path.write_text(exp2_table, encoding="utf-8")
    legacy_exp1_tables = (
        "experiment1_system_characterization.tex",
        "experiment1_static_occlusion_stability.tex",
        "experiment1_dynamic_6dof_fidelity.tex",
        "experiment1_performance.tex",
        "experiment1_anchor_behavior.tex",
    )
    for legacy_name in legacy_exp1_tables:
        legacy_exp1_table = table_root / legacy_name
        if legacy_exp1_table.exists():
            legacy_exp1_table.unlink()
    legacy_figure_root = tex_root / "figures"
    for legacy_name in ("figure2_experiment1.tex", "figure3_experiment2.tex"):
        legacy_figure = legacy_figure_root / legacy_name
        if legacy_figure.exists():
            legacy_figure.unlink()
    if legacy_figure_root.exists() and not any(legacy_figure_root.iterdir()):
        legacy_figure_root.rmdir()

    return {
        **exp1_table_paths,
        "exp2_table": exp2_table_path,
        "summary": summary_path,
        "capture": capture_path,
        "vcd_risk_coverage": vcd_curve_path,
        "vcd_aurc": vcd_aurc_path,
        "performance": performance_path,
        "plot_data": plot_data_path,
        "strategy_metrics": strategy_metrics_path,
        "strategy_summary": strategy_summary_path,
    }


__all__ = [
    "build_exp1_dynamic_table",
    "build_exp1_static_table",
    "build_exp1_transition_table",
    "build_exp2_attribution_table",
    "write_analysis_artifacts",
]
