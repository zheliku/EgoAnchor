"""实验三六页中文结果工作簿的写入与回读校验。"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any
from uuid import uuid4

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from .contracts import (
    AnalysisTables,
    EGOANCHOR,
    Exp3Data,
    MAIN_FAMILY,
    METHOD_LABELS,
    OBJECTS,
    ONE_EURO,
    PRIMARY_OUTCOMES,
    SCALE_FAMILY,
    SCALE_OUTCOMES,
)
from .settings import AnalysisSettings


INFO_SHEET = "说明"
"""来源、统计口径、解释边界和页面指南所在的首张工作表。"""

SAMPLE_QC_SHEET = "样本与质控"
"""样本流程、参与者概况、安全与设计平衡的合并工作表。"""

RESULTS_SHEET = "主结果"
"""十二项冻结结局的唯一推断结果工作表。"""

OBJECT_RESULTS_SHEET = "分物体描述"
"""七个主条目乘三个对象的纯描述结果工作表。"""

RELIABILITY_SHEET = "量表信度"
"""已发表量表家族五项结局在当前样本中的内部一致性工作表。"""

CHOICES_SHEET = "选择结果"
"""最终偏好、信任选择、强度与区分信心的描述工作表。"""

_SHEET_GUIDE: tuple[tuple[str, str], ...] = (
    (INFO_SHEET, "数据来源、统计口径、解释边界与六页索引"),
    (SAMPLE_QC_SHEET, "样本流、参与者概况、安全与设计平衡"),
    (RESULTS_SHEET, "七个主条目和已发表量表家族五项结局的 Holm 校正结果"),
    (OBJECT_RESULTS_SHEET, "七个主条目在三个对象上的方向检查；只作描述，不作推断"),
    (RELIABILITY_SHEET, "AQ、TiA 与 S-TIAS 按方法计算的当前样本信度"),
    (CHOICES_SHEET, "总体偏好、信任选择、强度、区分信心与选择交叉表"),
)
"""工作表固定顺序及其唯一职责，同时用于说明页和回读验证。"""

_NAVY = "213A50"
_TEAL = "2F6F73"
_PALE_TEAL = "E8F2F1"
_PALE_GREEN = "EAF3E8"
_PALE_RED = "F8E9E7"
_PALE_YELLOW = "FFF4D6"
_HEADER = "EDF1F4"
_BAND = "F8FAFB"
_WHITE = "FFFFFF"
_TEXT = "263643"
_MUTED = "60717D"
_RULE = "B8C3CB"
_FONT = "Microsoft YaHei"
"""结果工作簿使用的克制配色和中文字体。"""

_CORE_PROPERTIES = {
    "creator": "EgoAnchor",
    "lastModifiedBy": "EgoAnchor",
    "title": "EgoAnchor 实验三分析结果",
    "subject": "跨对象主观感知评价的精简统计结果",
    "description": "EgoAnchor 实验三统计分析结果工作簿。",
    "keywords": "EgoAnchor, Experiment 3, Wilcoxon, Holm",
    "category": "experiment-3-analysis",
}
"""结果簿写出和回读共同使用的固定核心属性。"""

_OUTCOME_NAMES = {
    "Q1": "Q1 静止稳定",
    "Q2": "Q2 运动附着",
    "Q3": "Q3 姿态一致",
    "Q4": "Q4 恢复一致",
    "Q5": "Q5 位置正确",
    "Q6": "Q6 依赖意愿",
    "Q7": "Q7 稳定—响应平衡",
    "AQ_EQ": "AQ 嵌入质量",
    "AQ_IQ": "AQ 交互质量",
    "TIA_RC": "TiA 可靠性/能力",
    "TIA_UP": "TiA 理解/可预测性",
    "STIAS": "S-TIAS 信任",
}
"""冻结结局键到中文短名称的映射。"""

_OBJECT_NAMES = {
    "blue_mouse": "蓝色鼠标",
    "stapler": "订书机",
    "gamepad": "游戏手柄",
}
"""正式对象键到中文显示名的映射。"""

_SAMPLE_NAMES = {
    "Preallocated_Slots": "预分配平衡单元",
    "Consented": "签署知情同意",
    "Started": "开始体验",
    "Completed_Session": "完成会话",
    "Included": "纳入分析",
    "Excluded": "排除",
    "Pending_Review": "待人工复核",
    "Age": "年龄（岁）",
    "Gender": "性别",
    "Handedness": "主手",
    "Vision": "视力",
    "VRMR_Experience": "VR/MR 经验",
    "PhysicalMR_Experience": "实物 AR/MR 经验",
    "Baseline_Discomfort": "实验前不适",
    "End_Discomfort": "实验后不适",
    "Discomfort_Change": "不适变化",
    "Exclusion_Reason": "排除原因",
    "Balance_Unit": "24 平衡单元",
    "Object_Order": "物体顺序",
    "Method_Sequence": "对象内方法序列",
    "A_Mapping": "匿名标签映射",
    "First_Method": "先行方法",
}
"""样本摘要内部键到中文显示名的映射。"""

def write_results_workbook(
    destination: Path,
    *,
    data: Exp3Data,
    tables: AnalysisTables,
    settings: AnalysisSettings,
    settings_sha256: str,
    batch_config_path: Path,
    paper_config_path: Path,
    validation: dict[str, Any],
) -> Path:
    """写入固定六页中文结果工作簿并回读关键契约。

    该接口只接收最终汇总表，不接收逐参与者评分。结果工作簿只承担阅读和论文汇报
    职责，不同时充当内部缓存、建模诊断或人工编码工作区。
    """

    output = destination.expanduser().resolve()
    input_digest = _require_sha256(data.source_sha256, "输入 SHA-256")
    settings_digest = _require_sha256(settings_sha256, "参数 SHA-256")
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, value in _CORE_PROPERTIES.items():
        setattr(workbook.properties, name, value)
    workbook.properties.contentStatus = None
    _write_info(
        workbook,
        data,
        settings,
        settings_digest,
        batch_config_path,
        paper_config_path,
        validation,
    )
    _write_sample_quality(workbook, tables.sample)
    _write_main_results(workbook, tables.results, settings.alpha)
    _write_object_descriptions(workbook, tables.objects)
    _write_reliability(workbook, tables.reliability)
    _write_choices(workbook, tables.choices)
    temporary = output.with_name(f".{output.stem}.{uuid4().hex}.tmp.xlsx")
    try:
        try:
            workbook.calculation.calcMode = "auto"
            workbook.calculation.fullCalcOnLoad = True
            workbook.save(temporary)
        finally:
            workbook.close()
        _verify_results_workbook(
            temporary,
            input_sha256=input_digest,
            settings_sha256=settings_digest,
            included_count=len(validation["included_participants"]),
        )
        temporary.replace(output)
    finally:
        temporary.unlink(missing_ok=True)
    return output


def _write_info(
    workbook: Workbook,
    data: Exp3Data,
    settings: AnalysisSettings,
    settings_digest: str,
    batch_config_path: Path,
    paper_config_path: Path,
    validation: dict[str, Any],
) -> None:
    """写入来源、推断口径、诚实边界和六页索引。"""

    worksheet = workbook.create_sheet(INFO_SHEET)
    _prepare_sheet(worksheet, tab_color=_NAVY, portrait=True)
    _write_title(
        worksheet,
        "EgoAnchor 实验三分析结果",
        "六页阅读版 · 所有结果由五张原始数据表独立重算",
        2,
    )
    row = 4
    row = _write_key_value_section(
        worksheet,
        row,
        "数据来源",
        (
            ("输入工作簿", data.source_path),
            ("输入 SHA-256", data.source_sha256),
            ("纳入参与者", validation["included_count"]),
            ("批处理配置", str(batch_config_path)),
            ("论文参数配置", str(paper_config_path)),
            ("参数 SHA-256", settings_digest),
            ("模板版本", settings.template_version),
        ),
    )
    row = _write_key_value_section(
        worksheet,
        row,
        "统计口径",
        (
            (
                "分析单位",
                "区块级结局先按参与者在三个对象上取均值；TiA 与 S-TIAS 使用方法级"
                "单次施测得分；随后计算 EgoAnchor−One-Euro 配对差",
            ),
            (
                "Wilcoxon",
                "配对差先四舍五入至 12 位；删除零差；并列绝对差使用中秩；"
                "按观测中秩枚举双侧条件精确符号分配",
            ),
            (
                "适用前提",
                "零假设下非零配对差的符号必须可交换；通常对应配对差分布关于 0 对称，"
                "“精确”不表示无分布前提",
            ),
            ("多重比较", "七个主条目与已发表量表家族五项结局分别做 Holm 校正"),
            ("效应量", f"匹配秩双列相关 r_rb；自举区间置信水平 {settings.confidence_level:.0%}"),
            ("逐对象结果", "只检查描述方向，不计算逐对象 p 值、星号、Holm 或效应推断"),
            ("显著性阈值", settings.alpha),
        ),
    )
    row = _write_key_value_section(
        worksheet,
        row,
        "解释边界",
        (
            ("信度", "只表示当前样本内部一致性；AQ 的三物体均值单位与单次方法级量表不可互比"),
            ("发布边界", "本工作簿只分析问卷与选择结果，不提供客观任务表现证据"),
        ),
    )
    guide = tuple((name, purpose) for name, purpose in _SHEET_GUIDE)
    _write_key_value_section(worksheet, row, "页面指南", guide)
    worksheet.column_dimensions["A"].width = 23
    worksheet.column_dimensions["B"].width = 100
    worksheet.freeze_panes = "A4"


def _write_sample_quality(
    workbook: Workbook,
    sample: pd.DataFrame,
) -> None:
    """把样本流程、背景、安全与设计平衡写入一张分节质控页。"""

    worksheet = workbook.create_sheet(SAMPLE_QC_SHEET)
    _prepare_sheet(worksheet, tab_color=_TEAL, portrait=False)
    _write_title(
        worksheet,
        "样本与质控",
        "确认纳入、缺失、参与者背景与实验顺序平衡",
        6,
    )
    row = 4
    row = _write_section_table(
        worksheet,
        row,
        "样本流程",
        "人数和比例均显示显式分母；待复核或排除不能被当作缺失参与者静默忽略。",
        _sample_rows(sample, ("Sample_Flow",)),
        _sample_columns(),
    )
    row = _write_section_table(
        worksheet,
        row,
        "参与者概况",
        "连续变量给出均值±SD、中位数[Q1,Q3]与范围；零缺失行不重复展示。",
        _sample_rows(sample, ("Participant_Profile",), omit_zero_missing=True),
        _sample_columns(),
    )
    row = _write_section_table(
        worksheet,
        row,
        "安全与排除",
        "安全分母为实际开始体验者；若无排除记录，本节保留空状态而不伪造零分母比例。",
        _sample_rows(sample, ("Safety", "Exclusion"), omit_zero_missing=True),
        _sample_columns(),
    )
    row = _write_section_table(
        worksheet,
        row,
        "设计平衡",
        "按冻结因子汇总水平数、实际人数范围和最大偏差；详细单元分配保留在原始 Participants 表。",
        _balance_summary(sample),
        (
            ("Factor", "因素", 21, "text"),
            ("Levels", "水平数", 9, "integer"),
            ("Count_Range", "实际人数范围", 16, "text"),
            ("Expected", "实际 N 下期望/水平", 18, "decimal"),
            ("Max_Deviation", "最大绝对偏差", 16, "decimal"),
            ("Status", "状态", 14, "text"),
        ),
    )
    worksheet.freeze_panes = "A7"


def _write_main_results(workbook: Workbook, results: pd.DataFrame, alpha: float) -> None:
    """写入十二项冻结结果，并突出家族校正后的结论。"""

    worksheet = workbook.create_sheet(RESULTS_SHEET)
    _prepare_sheet(worksheet, tab_color="D39D38", portrait=False)
    _write_title(
        worksheet,
        "主结果",
        "唯一推断页 · 确证结论只读取 Holm 校正 p；原始 p、dz 和探索性 AQ 单项不在此重复",
        12,
    )
    display = _main_result_rows(results, alpha)
    columns = (
        ("Family", "检验家族", 16, "text"),
        ("Outcome", "指标", 25, "text"),
        ("N", "配对 N", 8, "integer"),
        ("N_Nonzero", "非零差 N", 10, "integer"),
        ("OneEuro", "One-Euro 中位数 [Q1, Q3]", 25, "text"),
        ("EgoAnchor", "EgoAnchor 中位数 [Q1, Q3]", 25, "text"),
        ("Difference", "配对差中位数 [Q1, Q3]", 25, "text"),
        ("W", "W", 9, "decimal"),
        ("p_Holm", "Holm 校正 p", 14, "pvalue"),
        ("r_rb", "r_rb", 9, "effect"),
        ("Effect_CI", "r_rb 区间", 22, "text"),
        ("Conclusion", "简明结论", 25, "text"),
    )
    header_row, last_row = _write_table(worksheet, 4, display, columns, auto_filter=True)
    p_column = 9
    for row in range(header_row + 1, last_row + 1):
        value = worksheet.cell(row, p_column).value
        if isinstance(value, (int, float)) and value < alpha:
            worksheet.cell(row, p_column).fill = PatternFill("solid", fgColor=_PALE_YELLOW)
            worksheet.cell(row, 12).font = Font(name=_FONT, size=9, bold=True, color=_TEXT)
    worksheet.freeze_panes = f"A{header_row + 1}"


def _write_object_descriptions(workbook: Workbook, objects: pd.DataFrame) -> None:
    """写入七项主条目在三个对象上的配对描述统计。"""

    worksheet = workbook.create_sheet(OBJECT_RESULTS_SHEET)
    _prepare_sheet(worksheet, tab_color="7897B3", portrait=False)
    _write_title(
        worksheet,
        "分物体描述",
        "方向一致性检查 · 21 行均为描述统计，不计算逐物体 p 值、星号或显著性结论",
        7,
    )
    display = _object_rows(objects)
    header_row, _ = _write_table(
        worksheet,
        4,
        display,
        (
            ("Outcome", "指标", 25, "text"),
            ("Object", "物体", 14, "text"),
            ("N", "配对 N", 9, "integer"),
            ("OneEuro", "One-Euro 中位数 [Q1, Q3]", 25, "text"),
            ("EgoAnchor", "EgoAnchor 中位数 [Q1, Q3]", 25, "text"),
            ("Difference", "配对差中位数 [Q1, Q3]", 25, "text"),
            ("Direction", "描述方向", 22, "text"),
        ),
        auto_filter=True,
    )
    worksheet.freeze_panes = f"A{header_row + 1}"


def _write_reliability(workbook: Workbook, reliability: pd.DataFrame) -> None:
    """写入已发表量表家族五项结局按方法分开的当前样本信度。"""

    worksheet = workbook.create_sheet(RELIABILITY_SHEET)
    _prepare_sheet(worksheet, tab_color="7B8F5A", portrait=False)
    display = _reliability_rows(reliability)
    columns: list[tuple[str, str, float, str]] = [
        ("Scale", "量表", 25, "text"),
        ("Method", "方法", 22, "text"),
        ("Unit", "测量单位", 22, "text"),
        ("N", "N", 8, "integer"),
        ("Items", "条目数", 9, "integer"),
        ("Alpha", "Cronbach α", 13, "effect"),
        ("Omega", "McDonald ω", 13, "effect"),
    ]
    spearman_brown = pd.to_numeric(
        display.get("SpearmanBrown", pd.Series(dtype=float)),
        errors="coerce",
    )
    if spearman_brown.notna().any():
        columns.append(("SpearmanBrown", "Spearman–Brown", 17, "effect"))
    columns.append(("Note", "解释", 52, "text"))
    _write_title(
        worksheet,
        "量表信度",
        "当前样本内部一致性 · AQ 为三物体均值分析单位，TiA/S-TIAS 为单次方法级测量，两者不可直接比较",
        len(columns),
    )
    header_row, _ = _write_table(
        worksheet,
        4,
        display,
        tuple(columns),
        auto_filter=True,
    )
    worksheet.freeze_panes = f"A{header_row + 1}"


def _write_choices(workbook: Workbook, choices: pd.DataFrame) -> None:
    """把最终选择长表压缩为分布、评分摘要、一致性和三乘三交叉表。"""

    worksheet = workbook.create_sheet(CHOICES_SHEET)
    _prepare_sheet(worksheet, tab_color="8D7098", portrait=False)
    _write_title(
        worksheet,
        "选择结果",
        "最终问卷的描述统计 · 强制选择、强度和区分信心不进入十二项 Wilcoxon 家族",
        4,
    )
    row = 4
    row = _write_section_table(
        worksheet,
        row,
        "总体偏好与信任选择",
        "匿名标签已按参与者内稳定映射解码为实际方法。",
        _choice_distribution(choices),
        (
            ("Measure", "问题", 20, "text"),
            ("Category", "选择", 24, "text"),
            ("Count", "人数", 10, "integer"),
            ("Proportion", "比例", 12, "percent"),
        ),
    )
    row = _write_section_table(
        worksheet,
        row,
        "强度与区分信心",
        "偏好强度只对做出方法选择者有值；“无明显偏好”按 N/A 处理。",
        _choice_summaries(choices),
        (
            ("Measure", "指标", 20, "text"),
            ("N", "有效 N", 10, "integer"),
            ("MedianIQR", "中位数 [Q1, Q3]", 24, "text"),
            ("MeanSD", "均值 ± SD", 18, "text"),
        ),
    )
    row = _write_section_table(
        worksheet,
        row,
        "偏好—信任不一致",
        "仅描述同一参与者的两次最终选择是否不同。",
        _choice_disagreement(choices),
        (
            ("Metric", "指标", 28, "text"),
            ("Count", "人数", 10, "integer"),
            ("Proportion", "比例", 12, "percent"),
        ),
    )
    _write_section_table(
        worksheet,
        row,
        "总体偏好 × 信任选择",
        "行为为总体偏好，列为更愿意信任的方法；单元格为人数。",
        _choice_cross_table(choices),
        (
            ("Preference", "总体偏好＼信任选择", 26, "text"),
            (EGOANCHOR, METHOD_LABELS[EGOANCHOR], 14, "integer"),
            (ONE_EURO, METHOD_LABELS[ONE_EURO], 18, "integer"),
            ("No_Preference", "无明显偏好", 16, "integer"),
        ),
    )
    worksheet.freeze_panes = "A7"


def _sample_columns() -> tuple[tuple[str, str, float, str], ...]:
    """返回样本分节共用的紧凑列定义。"""

    return (
        ("Variable", "指标", 24, "text"),
        ("Category", "类别", 22, "text"),
        ("N", "人数", 9, "integer"),
        ("Denominator", "分母", 9, "integer"),
        ("Proportion", "比例", 11, "percent"),
        ("Summary", "描述统计", 52, "text"),
    )


def _sample_rows(
    sample: pd.DataFrame,
    sections: tuple[str, ...],
    *,
    omit_zero_missing: bool = False,
) -> pd.DataFrame:
    """把样本长表转成中文显示行，并可省略重复的零缺失行。"""

    selected = sample[sample["Section"].astype(str).isin(sections)]
    rows: list[dict[str, Any]] = []
    for _, source in selected.iterrows():
        category = str(source.get("Category", ""))
        count = _as_number(source.get("N"))
        if omit_zero_missing and category == "Missing" and count == 0:
            continue
        rows.append(
            {
                "Variable": _SAMPLE_NAMES.get(str(source.get("Variable")), str(source.get("Variable"))),
                "Category": _category_name(category),
                "N": count,
                "Denominator": _as_number(source.get("Denominator")),
                "Proportion": _as_number(source.get("Proportion")),
                "Summary": _continuous_summary(source),
            }
        )
    return pd.DataFrame(rows, columns=("Variable", "Category", "N", "Denominator", "Proportion", "Summary"))


def _continuous_summary(row: pd.Series) -> str:
    """把连续样本变量压缩为一段可直接阅读的描述统计。"""

    mean = _as_number(row.get("Mean"))
    if not math.isfinite(mean):
        return ""
    return (
        f"均值±SD {_format_number(mean)}±{_format_number(row.get('SD'))}；"
        f"中位数[Q1,Q3] {_summary_text(row.get('Q1'), row.get('Median'), row.get('Q3'))}；"
        f"范围 {_format_number(row.get('Min'))}–{_format_number(row.get('Max'))}"
    )


def _balance_summary(sample: pd.DataFrame) -> pd.DataFrame:
    """把逐水平设计平衡行压缩为每个冻结因子一行。"""

    balance = sample[sample["Section"].astype(str) == "Design_Balance"]
    rows: list[dict[str, Any]] = []
    for variable, group in balance.groupby("Variable", sort=False):
        counts = pd.to_numeric(group["N"], errors="coerce").dropna()
        expected = pd.to_numeric(group["Expected_At_Actual_N"], errors="coerce").dropna()
        deviations = pd.to_numeric(group["Deviation_From_Actual_Balance"], errors="coerce").abs().dropna()
        statuses = set(group["Status"].dropna().astype(str))
        status = "平衡" if statuses == {"balanced"} else (
            "部分覆盖" if "partial_coverage" in statuses else "需要复核"
        )
        count_range = "" if counts.empty else f"{int(counts.min())}–{int(counts.max())}"
        rows.append(
            {
                "Factor": _SAMPLE_NAMES.get(str(variable), str(variable)),
                "Levels": int(len(group)),
                "Count_Range": count_range,
                "Expected": float(expected.iloc[0]) if len(expected) else math.nan,
                "Max_Deviation": float(deviations.max()) if len(deviations) else math.nan,
                "Status": status,
            }
        )
    return pd.DataFrame(rows)


def _main_result_rows(results: pd.DataFrame, alpha: float) -> pd.DataFrame:
    """从内部统计表提取十二项论文可读结果，不复制原始 p 或 dz。"""

    rows: list[dict[str, Any]] = []
    for _, source in results.iterrows():
        family = str(source.get("Family"))
        if family not in {MAIN_FAMILY, SCALE_FAMILY}:
            continue
        effect = _as_number(source.get("r_rb"))
        adjusted_p = _as_number(source.get("p_Holm"))
        rows.append(
            {
                "Family": "主证实条目" if family == MAIN_FAMILY else "已发表量表",
                "Outcome": _OUTCOME_NAMES.get(str(source.get("Outcome")), str(source.get("Outcome"))),
                "N": source.get("N"),
                "N_Nonzero": source.get("N_Nonzero"),
                "OneEuro": _summary_text(
                    source.get("OneEuro_Q1"), source.get("OneEuro_Median"), source.get("OneEuro_Q3")
                ),
                "EgoAnchor": _summary_text(
                    source.get("EgoAnchor_Q1"), source.get("EgoAnchor_Median"), source.get("EgoAnchor_Q3")
                ),
                "Difference": _summary_text(
                    source.get("Difference_Q1"), source.get("Difference_Median"), source.get("Difference_Q3")
                ),
                "W": source.get("W"),
                "p_Holm": adjusted_p,
                "r_rb": effect,
                "Effect_CI": _effect_interval(source),
                "Conclusion": _result_conclusion(adjusted_p, effect, alpha),
            }
        )
    return pd.DataFrame(rows)


def _effect_interval(row: pd.Series) -> str:
    """按区间状态生成匹配秩双列相关的诚实显示文本。"""

    status = str(row.get("r_rb_CI_Status", ""))
    if status == "degenerate_at_bound":
        return "不报告（方向完全一致）"
    low = _as_number(row.get("r_rb_CI_Low"))
    high = _as_number(row.get("r_rb_CI_High"))
    if not (math.isfinite(low) and math.isfinite(high)):
        return "不可估计"
    return f"[{low:.2f}, {high:.2f}]"


def _result_conclusion(adjusted_p: float, effect: float, alpha: float) -> str:
    """依据家族校正 p 与效应方向生成不越界的简短结论。"""

    if not (math.isfinite(adjusted_p) and adjusted_p < alpha):
        return "校正后未检出差异"
    if effect > 0.0:
        return "EgoAnchor 评分更高"
    if effect < 0.0:
        return "One-Euro 评分更高"
    return "校正后有差异，方向待核查"


def _object_rows(objects: pd.DataFrame) -> pd.DataFrame:
    """把逐对象内部描述字段转换为七列中文阅读表。"""

    direction_names = {
        "EgoAnchor_higher": "EgoAnchor 较高",
        "OneEuro_higher": "One-Euro 较高",
        "median_tie": "配对差中位数为 0",
        "not_available": "数据不足",
    }
    rows: list[dict[str, Any]] = []
    for _, source in objects.iterrows():
        rows.append(
            {
                "Outcome": _OUTCOME_NAMES.get(str(source.get("Outcome")), str(source.get("Outcome"))),
                "Object": _OBJECT_NAMES.get(str(source.get("Object_Key")), str(source.get("Object_Key"))),
                "N": source.get("N"),
                "OneEuro": _summary_text(
                    source.get("OneEuro_Q1"), source.get("OneEuro_Median"), source.get("OneEuro_Q3")
                ),
                "EgoAnchor": _summary_text(
                    source.get("EgoAnchor_Q1"), source.get("EgoAnchor_Median"), source.get("EgoAnchor_Q3")
                ),
                "Difference": _summary_text(
                    source.get("Difference_Q1"), source.get("Difference_Median"), source.get("Difference_Q3")
                ),
                "Direction": direction_names.get(str(source.get("Direction")), str(source.get("Direction"))),
            }
        )
    return pd.DataFrame(rows)


def _reliability_rows(reliability: pd.DataFrame) -> pd.DataFrame:
    """把信度结果转换为测量单位边界明确的中文显示行。"""

    rows: list[dict[str, Any]] = []
    for _, source in reliability.iterrows():
        unit = str(source.get("Measurement_Unit"))
        items = int(_as_number(source.get("Items"))) if math.isfinite(_as_number(source.get("Items"))) else 0
        if unit == "block_mean":
            note = "三物体均值分的信度；不可与单次施测或原量表发表值直接比较"
            unit_name = "三物体均值"
        elif unit == "method_single":
            note = "每种方法单次施测的条目级内部一致性"
            unit_name = "方法级单次施测"
        else:
            note = "测量单位混合，需复核"
            unit_name = "混合"
        if items == 2:
            note += "；两条目量表主要读取 α 与 Spearman–Brown"
        rows.append(
            {
                "Scale": _OUTCOME_NAMES.get(str(source.get("Outcome")), str(source.get("Outcome"))),
                "Method": METHOD_LABELS.get(str(source.get("Condition")), str(source.get("Condition"))),
                "Unit": unit_name,
                "N": source.get("N"),
                "Items": source.get("Items"),
                "Alpha": source.get("Cronbach_Alpha"),
                "Omega": source.get("Omega_Total"),
                "SpearmanBrown": source.get("Spearman_Brown"),
                "Note": note,
            }
        )
    return pd.DataFrame(rows)


def _choice_distribution(choices: pd.DataFrame) -> pd.DataFrame:
    """提取总体偏好和信任选择的三分类分布。"""

    names = {"Method_Choice": "总体偏好", "Trust_Choice": "更愿意信任"}
    selected = choices[choices["Measure"].astype(str).isin(names)]
    rows: list[dict[str, Any]] = []
    for _, source in selected.iterrows():
        category = str(source.get("Category"))
        rows.append(
            {
                "Measure": names[str(source.get("Measure"))],
                "Category": METHOD_LABELS.get(category, category),
                "Count": source.get("Count"),
                "Proportion": source.get("Proportion"),
            }
        )
    return pd.DataFrame(rows)


def _choice_summaries(choices: pd.DataFrame) -> pd.DataFrame:
    """提取偏好强度和区分信心的两行集中趋势摘要。"""

    names = {
        "Preference_Strength": "偏好强度",
        "Discrimination_Confidence": "区分信心",
    }
    selected = choices[
        choices["Measure"].astype(str).isin(names)
        & (choices["Category"].astype(str) == "summary")
    ]
    rows: list[dict[str, Any]] = []
    for _, source in selected.iterrows():
        rows.append(
            {
                "Measure": names[str(source.get("Measure"))],
                "N": source.get("Count"),
                "MedianIQR": _summary_text(source.get("Q1"), source.get("Median"), source.get("Q3")),
                "MeanSD": f"{_format_number(source.get('Mean'))} ± {_format_number(source.get('SD'))}",
            }
        )
    return pd.DataFrame(rows)


def _choice_disagreement(choices: pd.DataFrame) -> pd.DataFrame:
    """提取总体偏好与信任选择不一致的人数和比例。"""

    selected = choices[choices["Measure"].astype(str) == "Choice_Disagreement"]
    if selected.empty:
        return pd.DataFrame(columns=("Metric", "Count", "Proportion"))
    source = selected.iloc[0]
    return pd.DataFrame(
        (
            {
                "Metric": "总体偏好与信任选择不同",
                "Count": source.get("Count"),
                "Proportion": source.get("Proportion"),
            },
        )
    )


def _choice_cross_table(choices: pd.DataFrame) -> pd.DataFrame:
    """把堆叠的偏好×信任行恢复为直观的三乘三计数矩阵。"""

    categories = (EGOANCHOR, ONE_EURO, "无明显偏好")
    selected = choices[choices["Measure"].astype(str) == "Method_By_Trust"]
    matrix = {row: {column: 0 for column in categories} for row in categories}
    for _, source in selected.iterrows():
        parts = str(source.get("Category", "")).split(" → ", maxsplit=1)
        if len(parts) == 2 and parts[0] in matrix and parts[1] in matrix[parts[0]]:
            matrix[parts[0]][parts[1]] = int(_as_number(source.get("Count")))
    rows = []
    for preference in categories:
        rows.append(
            {
                "Preference": METHOD_LABELS.get(preference, preference),
                EGOANCHOR: matrix[preference][EGOANCHOR],
                ONE_EURO: matrix[preference][ONE_EURO],
                "No_Preference": matrix[preference]["无明显偏好"],
            }
        )
    return pd.DataFrame(rows)


def _write_title(
    worksheet: Any,
    title: str,
    subtitle: str,
    max_column: int,
) -> None:
    """写入深蓝主标题和一行浅色副标题。"""

    last = get_column_letter(max_column)
    worksheet.merge_cells(f"A1:{last}1")
    worksheet["A1"] = title
    worksheet["A1"].fill = PatternFill("solid", fgColor=_NAVY)
    worksheet["A1"].font = Font(name=_FONT, size=15, bold=True, color=_WHITE)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 30
    worksheet.merge_cells(f"A2:{last}2")
    worksheet["A2"] = subtitle
    worksheet["A2"].font = Font(name=_FONT, size=9, italic=True, color=_MUTED)
    worksheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 25


def _write_key_value_section(
    worksheet: Any,
    start_row: int,
    title: str,
    rows: tuple[tuple[str, Any], ...],
) -> int:
    """在说明页写入一个两列事实分节，并返回下一可写行。"""

    _write_section_heading(worksheet, start_row, title, 2)
    header_row = start_row + 1
    for column, value in enumerate(("项目", "内容"), start=1):
        _style_header_cell(worksheet.cell(header_row, column, value), value)
    for offset, (key, value) in enumerate(rows, start=1):
        row = header_row + offset
        worksheet.cell(row, 1, key)
        worksheet.cell(row, 2, _excel_value(value))
        _style_body_row(worksheet, row, 2, band=offset % 2 == 0)
        worksheet.cell(row, 1).font = Font(name=_FONT, size=9, bold=True, color=_TEXT)
        worksheet.cell(row, 1).alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.cell(row, 2).font = Font(name=_FONT, size=9, color=_TEXT)
        worksheet.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
    return header_row + len(rows) + 2


def _write_section_table(
    worksheet: Any,
    start_row: int,
    title: str,
    note: str,
    frame: pd.DataFrame,
    columns: tuple[tuple[str, str, float, str], ...],
) -> int:
    """写入“分节标题—解释—表格”的一组内容并返回下一可写行。"""

    max_column = len(columns)
    _write_section_heading(worksheet, start_row, title, max_column)
    last = get_column_letter(max_column)
    worksheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=max_column)
    note_cell = worksheet.cell(start_row + 1, 1, note)
    note_cell.font = Font(name=_FONT, size=8, italic=True, color=_MUTED)
    note_cell.alignment = Alignment(vertical="center", wrap_text=True)
    worksheet.row_dimensions[start_row + 1].height = 24
    header_row, last_row = _write_table(worksheet, start_row + 2, frame, columns, auto_filter=False)
    if frame.empty:
        worksheet.merge_cells(f"A{header_row + 1}:{last}{header_row + 1}")
        worksheet.cell(header_row + 1, 1, "当前没有可展示记录")
        worksheet.cell(header_row + 1, 1).font = Font(name=_FONT, size=9, italic=True, color=_MUTED)
        worksheet.cell(header_row + 1, 1).alignment = Alignment(vertical="center")
        last_row = header_row + 1
    return last_row + 2


def _write_table(
    worksheet: Any,
    header_row: int,
    frame: pd.DataFrame,
    columns: tuple[tuple[str, str, float, str], ...],
    *,
    auto_filter: bool,
) -> tuple[int, int]:
    """把一个 DataFrame 写成无全格边框的浅色阅读表。"""

    for column_index, (key, title, width, _) in enumerate(columns, start=1):
        _style_header_cell(worksheet.cell(header_row, column_index, title), title)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    for offset, (_, source) in enumerate(frame.iterrows(), start=1):
        row = header_row + offset
        for column_index, (key, _, _, kind) in enumerate(columns, start=1):
            cell = worksheet.cell(row, column_index, _excel_value(source.get(key)))
            _style_value(cell, kind)
        _style_body_row(worksheet, row, len(columns), band=offset % 2 == 0)
    last_row = header_row + len(frame)
    worksheet.row_dimensions[header_row].height = 30
    if auto_filter and len(frame):
        last_column = get_column_letter(len(columns))
        worksheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    return header_row, last_row


def _write_section_heading(
    worksheet: Any,
    row: int,
    title: str,
    max_column: int,
) -> None:
    """写入跨列的浅青色分节标题。"""

    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_column)
    cell = worksheet.cell(row, 1, title)
    cell.fill = PatternFill("solid", fgColor=_PALE_TEAL)
    cell.font = Font(name=_FONT, size=11, bold=True, color=_TEAL)
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[row].height = 24


def _style_header_cell(cell: Any, title: str) -> None:
    """应用浅灰表头，并用冗余底色区分两种方法列。"""

    fill = _HEADER
    if title.startswith("One-Euro"):
        fill = _PALE_GREEN
    elif title.startswith("EgoAnchor"):
        fill = _PALE_RED
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name=_FONT, size=9, bold=True, color=_TEXT)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="medium", color=_RULE))


def _style_body_row(worksheet: Any, row: int, max_column: int, *, band: bool) -> None:
    """应用无竖向边框的轻量隔行底色和统一对齐。"""

    for cell in worksheet[row][:max_column]:
        if band and cell.fill.fill_type is None:
            cell.fill = PatternFill("solid", fgColor=_BAND)
    worksheet.row_dimensions[row].height = 21


def _style_value(cell: Any, kind: str) -> None:
    """按列语义设置对齐方式和稳定数字格式。"""

    cell.font = Font(name=_FONT, size=9, color=_TEXT)
    cell.alignment = Alignment(
        horizontal="left" if kind == "text" else "center",
        vertical="center",
        wrap_text=kind == "text",
    )
    if kind == "integer":
        cell.number_format = "0"
    elif kind == "percent":
        cell.number_format = "0.0%"
    elif kind == "pvalue":
        cell.number_format = '[<0.0001]"<0.0001";0.0000'
    elif kind == "effect":
        cell.number_format = "0.00"
    elif kind == "decimal":
        cell.number_format = "0.000"


def _prepare_sheet(worksheet: Any, *, tab_color: str, portrait: bool) -> None:
    """配置网格、打印、页边距和标签色。"""

    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = tab_color
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.orientation = "portrait" if portrait else "landscape"
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.45
    worksheet.page_margins.bottom = 0.45
    worksheet.sheet_format.defaultRowHeight = 19


def _summary_text(q1: Any, median: Any, q3: Any) -> str:
    """把中位数和四分位点写成统一的 ``Mdn [Q1, Q3]`` 文本。"""

    values = tuple(_as_number(value) for value in (median, q1, q3))
    if not all(math.isfinite(value) for value in values):
        return ""
    return f"{values[0]:.2f} [{values[1]:.2f}, {values[2]:.2f}]"


def _category_name(category: str) -> str:
    """把样本摘要中的稳定英文类别翻译为简洁中文。"""

    return {
        "all": "全部",
        "yes": "是",
        "Summary": "汇总",
        "Missing": "缺失",
        "Worsened": "加重",
    }.get(category, category)


def _require_sha256(value: object, label: str) -> str:
    """读取 64 位小写十六进制 SHA-256，拒绝伪摘要和隐式转换。"""

    hexadecimal = frozenset("0123456789abcdef")
    if (
        not isinstance(value, str)
        or len(value) != 64
        or not set(value).issubset(hexadecimal)
    ):
        raise ValueError(f"实验三{label}必须是 64 位小写十六进制 SHA-256")
    return value


def _as_number(value: Any) -> float:
    """把标量安全转换为有限性可检查的浮点数。"""

    try:
        return float(value)
    except (TypeError, ValueError):
        return math.nan


def _format_number(value: Any) -> str:
    """用最多三位小数显示一个数值，缺失时返回短横线。"""

    number = _as_number(value)
    if not math.isfinite(number):
        return "—"
    return f"{number:.3f}".rstrip("0").rstrip(".")


def _format_p(value: Any) -> str:
    """把 p 值格式化为四位小数，并保留极小值边界。"""

    number = _as_number(value)
    if not math.isfinite(number):
        return "—"
    return "<0.0001" if number < 0.0001 else f"{number:.4f}"


def _excel_value(value: Any) -> Any:
    """把 pandas/numpy 值转换为 openpyxl 可安全写入的标量。"""

    if value is None:
        return None
    if isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, (tuple, list, dict)):
        return str(value)
    return value


def _verify_results_workbook(
    path: Path,
    *,
    input_sha256: str,
    settings_sha256: str,
    included_count: int,
) -> None:
    """回读六页、输入与参数摘要、样本 N 及两张核心结果表的冻结键。"""

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        mismatched_properties = tuple(
            name
            for name, expected_value in _CORE_PROPERTIES.items()
            if getattr(workbook.properties, name, None) != expected_value
        )
        fields = mismatched_properties
        if workbook.properties.contentStatus is not None:
            fields += ("contentStatus",)
        if fields:
            raise ValueError(
                "实验三结果工作簿的核心属性回读不一致：" + "、".join(fields)
            )
        expected = [name for name, _ in _SHEET_GUIDE]
        if workbook.sheetnames != expected:
            raise ValueError(f"实验三结果工作簿必须固定为六张中文页：{workbook.sheetnames}")
        info = workbook[INFO_SHEET]
        facts = {
            str(info.cell(row, 1).value): info.cell(row, 2).value
            for row in range(1, info.max_row + 1)
            if info.cell(row, 1).value is not None
        }
        expected_facts = {
            "输入 SHA-256": _require_sha256(input_sha256, "输入 SHA-256"),
            "参数 SHA-256": _require_sha256(settings_sha256, "参数 SHA-256"),
        }
        mismatched = tuple(
            key
            for key, expected_value in expected_facts.items()
            if facts.get(key) != expected_value
        )
        if mismatched:
            raise ValueError(
                "实验三结果工作簿的输入或参数摘要回读不一致："
                + "、".join(mismatched)
            )
        if int(facts.get("纳入参与者") or 0) != included_count:
            raise ValueError("实验三结果工作簿的纳入人数回读不一致")
        expected_outcomes = tuple(
            _OUTCOME_NAMES[outcome]
            for outcome in (*PRIMARY_OUTCOMES, *SCALE_OUTCOMES)
        )
        actual_outcomes = _table_column_values(workbook[RESULTS_SHEET], "指标")
        if actual_outcomes != expected_outcomes:
            raise ValueError(
                "主结果必须按冻结顺序恰好包含七个主条目和已发表量表家族五项结局："
                f"{actual_outcomes}"
            )
        expected_object_keys = tuple(
            (_OUTCOME_NAMES[outcome], _OBJECT_NAMES[object_key])
            for outcome in PRIMARY_OUTCOMES
            for object_key in OBJECTS
        )
        actual_object_keys = _table_key_values(
            workbook[OBJECT_RESULTS_SHEET],
            ("指标", "物体"),
        )
        if actual_object_keys != expected_object_keys:
            raise ValueError(
                "分物体描述必须按冻结顺序恰好包含七个主条目乘三个对象："
                f"{actual_object_keys}"
            )
    finally:
        workbook.close()


def _table_column_values(worksheet: Any, header: str) -> tuple[str, ...]:
    """返回指定表头下方的连续文本值，用于验证冻结行键与顺序。"""

    location = _find_table_header(worksheet, header)
    if location is None:
        return ()
    header_row, header_column = location
    values: list[str] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        value = worksheet.cell(row, header_column).value
        if value is None:
            break
        values.append(str(value))
    return tuple(values)


def _table_key_values(
    worksheet: Any,
    headers: tuple[str, ...],
) -> tuple[tuple[str, ...], ...]:
    """返回同一表头行下的连续复合键，保留冻结行顺序。"""

    locations = tuple(_find_table_header(worksheet, header) for header in headers)
    if any(location is None for location in locations):
        return ()
    checked_locations = tuple(location for location in locations if location is not None)
    header_rows = {row for row, _ in checked_locations}
    if len(header_rows) != 1:
        return ()
    header_row = checked_locations[0][0]
    columns = tuple(column for _, column in checked_locations)
    keys: list[tuple[str, ...]] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        values = tuple(worksheet.cell(row, column).value for column in columns)
        if all(value is None for value in values):
            break
        keys.append(tuple("" if value is None else str(value) for value in values))
    return tuple(keys)


def _find_table_header(worksheet: Any, header: str) -> tuple[int, int] | None:
    """在一个阅读页中定位唯一表头文本。"""

    return next(
        (
            (row, column)
            for row in range(1, worksheet.max_row + 1)
            for column in range(1, worksheet.max_column + 1)
            if str(worksheet.cell(row, column).value) == header
        ),
        None,
    )


__all__ = [
    "CHOICES_SHEET",
    "INFO_SHEET",
    "OBJECT_RESULTS_SHEET",
    "RELIABILITY_SHEET",
    "RESULTS_SHEET",
    "SAMPLE_QC_SHEET",
    "write_results_workbook",
]
