"""实验三五表原始工作簿的严格只读 reader。"""

from __future__ import annotations

import hashlib
import math
from collections.abc import Iterable
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]

from ...common import file_sha256
from .contracts import (
    BLOCK_ITEMS,
    EXCLUSION_REASONS,
    Exp3Data,
    MINIMUM_PARTICIPANTS,
    METHOD_ITEM_COLUMNS,
    METHODS,
    OBJECT_RAW_LABELS,
    OBJECTS,
    PARTICIPANT_BACKGROUND_COLUMNS,
    PARTICIPANT_CATEGORIES,
    WORKBOOK_CONTRACT_ID,
    WORKBOOK_DATA_CATEGORY,
    required_block_items,
)


_REQUIRED_SHEETS = frozenset(
    {"Questionnaire", "Participants", "Block", "Method", "Final"}
)
"""v5.3 原始工作簿必须包含的五张工作表。"""

_WORKBOOK_DESCRIPTION = "EgoAnchor 实验三 v5.3 五表空白数据模板。"
"""当前 v5.3 工作簿在文档属性中保存的版本说明。"""

_V53_QUESTIONNAIRE_TEXT = {
    "AQ_EQ2": "虚拟内容看起来真实、自然地融入了真实物体及其周围环境。",
    "TIA_RC1": "这种对象锚定方法能够根据当前情况做出正确的锚定反应。",
    "TIA_RC4": "这种对象锚定方法能够处理复杂的对象锚定任务。",
    "TIA_UP1": "这种对象锚定方法当前的工作状态对我来说始终清楚。",
}
"""identifier 缺失时用于确认 v5.3 版本的关键施测文本。"""

_QUESTIONNAIRE_ITEM_ORDER = (
    "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7",
    "AQ_EQ1", "AQ_EQ2", "AQ_EQ3", "AQ_IQ1", "AQ_IQ2", "AQ_IQ3",
)
"""区块问卷按研究定制条目、AQ-EQ、AQ-IQ 分组的固定顺序。"""

_PARTICIPANT_DESIGN_COLUMNS = (
    "Participant_ID",
    "平衡单元",
    "物体排列ID",
    "标签序列",
    "物体1",
    "物体2",
    "物体3",
    "区块内标签顺序",
    "方法A=（保密）",
    "方法B=（保密）",
    "先行实际方法",
)
"""24 平衡单元不得缺失或被公式替换的设计列。"""

_BLOCK_DESIGN_COLUMNS = (
    "Participant_ID",
    "Block_Index",
    "平衡单元",
    "物体位置",
    "物体",
    "Object_Key",
    "Shown_Label",
    "Condition(保密)",
    "物体内先后",
    "该方法第几次",
)
"""``Block`` 工作表的固定设计列。"""

_METHOD_DESIGN_COLUMNS = (
    "Participant_ID",
    "Rating_Order",
    "Shown_Label",
    "Condition(保密)",
)
"""``Method`` 工作表的固定设计列。"""

_METHOD_STATUS_COLUMNS = ("尺度切换确认", "技术问题")
"""方法级问卷的完成状态列。"""

_OBJECT_ORDERS = {
    1: ("鼠标", "固定订书机", "游戏手柄"),
    2: ("鼠标", "游戏手柄", "固定订书机"),
    3: ("固定订书机", "鼠标", "游戏手柄"),
    4: ("固定订书机", "游戏手柄", "鼠标"),
    5: ("游戏手柄", "鼠标", "固定订书机"),
    6: ("游戏手柄", "固定订书机", "鼠标"),
}
"""物体排列 ID 与三个正式对象全排列的冻结绑定。"""

def read_workbook(path: Path) -> Exp3Data:
    """从同一字节快照读取五表原始值并验证结构身份。"""

    source = path.expanduser().resolve()
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise FileNotFoundError(f"实验三输入必须是现存 XLSX：{source}")
    payload = source.read_bytes()
    source_digest = hashlib.sha256(payload).hexdigest()
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    try:
        actual_sheets = frozenset(workbook.sheetnames)
        if actual_sheets != _REQUIRED_SHEETS:
            missing = _REQUIRED_SHEETS.difference(actual_sheets)
            unexpected = actual_sheets.difference(_REQUIRED_SHEETS)
            raise ValueError(
                "实验三工作簿必须恰好包含五张固定工作表："
                f"missing={sorted(missing)}, unexpected={sorted(unexpected)}"
            )
        _validate_workbook_identity(workbook)
        _validate_questionnaire_sheet(workbook["Questionnaire"])
        participants = _read_table(workbook["Participants"], 1, 2, workbook["Participants"].max_row)
        blocks = _read_table(workbook["Block"], 1, 2, workbook["Block"].max_row)
        methods = _read_table(workbook["Method"], 1, 2, workbook["Method"].max_row)
        finals = _read_table(workbook["Final"], 1, 2, workbook["Final"].max_row)
    finally:
        workbook.close()
    if file_sha256(source) != source_digest:
        raise ValueError("实验三输入工作簿在读取期间发生变化，请保存完成后重试")
    data = Exp3Data(
        participants=participants,
        blocks=blocks,
        methods=methods,
        finals=finals,
        source_path=str(source),
        source_sha256=source_digest,
    )
    _validate_structure(data)
    return data


def validate_for_analysis(
    data: Exp3Data,
    *,
    aq_mode: str,
) -> dict[str, Any]:
    """对所有输入统一检查采集完成状态、平衡设计与合法值。"""

    participants = data.participants.copy()
    included = participants[participants["纳入分析"].map(_is_yes)]
    included_ids = frozenset(included["Participant_ID"].astype(str))
    safety_ids = frozenset(
        participants.loc[participants["签署同意"].map(_is_yes), "Participant_ID"].astype(str)
    )
    warnings: list[str] = []
    if len(included_ids) < MINIMUM_PARTICIPANTS:
        raise ValueError(
            f"纳入分析且已确认的参与者只有 {len(included_ids)} 人，"
            f"少于冻结下限 {MINIMUM_PARTICIPANTS}"
        )
    _validate_participant_values(
        data.participants,
        included_ids,
    )
    _validate_block_values(
        data.blocks,
        included_ids,
        aq_mode=aq_mode,
    )
    _validate_method_values(
        data.methods,
        included_ids,
    )
    _validate_final_values(
        data.finals,
        included_ids,
        safety_ids,
    )
    if len(included_ids) < 24:
        warnings.append(f"当前纳入 {len(included_ids)} 人，少于目标 N=24；结果按实际配对 N 报告")
    return {
        "included_participants": tuple(sorted(included_ids)),
        "included_count": len(included_ids),
        "warnings": tuple(warnings),
    }


def describe_workbook(data: Exp3Data) -> dict[str, Any]:
    """返回不要求采集完成的结构与填表进度摘要。"""

    participants = data.participants
    blocks = data.blocks
    methods = data.methods
    finals = data.finals
    return {
        "passed": True,
        "source": data.source_path,
        "source_sha256": data.source_sha256,
        "participant_rows": len(participants),
        "included_confirmed": int(participants["纳入分析"].map(_is_yes).sum()),
        "block_rows": len(blocks),
        "block_scores_filled": int(blocks[list(BLOCK_ITEMS.values())].notna().sum().sum()),
        "method_rows": len(methods),
        "method_scores_filled": int(methods[list(METHOD_ITEM_COLUMNS)].notna().sum().sum()),
        "final_rows": len(finals),
        "final_choices_filled": int(finals["方法选择(标签)"].notna().sum()),
    }


def _read_table(worksheet: Any, header_row: int, first_row: int, last_row: int) -> pd.DataFrame:
    """按指定表头和行边界读取一张普通值表。"""

    header_cells = next(
        worksheet.iter_rows(
            min_row=header_row,
            max_row=header_row,
            min_col=1,
            max_col=worksheet.max_column,
        )
    )
    headers = [cell.value for cell in header_cells]
    while headers and headers[-1] is None:
        headers.pop()
    if not headers or any(value is None for value in headers):
        raise ValueError(f"{worksheet.title}!{header_row} 表头包含空列")
    rows: list[dict[str, Any]] = []
    for row_number, cells in enumerate(
        worksheet.iter_rows(
            min_row=first_row,
            max_row=last_row,
            min_col=1,
            max_col=len(headers),
        ),
        start=first_row,
    ):
        if all(cell.value is None for cell in cells):
            continue
        _reject_formulas(worksheet.title, row_number, cells)
        rows.append({str(header): cell.value for header, cell in zip(headers, cells, strict=True)})
    return pd.DataFrame(rows, columns=[str(value) for value in headers])


def _reject_formulas(sheet_name: str, row_number: int, cells: Iterable[Cell]) -> None:
    """拒绝在原始值区域放入公式，避免评分来源被派生值覆盖。"""

    for cell in cells:
        if cell.data_type == "f":
            raise ValueError(f"原始数据区域不得含公式：{sheet_name}!{cell.coordinate}（第 {row_number} 行）")


def _validate_questionnaire_sheet(worksheet: Any) -> None:
    """核对区块问卷页序、连续题号和已删除可选题状态。"""

    header_row: int | None = None
    for row in range(1, worksheet.max_row + 1):
        if worksheet.cell(row, 1).value == "页序" and worksheet.cell(row, 2).value == "Item_ID":
            header_row = row
            break
    if header_row is None:
        raise ValueError("Questionnaire 缺少区块条目表头")
    items: list[str] = []
    pages: list[int] = []
    for row in range(header_row + 1, worksheet.max_row + 1):
        page = worksheet.cell(row, 1).value
        item = worksheet.cell(row, 2).value
        if page == "量尺":
            break
        if page is None and item is None:
            continue
        try:
            pages.append(int(page))
        except (TypeError, ValueError) as error:
            raise ValueError(f"Questionnaire 第 {row} 行页序必须是整数") from error
        items.append(str(item))
    if tuple(pages) != tuple(range(1, 14)) or tuple(items) != _QUESTIONNAIRE_ITEM_ORDER:
        raise ValueError(
            "Questionnaire 区块条目必须按 1--13 页固定排列："
            f"pages={pages}, items={items}"
        )


def _validate_workbook_identity(workbook: Any) -> None:
    """验证 v5.3 契约；兼容 WPS/Excel 保存时清除 identifier 的情况。"""

    identifier = workbook.properties.identifier
    if identifier is not None:
        if identifier != WORKBOOK_CONTRACT_ID:
            raise ValueError(
                "实验三工作簿契约标识不匹配："
                f"{identifier!r} != {WORKBOOK_CONTRACT_ID!r}"
            )
        return

    metadata = (workbook.properties.category, workbook.properties.description)
    expected_metadata = (WORKBOOK_DATA_CATEGORY, _WORKBOOK_DESCRIPTION)
    if metadata != expected_metadata:
        raise ValueError(
            "实验三工作簿缺少契约标识，且 v5.3 文档属性不匹配："
            f"{metadata!r} != {expected_metadata!r}"
        )

    questionnaire = workbook["Questionnaire"]
    actual_text: dict[str, Any] = {}
    for row in range(1, questionnaire.max_row + 1):
        item_candidates = (
            questionnaire.cell(row, 1).value,
            questionnaire.cell(row, 2).value,
        )
        for item in item_candidates:
            if item in _V53_QUESTIONNAIRE_TEXT:
                actual_text[str(item)] = questionnaire.cell(row, 4).value
                break
    if actual_text != _V53_QUESTIONNAIRE_TEXT:
        raise ValueError(
            "实验三工作簿缺少契约标识，且关键施测文本不是 v5.3："
            f"{actual_text!r}"
        )


def _validate_structure(data: Exp3Data) -> None:
    """验证 24 平衡单元和五表记录的固定身份。"""

    _require_columns(
        data.participants,
        _PARTICIPANT_DESIGN_COLUMNS
        + tuple(PARTICIPANT_BACKGROUND_COLUMNS.values())
        + ("签署同意", "基线不适", "纳入分析", "退出/技术问题"),
        "Participants",
    )
    _require_columns(
        data.blocks,
        _BLOCK_DESIGN_COLUMNS + tuple(BLOCK_ITEMS.values()) + (
            "任务完成",
            "问卷完成",
            "区块有效",
            "技术问题",
        ),
        "Block",
    )
    _require_columns(
        data.methods,
        _METHOD_DESIGN_COLUMNS + METHOD_ITEM_COLUMNS + _METHOD_STATUS_COLUMNS,
        "Method",
    )
    _require_columns(
        data.finals,
        (
            "Participant_ID",
            "方法选择(标签)",
            "偏好强度(1-7/NA)",
            "信任选择(标签)",
            "区分信心(1-7)",
            "开放:最明显区别",
            "开放:最破坏信任的现象",
            "结束不适",
        ),
        "Final",
    )
    if len(data.participants) != 24 or len(data.blocks) != 144 or len(data.methods) != 48 or len(data.finals) != 24:
        raise ValueError(
            "实验三固定结构必须是 Participants=24、Block=144、Method=48、Final=24"
        )
    participant_ids = tuple(data.participants["Participant_ID"].astype(str))
    if len(set(participant_ids)) != 24:
        raise ValueError("Participants 的 Participant_ID 必须唯一")
    expected = frozenset(participant_ids)
    for name, table, repetitions in (
        ("Block", data.blocks, 6),
        ("Method", data.methods, 2),
        ("Final", data.finals, 1),
    ):
        counts = table["Participant_ID"].astype(str).value_counts()
        if frozenset(counts.index) != expected or not (counts == repetitions).all():
            raise ValueError(f"{name} 必须为每位参与者恰好保留 {repetitions} 行")
    _validate_design_mapping(data)


def _validate_design_mapping(data: Exp3Data) -> None:
    """核对标签映射、对象覆盖和方法级评分顺序。"""

    design = data.participants.copy()
    if design["平衡单元"].astype(str).duplicated().any():
        raise ValueError("Participants 的 24 个平衡单元必须唯一")
    order_ids = tuple(_exact_integer(value) for value in design["物体排列ID"])
    if any(order_id not in _OBJECT_ORDERS for order_id in order_ids):
        raise ValueError("Participants 的物体排列ID必须是 1--6 整数")
    if set(order_ids) != set(_OBJECT_ORDERS):
        raise ValueError("Participants 的物体排列ID必须覆盖 1--6")
    actual_cells = {
        (order_id, str(sequence), str(mapping))
        for order_id, sequence, mapping in zip(
            order_ids,
            design["标签序列"],
            design["方法A=（保密）"],
            strict=True,
        )
    }
    expected_cells = {
        (order_id, sequence, mapping)
        for order_id in range(1, 7)
        for sequence in ("S1", "S2")
        for mapping in METHODS
    }
    if actual_cells != expected_cells:
        raise ValueError("Participants 必须恰好覆盖 6 物体排列 × S1/S2 × A 标签映射的 24 单元")
    expected_label_order = {
        "S1": "A-B / B-A / A-B",
        "S2": "B-A / A-B / B-A",
    }
    for _, row in design.iterrows():
        order_id = _exact_integer(row["物体排列ID"])
        if order_id is None:
            raise ValueError("Participants 的物体排列ID必须是 1--6 整数")
        actual_object_order = tuple(str(row[f"物体{position}"]) for position in range(1, 4))
        if actual_object_order != _OBJECT_ORDERS[order_id]:
            raise ValueError(
                f"{row['Participant_ID']} 的物体排列ID {order_id} 未绑定冻结的物体1--3顺序"
            )
        sequence = str(row["标签序列"])
        if str(row["区块内标签顺序"]) != expected_label_order.get(sequence):
            raise ValueError(f"{row['Participant_ID']} 的区块内标签顺序与 {sequence} 不一致")
        first_label = "方法A" if sequence == "S1" else "方法B"
        first_method = row["方法A=（保密）"] if first_label == "方法A" else row["方法B=（保密）"]
        if str(row["先行实际方法"]) != str(first_method):
            raise ValueError(f"{row['Participant_ID']} 的先行实际方法与标签序列不一致")

    participants = design.set_index("Participant_ID")
    for participant_id, rows in data.blocks.groupby("Participant_ID", sort=False):
        if frozenset(rows["Condition(保密)"].astype(str)) != frozenset(METHODS):
            raise ValueError(f"{participant_id} 的区块未覆盖两种方法")
        if frozenset(rows["Object_Key"].astype(str)) != frozenset(OBJECTS):
            raise ValueError(f"{participant_id} 的区块未覆盖三个正式对象")
        if len(rows.groupby(["Object_Key", "Condition(保密)"], dropna=False)) != 6:
            raise ValueError(f"{participant_id} 的对象×方法区块不唯一")
        mapping = participants.loc[participant_id]
        expected_method = {
            "方法A": str(mapping["方法A=（保密）"]),
            "方法B": str(mapping["方法B=（保密）"]),
        }
        _validate_participant_block_order(
            str(participant_id),
            rows,
            mapping,
            expected_method,
        )
    for participant_id, rows in data.methods.groupby("Participant_ID", sort=False):
        if frozenset(rows["Condition(保密)"].astype(str)) != frozenset(METHODS):
            raise ValueError(f"{participant_id} 的方法级记录未覆盖两种方法")
        mapping = participants.loc[participant_id]
        expected_conditions = {
            "方法A": str(mapping["方法A=（保密）"]),
            "方法B": str(mapping["方法B=（保密）"]),
        }
        sequence = str(mapping["标签序列"])
        expected_labels = ("方法A", "方法B") if sequence == "S1" else ("方法B", "方法A")
        ordered = rows.sort_values("Rating_Order")
        if tuple(pd.to_numeric(ordered["Rating_Order"], errors="coerce")) != (1, 2):
            raise ValueError(f"{participant_id} 的方法级 Rating_Order 必须为 1、2")
        if tuple(ordered["Shown_Label"].astype(str)) != expected_labels:
            raise ValueError(f"{participant_id} 的方法级评分顺序与 {sequence} 不一致")
        for _, row in ordered.iterrows():
            label = str(row["Shown_Label"])
            if expected_conditions.get(label) != str(row["Condition(保密)"]):
                raise ValueError(f"{participant_id} 的方法级 {label} 映射与 Participants 不一致")


def _validate_participant_block_order(
    participant_id: str,
    rows: pd.DataFrame,
    mapping: pd.Series,
    expected_method: dict[str, str],
) -> None:
    """核对六个实际区块与 Participants 冻结计划逐位置一致。"""

    block_indices = pd.to_numeric(rows["Block_Index"], errors="coerce")
    if block_indices.isna().any() or (block_indices % 1 != 0).any():
        raise ValueError(f"{participant_id} 的 Block_Index 必须是 1--6 整数")
    ordered = rows.assign(_block_index=block_indices.astype(int)).sort_values("_block_index")
    if tuple(ordered["_block_index"]) != tuple(range(1, 7)):
        raise ValueError(f"{participant_id} 的实际区块顺序必须恰好覆盖 1--6")

    sequence = str(mapping["标签序列"])
    label_pairs = {
        "S1": (("方法A", "方法B"), ("方法B", "方法A"), ("方法A", "方法B")),
        "S2": (("方法B", "方法A"), ("方法A", "方法B"), ("方法B", "方法A")),
    }
    expected_pairs = label_pairs.get(sequence)
    if expected_pairs is None:
        raise ValueError(f"{participant_id} 的标签序列不是 S1 或 S2")
    raw_key_by_label = {label: key for key, label in OBJECT_RAW_LABELS.items()}
    method_occurrences = {method: 0 for method in METHODS}
    for row_offset, (_, row) in enumerate(ordered.iterrows()):
        object_position = row_offset // 2 + 1
        within_object = row_offset % 2 + 1
        planned_object = str(mapping[f"物体{object_position}"])
        expected_object_key = raw_key_by_label.get(planned_object)
        if expected_object_key is None:
            raise ValueError(f"{participant_id} 的 Participants 物体计划含未知标签：{planned_object}")
        expected_label = expected_pairs[object_position - 1][within_object - 1]
        expected_condition = expected_method[expected_label]
        method_occurrences[expected_condition] += 1
        expected_values = {
            "平衡单元": str(mapping["平衡单元"]),
            "物体位置": object_position,
            "物体": planned_object,
            "Object_Key": expected_object_key,
            "Shown_Label": expected_label,
            "Condition(保密)": expected_condition,
            "物体内先后": within_object,
            "该方法第几次": method_occurrences[expected_condition],
        }
        for column, expected in expected_values.items():
            actual = row[column]
            matches = (
                _exact_integer(actual) == expected
                if isinstance(expected, int)
                else str(actual) == expected
            )
            if not matches:
                raise ValueError(
                    f"{participant_id} 的实际区块顺序与 Participants 计划不一致："
                    f"Block_Index={row['_block_index']}，{column}={actual!r}，期望 {expected!r}"
                )


def _require_columns(table: pd.DataFrame, columns: Iterable[str], name: str) -> None:
    """要求逻辑表包含指定列。"""

    missing = set(columns).difference(table.columns)
    if missing:
        raise ValueError(f"{name} 缺少列：{sorted(missing)}")


def _validate_block_values(
    blocks: pd.DataFrame,
    included_ids: frozenset[str],
    *,
    aq_mode: str,
) -> None:
    """检查纳入参与者的区块状态与 1--7 原始评分。"""

    selected = blocks[blocks["Participant_ID"].astype(str).isin(included_ids)]
    for participant_id, rows in selected.groupby("Participant_ID", sort=False):
        valid = rows.apply(_block_is_valid, axis=1)
        if int(valid.sum()) != 6:
            raise ValueError(f"{participant_id} 必须有 6 个明确完成且有效的正式区块")
    required = [BLOCK_ITEMS[item] for item in required_block_items(aq_mode)]
    _validate_numeric_range(selected, required, 1.0, 7.0, "区块评分")


def _validate_participant_values(
    participants: pd.DataFrame,
    included_ids: frozenset[str],
) -> None:
    """校验纳入样本的背景字段，并保留原始分类语义。"""

    selected = participants[participants["Participant_ID"].astype(str).isin(included_ids)]
    if not selected["签署同意"].map(_is_yes).all():
        raise ValueError("纳入参与者必须全部明确签署同意")
    pending = participants[
        participants["签署同意"].map(_is_yes)
        & ~participants["纳入分析"].map(_is_yes)
        & ~participants["纳入分析"].map(_is_no)
    ]
    if not pending.empty:
        raise ValueError("正式分析前，所有已签署同意的参与者都必须明确标记纳入或排除")
    age = pd.to_numeric(selected["B1_年龄"], errors="coerce")
    if age.isna().any() or ((age <= 0) | (age > 120) | (age % 1 != 0)).any():
        raise ValueError("纳入参与者的 B1_年龄必须是 1--120 的正整数；此范围只作录入合法性检查")
    for output_column, source_column in PARTICIPANT_BACKGROUND_COLUMNS.items():
        if output_column == "Age":
            continue
        allowed = set(PARTICIPANT_CATEGORIES[output_column])
        values = selected[source_column]
        if values.isna().any() or not values.astype(str).isin(allowed).all():
            raise ValueError(f"纳入参与者的 {source_column} 必须使用冻结下拉选项")
    exposed = participants[participants["签署同意"].map(_is_yes)]
    baseline_allowed = set(PARTICIPANT_CATEGORIES["Baseline_Discomfort"])
    baseline = exposed["基线不适"]
    invalid_baseline = ~baseline.map(_is_missing_or_na) & ~baseline.astype(str).isin(baseline_allowed)
    if invalid_baseline.any():
        raise ValueError("已开始参与者的基线不适必须留空或使用冻结选项")
    excluded = participants[
        participants["签署同意"].map(_is_yes)
        & participants["纳入分析"].map(_is_no)
    ]
    if excluded["退出/技术问题"].map(_is_missing_or_na).any():
        raise ValueError("已签署同意但不纳入分析的参与者必须记录退出/技术问题")
    recorded_reasons = excluded.loc[~excluded["退出/技术问题"].map(_is_missing_or_na), "退出/技术问题"]
    if not recorded_reasons.astype(str).isin(EXCLUSION_REASONS).all():
        raise ValueError("退出/技术问题必须使用冻结主原因；补充细节只写备注")


def _validate_method_values(
    methods: pd.DataFrame,
    included_ids: frozenset[str],
) -> None:
    """检查纳入参与者的方法级原始评分与完成状态。"""

    selected = methods[methods["Participant_ID"].astype(str).isin(included_ids)]
    for column in METHOD_ITEM_COLUMNS[:10]:
        if selected[column].map(_is_blank_response).any():
            raise ValueError(f"纳入参与者的 TiA 条目 {column} 必须填写评分或“无法回答”")
    _validate_numeric_range(selected, METHOD_ITEM_COLUMNS[:10], 1.0, 5.0, "TiA 原始评分", allow_missing=True)
    _validate_numeric_range(selected, METHOD_ITEM_COLUMNS[10:], 1.0, 7.0, "S-TIAS 原始评分")
    if not method_assessment_complete_mask(selected).all():
        raise ValueError("纳入参与者的两次方法级问卷都必须完整作答并确认尺度切换")
    if not method_record_valid_mask(selected).all():
        raise ValueError("纳入参与者的方法级记录必须完整且没有技术故障")


def _validate_final_values(
    finals: pd.DataFrame,
    included_ids: frozenset[str],
    safety_ids: frozenset[str],
) -> None:
    """检查最终选择、条件跳题与区分信心。"""

    selected = finals[finals["Participant_ID"].astype(str).isin(included_ids)]
    choices = {"方法A", "方法B", "无明显偏好"}
    for column in ("方法选择(标签)", "信任选择(标签)"):
        invalid = selected[column].dropna().astype(str).map(lambda value: value not in choices)
        if selected[column].isna().any() or invalid.any():
            raise ValueError(f"纳入参与者的 {column} 必须使用方法A/方法B/无明显偏好")
    _validate_numeric_range(selected, ("区分信心(1-7)",), 1.0, 7.0, "区分信心")
    for _, row in selected.iterrows():
        choice = row["方法选择(标签)"]
        strength = row["偏好强度(1-7/NA)"]
        if str(choice) == "无明显偏好":
            if not _is_missing_or_na(strength):
                raise ValueError("选择无明显偏好时，偏好强度必须留空或写 N/A")
        elif _number_or_none(strength) is None or not 1.0 <= float(strength) <= 7.0:
            raise ValueError("做出方法选择时，偏好强度必须为 1--7")
    for column in ("开放:最明显区别", "开放:最破坏信任的现象", "结束不适"):
        if selected[column].map(_is_missing_or_na).any():
            raise ValueError(f"纳入参与者的 {column} 不得缺失")
    allowed_discomfort = set(PARTICIPANT_CATEGORIES["End_Discomfort"])
    if not selected["结束不适"].astype(str).isin(allowed_discomfort).all():
        raise ValueError("纳入参与者的结束不适必须使用冻结选项")
    safety = finals[finals["Participant_ID"].astype(str).isin(safety_ids)]
    recorded = safety.loc[~safety["结束不适"].map(_is_missing_or_na), "结束不适"]
    if not recorded.astype(str).isin(allowed_discomfort).all():
        raise ValueError("已开始参与者的非空结束不适必须使用冻结选项")


def method_assessment_complete_mask(methods: pd.DataFrame) -> pd.Series:
    """返回方法级问卷已作答且完成尺度切换确认的掩码。"""

    responses = methods.loc[:, list(METHOD_ITEM_COLUMNS)].apply(
        lambda column: ~column.map(_is_blank_response)
    )
    return responses.all(axis=1) & methods["尺度切换确认"].map(_is_yes)


def method_record_valid_mask(methods: pd.DataFrame) -> pd.Series:
    """返回完成作答且没有记录技术故障的方法级记录掩码。"""

    valid = method_assessment_complete_mask(methods)
    valid &= methods["技术问题"].map(
        lambda value: _is_missing_or_na(value) or str(value).strip() == "无"
    )
    return valid


def block_valid_mask(blocks: pd.DataFrame) -> pd.Series:
    """返回同时满足完成、问卷与技术状态的区块有效掩码。"""

    return blocks.apply(_block_is_valid, axis=1)


def _block_is_valid(row: pd.Series) -> bool:
    """判断一个区块是否满足冻结的技术有效条件。"""

    return (
        _is_yes(row.get("任务完成"))
        and _is_yes(row.get("问卷完成"))
        and _is_yes(row.get("区块有效"))
    )


def included_participant_ids(participants: pd.DataFrame) -> frozenset[str]:
    """返回明确标记为纳入分析的参与者 ID。"""

    return frozenset(
        participants.loc[participants["纳入分析"].map(_is_yes), "Participant_ID"].astype(str)
    )


def _validate_numeric_range(
    table: pd.DataFrame,
    columns: Iterable[str],
    minimum: float,
    maximum: float,
    label: str,
    *,
    allow_missing: bool = False,
) -> None:
    """检查若干评分列是整数刻度值，并按契约处理缺失。"""

    for column in columns:
        numeric = pd.to_numeric(table[column], errors="coerce")
        invalid_text = table[column].notna() & numeric.isna() & ~table[column].map(_is_missing_or_na)
        if invalid_text.any():
            raise ValueError(f"{label}列 {column} 含非数值响应")
        valid = numeric.dropna()
        if ((valid < minimum) | (valid > maximum) | (valid % 1 != 0)).any():
            raise ValueError(f"{label}列 {column} 必须是 {minimum:.0f}--{maximum:.0f} 整数")
        if not allow_missing and numeric.isna().any():
            raise ValueError(f"{label}列 {column} 不允许缺失")


def _number_or_none(value: Any) -> float | None:
    """把普通数字转换为 float，缺失或 N/A 返回 None。"""

    if _is_missing_or_na(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _exact_integer(value: Any) -> int | None:
    """把有限整数值规范化为 int，拒绝小数、无穷和缺失。"""

    number = _number_or_none(value)
    if number is None or not math.isfinite(number) or number % 1 != 0:
        return None
    return int(number)


def _is_missing_or_na(value: Any) -> bool:
    """识别工作簿允许的空值或 N/A 文本。"""

    if value is None or (isinstance(value, float) and pd.isna(value)):
        return True
    return str(value).strip().lower() in {"", "n/a", "na", "无法回答"}


def _is_yes(value: Any) -> bool:
    """识别人工确认字段中的肯定值。"""

    return str(value or "").strip().lower() in {"是", "yes", "true", "1"}


def _is_no(value: Any) -> bool:
    """识别人工确认字段中的否定值。"""

    return str(value or "").strip().lower() in {"否", "no", "false", "0"}


def _is_blank_response(value: Any) -> bool:
    """区分真正空白与 TiA 的显式“无法回答”响应。"""

    if isinstance(value, str) and value.strip() == "无法回答":
        return False
    return _is_missing_or_na(value)


__all__ = [
    "block_valid_mask",
    "describe_workbook",
    "included_participant_ids",
    "method_assessment_complete_mask",
    "method_record_valid_mask",
    "read_workbook",
    "validate_for_analysis",
]
