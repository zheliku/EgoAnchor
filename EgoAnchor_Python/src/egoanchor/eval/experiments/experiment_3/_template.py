"""从当前 v5.3 五表工作簿生成空白实验三数据模板。"""

from __future__ import annotations

from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]

from .analysis import (
    BLOCK_ITEMS,
    EXCLUSION_REASONS,
    METHOD_ITEM_COLUMNS,
    PARTICIPANT_BACKGROUND_COLUMNS,
    PARTICIPANT_CATEGORIES,
    VRMR_EXPERIENCE_TEMPLATE_OPTIONS,
    WORKBOOK_CONTRACT_ID,
    WORKBOOK_DATA_CATEGORY,
    AnalysisSettings,
)


_SHEET_ROWS = {
    "Questionnaire": 66,
    "Participants": 25,
    "Block": 145,
    "Method": 49,
    "Final": 25,
}
"""五表模板各自允许保留的最后一行。"""

_BLOCK_STATUS_COLUMNS = ("任务完成", "问卷完成", "区块有效", "技术问题")
"""区块评分之后由实验员填写的状态列。"""

_METHOD_STATUS_COLUMNS = ("尺度切换确认", "技术问题", "备注")
"""方法级问卷的状态与备注列。"""

_FINAL_INPUT_COLUMNS = (
    "方法选择(标签)",
    "偏好强度(1-7/NA)",
    "信任选择(标签)",
    "区分信心(1-7)",
    "开放:最明显区别",
    "开放:最破坏信任的现象",
    "结束不适",
    "方法选择",
    "信任选择",
    "访谈备注",
)
"""最终问卷中除参与者编号外的全部可填写列。"""


def build_raw_template(
    settings: AnalysisSettings,
    destination: Path,
    *,
    source_template: Path,
) -> Path:
    """复制正式五表结构，只清空人工填写区域并重建输入校验。"""

    source = source_template.expanduser().resolve()
    output = destination.expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"原始工作簿结构来源不存在：{source}")
    if output == source:
        raise ValueError("新原始模板不得覆盖结构来源文件")
    if output.exists():
        raise FileExistsError(f"拒绝覆盖已有实验三原始工作簿：{output}")
    if settings.template_version != "v5.3":
        raise ValueError("实验三空白模板只接受当前 v5.3 配置")

    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.stem}.{uuid4().hex}.tmp.xlsx")
    workbook = load_workbook(source)
    try:
        _validate_layout(workbook)
        _set_properties(workbook)
        _ensure_v53_questionnaire_text(workbook)
        _clear_inputs(workbook)
        _rebuild_validations(workbook)
        _prepare_navigation(workbook)
        workbook.save(temporary)
    finally:
        workbook.close()
    try:
        temporary.replace(output)
    finally:
        temporary.unlink(missing_ok=True)
    return output


def _validate_layout(workbook: Any) -> None:
    """拒绝旧版七表工作簿或缺少关键列的结构来源。"""

    expected = tuple(_SHEET_ROWS)
    if tuple(workbook.sheetnames) != expected:
        raise ValueError(
            f"实验三模板必须按固定顺序包含五表：{expected}，实际为 {tuple(workbook.sheetnames)}"
        )
    _require_headers(
        workbook["Block"],
        (*BLOCK_ITEMS.values(), *_BLOCK_STATUS_COLUMNS),
    )
    _require_headers(
        workbook["Method"],
        (*METHOD_ITEM_COLUMNS, *_METHOD_STATUS_COLUMNS),
    )
    _require_headers(workbook["Final"], _FINAL_INPUT_COLUMNS)


def _set_properties(workbook: Any) -> None:
    """写入可由严格 reader 核对的 v5.3 后续模板身份。"""

    workbook.properties.creator = "EgoAnchor"
    workbook.properties.lastModifiedBy = "EgoAnchor"
    workbook.properties.title = "EgoAnchor 实验三空白原始数据模板"
    workbook.properties.subject = "跨对象主观评价采集模板"
    workbook.properties.description = "EgoAnchor 实验三 v5.3 五表空白数据模板。"
    workbook.properties.keywords = "EgoAnchor, Experiment 3, questionnaire"
    workbook.properties.identifier = WORKBOOK_CONTRACT_ID
    workbook.properties.category = WORKBOOK_DATA_CATEGORY
    workbook.properties.version = "v5.3"


def _ensure_v53_questionnaire_text(workbook: Any) -> None:
    """确保派生空白模板使用当前 v5.3 问卷措辞。"""

    sheet = workbook["Questionnaire"]
    sheet.cell(7, 3).value = "从未 / 1–5 次 / 6–20 次 / 21 次及以上"
    block_replacements = {
        "AQ_EQ2": "虚拟内容看起来真实、自然地融入了真实物体及其周围环境。",
    }
    method_replacements = {
        "TIA_RC1": "这种对象锚定方法能够根据当前情况做出正确的锚定反应。",
        "TIA_RC4": "这种对象锚定方法能够处理复杂的对象锚定任务。",
        "TIA_UP1": "这种对象锚定方法当前的工作状态对我来说始终清楚。",
    }
    for row in range(1, sheet.max_row + 1):
        item_id = sheet.cell(row, 2).value
        if item_id in block_replacements:
            sheet.cell(row, 4).value = block_replacements[item_id]
            sheet.cell(row, 6).value = (
                "The virtual content looks realistic and naturally integrated with the "
                "physical object and its surrounding environment."
            )
            sheet.cell(row, 9).value = "v5.3 情境化改写；正式使用前经认知访谈确认"
        method_id = sheet.cell(row, 1).value
        if method_id in method_replacements:
            sheet.cell(row, 4).value = method_replacements[method_id]
    sheet.cell(40, 1).value = (
        "v5.3 情境化措辞须经认知访谈确认：TIA_RC1、TIA_RC4、TIA_UP1；"
        "S-TIAS 三项仅更换测量对象，继续标为 adapted。"
    )


def _clear_inputs(workbook: Any) -> None:
    """保留平衡设计和匿名方法映射，清空全部参与者回答。"""

    participant_inputs = (
        *PARTICIPANT_BACKGROUND_COLUMNS.values(),
        "签署同意",
        "基线不适",
        "纳入分析",
        "退出/技术问题",
        "备注",
    )
    _clear_columns(workbook["Participants"], participant_inputs, 2, 25)
    _clear_columns(
        workbook["Block"],
        (*BLOCK_ITEMS.values(), *_BLOCK_STATUS_COLUMNS),
        2,
        145,
    )
    _clear_columns(
        workbook["Method"],
        (*METHOD_ITEM_COLUMNS, *_METHOD_STATUS_COLUMNS),
        2,
        49,
    )
    _clear_columns(workbook["Final"], _FINAL_INPUT_COLUMNS, 2, 25)


def _clear_columns(
    worksheet: Any,
    headers: tuple[str, ...] | Any,
    first_row: int,
    last_row: int,
) -> None:
    """按表头名清空一组输入列，避免依赖易漂移的 Excel 字母。"""

    columns = _header_columns(worksheet)
    for header in headers:
        column = columns[str(header)]
        for row in range(first_row, last_row + 1):
            worksheet.cell(row, column).value = None


def _rebuild_validations(workbook: Any) -> None:
    """为五张数据表重建与当前列顺序一致的输入校验。"""

    participants = workbook["Participants"]
    participants.data_validations.dataValidation = []
    _whole(participants, "B1_年龄", 2, 25, 1, 120)
    _list(participants, "B2_性别", 2, 25, PARTICIPANT_CATEGORIES["Gender"])
    _list(participants, "B3_主手", 2, 25, PARTICIPANT_CATEGORIES["Handedness"])
    _list(participants, "B4_视力", 2, 25, PARTICIPANT_CATEGORIES["Vision"])
    _list(participants, "B5_VR/MR经验", 2, 25, VRMR_EXPERIENCE_TEMPLATE_OPTIONS)
    _list(participants, "B6_实物MR经验", 2, 25, PARTICIPANT_CATEGORIES["PhysicalMR_Experience"])
    _list(participants, "签署同意", 2, 25, ("是", "否"))
    _list(participants, "基线不适", 2, 25, PARTICIPANT_CATEGORIES["Baseline_Discomfort"])
    _list(participants, "纳入分析", 2, 25, ("是", "否"))
    _list(participants, "退出/技术问题", 2, 25, ("无", *EXCLUSION_REASONS))

    blocks = workbook["Block"]
    blocks.data_validations.dataValidation = []
    for header in BLOCK_ITEMS.values():
        _whole(blocks, header, 2, 145, 1, 7)
    for header in _BLOCK_STATUS_COLUMNS[:3]:
        _list(blocks, header, 2, 145, ("是", "否"))

    methods = workbook["Method"]
    methods.data_validations.dataValidation = []
    for header in METHOD_ITEM_COLUMNS[:10]:
        _whole(methods, header, 2, 49, 1, 5)
    for header in METHOD_ITEM_COLUMNS[10:]:
        _whole(methods, header, 2, 49, 1, 7)
    _list(methods, "尺度切换确认", 2, 49, ("是", "否"))

    final = workbook["Final"]
    final.data_validations.dataValidation = []
    for header in ("方法选择(标签)", "信任选择(标签)"):
        _list(final, header, 2, 25, ("方法A", "方法B", "无明显偏好"))
    _whole(final, "区分信心(1-7)", 2, 25, 1, 7)
    _list(final, "结束不适", 2, 25, PARTICIPANT_CATEGORIES["End_Discomfort"])
    strength_column = get_column_letter(_header_columns(final)["偏好强度(1-7/NA)"])
    validation = DataValidation(
        type="custom",
        formula1=(
            f'OR({strength_column}2="",{strength_column}2="N/A",'
            f'AND(ISNUMBER({strength_column}2),{strength_column}2>=1,'
            f'{strength_column}2<=7,MOD({strength_column}2,1)=0))'
        ),
        allow_blank=True,
    )
    validation.error = "请输入 1--7 的整数、N/A 或留空。"
    final.add_data_validation(validation)
    validation.add(f"{strength_column}2:{strength_column}25")


def _prepare_navigation(workbook: Any) -> None:
    """收紧空白使用范围，并为长表启用表头冻结和筛选。"""

    for sheet_name, last_row in _SHEET_ROWS.items():
        worksheet = workbook[sheet_name]
        if worksheet.max_row > last_row:
            worksheet.delete_rows(last_row + 1, worksheet.max_row - last_row)
        last_column = max(_header_columns(worksheet).values()) if sheet_name != "Questionnaire" else 9
        if worksheet.max_column > last_column:
            worksheet.delete_cols(last_column + 1, worksheet.max_column - last_column)
    for sheet_name, pane in (
        ("Participants", "L2"),
        ("Block", "K2"),
        ("Method", "E2"),
        ("Final", "B2"),
    ):
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = pane
        worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"


def _require_headers(worksheet: Any, headers: tuple[str, ...] | Any) -> None:
    """要求工作表第一行包含指定列名。"""

    columns = _header_columns(worksheet)
    missing = set(map(str, headers)).difference(columns)
    if missing:
        raise ValueError(f"{worksheet.title} 缺少模板列：{sorted(missing)}")


def _header_columns(worksheet: Any) -> dict[str, int]:
    """返回第一行非空表头到一基列号的唯一映射。"""

    columns: dict[str, int] = {}
    for cell in worksheet[1]:
        if cell.value is None:
            continue
        header = str(cell.value)
        if header in columns:
            raise ValueError(f"{worksheet.title} 含重复表头：{header}")
        columns[header] = cell.column
    return columns


def _whole(
    worksheet: Any,
    header: str,
    first_row: int,
    last_row: int,
    minimum: int,
    maximum: int,
) -> None:
    """为一列添加允许空白的整数范围校验。"""

    column = get_column_letter(_header_columns(worksheet)[header])
    validation = DataValidation(
        type="whole",
        operator="between",
        formula1=str(minimum),
        formula2=str(maximum),
        allow_blank=True,
    )
    validation.error = f"请输入 {minimum}--{maximum} 的整数。"
    worksheet.add_data_validation(validation)
    validation.add(f"{column}{first_row}:{column}{last_row}")


def _list(
    worksheet: Any,
    header: str,
    first_row: int,
    last_row: int,
    values: tuple[str, ...],
) -> None:
    """为一列添加允许空白的冻结选项下拉。"""

    column = get_column_letter(_header_columns(worksheet)[header])
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "请从下拉列表中选择。"
    worksheet.add_data_validation(validation)
    validation.add(f"{column}{first_row}:{column}{last_row}")


__all__ = ["build_raw_template"]
