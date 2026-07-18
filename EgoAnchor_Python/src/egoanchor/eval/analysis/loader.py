"""Stage 2 的 XLSX-only 读取、联接和 typed 分析输入。"""

from __future__ import annotations

import hashlib
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping

import numpy as np
from openpyxl import load_workbook as load_xlsx  # type: ignore[import-untyped]

from ..contracts import SHEET_NAMES
from .exp1 import EXP1_VARIANTS, Exp1Admission, Exp1RenderSeries, Exp1Trial
from .exp2 import Exp2VariantDefinition, validate_exp2_variant_definitions
from .vcd import VCD_FULL_VARIANT_ID, VCD_SCENARIO_ID, VcdCandidate
from .windows import EventMarker, parse_event_markers


_REQUIRED_SHEETS = (
    "manifest",
    "variants",
    "completed_trials",
    "events",
    "event_payload",
    "unity_reference",
    "unity_admission",
    "unity_render",
    "qc_checks",
)
"""指标计算必须存在的 Stage 1 逻辑 sheet。"""


@dataclass(frozen=True, slots=True)
class WorkbookInput:
    """描述一个已验证的 Stage 1 workbook 输入及其 lineage。"""

    path: Path
    """输入 XLSX 的绝对路径。"""

    sha256: str
    """输入 XLSX 二进制 SHA-256。"""

    session_id: str
    """workbook 内唯一 session。"""

    row_count: int
    """已读取分析 sheet 的事实行总数。"""

    object_id: str
    """manifest 冻结对象标识。"""

    run_kind: str
    """manifest 运行类型，正式批次必须一致。"""

    protocol_version: str
    """manifest 协议版本。"""

    config_hash: str
    """manifest 整体 runtime 配置 hash。"""

    frozen_parameter_set_id: str
    """manifest 冻结参数集标识。"""

    object_model_id: str
    """manifest 目标模型标识。"""


@dataclass(frozen=True, slots=True)
class LoadedBatch:
    """保存分析层所需的全部 workbook 联接结果。"""

    inputs: tuple[WorkbookInput, ...]
    """批次输入 workbook 及 hash。"""

    trials: tuple[Exp1Trial, ...]
    """最终完成且未被后续作废的正式 trial。"""

    variant_definitions: tuple[Exp2VariantDefinition, ...]
    """所有 workbook 中去重后的八个 runtime 定义。"""

    vcd_candidates: tuple[VcdCandidate, ...]
    """完整 EgoAnchor 遮挡候选与同帧 reference 联接。"""

    trial_windows: tuple[dict[str, Any], ...]
    """供 CSV common/trial_windows 发布的 marker 窗口行。"""

    frame_rows: tuple[dict[str, Any], ...]
    """供 CSV common/frame_metrics 发布的标准化 render 行。"""

    candidate_rows: tuple[dict[str, Any], ...]
    """供 CSV common/candidate_metrics 发布的标准化 admission 行。"""


def workbook_sha256(path: Path) -> str:
    """流式计算 XLSX 文件 SHA-256。"""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _clean(value: Any) -> Any:
    """把 workbook 空文本 marker 还原为 Python 空值。"""

    return None if value in (None, "@empty-text") else value


def _text(row: Mapping[str, Any], name: str, *, required: bool = True) -> str:
    """读取 workbook 行中的稳定文本字段。"""

    value = _clean(row.get(name))
    result = "" if value is None else str(value).strip()
    if required and not result:
        raise ValueError(f"workbook 行缺少 {name}")
    return result


def _float(row: Mapping[str, Any], name: str, *, required: bool = False) -> float:
    """读取可空有限浮点字段，空值统一为 NaN。"""

    value = _clean(row.get(name))
    if value is None or value == "":
        if required:
            raise ValueError(f"workbook 行缺少 {name}")
        return float("nan")
    if isinstance(value, bool):
        raise ValueError(f"workbook 字段 {name} 不能是 bool")
    result = float(value)
    if required and not math.isfinite(result):
        raise ValueError(f"workbook 字段 {name} 必须有限")
    return result


def _int(row: Mapping[str, Any], name: str, *, required: bool = False) -> int | None:
    """读取可空整数，并拒绝非整数浮点。"""

    value = _clean(row.get(name))
    if value is None or value == "":
        if required:
            raise ValueError(f"workbook 行缺少 {name}")
        return None
    if isinstance(value, bool):
        raise ValueError(f"workbook 字段 {name} 不能是 bool")
    result = int(value)
    if float(value) != result:
        raise ValueError(f"workbook 字段 {name} 不是整数")
    return result


def _bool(row: Mapping[str, Any], name: str, *, required: bool = True) -> bool:
    """读取严格布尔字段，兼容 Excel 回读的 bool 和规范文本。"""

    value = _clean(row.get(name))
    if isinstance(value, bool):
        return value
    if isinstance(value, str) and value.lower() in {"true", "false"}:
        return value.lower() == "true"
    if value is None and not required:
        return False
    raise ValueError(f"workbook 字段 {name} 不是 bool")


def _read_rows(workbook: Any, logical_name: str, physical_names: Mapping[str, tuple[str, ...]]) -> list[dict[str, Any]]:
    """读取一个逻辑 sheet 的全部物理分片并返回行映射。"""

    rows: list[dict[str, Any]] = []
    for physical_name in physical_names.get(logical_name, (logical_name,)):
        if physical_name not in workbook.sheetnames:
            raise ValueError(f"workbook 缺少物理 sheet：{physical_name}")
        iterator = workbook[physical_name].iter_rows(values_only=True)
        try:
            header = tuple(next(iterator))
        except StopIteration:
            continue
        if any(value is None for value in header) or len(set(header)) != len(header):
            raise ValueError(f"workbook sheet 表头非法：{physical_name}")
        for values in iterator:
            if all(value is None for value in values):
                continue
            if len(values) > len(header):
                raise ValueError(f"workbook 行列数超过表头：{physical_name}")
            padded = tuple(values) + (None,) * (len(header) - len(values))
            rows.append({str(key): _clean(value) for key, value in zip(header, padded, strict=True)})
    return rows


def _physical_sheet_map(workbook: Any) -> dict[str, tuple[str, ...]]:
    """从 sheet_index 构造逻辑 sheet 到物理分片的映射。"""

    if "sheet_index" not in workbook.sheetnames:
        return {name: (name,) for name in SHEET_NAMES}
    rows = _read_rows(workbook, "sheet_index", {"sheet_index": ("sheet_index",)})
    mapping: dict[str, list[str]] = {}
    for row in rows:
        logical = _text(row, "logical_sheet")
        physical = _text(row, "physical_sheet")
        mapping.setdefault(logical, []).append(physical)
    return {key: tuple(value) for key, value in mapping.items()}


def _array(rows: Iterable[Mapping[str, Any]], names: tuple[str, ...]) -> np.ndarray:
    """将多列标量转换为二维 double 数组。"""

    return np.asarray([[_float(row, name) for name in names] for row in rows], dtype=np.float64)


def _vector3(row: Mapping[str, Any], names: tuple[str, str, str]) -> tuple[float, float, float]:
    """读取三维位置并返回固定长度 tuple，便于 typed pose 联接。"""

    return (_float(row, names[0]), _float(row, names[1]), _float(row, names[2]))


def _marker_groups(
    events: Iterable[Mapping[str, Any]],
    payload: Iterable[Mapping[str, Any]],
) -> dict[tuple[str, str, str, str], tuple[EventMarker, ...]]:
    """按 trial 分组解析显式 marker。"""

    markers = parse_event_markers(events, payload)
    grouped: dict[tuple[str, str, str, str], list[EventMarker]] = {}
    for marker in markers:
        grouped.setdefault(marker.trial_key, []).append(marker)
    return {key: tuple(value) for key, value in grouped.items()}


def _final_trials(
    completed_rows: Iterable[Mapping[str, Any]],
    events: Iterable[Mapping[str, Any]],
) -> dict[tuple[str, str, str, str], tuple[Mapping[str, Any], float]]:
    """依据 lifecycle event 投影最终完成且未后续作废的 trial。"""

    lifecycle: dict[tuple[str, str, str, str], list[tuple[str, float]]] = {}
    for row in events:
        event = _text(row, "event", required=False)
        if event not in {"trial_ended", "trial_rejected"}:
            continue
        key = (
            _text(row, "session_id"),
            _text(row, "experiment_id"),
            _text(row, "scenario_id"),
            _text(row, "trial_id"),
        )
        lifecycle.setdefault(key, []).append((event, _float(row, "mono_ms", required=True)))
    result: dict[tuple[str, str, str, str], tuple[Mapping[str, Any], float]] = {}
    for row in completed_rows:
        key = (
            _text(row, "session_id"),
            _text(row, "experiment_id"),
            _text(row, "scenario_id"),
            _text(row, "trial_id"),
        )
        records = lifecycle.get(key, [])
        ended = [time for event, time in records if event == "trial_ended"]
        rejected = [time for event, time in records if event == "trial_rejected"]
        if len(ended) != 1 or any(time > ended[0] for time in rejected):
            raise ValueError(f"completed_trials 与 lifecycle 不一致：{key}")
        result[key] = (row, ended[0])
    return result


def _render_series(
    rows: Iterable[Mapping[str, Any]],
    key: tuple[str, str, str, str],
) -> tuple[Exp1RenderSeries, ...]:
    """将一个 final trial 的 render 行分组为四条系统序列。"""

    groups: dict[str, list[Mapping[str, Any]]] = {}
    for row in rows:
        row_key = (
            _text(row, "session_id", required=False),
            _text(row, "experiment_id", required=False),
            _text(row, "scenario_id", required=False),
            _text(row, "trial_id", required=False),
        )
        if row_key == key:
            groups.setdefault(_text(row, "variant_id"), []).append(row)
    series: list[Exp1RenderSeries] = []
    for variant_id, group in groups.items():
        group.sort(key=lambda row: (_float(row, "render_mono_ms", required=True), _int(row, "render_tick_id", required=True) or -1))
        if len(group) < 2:
            continue
        series.append(
            Exp1RenderSeries(
                variant_id=variant_id,
                render_tick_ids=np.asarray([_int(row, "render_tick_id", required=True) for row in group], dtype=np.int64),
                times_ms=np.asarray([_float(row, "render_mono_ms", required=True) for row in group]),
                head_positions_m=_array(group, ("head_pos_x_m", "head_pos_y_m", "head_pos_z_m")),
                head_rotations=_array(group, ("head_rot_x", "head_rot_y", "head_rot_z", "head_rot_w")),
                reference_pose_valid=np.asarray([_bool(row, "reference_pose_valid") for row in group], dtype=np.bool_),
                reference_positions_m=_array(group, ("reference_pos_x_m", "reference_pos_y_m", "reference_pos_z_m")),
                reference_rotations=_array(group, ("reference_rot_x", "reference_rot_y", "reference_rot_z", "reference_rot_w")),
                reference_linear_speed_m_s=np.asarray([_float(row, "reference_linear_speed_m_s") for row in group]),
                reference_angular_speed_deg_s=np.asarray([_float(row, "reference_angular_speed_deg_s") for row in group]),
                has_output_pose=np.asarray([_bool(row, "has_output_pose") for row in group], dtype=np.bool_),
                has_source_capture_timing=np.asarray([_bool(row, "has_source_capture_timing") for row in group], dtype=np.bool_),
                source_capture_mono_ms=np.asarray([_float(row, "source_capture_mono_ms") for row in group]),
                has_display_pose=np.asarray([_bool(row, "has_display_pose") for row in group], dtype=np.bool_),
                display_positions_m=_array(group, ("display_pos_x_m", "display_pos_y_m", "display_pos_z_m")),
                display_rotations=_array(group, ("display_rot_x", "display_rot_y", "display_rot_z", "display_rot_w")),
                latest_static_locked=np.asarray([_bool(row, "latest_static_locked") for row in group], dtype=np.bool_),
            )
        )
    return tuple(sorted(series, key=lambda item: item.variant_id))


def load_workbook(path: Path) -> tuple[WorkbookInput, LoadedBatch]:
    """读取一个完整 XLSX，并返回来源信息与单 workbook batch。"""

    normalized = path.expanduser().resolve()
    if normalized.suffix.lower() != ".xlsx":
        raise ValueError(f"Stage 2 只接受 XLSX 输入：{path}")
    if not normalized.is_file():
        raise FileNotFoundError(f"XLSX 输入不存在：{normalized}")
    workbook = load_xlsx(normalized, read_only=True, data_only=True)
    try:
        missing = [name for name in _REQUIRED_SHEETS if name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"workbook 缺少必需 sheet：{', '.join(missing)}")
        physical = _physical_sheet_map(workbook)
        manifest = _read_rows(workbook, "manifest", physical)
        if len(manifest) != 1:
            raise ValueError("manifest 必须恰好一行")
        session_id = _text(manifest[0], "session_id")
        variants = _read_rows(workbook, "variants", physical)
        events = _read_rows(workbook, "events", physical)
        payload = _read_rows(workbook, "event_payload", physical)
        completed = _read_rows(workbook, "completed_trials", physical)
        references = _read_rows(workbook, "unity_reference", physical)
        admissions = _read_rows(workbook, "unity_admission", physical)
        renders = _read_rows(workbook, "unity_render", physical)
        qc_checks = _read_rows(workbook, "qc_checks", physical)
        failed_qc = [
            _text(row, "check_id")
            for row in qc_checks
            if _text(row, "status", required=False).lower() not in {"pass", "passed"}
            and _text(row, "severity", required=False).lower() == "error"
        ]
        if failed_qc:
            raise ValueError(f"Stage 1 workbook QC 未通过：{failed_qc}")
        final = _final_trials(completed, events)
        marker_groups = _marker_groups(events, payload)
        workbook_digest = workbook_sha256(normalized)

        trials: list[Exp1Trial] = []
        trial_windows: list[dict[str, Any]] = []
        frame_rows: list[dict[str, Any]] = []
        candidate_rows: list[dict[str, Any]] = []
        vcd_candidates: list[VcdCandidate] = []
        reference_by_frame = {
            (_text(row, "session_id"), _int(row, "frame_id", required=True)): row for row in references
        }
        for key, (completed_row, trial_end_ms) in final.items():
            session, experiment, scenario, trial_id = key
            if session != session_id:
                raise ValueError("workbook 内 session_id 不一致")
            markers = marker_groups.get(key, ())
            if not markers:
                raise ValueError(f"完成 trial 缺少 marker：{key}")
            render_series = _render_series(renders, key)
            if not set(EXP1_VARIANTS).issubset({item.variant_id for item in render_series}):
                raise ValueError(f"完成 trial 缺少实验一四系统 render：{key}")
            trial_admissions = [
                row
                for row in admissions
                if (
                    _text(row, "session_id", required=False),
                    _text(row, "experiment_id", required=False),
                    _text(row, "scenario_id", required=False),
                    _text(row, "trial_id", required=False),
                ) == key
            ]
            trials.append(
                Exp1Trial(
                    session_id=session,
                    experiment_id=experiment,
                    scenario_id=scenario,
                    trial_id=trial_id,
                    condition_id=_text(completed_row, "condition_id"),
                    workbook_sha256=workbook_digest,
                    trial_end_ms=trial_end_ms,
                    markers=markers,
                    render_series=render_series,
                    admissions=tuple(
                        Exp1Admission(
                            candidate_id=_text(row, "candidate_id"),
                            variant_id=_text(row, "variant_id"),
                            source_capture_mono_ms=_float(row, "source_capture_mono_ms", required=True),
                            admission_decision=_text(row, "admission_decision"),
                        )
                        for row in trial_admissions
                        if _text(row, "candidate_id", required=False)
                    ),
                )
            )
            for marker in markers:
                trial_windows.append(
                    {
                        "session_id": session,
                        "experiment_id": experiment,
                        "scenario_id": scenario,
                        "trial_id": trial_id,
                        "event_id": marker.event_id,
                        "condition_id": _text(completed_row, "condition_id"),
                        "variant_id": "",
                        "metric_key": "event_window",
                        "metric_value": None,
                        "metric_unit": "",
                        "aggregation_level": "event_window",
                        "input_workbook_sha256": workbook_digest,
                    }
                )
            for row in renders:
                row_key = (
                    _text(row, "session_id", required=False),
                    _text(row, "experiment_id", required=False),
                    _text(row, "scenario_id", required=False),
                    _text(row, "trial_id", required=False),
                )
                if row_key == key:
                    frame_rows.append({**row, "input_workbook_sha256": workbook_digest})
            for row in trial_admissions:
                if _text(row, "candidate_id", required=False):
                    candidate_rows.append({**row, "input_workbook_sha256": workbook_digest})
                    if scenario == VCD_SCENARIO_ID and _text(row, "variant_id") == VCD_FULL_VARIANT_ID:
                        frame_id = _int(row, "frame_id", required=True)
                        if frame_id is None:
                            raise ValueError("VCD admission 缺少 frame_id")
                        reference = reference_by_frame.get((session, frame_id))
                        vcd_candidates.append(
                            VcdCandidate(
                                session_id=session,
                                scenario_id=scenario,
                                trial_id=trial_id,
                                candidate_id=_text(row, "candidate_id"),
                                frame_id=frame_id,
                                source_capture_mono_ms=_float(row, "source_capture_mono_ms", required=True),
                                variant_id=_text(row, "variant_id"),
                                admission_decision=_text(row, "admission_decision"),
                                vcd_score=_float(row, "vcd_score", required=True),
                                has_aligned_raw=_bool(row, "has_aligned_raw"),
                                aligned_raw_position_m=_vector3(row, ("aligned_raw_pos_x_m", "aligned_raw_pos_y_m", "aligned_raw_pos_z_m")),
                                reference_frame_id=frame_id if reference is not None else None,
                                reference_session_id=session if reference is not None else None,
                                reference_pose_valid=_bool(reference, "reference_pose_valid") if reference is not None else None,
                                reference_position_m=_vector3(reference, ("reference_pos_x_m", "reference_pos_y_m", "reference_pos_z_m")) if reference is not None else (float("nan"), float("nan"), float("nan")),
                                input_workbook_sha256=workbook_digest,
                            )
                        )
        definitions = tuple(
            Exp2VariantDefinition(
                variant_id=_text(row, "variant_id"),
                uses_capture_time_alignment=_bool(row, "uses_capture_time_alignment"),
                uses_vcd_admission=_bool(row, "uses_vcd_admission"),
                uses_temporal_synthesis=_bool(row, "uses_temporal_synthesis"),
                uses_static_lock=_bool(row, "uses_static_lock"),
                uses_low_score_reacquire=_bool(row, "uses_low_score_reacquire"),
                uses_server_reacquire=_bool(row, "uses_server_reacquire"),
                world_alignment_mode=_text(row, "world_alignment_mode"),
                quality_gate=_text(row, "quality_gate"),
                motion_model=_text(row, "motion_model"),
                smoothing_strategy=_text(row, "smoothing_strategy"),
            )
            for row in variants
        )
        validate_exp2_variant_definitions(definitions)
        row_count = sum(
            len(rows)
            for rows in (
                manifest,
                variants,
                events,
                payload,
                completed,
                references,
                admissions,
                renders,
                qc_checks,
            )
        )
        source = WorkbookInput(
            normalized,
            workbook_digest,
            session_id,
            row_count,
            _text(manifest[0], "object_id"),
            _text(manifest[0], "run_kind"),
            _text(manifest[0], "protocol_version"),
            _text(manifest[0], "config_hash"),
            _text(manifest[0], "frozen_parameter_set_id"),
            _text(manifest[0], "object_model_id"),
        )
        loaded = LoadedBatch(
            inputs=(source,),
            trials=tuple(sorted(trials, key=lambda item: (item.session_id, item.trial_id))),
            variant_definitions=definitions,
            vcd_candidates=tuple(vcd_candidates),
            trial_windows=tuple(trial_windows),
            frame_rows=tuple(frame_rows),
            candidate_rows=tuple(candidate_rows),
        )
        return source, loaded
    finally:
        workbook.close()


def load_workbook_batch(paths: Iterable[Path]) -> LoadedBatch:
    """读取 XLSX 批次并拒绝重复 session 或不一致 runtime 定义。"""

    normalized_paths = tuple(
        sorted(
            (Path(path) for path in paths),
            key=lambda path: str(path.expanduser().resolve()).casefold(),
        )
    )
    if not normalized_paths:
        raise ValueError("Stage 2 workbook 批次不能为空")
    loaded_batches = [load_workbook(path) for path in normalized_paths]
    sessions = [source.session_id for source, _ in loaded_batches]
    if len(sessions) != len(set(sessions)):
        raise ValueError(f"分析批次包含重复 session_id：{sessions}")
    definitions = loaded_batches[0][1].variant_definitions
    reference = loaded_batches[0][0]
    batch_signature = (
        reference.object_id,
        reference.run_kind,
        reference.protocol_version,
        reference.config_hash,
        reference.frozen_parameter_set_id,
        reference.object_model_id,
    )
    for _, batch in loaded_batches[1:]:
        source = batch.inputs[0]
        if (
            source.object_id,
            source.run_kind,
            source.protocol_version,
            source.config_hash,
            source.frozen_parameter_set_id,
            source.object_model_id,
        ) != batch_signature:
            raise ValueError("分析批次的 manifest 对象、协议、配置或参数集不一致")
        if batch.variant_definitions != definitions:
            raise ValueError("分析批次的 runtime variant 定义不一致")
    return LoadedBatch(
        inputs=tuple(source for source, _ in loaded_batches),
        trials=tuple(trial for _, batch in loaded_batches for trial in batch.trials),
        variant_definitions=definitions,
        vcd_candidates=tuple(candidate for _, batch in loaded_batches for candidate in batch.vcd_candidates),
        trial_windows=tuple(row for _, batch in loaded_batches for row in batch.trial_windows),
        frame_rows=tuple(row for _, batch in loaded_batches for row in batch.frame_rows),
        candidate_rows=tuple(row for _, batch in loaded_batches for row in batch.candidate_rows),
    )


__all__ = ["LoadedBatch", "WorkbookInput", "load_workbook", "load_workbook_batch", "workbook_sha256"]
