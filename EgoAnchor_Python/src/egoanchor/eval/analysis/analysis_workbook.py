"""Stage 2 CSV 同源审阅工作簿的确定性构建与回读验证。"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import hashlib
import json
from pathlib import Path
import time
from typing import Callable, Mapping, Sequence
import xml.etree.ElementTree as ET
import zipfile

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from ..contracts import CSV_TABLE_CONTRACTS, CsvTableContract


_FIXED_DATETIME = datetime(1980, 1, 1, 0, 0, 0)
"""XLSX core properties 使用的固定时间。"""

_HEADER_FILL = PatternFill("solid", fgColor="2F6F6D")
"""审阅表头的低饱和青绿色填充。"""

_HEADER_FONT = Font(name="Arial", size=9, bold=True, color="FFFFFF")
"""审阅表头字体。"""

_BODY_FONT = Font(name="Arial", size=9, color="202020")
"""审阅数据字体。"""

_HEADER_BORDER = Border(bottom=Side(style="thin", color="8FA8A7"))
"""表头与数据之间的细分隔线。"""

_WINDOWS_RETRY_ATTEMPTS = 6
"""Windows 共享锁释放的最多尝试次数。"""

_WINDOWS_RETRY_DELAY_SECONDS = 0.1
"""Windows 共享锁重试的基础等待时间。"""


@dataclass(frozen=True, slots=True)
class AnalysisSheetSpec:
    """描述分析工作簿中一个 CSV 来源 sheet。"""

    sheet_name: str
    """不超过 31 字符的稳定工作表名。"""

    source_csv: str
    """相对于 Stage 2 根目录的 CSV 路径。"""

    table_name: str
    """复用列类型的 CSV 逻辑契约名。"""

    row_filter: Callable[[Mapping[str, str]], bool] | None = None
    """可选的实验投影筛选函数。"""


def _is_exp1(row: Mapping[str, str]) -> bool:
    """判断 paper 行是否属于实验一。"""

    return row.get("experiment") == "exp1_system_characterization"


def _is_exp2(row: Mapping[str, str]) -> bool:
    """判断 paper 行是否属于实验二。"""

    return row.get("experiment") == "exp2_design_attribution"


def _is_exp1_lineage(row: Mapping[str, str]) -> bool:
    """保留实验一、共享窗口和实验一 paper 行的 lineage。"""

    output = str(row.get("output_path") or "")
    key = str(row.get("source_row_key") or "")
    return (
        output.startswith("exp1/")
        or output.startswith("plots/exp1_")
        or output == "common/trial_windows.csv"
        or (output.startswith("paper/") and "experiment=exp1_system_characterization" in key)
    )


def _is_exp2_lineage(row: Mapping[str, str]) -> bool:
    """保留实验二、共享窗口和实验二 paper 行的 lineage。"""

    output = str(row.get("output_path") or "")
    key = str(row.get("source_row_key") or "")
    return (
        output.startswith("exp2/")
        or output.startswith("plots/exp2_")
        or output == "common/trial_windows.csv"
        or (output.startswith("paper/") and "experiment=exp2_design_attribution" in key)
    )


_EXP1_SHEETS = (
    AnalysisSheetSpec("inputs", "audit/inputs.csv", "inputs"),
    AnalysisSheetSpec("analysis_qc", "audit/analysis_qc.csv", "analysis_qc"),
    AnalysisSheetSpec("metric_catalog", "audit/metric_catalog.csv", "metric_catalog"),
    AnalysisSheetSpec("trial_windows", "common/trial_windows.csv", "trial_windows"),
    AnalysisSheetSpec("event_metrics", "exp1/event_metrics.csv", "event_metrics"),
    AnalysisSheetSpec("trial_metrics", "exp1/trial_metrics.csv", "trial_metrics"),
    AnalysisSheetSpec("session_metrics", "exp1/session_metrics.csv", "session_metrics"),
    AnalysisSheetSpec("scenario_summary", "exp1/scenario_summary.csv", "scenario_summary"),
    AnalysisSheetSpec("head_motion_trace", "plots/exp1_head_motion_trace.csv", "exp1_head_motion_trace"),
    AnalysisSheetSpec("start_stop_trace", "plots/exp1_start_stop_trace.csv", "exp1_start_stop_trace"),
    AnalysisSheetSpec("lag_tradeoff", "plots/exp1_lag_tradeoff.csv", "exp1_lag_tradeoff"),
    AnalysisSheetSpec("occlusion_trace", "plots/exp1_occlusion_trace.csv", "exp1_occlusion_trace"),
    AnalysisSheetSpec("paper_numbers", "paper/numbers.csv", "numbers", _is_exp1),
    AnalysisSheetSpec("paper_tables", "paper/tables.csv", "tables", _is_exp1),
    AnalysisSheetSpec("lineage", "audit/lineage.csv", "lineage", _is_exp1_lineage),
)
"""实验一审阅工作簿的稳定 sheet 顺序和 CSV 来源。"""

_EXP2_SHEETS = (
    AnalysisSheetSpec("inputs", "audit/inputs.csv", "inputs"),
    AnalysisSheetSpec("analysis_qc", "audit/analysis_qc.csv", "analysis_qc"),
    AnalysisSheetSpec("metric_catalog", "audit/metric_catalog.csv", "metric_catalog"),
    AnalysisSheetSpec("trial_windows", "common/trial_windows.csv", "trial_windows"),
    AnalysisSheetSpec("event_metrics", "exp2/event_metrics.csv", "event_metrics"),
    AnalysisSheetSpec("trial_metrics", "exp2/trial_metrics.csv", "trial_metrics"),
    AnalysisSheetSpec("session_metrics", "exp2/session_metrics.csv", "session_metrics"),
    AnalysisSheetSpec("paired_deltas", "exp2/paired_deltas.csv", "paired_deltas"),
    AnalysisSheetSpec("paired_summary", "exp2/paired_summary.csv", "paired_summary"),
    AnalysisSheetSpec("vcd_risk_points", "exp2/vcd_risk_points.csv", "vcd_risk_points"),
    AnalysisSheetSpec("vcd_curve", "exp2/vcd_curve.csv", "vcd_curve"),
    AnalysisSheetSpec("vcd_aurc", "exp2/vcd_aurc.csv", "vcd_aurc"),
    AnalysisSheetSpec("mechanism_attribution", "plots/exp2_mechanism_attribution.csv", "exp2_mechanism_attribution"),
    AnalysisSheetSpec("vcd_plot", "plots/exp2_vcd_curve.csv", "exp2_vcd_curve"),
    AnalysisSheetSpec("paper_numbers", "paper/numbers.csv", "numbers", _is_exp2),
    AnalysisSheetSpec("paper_tables", "paper/tables.csv", "tables", _is_exp2),
    AnalysisSheetSpec("lineage", "audit/lineage.csv", "lineage", _is_exp2_lineage),
)
"""实验二审阅工作簿的稳定 sheet 顺序和 CSV 来源。"""


def _contracts() -> dict[str, CsvTableContract]:
    """返回 CSV 逻辑名到列契约的映射。"""

    return {contract.name: contract for contract in CSV_TABLE_CONTRACTS}


def _safe_text(value: str) -> str:
    """返回原始文本；调用方以显式 string 类型写入，避免 Excel 公式解析。"""

    return value


def _typed_value(value: str, dtype: str) -> object:
    """按 CSV 列契约把文本回读为 XLSX 单元格类型。"""

    if value == "":
        return None
    if dtype == "int":
        return int(value)
    if dtype == "float":
        return float(value)
    if dtype == "bool":
        lowered = value.lower()
        if lowered not in {"true", "false"}:
            raise ValueError(f"CSV 布尔值非法：{value}")
        return lowered == "true"
    return _safe_text(value)


def _header_sha256(header: Sequence[str]) -> str:
    """返回稳定列名序列的 SHA-256。"""

    encoded = json.dumps(list(header), ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _write_table_sheet(
    workbook: Workbook,
    spec: AnalysisSheetSpec,
    header: tuple[str, ...],
    rows: Sequence[Mapping[str, str]],
) -> None:
    """写入一个带类型、冻结表头和稳定列宽的数据 sheet。"""

    contract = _contracts()[spec.table_name]
    columns = {column.name: column for column in contract.columns}
    if header != contract.column_names():
        raise ValueError(f"分析 XLSX 来源 CSV 表头不符合契约：{spec.source_csv}")
    sheet = workbook.create_sheet(spec.sheet_name)
    sheet.sheet_view.showGridLines = False
    sheet.append(list(header))
    for row_index, row in enumerate(rows, start=2):
        for column_index, name in enumerate(header, start=1):
            value = _typed_value(str(row.get(name) or ""), columns[name].dtype)
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            if columns[name].dtype == "text":
                cell.data_type = "s"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 22
    for column_index, name in enumerate(header, 1):
        width = len(name) + 2
        for cell in tuple(sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2))[0]:
            cell.font = _BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 40))
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(width, 10), 40)


def _write_metadata_sheet(workbook: Workbook, name: str, header: tuple[str, ...], rows: Sequence[Sequence[object]]) -> None:
    """写入工作簿说明或 sheet 索引。"""

    sheet = workbook.create_sheet(name)
    sheet.sheet_view.showGridLines = False
    sheet.append(list(header))
    for row in rows:
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER
    for column_index in range(1, len(header) + 1):
        values = [sheet.cell(row, column_index).value for row in range(1, sheet.max_row + 1)]
        width = min(max(max(len(str(value or "")) for value in values) + 2, 12), 64)
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def _normalize_core_properties(data: bytes) -> bytes:
    """固定 OpenPyXL 自动写入的 core 创建和修改时间。"""

    namespace = "http://purl.org/dc/terms/"
    ET.register_namespace("dcterms", namespace)
    root = ET.fromstring(data)
    for tag in ("created", "modified"):
        node = root.find(f"{{{namespace}}}{tag}")
        if node is not None:
            node.text = "1980-01-01T00:00:00Z"
    return ET.tostring(root, encoding="utf-8", xml_declaration=False)


def _normalize_archive(source: Path, destination: Path) -> None:
    """固定 XLSX ZIP 条目顺序和时间戳。"""

    timestamp = (1980, 1, 1, 0, 0, 0)
    with zipfile.ZipFile(source, "r") as input_archive, zipfile.ZipFile(
        destination,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as output_archive:
        for original in sorted(input_archive.infolist(), key=lambda item: item.filename):
            normalized = zipfile.ZipInfo(original.filename, timestamp)
            normalized.compress_type = zipfile.ZIP_DEFLATED
            normalized.external_attr = original.external_attr
            normalized.create_system = original.create_system
            normalized.flag_bits = original.flag_bits
            data = input_archive.read(original.filename)
            if original.filename == "docProps/core.xml":
                data = _normalize_core_properties(data)
            output_archive.writestr(normalized, data)


def _unlink_with_retry(path: Path) -> None:
    """删除 OpenPyXL 临时文件，并对 Windows 短暂共享锁有界重试。"""

    for attempt in range(_WINDOWS_RETRY_ATTEMPTS):
        try:
            path.unlink(missing_ok=True)
            return
        except PermissionError:
            if attempt + 1 == _WINDOWS_RETRY_ATTEMPTS:
                raise
            time.sleep(_WINDOWS_RETRY_DELAY_SECONDS * (attempt + 1))


def _validate_workbook(
    path: Path,
    specs: Sequence[AnalysisSheetSpec],
    expected_rows: Mapping[str, int],
    expected_data: Mapping[str, Sequence[Mapping[str, str]]],
) -> None:
    """独立回读工作簿，拒绝 sheet、行数、单元格或公式偏差。"""

    expected_names = ["workbook_info", "sheet_index", *(spec.sheet_name for spec in specs)]
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != expected_names:
            raise ValueError("分析 XLSX sheet 集合或顺序不符合契约")
        for spec in specs:
            sheet = workbook[spec.sheet_name]
            contract = _contracts()[spec.table_name]
            header = tuple(cell.value for cell in sheet[1])
            if header != contract.column_names():
                raise ValueError(f"分析 XLSX 表头回读不一致：{spec.sheet_name}")
            if sheet.max_row - 1 != expected_rows[spec.sheet_name]:
                raise ValueError(f"分析 XLSX 行数回读不一致：{spec.sheet_name}")
            columns = {column.name: column for column in _contracts()[spec.table_name].columns}
            for row_index, expected in enumerate(expected_data[spec.sheet_name], start=2):
                expected_values = tuple(
                    _typed_value(str(expected.get(name) or ""), columns[name].dtype)
                    for name in contract.column_names()
                )
                actual_values = tuple(
                    sheet.cell(row=row_index, column=column_index).value
                    for column_index in range(1, len(contract.column_names()) + 1)
                )
                if actual_values != expected_values:
                    raise ValueError(f"分析 XLSX 单元格回读不一致：{spec.sheet_name}/{row_index}")
        for sheet in workbook.worksheets:
            if any(cell.data_type == "f" for row in sheet.iter_rows() for cell in row):
                raise ValueError(f"分析 XLSX 不得包含公式：{sheet.title}")
    finally:
        workbook.close()


def _write_analysis_workbook(
    root: Path,
    csv_sha256: Mapping[str, str],
    finalized_rows: Mapping[str, Sequence[Mapping[str, str]]],
    *,
    specs: Sequence[AnalysisSheetSpec],
    experiment_id: str,
    output_name: str,
    code_version: str,
    parameter_set_id: str,
) -> Path:
    """从 finalized rows 构建一个确定性的实验审阅工作簿。"""

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "egoanchor.eval"
    workbook.properties.created = _FIXED_DATETIME
    workbook.properties.modified = _FIXED_DATETIME
    source_data: list[tuple[AnalysisSheetSpec, tuple[str, ...], list[dict[str, str]], str]] = []
    for spec in specs:
        source_key = spec.source_csv.removesuffix(".csv")
        logical_name = Path(source_key).name
        raw_rows = finalized_rows.get(source_key)
        if raw_rows is None:
            raw_rows = finalized_rows.get(logical_name)
        if raw_rows is None:
            raise ValueError(f"缺少同源 Stage 2 表行：{source_key}")
        header = _contracts()[spec.table_name].column_names()
        rows = [dict(row) for row in raw_rows]
        if spec.row_filter is not None:
            rows = [row for row in rows if spec.row_filter(row)]
        table_key = spec.source_csv.removesuffix(".csv")
        source_hash = csv_sha256.get(table_key)
        if source_hash is None:
            logical = Path(spec.source_csv).stem
            matches = [value for key, value in csv_sha256.items() if key.endswith(f"/{logical}") or key == logical]
            if len(matches) != 1:
                raise ValueError(f"缺少同源 Stage 2 表 hash：{source_key}")
            source_hash = matches[0]
        source_data.append((spec, header, rows, source_hash))
    _write_metadata_sheet(
        workbook,
        "workbook_info",
        ("key", "value"),
        (
            ("experiment_id", experiment_id),
            ("contract", "analysis_workbook-v1"),
            ("code_version", code_version),
            ("parameter_set_id", parameter_set_id),
            ("source_sheet_count", len(source_data)),
        ),
    )
    _write_metadata_sheet(
        workbook,
        "sheet_index",
        ("sheet_name", "source_csv", "source_csv_sha256", "row_count", "header_sha256"),
        tuple(
            (spec.sheet_name, spec.source_csv, source_hash, len(rows), _header_sha256(header))
            for spec, header, rows, source_hash in source_data
        ),
    )
    for spec, header, rows, _ in source_data:
        _write_table_sheet(workbook, spec, header, rows)
    stem = Path(output_name).stem
    raw = root / f".{stem}.raw.xlsx"
    destination = root / output_name
    try:
        workbook.save(raw)
    finally:
        workbook.close()
    _normalize_archive(raw, destination)
    _unlink_with_retry(raw)
    _validate_workbook(
        destination,
        specs,
        {spec.sheet_name: len(rows) for spec, _, rows, _ in source_data},
        {spec.sheet_name: rows for spec, _, rows, _ in source_data},
    )
    return destination


def write_analysis_workbooks(
    root: Path,
    csv_sha256: Mapping[str, str],
    finalized_rows: Mapping[str, Sequence[Mapping[str, str]]],
    *,
    code_version: str,
    parameter_set_id: str,
) -> dict[str, str]:
    """在 Stage 2 staging 根目录写入并回读当前实验审阅工作簿。"""

    paths = {}
    exp1_keys = ("exp1/event_metrics", "exp1/scenario_summary")
    exp2_keys = ("exp2/event_metrics", "exp2/paired_deltas", "exp2/paired_summary")
    has_exp2 = any(finalized_rows.get(key) for key in exp2_keys)
    if any(finalized_rows.get(key) for key in exp1_keys) or not has_exp2:
        exp1_path = _write_analysis_workbook(
            root,
            csv_sha256,
            finalized_rows,
            specs=_EXP1_SHEETS,
            experiment_id="exp1_system_characterization",
            output_name="exp1_analysis.xlsx",
            code_version=code_version,
            parameter_set_id=parameter_set_id,
        )
        paths[exp1_path.name] = hashlib.sha256(exp1_path.read_bytes()).hexdigest()
    if has_exp2:
        exp2_path = _write_analysis_workbook(
            root,
            csv_sha256,
            finalized_rows,
            specs=_EXP2_SHEETS,
            experiment_id="exp2_design_attribution",
            output_name="exp2_analysis.xlsx",
            code_version=code_version,
            parameter_set_id=parameter_set_id,
        )
        paths[exp2_path.name] = hashlib.sha256(exp2_path.read_bytes()).hexdigest()
    if not paths:
        raise ValueError("Stage 2 没有可发布的实验审阅工作簿")
    return paths


__all__ = ["AnalysisSheetSpec", "write_analysis_workbooks"]
