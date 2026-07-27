"""实验三三段堆叠原始工作簿的严格只读 reader。"""

from __future__ import annotations

import hashlib
from collections.abc import Iterable
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]

from .contracts import (
    BLOCK_ITEMS,
    EGOANCHOR,
    Exp3Data,
    METHOD_ITEM_COLUMNS,
    METHODS,
    OBJECTS,
    ONE_EURO,
    WORKBOOK_CONTRACT_ID,
    WORKBOOK_SOURCE_CATEGORY,
    required_block_items,
)


_REQUIRED_SHEETS = frozenset({"README", "Participants", "Records"})
"""分析输入必须包含的最小工作表集合。"""

_PARTICIPANT_DESIGN_COLUMNS = (
    "Participant_ID",
    "平衡单元",
    "物体排列ID",
    "标签序列",
    "物体1",
    "物体2",
    "物体3",
    "区块内标签顺序",
    "方法A=（保密）",
    "方法B=（保密）",
    "先行实际方法",
)
"""24 平衡单元不得缺失或被公式替换的设计列。"""

_BLOCK_DESIGN_COLUMNS = (
    "Participant_ID",
    "Block_Index",
    "平衡单元",
    "物体位置",
    "物体",
    "Object_Key",
    "Shown_Label",
    "Condition(保密)",
    "物体内先后",
    "该方法第几次",
)
"""Records A 段的固定设计列。"""

_METHOD_DESIGN_COLUMNS = (
    "Participant_ID",
    "Rating_Order",
    "Shown_Label",
    "Condition(保密)",
)
"""Records B 段的固定设计列。"""


def workbook_sha256(path: Path) -> str:
    """返回原始工作簿的 SHA-256。"""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while block := handle.read(1024 * 1024):
            digest.update(block)
    return digest.hexdigest()


def read_workbook(path: Path) -> Exp3Data:
    """按值读取 Participants 与 Records 三段，并验证结构身份。"""

    source = path.expanduser().resolve()
    if source.suffix.lower() != ".xlsx" or not source.is_file():
        raise FileNotFoundError(f"实验三输入必须是现存 XLSX：{source}")
    workbook = load_workbook(source, read_only=False, data_only=False)
    try:
        missing = _REQUIRED_SHEETS.difference(workbook.sheetnames)
        if missing:
            raise ValueError(f"实验三工作簿缺少工作表：{sorted(missing)}")
        participants = _read_table(workbook["Participants"], 2, 3, 26)
        records = workbook["Records"]
        block_header = _find_header_after(records, "A. 区块记录")
        method_header = _find_header_after(records, "B. 方法级记录")
        final_header = _find_header_after(records, "C. 最终问卷记录")
        blocks = _read_table(records, block_header, block_header + 1, method_header - 2)
        methods = _read_table(records, method_header, method_header + 1, final_header - 2)
        finals = _read_table(records, final_header, final_header + 1, records.max_row)
        source_kind = _source_kind(workbook, source.name)
    finally:
        workbook.close()
    data = Exp3Data(
        participants=participants,
        blocks=blocks,
        methods=methods,
        finals=finals,
        source_kind=source_kind,
        source_path=str(source),
        source_sha256=workbook_sha256(source),
    )
    _validate_structure(data)
    return data


def validate_for_analysis(
    data: Exp3Data,
    *,
    minimum_participants: int,
    aq_mode: str,
    q10_enabled: bool,
    allow_synthetic: bool = False,
) -> dict[str, Any]:
    """检查采集完成状态、合法值和正式来源边界。"""

    if data.source_kind != "formal":
        if data.source_kind != "synthetic" or not allow_synthetic:
            raise ValueError("输入缺少正式工作簿契约标识，或属于合成/模拟数据")
    participants = data.participants.copy()
    included = participants[participants["纳入分析"].map(_is_yes)]
    included_ids = frozenset(included["Participant_ID"].astype(str))
    warnings: list[str] = []
    if len(included_ids) < minimum_participants:
        raise ValueError(
            f"纳入分析且已确认的参与者只有 {len(included_ids)} 人，少于冻结下限 {minimum_participants}"
        )
    _validate_block_values(
        data.blocks,
        included_ids,
        aq_mode=aq_mode,
        q10_enabled=q10_enabled,
    )
    _validate_method_values(data.methods, included_ids)
    _validate_final_values(data.finals, included_ids)
    if len(included_ids) < 24:
        warnings.append(f"当前纳入 {len(included_ids)} 人，少于目标 N=24；结果按实际配对 N 报告")
    return {
        "included_participants": tuple(sorted(included_ids)),
        "included_count": len(included_ids),
        "warnings": tuple(warnings),
        "source_kind": data.source_kind,
    }


def describe_workbook(data: Exp3Data) -> dict[str, Any]:
    """返回不要求采集完成的结构与填表进度摘要。"""

    participants = data.participants
    blocks = data.blocks
    methods = data.methods
    finals = data.finals
    return {
        "passed": True,
        "source": data.source_path,
        "source_sha256": data.source_sha256,
        "source_kind": data.source_kind,
        "participant_rows": len(participants),
        "included_confirmed": int(participants["纳入分析"].map(_is_yes).sum()),
        "block_rows": len(blocks),
        "block_scores_filled": int(blocks[list(BLOCK_ITEMS.values())].notna().sum().sum()),
        "method_rows": len(methods),
        "method_scores_filled": int(methods[list(METHOD_ITEM_COLUMNS)].notna().sum().sum()),
        "final_rows": len(finals),
        "final_choices_filled": int(finals["方法选择(标签)"].notna().sum()),
    }


def _read_table(worksheet: Any, header_row: int, first_row: int, last_row: int) -> pd.DataFrame:
    """按指定表头和行边界读取一张普通值表。"""

    headers = [worksheet.cell(header_row, column).value for column in range(1, worksheet.max_column + 1)]
    while headers and headers[-1] is None:
        headers.pop()
    if not headers or any(value is None for value in headers):
        raise ValueError(f"{worksheet.title}!{header_row} 表头包含空列")
    rows: list[dict[str, Any]] = []
    for row_number in range(first_row, last_row + 1):
        cells = [worksheet.cell(row_number, column) for column in range(1, len(headers) + 1)]
        if all(cell.value is None for cell in cells):
            continue
        _reject_formulas(worksheet.title, row_number, cells)
        rows.append({str(header): cell.value for header, cell in zip(headers, cells, strict=True)})
    return pd.DataFrame(rows, columns=[str(value) for value in headers])


def _reject_formulas(sheet_name: str, row_number: int, cells: Iterable[Cell]) -> None:
    """拒绝在原始值区域放入公式，避免评分来源被派生值覆盖。"""

    for cell in cells:
        if cell.data_type == "f":
            raise ValueError(f"原始数据区域不得含公式：{sheet_name}!{cell.coordinate}（第 {row_number} 行）")


def _find_header_after(worksheet: Any, marker: str) -> int:
    """按 A 列节标题定位其下一行表头。"""

    for row_number in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row_number, 1).value
        if isinstance(value, str) and value.startswith(marker):
            return row_number + 1
    raise ValueError(f"Records 缺少节标题：{marker}")


def _source_kind(workbook: Any, filename: str) -> str:
    """优先按核心属性识别正式输入，只用文本识别合成演练。"""

    if (
        workbook.properties.identifier == WORKBOOK_CONTRACT_ID
        and workbook.properties.category == WORKBOOK_SOURCE_CATEGORY
    ):
        return "formal"
    readme = workbook["README"]

    text = " ".join(
        str(readme.cell(row, column).value or "")
        for row in range(1, min(readme.max_row, 12) + 1)
        for column in range(1, min(readme.max_column, 5) + 1)
    ).lower()
    combined = f"{filename} {text}".lower()
    synthetic_terms = ("合成数据", "模拟演练", "synthetic", "simulated", "claude-opus")
    if any(term in combined for term in synthetic_terms):
        return "synthetic"
    return "unknown"


def _validate_structure(data: Exp3Data) -> None:
    """验证 24 平衡单元和三段记录的固定身份。"""

    _require_columns(data.participants, _PARTICIPANT_DESIGN_COLUMNS + ("纳入分析",), "Participants")
    _require_columns(
        data.blocks,
        _BLOCK_DESIGN_COLUMNS + tuple(BLOCK_ITEMS.values()) + (
            "任务完成",
            "问卷完成",
            "区块有效",
            "技术问题",
        ),
        "Records A",
    )
    _require_columns(
        data.methods,
        _METHOD_DESIGN_COLUMNS + METHOD_ITEM_COLUMNS + ("尺度切换确认", "技术问题"),
        "Records B",
    )
    _require_columns(
        data.finals,
        (
            "Participant_ID",
            "方法选择(标签)",
            "偏好强度(1-7/NA)",
            "信任选择(标签)",
            "区分信心(1-7)",
        ),
        "Records C",
    )
    if len(data.participants) != 24 or len(data.blocks) != 144 or len(data.methods) != 48 or len(data.finals) != 24:
        raise ValueError(
            "实验三固定结构必须是 Participants=24、Records A=144、B=48、C=24"
        )
    participant_ids = tuple(data.participants["Participant_ID"].astype(str))
    if len(set(participant_ids)) != 24:
        raise ValueError("Participants 的 Participant_ID 必须唯一")
    expected = frozenset(participant_ids)
    for name, table, repetitions in (
        ("Records A", data.blocks, 6),
        ("Records B", data.methods, 2),
        ("Records C", data.finals, 1),
    ):
        counts = table["Participant_ID"].astype(str).value_counts()
        if frozenset(counts.index) != expected or not (counts == repetitions).all():
            raise ValueError(f"{name} 必须为每位参与者恰好保留 {repetitions} 行")
    _validate_design_mapping(data)


def _validate_design_mapping(data: Exp3Data) -> None:
    """核对标签映射、对象覆盖和方法级评分顺序。"""

    participants = data.participants.set_index("Participant_ID")
    for participant_id, rows in data.blocks.groupby("Participant_ID", sort=False):
        if frozenset(rows["Condition(保密)"].astype(str)) != frozenset(METHODS):
            raise ValueError(f"{participant_id} 的区块未覆盖两种方法")
        if frozenset(rows["Object_Key"].astype(str)) != frozenset(OBJECTS):
            raise ValueError(f"{participant_id} 的区块未覆盖三个正式对象")
        if len(rows.groupby(["Object_Key", "Condition(保密)"], dropna=False)) != 6:
            raise ValueError(f"{participant_id} 的对象×方法区块不唯一")
        mapping = participants.loc[participant_id]
        expected_label = {
            "方法A": str(mapping["方法A=（保密）"]),
            "方法B": str(mapping["方法B=（保密）"]),
        }
        for _, row in rows.iterrows():
            label = str(row["Shown_Label"])
            condition = str(row["Condition(保密)"])
            if expected_label.get(label) != condition:
                raise ValueError(f"{participant_id} 的 {label} 映射与 Participants 不一致")
    for participant_id, rows in data.methods.groupby("Participant_ID", sort=False):
        if frozenset(rows["Condition(保密)"].astype(str)) != frozenset(METHODS):
            raise ValueError(f"{participant_id} 的方法级记录未覆盖两种方法")


def _require_columns(table: pd.DataFrame, columns: Iterable[str], name: str) -> None:
    """要求逻辑表包含指定列。"""

    missing = set(columns).difference(table.columns)
    if missing:
        raise ValueError(f"{name} 缺少列：{sorted(missing)}")


def _validate_block_values(
    blocks: pd.DataFrame,
    included_ids: frozenset[str],
    *,
    aq_mode: str,
    q10_enabled: bool,
) -> None:
    """检查纳入参与者的区块状态与 1--7 原始评分。"""

    selected = blocks[blocks["Participant_ID"].astype(str).isin(included_ids)]
    for participant_id, rows in selected.groupby("Participant_ID", sort=False):
        valid = rows.apply(_block_is_valid, axis=1)
        if int(valid.sum()) != 6:
            raise ValueError(f"{participant_id} 必须有 6 个明确完成且有效的正式区块")
    required = [BLOCK_ITEMS[item] for item in required_block_items(aq_mode)]
    if q10_enabled:
        required.append(BLOCK_ITEMS["Q10"])
    _validate_numeric_range(selected, required, 1.0, 7.0, "区块评分")
    optional = set(BLOCK_ITEMS.values()).difference(required)
    _validate_numeric_range(
        selected,
        tuple(sorted(optional)),
        1.0,
        7.0,
        "可选区块评分",
        allow_missing=True,
    )


def _validate_method_values(methods: pd.DataFrame, included_ids: frozenset[str]) -> None:
    """检查纳入参与者的方法级原始评分与审核状态。"""

    selected = methods[methods["Participant_ID"].astype(str).isin(included_ids)]
    _validate_numeric_range(selected, METHOD_ITEM_COLUMNS[:10], 1.0, 5.0, "TiA 原始评分", allow_missing=True)
    _validate_numeric_range(selected, METHOD_ITEM_COLUMNS[10:], 1.0, 7.0, "S-TIAS 原始评分")
    if not selected["尺度切换确认"].map(_is_yes).all():
        raise ValueError("纳入参与者的两次方法级问卷都必须确认尺度切换")
    for optional_column in ("A/B归属回忆确认", "方法级记录有效"):
        if optional_column in selected and not selected[optional_column].map(_is_yes).all():
            raise ValueError(f"纳入参与者的 {optional_column} 必须全部确认")


def _validate_final_values(finals: pd.DataFrame, included_ids: frozenset[str]) -> None:
    """检查最终选择、条件跳题与区分信心。"""

    selected = finals[finals["Participant_ID"].astype(str).isin(included_ids)]
    choices = {"方法A", "方法B", "无明显偏好"}
    for column in ("方法选择(标签)", "信任选择(标签)"):
        invalid = selected[column].dropna().astype(str).map(lambda value: value not in choices)
        if selected[column].isna().any() or invalid.any():
            raise ValueError(f"纳入参与者的 {column} 必须使用方法A/方法B/无明显偏好")
    _validate_numeric_range(selected, ("区分信心(1-7)",), 1.0, 7.0, "区分信心")
    for _, row in selected.iterrows():
        choice = row["方法选择(标签)"]
        strength = row["偏好强度(1-7/NA)"]
        if str(choice) == "无明显偏好":
            if not _is_missing_or_na(strength):
                raise ValueError("选择无明显偏好时，偏好强度必须留空或写 N/A")
        elif _number_or_none(strength) is None or not 1.0 <= float(strength) <= 7.0:
            raise ValueError("做出方法选择时，偏好强度必须为 1--7")


def block_valid_mask(blocks: pd.DataFrame) -> pd.Series:
    """返回同时满足完成、问卷与技术状态的区块有效掩码。"""

    return blocks.apply(_block_is_valid, axis=1)


def _block_is_valid(row: pd.Series) -> bool:
    """判断一个区块是否满足冻结的技术有效条件。"""

    return (
        _is_yes(row.get("任务完成"))
        and _is_yes(row.get("问卷完成"))
        and _is_yes(row.get("区块有效"))
    )


def included_participant_ids(participants: pd.DataFrame) -> frozenset[str]:
    """返回明确标记为纳入分析的参与者 ID。"""

    return frozenset(
        participants.loc[participants["纳入分析"].map(_is_yes), "Participant_ID"].astype(str)
    )


def _validate_numeric_range(
    table: pd.DataFrame,
    columns: Iterable[str],
    minimum: float,
    maximum: float,
    label: str,
    *,
    allow_missing: bool = False,
) -> None:
    """检查若干评分列是整数刻度值，并按契约处理缺失。"""

    for column in columns:
        numeric = pd.to_numeric(table[column], errors="coerce")
        invalid_text = table[column].notna() & numeric.isna() & ~table[column].map(_is_missing_or_na)
        if invalid_text.any():
            raise ValueError(f"{label}列 {column} 含非数值响应")
        valid = numeric.dropna()
        if ((valid < minimum) | (valid > maximum) | (valid % 1 != 0)).any():
            raise ValueError(f"{label}列 {column} 必须是 {minimum:.0f}--{maximum:.0f} 整数")
        if not allow_missing and numeric.isna().any():
            raise ValueError(f"{label}列 {column} 不允许缺失")


def _number_or_none(value: Any) -> float | None:
    """把普通数字转换为 float，缺失或 N/A 返回 None。"""

    if _is_missing_or_na(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _is_missing_or_na(value: Any) -> bool:
    """识别工作簿允许的空值或 N/A 文本。"""

    if value is None or (isinstance(value, float) and pd.isna(value)):
        return True
    return str(value).strip().lower() in {"", "n/a", "na", "无法回答"}


def _is_yes(value: Any) -> bool:
    """识别人工确认字段中的肯定值。"""

    return str(value or "").strip().lower() in {"是", "yes", "true", "1"}


__all__ = [
    "block_valid_mask",
    "describe_workbook",
    "included_participant_ids",
    "read_workbook",
    "validate_for_analysis",
    "workbook_sha256",
]
