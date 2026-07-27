"""从 v5.1 美化定稿生成正式空白原始数据模板。"""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.formatting.rule import FormulaRule  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[import-untyped]

from .contracts import (
    BLOCK_RECORD_COLUMNS,
    METHOD_RECORD_COLUMNS,
    METHOD_SCALE_ITEMS,
    OUTCOME_LABELS,
    PRIMARY_OUTCOMES,
    REVERSED_TIA_ITEMS,
    SCALE_OUTCOMES,
    WORKBOOK_CONTRACT_ID,
    WORKBOOK_SOURCE_CATEGORY,
    aq_scale_items,
    required_block_items,
)
from .settings import Exp3Settings


_NAVY = "18324A"
_TEAL = "2F6F73"
_PALE_BLUE = "DCEAF5"
_GREEN = "E2F0D9"
_YELLOW = "FFF2CC"
_PALE_RED = "FCE8E6"
_TEXT = "22313F"
_WHITE = "FFFFFF"
_GRID = "C8D1D9"
"""正式模板沿用的克制配色。"""


def build_raw_template(
    settings: Exp3Settings,
    destination: Path | None = None,
    *,
    overwrite: bool = False,
) -> Path:
    """复制美化来源，清空采集值并加入实时公式分析区。"""

    source = settings.paths.source_template
    output = (destination or settings.paths.input_workbook).expanduser().resolve()
    if not source.is_file():
        raise FileNotFoundError(f"原始美化模板不存在：{source}")
    if output == source.resolve():
        raise ValueError("新原始模板不得覆盖美化来源文件")
    if output.exists() and not overwrite:
        raise FileExistsError(f"拒绝覆盖已有实验三原始工作簿：{output}")
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.stem}.{uuid4().hex}.tmp.xlsx")
    workbook = load_workbook(source)
    try:
        _clean_front_matter(workbook)
        _clear_collection_fields(workbook)
        _repair_records_validations(workbook)
        _replace_formula_sheets(workbook, settings)
        _enable_recalculation(workbook)
        _assert_no_synthetic_boilerplate(workbook)
        workbook.save(temporary)
    finally:
        workbook.close()
    try:
        temporary.replace(output)
    finally:
        temporary.unlink(missing_ok=True)
    return output


def _clean_front_matter(workbook: Any) -> None:
    """删除过期过程说明，并把首屏改成正式采集事实。"""

    workbook.properties.identifier = WORKBOOK_CONTRACT_ID
    workbook.properties.category = WORKBOOK_SOURCE_CATEGORY
    workbook.properties.version = "v5.1"
    readme = workbook["README"]
    readme["A1"] = "EgoAnchor 实验三正式原始数据工作簿（v5.1）"
    readme["A2"] = (
        "唯一权威规格：2026-EgoAnchor/experiment_3_questionnaire_design_zh.md（v5.1）。"
        "Participants 与 Records 是唯一人工输入；Derived 和 Analysis 只读取原始值并实时复算，"
        "不得把公式结果粘回原始评分区。TiA 反向计分仅在派生层执行。"
    )
    readme["A19"] = (
        "中文施测措辞已于 2026-07-26 冻结为 v5.1 Final wording。正式采集前完成确认性认知访谈；"
        "只有发现理解不良时，才按修订、重测、再冻结的顺序处理。"
    )
    audit = workbook["Verification_Audit"]
    audit["A1"] = "内容核验记录（v5.1；正式采集前审计）"
    audit["A2"] = (
        "官方原文、对象化改编和当前样本信度必须分开表述；正式采集前剩余前置为确认性认知访谈与预留参数冻结。"
    )
    audit["C18"] = (
        "Q2 已按 v5.1 冻结为“始终附着在真实物体上的同一位置”，用于与 AQ-IQ3 的运动平滑构念区分。"
    )
    questionnaire = workbook["Questionnaire"]
    questionnaire["A2"] = (
        "中文施测版为头显或纸面呈现文本；官方英文原文仅用于核对，对象化英文施测版用于附录与英文施测。"
        "正式中文措辞已经逐条语义审计，预实验继续做确认性认知访谈。"
    )


def _clear_collection_fields(workbook: Any) -> None:
    """保留设计映射，清空所有需要人工或运行时写入的字段。"""

    participants = workbook["Participants"]
    for row in range(3, 27):
        for column in range(12, 25):
            participants.cell(row, column).value = None

    records = workbook["Records"]
    for row in range(5, 149):
        for column in range(11, 42):
            records.cell(row, column).value = None
    for row in range(152, 200):
        for column in range(5, 26):
            records.cell(row, column).value = None
    for row in range(203, 227):
        for column in range(2, 12):
            records.cell(row, column).value = None

    # v5.1 预实验要求的方法归属回忆与记录有效性审核补在 B 段末尾。
    records["X151"] = "A/B归属回忆确认"
    records["Y151"] = "方法级记录有效"
    for column in (24, 25):
        records.cell(151, column)._style = copy(records["W151"]._style)
        records.cell(151, column).alignment = copy(records["W151"].alignment)
        records.column_dimensions[get_column_letter(column)].width = 16
        for row in range(152, 200):
            records.cell(row, column)._style = copy(records["W152"]._style)

    # 原版 P024 最终问卷行缺少完整样式，这里与前一行对齐。
    for column in range(2, 12):
        records.cell(226, column)._style = copy(records.cell(225, column)._style)


def _repair_records_validations(workbook: Any) -> None:
    """重建原始记录的输入校验，并允许冻结规则中的合法缺失。"""

    records = workbook["Records"]
    records.data_validations.dataValidation = []
    _add_whole_number_validation(records, "K5:M148", 1, 7, allow_blank=False)
    _add_whole_number_validation(records, "O5:X148", 1, 7, allow_blank=False)
    _add_whole_number_validation(records, "N5:N148", 1, 7, allow_blank=True)
    for range_reference in ("AB5:AD148", "AL5:AL148", "R152:R199", "X152:Y199"):
        _add_list_validation(records, range_reference, '"是,否"', allow_blank=False)
    _add_list_validation(
        records,
        "AE5:AE148",
        '"无,设备故障,网络故障,追踪异常,问卷中断,其他"',
        allow_blank=True,
    )
    _add_list_validation(
        records,
        "AK5:AK148",
        '"Coasting,FrozenUncertain,Lost,不适用"',
        allow_blank=True,
    )
    _add_list_validation(
        records,
        "V152:V199",
        '"无,设备故障,问卷中断,尺度误用,其他"',
        allow_blank=True,
    )
    _add_custom_scale_validation(records, "E152:N199", 1, 5)
    _add_custom_scale_validation(records, "O152:Q199", 1, 7)
    _add_list_validation(records, "B203:B226", '"方法A,方法B,无明显偏好"', allow_blank=False)
    _add_list_validation(records, "D203:D226", '"方法A,方法B,无明显偏好"', allow_blank=False)
    _add_custom_na_validation(records, "C203:C226")
    _add_whole_number_validation(records, "E203:E226", 1, 7, allow_blank=False)


def _replace_formula_sheets(workbook: Any, settings: Exp3Settings) -> None:
    """用正式实时派生表和分析面板替换旧的空回填壳。"""

    if "Derived" in workbook.sheetnames:
        workbook.remove(workbook["Derived"])
    analysis_index = workbook.sheetnames.index("Analysis")
    workbook.remove(workbook["Analysis"])
    derived = workbook.create_sheet("Derived", analysis_index)
    analysis = workbook.create_sheet("Analysis", analysis_index + 1)
    _build_derived_sheet(derived, settings)
    _build_analysis_sheet(analysis)


def _build_derived_sheet(worksheet: Any, settings: Exp3Settings) -> None:
    """构建原始值到小分、参与者均值与配对差的透明公式链。"""

    _title_row(
        worksheet,
        1,
        26,
        "实时派生（仅公式）：区块有效性与 AQ → TiA 换向与方法分 → 三物体均值 → 配对差",
    )
    _section_row(worksheet, 3, 26, "D1. 区块级派生：144 行与 Records A 段逐行对应")
    d1_headers = (
        "Participant_ID", "Block_Index", "Condition", "Object_Key", "Valid_Block",
        "Q1", "Q8", "Q2", "Q9", "Q3", "Q6", "Q7", "AQ_EQ", "AQ_IQ", "Q10",
        "Duration_Over_150s", "Straightline_5Plus", "Candidate_Rate_Hz", "VCD_Median",
        "VCD_Admission_Rate", "Output_Availability", "Occlusion_Seconds", "Lifecycle_State",
        "Entered_Lost", "Reacquisition_Count", "StaticLock_Count",
    )
    _header_row(worksheet, 4, d1_headers)
    source_columns = {
        "Participant_ID": "A", "Block_Index": "B", "Condition": "H", "Object_Key": "F",
        **BLOCK_RECORD_COLUMNS,
        "Candidate_Rate_Hz": "AF", "VCD_Median": "AG",
        "VCD_Admission_Rate": "AH", "Output_Availability": "AI", "Occlusion_Seconds": "AJ",
        "Lifecycle_State": "AK", "Entered_Lost": "AL", "Reacquisition_Count": "AM", "StaticLock_Count": "AN",
    }
    required_items = required_block_items(settings.aq_mode)
    aq_items = aq_scale_items(settings.aq_mode)
    for derived_row, source_row in zip(range(5, 149), range(5, 149), strict=True):
        for column_index, header in enumerate(d1_headers, start=1):
            cell = worksheet.cell(derived_row, column_index)
            if header in source_columns:
                source_column = source_columns[header]
                cell.value = f'=IF(Records!{source_column}{source_row}="","",Records!{source_column}{source_row})'
            elif header == "Valid_Block":
                required_references = ",".join(
                    f"Records!{BLOCK_RECORD_COLUMNS[item]}{source_row}"
                    for item in required_items
                )
                cell.value = (
                    f'=IF(AND(COUNTIFS(Participants!$A$3:$A$26,Records!A{source_row},Participants!$V$3:$V$26,"是")=1,'
                    f'Records!AB{source_row}="是",Records!AC{source_row}="是",'
                    f'Records!AD{source_row}="是",'
                    f'COUNT({required_references})={len(required_items)}),"是","否")'
                )
            elif header in aq_items:
                references = tuple(
                    f"Records!{BLOCK_RECORD_COLUMNS[item]}{source_row}"
                    for item in aq_items[header]
                )
                cell.value = _complete_average_formula(references)
            elif header == "Duration_Over_150s":
                cell.value = f'=IF(Records!AA{source_row}="","",IF(Records!AA{source_row}>150,"是","否"))'
            elif header == "Straightline_5Plus":
                cell.value = _straightline_formula(source_row, required_items)
            _formula_style(cell)

    _section_row(worksheet, 150, 20, "D2. 方法级派生：TiA 原始分换向后计算两个分量表；S-TIAS 保持 1–7")
    d2_headers = (
        "Participant_ID", "Condition", "Valid_Method", "TIA_RC1", "TIA_RC2", "TIA_RC3",
        "TIA_RC4", "TIA_RC5", "TIA_RC6", "TIA_UP1", "TIA_UP2", "TIA_UP3", "TIA_UP4",
        "STIAS1", "STIAS2", "STIAS3", "TIA_RC", "TIA_UP", "STIAS", "Duration_Seconds",
    )
    _header_row(worksheet, 151, d2_headers)
    method_references = {
        scale: ",".join(f"Records!{METHOD_RECORD_COLUMNS[item]}{{row}}" for item in items)
        for scale, items in METHOD_SCALE_ITEMS.items()
    }
    for derived_row, source_row in zip(range(152, 200), range(152, 200), strict=True):
        worksheet.cell(derived_row, 1).value = f'=Records!A{source_row}'
        worksheet.cell(derived_row, 2).value = f'=Records!D{source_row}'
        worksheet.cell(derived_row, 3).value = (
            f'=IF(AND(COUNTIFS(Participants!$A$3:$A$26,Records!A{source_row},Participants!$V$3:$V$26,"是")=1,'
            f'Records!R{source_row}="是",OR(Records!V{source_row}="",Records!V{source_row}="无"),'
            f'Records!X{source_row}="是",Records!Y{source_row}="是",'
            f'COUNT({method_references["TIA_RC"].format(row=source_row)})>={settings.tia_rc_min_items},'
            f'COUNT({method_references["TIA_UP"].format(row=source_row)})>={settings.tia_up_min_items},'
            f'COUNT({method_references["STIAS"].format(row=source_row)})>={settings.stias_min_items}),"是","否")'
        )
        for offset, (item, source_column) in enumerate(METHOD_RECORD_COLUMNS.items()):
            target = worksheet.cell(derived_row, 4 + offset)
            if item in REVERSED_TIA_ITEMS:
                target.value = f'=IF(ISNUMBER(Records!{source_column}{source_row}),6-Records!{source_column}{source_row},"")'
            else:
                target.value = f'=IF(ISNUMBER(Records!{source_column}{source_row}),Records!{source_column}{source_row},"")'
        worksheet.cell(derived_row, 17).value = f'=IF(COUNT(D{derived_row}:I{derived_row})>={settings.tia_rc_min_items},AVERAGE(D{derived_row}:I{derived_row}),"")'
        worksheet.cell(derived_row, 18).value = f'=IF(COUNT(J{derived_row}:M{derived_row})>={settings.tia_up_min_items},AVERAGE(J{derived_row}:M{derived_row}),"")'
        worksheet.cell(derived_row, 19).value = f'=IF(COUNT(N{derived_row}:P{derived_row})>={settings.stias_min_items},AVERAGE(N{derived_row}:P{derived_row}),"")'
        worksheet.cell(derived_row, 20).value = f'=IF(Records!U{source_row}="","",Records!U{source_row})'
        for column in range(1, 21):
            _formula_style(worksheet.cell(derived_row, column))

    _section_row(worksheet, 201, 21, "D3. 每人每方法：三个对象取均值；前 24 行 EgoAnchor，后 24 行 One-Euro")
    d3_headers = (
        "Participant_ID", "Condition", "Q1", "Q8", "Q2", "Q9", "Q3", "Q6", "Q7", "AQ_EQ",
        "AQ_IQ", "TIA_RC", "TIA_UP", "STIAS", "Q10", "Candidate_Rate_Hz", "VCD_Median",
        "VCD_Admission_Rate", "Output_Availability", "Occlusion_Seconds", "Valid_Objects",
    )
    _header_row(worksheet, 202, d3_headers)
    d1_columns = {
        "Q1": "F", "Q8": "G", "Q2": "H", "Q9": "I", "Q3": "J", "Q6": "K", "Q7": "L",
        "AQ_EQ": "M", "AQ_IQ": "N", "Q10": "O", "Candidate_Rate_Hz": "R", "VCD_Median": "S",
        "VCD_Admission_Rate": "T", "Output_Availability": "U", "Occlusion_Seconds": "V",
    }
    d2_columns = {"TIA_RC": "Q", "TIA_UP": "R", "STIAS": "S"}
    for method_index, method in enumerate(("EgoAnchor", "One-Euro")):
        for participant_index in range(24):
            row = 203 + method_index * 24 + participant_index
            participant_row = 3 + participant_index
            worksheet.cell(row, 1).value = f'=Participants!A{participant_row}'
            worksheet.cell(row, 2).value = method
            for column_index, outcome in enumerate(d3_headers[2:20], start=3):
                if outcome in d1_columns:
                    source_column = d1_columns[outcome]
                    worksheet.cell(row, column_index).value = (
                        f'=IF(COUNTIFS($A$5:$A$148,$A{row},$C$5:$C$148,$B{row},$E$5:$E$148,"是",'
                        f'${source_column}$5:${source_column}$148,">=0")=3,AVERAGEIFS(${source_column}$5:${source_column}$148,'
                        f'$A$5:$A$148,$A{row},$C$5:$C$148,$B{row},$E$5:$E$148,"是"),"")'
                    )
                else:
                    source_column = d2_columns[outcome]
                    worksheet.cell(row, column_index).value = (
                        f'=IF(COUNTIFS($A$152:$A$199,$A{row},$B$152:$B$199,$B{row},$C$152:$C$199,"是",'
                        f'${source_column}$152:${source_column}$199,">=0")=1,AVERAGEIFS(${source_column}$152:${source_column}$199,'
                        f'$A$152:$A$199,$A{row},$B$152:$B$199,$B{row},$C$152:$C$199,"是"),"")'
                    )
            worksheet.cell(row, 21).value = f'=COUNTIFS($A$5:$A$148,$A{row},$C$5:$C$148,$B{row},$E$5:$E$148,"是")'
            for column in range(1, 22):
                _formula_style(worksheet.cell(row, column))

    _section_row(worksheet, 252, 13, "D4. 每人配对差：EgoAnchor − One-Euro；Participant_ID 校验应全部为 OK")
    d4_outcomes = (*PRIMARY_OUTCOMES, *SCALE_OUTCOMES)
    _header_row(worksheet, 253, ("Participant_ID", *d4_outcomes, "Pair_Check"))
    d3_outcome_columns = {header: get_column_letter(index) for index, header in enumerate(d3_headers, start=1)}
    for participant_index in range(24):
        row = 254 + participant_index
        ea_row = 203 + participant_index
        oe_row = 227 + participant_index
        worksheet.cell(row, 1).value = f'=A{ea_row}'
        for column_index, outcome in enumerate(d4_outcomes, start=2):
            source_column = d3_outcome_columns[outcome]
            worksheet.cell(row, column_index).value = (
                f'=IF(OR({source_column}{ea_row}="",{source_column}{oe_row}=""),"",{source_column}{ea_row}-{source_column}{oe_row})'
            )
        worksheet.cell(row, 14).value = f'=IF(A{ea_row}=A{oe_row},"OK","ERROR")'
        for column in range(1, 15):
            _formula_style(worksheet.cell(row, column))

    _section_row(worksheet, 279, 5, "D5. 最终问卷派生：按 Participants 的纳入状态过滤")
    _header_row(
        worksheet,
        280,
        ("Participant_ID", "Included", "Preference_Strength", "Discrimination_Confidence", "Final_Complete"),
    )
    for participant_index in range(24):
        row = 281 + participant_index
        participant_row = 3 + participant_index
        final_row = 203 + participant_index
        worksheet.cell(row, 1).value = f'=Participants!A{participant_row}'
        worksheet.cell(row, 2).value = f'=Participants!V{participant_row}'
        worksheet.cell(row, 3).value = (
            f'=IF(AND(B{row}="是",ISNUMBER(Records!C{final_row})),Records!C{final_row},"")'
        )
        worksheet.cell(row, 4).value = (
            f'=IF(AND(B{row}="是",ISNUMBER(Records!E{final_row})),Records!E{final_row},"")'
        )
        worksheet.cell(row, 5).value = (
            f'=IF(B{row}<>"是","",IF(AND(Records!B{final_row}<>"",Records!D{final_row}<>"",'
            f'ISNUMBER(Records!E{final_row})),"是","否"))'
        )
        for column in range(1, 6):
            _formula_style(worksheet.cell(row, column))

    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = "A4:Z148"
    widths = {1: 15, 2: 13, 3: 13, 4: 15, 5: 12}
    for column in range(1, 27):
        worksheet.column_dimensions[get_column_letter(column)].width = widths.get(column, 14)
    worksheet.sheet_view.showGridLines = False


def _build_analysis_sheet(worksheet: Any) -> None:
    """构建绿色实时描述统计与黄色离线推断占位区。"""

    _title_row(worksheet, 1, 20, "实验三实时数据检查与分析概览")
    worksheet["A2"] = "绿色单元格由 Excel 公式实时更新；黄色列由 Python 结果工作簿提供。正式推断只以 Python 输出为准。"
    worksheet.merge_cells("A2:T2")
    worksheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    worksheet["A2"].font = Font(color="5B6570", italic=True, size=10)
    _section_row(worksheet, 4, 20, "A. 采集完整性与问卷负担")
    _header_row(worksheet, 5, ("检查项", "实时值", "目标/规则", "状态"))
    qc_rows = (
        ("确认纳入参与者", '=COUNTIF(Participants!V3:V26,"是")', "目标 24；硬下限 18", '=IF(B6>=24,"完成",IF(B6>=18,"可分析但未达目标","未达下限"))'),
        ("有效区块", '=COUNTIF(Derived!E5:E148,"是")', "每位纳入者 6 个", '=IF(B7=B6*6,"通过","检查缺失或审核")'),
        ("有效方法级记录", '=COUNTIF(Derived!C152:C199,"是")', "每位纳入者 2 个", '=IF(B8=B6*2,"通过","检查缺失或审核")'),
        ("最终问卷已填", '=COUNTIF(Derived!E281:E304,"是")', "每位纳入者 1 个", '=IF(B9=B6,"通过","检查缺失")'),
        ("区块问卷 >150 s", '=IFERROR(COUNTIF(Derived!P5:P148,"是")/COUNTIF(Derived!E5:E148,"是"),"")', "预实验负担诊断", '=IF(B10="","待填",IF(B10>=0.2,"触发复核","未触发"))'),
        (">=5 连续同分", '=IFERROR(COUNTIF(Derived!Q5:Q148,"是")/COUNTIF(Derived!E5:E148,"是"),"")', "预实验负担诊断", '=IF(B11="","待填",IF(B11>=0.05,"触发复核","未触发"))'),
    )
    for row_index, values in enumerate(qc_rows, start=6):
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row_index, column_index).value = value
            _formula_style(worksheet.cell(row_index, column_index))
    worksheet["B10"].number_format = "0.0%"
    worksheet["B11"].number_format = "0.0%"

    main_start = 14
    scale_start = 25
    _section_row(worksheet, main_start, 20, "B. 主证实家族：三物体均值后的完整配对统计")
    result_headers = (
        "Outcome", "简称", "N", "OE_Q1", "OE_Median", "OE_Q3", "EA_Q1", "EA_Median", "EA_Q3",
        "Diff_Median", "Diff_Mean", "Diff_SD", "dz", "W", "p_raw", "p_Holm", "r_rb", "r_CI_Low",
        "r_CI_High", "结论",
    )
    _header_row(worksheet, main_start + 1, result_headers)
    d3_columns = {
        "Q1": "C", "Q8": "D", "Q2": "E", "Q9": "F", "Q3": "G", "Q6": "H", "Q7": "I",
        "AQ_EQ": "J", "AQ_IQ": "K", "TIA_RC": "L", "TIA_UP": "M", "STIAS": "N",
    }
    d4_columns = {
        outcome: get_column_letter(index)
        for index, outcome in enumerate((*PRIMARY_OUTCOMES, *SCALE_OUTCOMES), start=2)
    }
    for row, outcome in enumerate(PRIMARY_OUTCOMES, start=main_start + 2):
        _analysis_result_row(worksheet, row, outcome, d3_columns[outcome], d4_columns[outcome])

    _section_row(worksheet, scale_start, 20, "C. 已发表量表家族：当前样本信度由 Python 另行报告")
    _header_row(worksheet, scale_start + 1, result_headers)
    for row, outcome in enumerate(SCALE_OUTCOMES, start=scale_start + 2):
        _analysis_result_row(worksheet, row, outcome, d3_columns[outcome], d4_columns[outcome])

    manipulation_start = 34
    _section_row(worksheet, manipulation_start, 20, "D. 操纵与运行时描述：每位参与者先在三个物体上取均值")
    _header_row(worksheet, manipulation_start + 1, ("Metric", "N", "One-Euro Mean", "EgoAnchor Mean", "Difference Mean", "TOST Margin", "p_TOST", "Status"))
    manipulation_columns = {
        "Candidate_Rate_Hz": "P", "VCD_Median": "Q", "VCD_Admission_Rate": "R",
        "Output_Availability": "S", "Occlusion_Seconds": "T",
    }
    for row, (metric, column) in enumerate(manipulation_columns.items(), start=manipulation_start + 2):
        worksheet.cell(row, 1).value = metric
        worksheet.cell(row, 2).value = f'=MIN(COUNT(Derived!{column}203:{column}226),COUNT(Derived!{column}227:{column}250))'
        worksheet.cell(row, 3).value = f'=IF(B{row}=0,"",AVERAGE(Derived!{column}227:{column}250))'
        worksheet.cell(row, 4).value = f'=IF(B{row}=0,"",AVERAGE(Derived!{column}203:{column}226))'
        worksheet.cell(row, 5).value = f'=IF(B{row}=0,"",D{row}-C{row})'
        for column_index in range(1, 6):
            _formula_style(worksheet.cell(row, column_index))
        for column_index in range(6, 9):
            _offline_style(worksheet.cell(row, column_index))

    choice_start = 43
    _section_row(worksheet, choice_start, 20, "E. 最终测量完成度")
    _header_row(worksheet, choice_start + 1, ("Measure", "Count", "Median", "Q1", "Q3", "Mean", "SD"))
    final_metrics = (
        ("Preference_Strength", "C"),
        ("Discrimination_Confidence", "D"),
    )
    for row, (metric, column) in enumerate(final_metrics, start=choice_start + 2):
        value_range = f'Derived!{column}281:{column}304'
        worksheet.cell(row, 1).value = metric
        worksheet.cell(row, 2).value = f'=COUNT({value_range})'
        worksheet.cell(row, 3).value = f'=IF(B{row}=0,"",MEDIAN({value_range}))'
        worksheet.cell(row, 4).value = f'=IF(B{row}=0,"",QUARTILE.INC({value_range},1))'
        worksheet.cell(row, 5).value = f'=IF(B{row}=0,"",QUARTILE.INC({value_range},3))'
        worksheet.cell(row, 6).value = f'=IF(B{row}=0,"",AVERAGE({value_range}))'
        worksheet.cell(row, 7).value = f'=IF(B{row}<2,"",STDEV.S({value_range}))'
        for column_index in range(1, 8):
            _formula_style(worksheet.cell(row, column_index))

    worksheet.freeze_panes = "C16"
    worksheet.sheet_view.showGridLines = False
    widths = (18, 28, 10, 11, 12, 11, 11, 12, 11, 13, 12, 11, 10, 10, 11, 11, 10, 11, 12, 20)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.auto_filter.ref = "A15:T21"
    worksheet.conditional_formatting.add(
        "D6:D11",
        FormulaRule(formula=['OR(D6="未达下限",D6="检查缺失或审核",D6="检查缺失")'], fill=PatternFill("solid", fgColor=_PALE_RED)),
    )


def _analysis_result_row(
    worksheet: Any,
    row: int,
    outcome: str,
    d3_column: str,
    d4_column: str,
) -> None:
    """写一行与 Python ``paired_result`` 定义完全一致的实时公式。"""

    ea_range = f'Derived!${d3_column}$203:${d3_column}$226'
    oe_range = f'Derived!${d3_column}$227:${d3_column}$250'
    diff_range = f'Derived!${d4_column}$254:${d4_column}$277'
    formulas = (
        outcome,
        OUTCOME_LABELS[outcome],
        f'=MIN(COUNT({oe_range}),COUNT({ea_range}),COUNT({diff_range}))',
        f'=IF(C{row}=0,"",QUARTILE.INC({oe_range},1))',
        f'=IF(C{row}=0,"",MEDIAN({oe_range}))',
        f'=IF(C{row}=0,"",QUARTILE.INC({oe_range},3))',
        f'=IF(C{row}=0,"",QUARTILE.INC({ea_range},1))',
        f'=IF(C{row}=0,"",MEDIAN({ea_range}))',
        f'=IF(C{row}=0,"",QUARTILE.INC({ea_range},3))',
        f'=IF(C{row}=0,"",MEDIAN({diff_range}))',
        f'=IF(C{row}=0,"",AVERAGE({diff_range}))',
        f'=IF(C{row}<2,"",STDEV.S({diff_range}))',
        f'=IF(OR(C{row}<2,L{row}=0),"",K{row}/L{row})',
    )
    for column_index, value in enumerate(formulas, start=1):
        worksheet.cell(row, column_index).value = value
        _formula_style(worksheet.cell(row, column_index))
    for column_index in range(14, 21):
        _offline_style(worksheet.cell(row, column_index))


def _complete_average_formula(references: tuple[str, ...]) -> str:
    """返回只有全部条目有数值时才计算均值的公式。"""

    arguments = ",".join(references)
    return f'=IF(COUNT({arguments})={len(references)},AVERAGE({arguments}),"")'


def _straightline_formula(source_row: int, items: tuple[str, ...]) -> str:
    """返回当前冻结问卷条目上的五项连续同分门禁公式。"""

    columns = tuple(BLOCK_RECORD_COLUMNS[item] for item in items)
    windows = []
    for start in range(len(columns) - 4):
        window = columns[start : start + 5]
        equality = ",".join(
            f'Records!{left}{source_row}=Records!{right}{source_row}'
            for left, right in zip(window, window[1:])
        )
        windows.append(f"AND({equality})")
    return (
        f'=IF(COUNT({",".join(f"Records!{column}{source_row}" for column in columns)})<{len(columns)},"",'
        f'IF(OR({",".join(windows)}),"是","否"))'
    )


def _title_row(worksheet: Any, row: int, last_column: int, text: str) -> None:
    """写深色首行标题。"""

    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
    cell = worksheet.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor=_NAVY)
    cell.font = Font(color=_WHITE, bold=True, size=15)
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[row].height = 28


def _section_row(worksheet: Any, row: int, last_column: int, text: str) -> None:
    """写青绿色分节标题。"""

    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
    cell = worksheet.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor=_TEAL)
    cell.font = Font(color=_WHITE, bold=True, size=11)
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[row].height = 23


def _header_row(worksheet: Any, row: int, headers: tuple[str, ...]) -> None:
    """写浅蓝色表头并设置自动换行。"""

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row, column, header)
        cell.fill = PatternFill("solid", fgColor=_PALE_BLUE)
        cell.font = Font(color=_TEXT, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    worksheet.row_dimensions[row].height = 30


def _formula_style(cell: Any) -> None:
    """设置实时公式区域的绿色底纹。"""

    cell.fill = PatternFill("solid", fgColor=_GREEN)
    cell.font = Font(color=_TEXT, size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=False)
    cell.border = _thin_border()
    if cell.column >= 3:
        cell.number_format = "0.000"


def _offline_style(cell: Any) -> None:
    """设置 Python 离线推断列的黄色占位底纹。"""

    cell.fill = PatternFill("solid", fgColor=_YELLOW)
    cell.font = Font(color=_TEXT, size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _thin_border()
    cell.number_format = "0.0000"


def _thin_border() -> Border:
    """返回统一的浅灰细边框。"""

    side = Side(style="thin", color=_GRID)
    return Border(left=side, right=side, top=side, bottom=side)


def _add_whole_number_validation(
    worksheet: Any,
    range_reference: str,
    minimum: int,
    maximum: int,
    *,
    allow_blank: bool,
) -> None:
    """添加整数刻度输入校验。"""

    validation = DataValidation(
        type="whole",
        operator="between",
        formula1=str(minimum),
        formula2=str(maximum),
        allow_blank=allow_blank,
    )
    validation.error = f"请输入 {minimum}–{maximum} 的整数" + ("，或留空" if allow_blank else "")
    validation.errorTitle = "评分范围错误"
    validation.showErrorMessage = True
    worksheet.add_data_validation(validation)
    validation.add(range_reference)


def _add_list_validation(
    worksheet: Any,
    range_reference: str,
    formula: str,
    *,
    allow_blank: bool,
) -> None:
    """添加有限选项下拉校验。"""

    validation = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
    validation.error = "请从下拉列表中选择"
    validation.errorTitle = "选项错误"
    validation.showErrorMessage = True
    worksheet.add_data_validation(validation)
    validation.add(range_reference)


def _add_custom_scale_validation(
    worksheet: Any,
    range_reference: str,
    minimum: int,
    maximum: int,
) -> None:
    """允许整数评分、空白或“无法回答”。"""

    first = range_reference.split(":", maxsplit=1)[0]
    formula = f'=OR({first}="",{first}="无法回答",AND(ISNUMBER({first}),{first}>={minimum},{first}<={maximum},MOD({first},1)=0))'
    validation = DataValidation(type="custom", formula1=formula, allow_blank=True)
    validation.error = f"请输入 {minimum}–{maximum} 的整数、无法回答，或留空"
    validation.errorTitle = "评分范围错误"
    validation.showErrorMessage = True
    worksheet.add_data_validation(validation)
    validation.add(range_reference)


def _add_custom_na_validation(worksheet: Any, range_reference: str) -> None:
    """允许偏好强度 1–7、N/A 或空白。"""

    first = range_reference.split(":", maxsplit=1)[0]
    formula = f'=OR({first}="",{first}="N/A",AND(ISNUMBER({first}),{first}>=1,{first}<=7,MOD({first},1)=0))'
    validation = DataValidation(type="custom", formula1=formula, allow_blank=True)
    validation.error = "请输入 1–7 的整数；无明显偏好时填 N/A 或留空"
    validation.errorTitle = "偏好强度错误"
    validation.showErrorMessage = True
    worksheet.add_data_validation(validation)
    validation.add(range_reference)


def _enable_recalculation(workbook: Any) -> None:
    """要求 Excel 打开时完整重算所有实时公式。"""

    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def _assert_no_synthetic_boilerplate(workbook: Any) -> None:
    """拒绝把模拟模型标语带入正式原始模板。"""

    forbidden = ("synthetic /", "模拟分析结果", "合成数据：", "gpt-5.6-thinking", "claude-opus")
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and any(term in cell.value.lower() for term in forbidden):
                    raise ValueError(f"正式原始模板仍含模拟标语：{worksheet.title}!{cell.coordinate}")


__all__ = ["build_raw_template"]
