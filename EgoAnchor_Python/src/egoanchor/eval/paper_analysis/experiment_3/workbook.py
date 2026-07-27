"""实验三离线分析结果工作簿的写入与回读校验。"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.formatting.rule import CellIsRule  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from .contracts import AnalysisTables, Exp3Data, ScoreData
from .settings import Exp3Settings


_NAVY = "18324A"
_TEAL = "2F6F73"
_PALE_BLUE = "DCEAF5"
_GREEN = "E2F0D9"
_YELLOW = "FFF2CC"
_RED = "FCE8E6"
_WHITE = "FFFFFF"
_TEXT = "22313F"
_GRID = "C8D1D9"
"""结果工作簿配色。"""


def write_results_workbook(
    destination: Path,
    *,
    data: Exp3Data,
    scores: ScoreData,
    tables: AnalysisTables,
    clmm_coefficients: pd.DataFrame,
    clmm_contrasts: pd.DataFrame,
    settings: Exp3Settings,
    settings_sha256: str,
    validation: dict[str, Any],
) -> Path:
    """写入包含描述统计、推断、模型、QC 与绘图数据的结果 XLSX。"""

    output = destination.expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    readme_rows = _readme_rows(data, settings, settings_sha256, validation)
    _write_key_value_sheet(workbook, "README", readme_rows)
    _write_frame(workbook, "Main_Results", tables.primary, significant_column="p_Holm")
    _write_frame(workbook, "Scale_Results", tables.scales, significant_column="p_Holm")
    _write_frame(workbook, "Secondary", tables.secondary, significant_column="p_raw")
    _write_frame(workbook, "Reliability", tables.reliability)
    _write_frame(workbook, "By_Object", tables.objects)
    _write_frame(workbook, "CLMM", clmm_coefficients, significant_column="p_raw")
    _write_frame(workbook, "CLMM_Contrasts", clmm_contrasts, significant_column="p_Holm_Conditional")
    _write_frame(workbook, "Manipulation", tables.manipulation, significant_column="p_TOST")
    _write_frame(workbook, "Choices", tables.choices)
    _write_frame(workbook, "Choice_Cross", tables.choice_cross)
    _write_frame(workbook, "Open_Coding", tables.open_coding)
    _write_frame(workbook, "Plot_Paired", tables.plot_paired)
    _write_frame(workbook, "Plot_Scales", tables.plot_scales)
    _write_frame(workbook, "Derived_Block", scores.block_scores)
    _write_frame(workbook, "Derived_Method", scores.method_scores)
    _write_frame(workbook, "Derived_Paired", scores.paired_scores)
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.save(output)
    workbook.close()
    _verify_results_workbook(output, data.source_sha256, len(validation["included_participants"]))
    return output


def _readme_rows(
    data: Exp3Data,
    settings: Exp3Settings,
    settings_digest: str,
    validation: dict[str, Any],
) -> tuple[tuple[str, Any], ...]:
    """生成结果首屏的来源、规则与诚实边界。"""

    warnings = "；".join(validation.get("warnings", ())) or "无"
    return (
        ("EgoAnchor 实验三分析结果", "由原始 Records 独立重算；不读取原始模板的公式缓存"),
        ("输入工作簿", data.source_path),
        ("输入 SHA-256", data.source_sha256),
        ("来源类型", data.source_kind),
        ("批处理配置", str(settings.paths.batch_config_path)),
        ("论文参数配置", str(settings.paths.paper_config_path)),
        ("参数 SHA-256", settings_digest),
        ("模板版本", settings.template_version),
        ("纳入参与者", validation["included_count"]),
        ("AQ 模式", settings.aq_mode),
        ("Q10", "启用" if settings.q10_enabled else "未启用"),
        ("Wilcoxon", "双侧；零差删除；并列秩精确符号置换；家族内 Holm"),
        ("汇总单位", "每位参与者先在三个对象上取均值，再形成 EgoAnchor−One-Euro 配对差"),
        ("CLMM", "逐条目随机截距累积 logit；报告收敛、梯度和交互 LRT"),
        ("TOST", "仅在预实验冻结正等价界后启用" if settings.equivalence_enabled else "未运行：等价界尚未冻结"),
        ("当前样本信度", "只对 AQ、TiA 与 S-TIAS 报告；不宣称验证改编量表"),
        ("警告", warnings),
        ("发布边界", "实验三只报告主观评价和无需真值的自参考日志，不提供客观任务表现证据"),
    )


def _write_key_value_sheet(
    workbook: Workbook,
    name: str,
    rows: tuple[tuple[str, Any], ...],
) -> None:
    """写入带深色标题和分隔线的首屏事实表。"""

    worksheet = workbook.create_sheet(name)
    worksheet.merge_cells("A1:B1")
    worksheet["A1"] = str(rows[0][0])
    worksheet["A1"].fill = PatternFill("solid", fgColor=_NAVY)
    worksheet["A1"].font = Font(color=_WHITE, bold=True, size=15)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 28
    worksheet["A2"] = "说明"
    worksheet["B2"] = rows[0][1]
    for row_index, (key, value) in enumerate(rows[1:], start=3):
        worksheet.cell(row_index, 1, key)
        worksheet.cell(row_index, 2, _excel_value(value))
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=2):
        row[0].fill = PatternFill("solid", fgColor=_PALE_BLUE)
        row[0].font = Font(color=_TEXT, bold=True, size=10)
        for cell in row:
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 98
    worksheet.freeze_panes = "A3"
    worksheet.sheet_view.showGridLines = False


def _write_frame(
    workbook: Workbook,
    name: str,
    frame: pd.DataFrame,
    *,
    significant_column: str | None = None,
) -> None:
    """把 DataFrame 写为普通单元格安全表，并应用统一格式。"""

    worksheet = workbook.create_sheet(name)
    columns = [str(column) for column in frame.columns]
    if not columns:
        worksheet["A1"] = "No rows"
        return
    for column_index, column in enumerate(columns, start=1):
        cell = worksheet.cell(1, column_index, column)
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = Font(color=_WHITE, bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
    for row_index, values in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column_index, _excel_value(value))
            cell.fill = PatternFill("solid", fgColor=_GREEN if row_index % 2 == 0 else "F4F8F7")
            cell.font = Font(color=_TEXT, size=9)
            cell.border = _border()
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(value, str) and len(value) > 35)
            if isinstance(value, (float, np.floating)):
                cell.number_format = "0.0000"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(frame) + 1)}"
    worksheet.row_dimensions[1].height = 32
    for column_index, column in enumerate(columns, start=1):
        sample = [len(column)]
        if not frame.empty:
            sample.extend(len(str(value)) for value in frame.iloc[:100, column_index - 1] if pd.notna(value))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(42, max(10, max(sample) + 2))
    worksheet.sheet_view.showGridLines = False
    if significant_column in columns and len(frame):
        column_letter = get_column_letter(columns.index(significant_column) + 1)
        cells = f"{column_letter}2:{column_letter}{len(frame) + 1}"
        worksheet.conditional_formatting.add(
            cells,
            CellIsRule(operator="lessThan", formula=["0.05"], fill=PatternFill("solid", fgColor=_YELLOW)),
        )
    if "Converged" in columns and len(frame):
        column_letter = get_column_letter(columns.index("Converged") + 1)
        worksheet.conditional_formatting.add(
            f"{column_letter}2:{column_letter}{len(frame) + 1}",
            CellIsRule(operator="equal", formula=["FALSE"], fill=PatternFill("solid", fgColor=_RED)),
        )


def _excel_value(value: Any) -> Any:
    """把 pandas/numpy 值转换为 openpyxl 可安全写入的标量。"""

    if value is None or (isinstance(value, (float, np.floating)) and not math.isfinite(float(value))):
        return None
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, (tuple, list, dict)):
        return str(value)
    return value


def _border() -> Border:
    """返回统一浅灰细边框。"""

    side = Side(style="thin", color=_GRID)
    return Border(left=side, right=side, top=side, bottom=side)


def _verify_results_workbook(path: Path, input_sha256: str, included_count: int) -> None:
    """回读关键表和来源字段，避免静默写出不完整结果。"""

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        required = {
            "README", "Main_Results", "Scale_Results", "Reliability", "By_Object", "CLMM",
            "Manipulation", "Choices", "Open_Coding", "Plot_Paired", "Plot_Scales",
        }
        missing = required.difference(workbook.sheetnames)
        if missing:
            raise ValueError(f"实验三结果工作簿缺少工作表：{sorted(missing)}")
        readme = workbook["README"]
        values = {str(readme.cell(row, 1).value): readme.cell(row, 2).value for row in range(2, readme.max_row + 1)}
        if values.get("输入 SHA-256") != input_sha256:
            raise ValueError("实验三结果工作簿的输入摘要回读不一致")
        if int(values.get("纳入参与者") or 0) != included_count:
            raise ValueError("实验三结果工作簿的纳入人数回读不一致")
    finally:
        workbook.close()


__all__ = ["write_results_workbook"]
