"""Stage 1 schema-v2 task 到完整 XLSX 的原子写出和独立回读验证。"""

from __future__ import annotations

import hashlib
import json
import math
import os
import time
import uuid
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.cell import WriteOnlyCell  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from ..contracts import SHEET_CONTRACTS, SHEET_NAMES, SheetContract
from .provenance import (
    SourceFileRecord,
    collect_source_files,
    file_sha256,
    reproducible_generated_at,
    source_set_sha256,
    stable_workbook_id,
)
from .qc import StageOneQcReport, run_task_qc
from .reader import SourceRow, TaskDataset, flatten_json, read_task


EXCEL_MAX_DATA_ROWS = 1_048_575
"""Excel 单 sheet 扣除表头后允许的最大事实行数。"""

EXCEL_MAX_CELL_CHARS = 32_767
"""Excel 单元格字符串的硬上限。"""

LARGE_VALUE_MARKER = "@large:"
"""工作簿内引用 large_values 分片的稳定前缀。"""

EMPTY_TEXT_MARKER = "@empty-text"
"""Excel 无法区分空字符串和空单元格时使用的稳定标记。"""

LITERAL_TEXT_MARKER = "@literal-text:"
"""原始文本与内部标记前缀冲突时使用的转义前缀。"""

_LITERAL_TEXT_COLUMNS_KEY = "@literal_text_columns"
"""回读时记录经过字面量转义的列，避免把原始文本误判为内部 marker。"""

_WINDOWS_FILE_RETRY_ATTEMPTS = 12
"""Windows XLSX 文件操作遇到短暂共享锁时允许的最大尝试次数。"""

_WINDOWS_FILE_RETRY_DELAY_SECONDS = 0.25
"""Windows XLSX 文件操作每次重试前的基础等待秒数。"""


class WorkbookValidationError(ValueError):
    """工作簿输入、输出或回读契约不满足时抛出的错误。"""


@dataclass(frozen=True, slots=True)
class WorkbookVerification:
    """一次独立 XLSX 回读验证的可审计结果。"""

    path: Path
    """已经验证的工作簿路径。"""

    passed: bool
    """所有结构、主外键、类型和摘要检查是否通过。"""

    logical_row_counts: Mapping[str, int]
    """按逻辑 sheet 汇总的事实行数。"""

    physical_sheets: Mapping[str, tuple[str, ...]]
    """逻辑 sheet 到物理分片 sheet 名称的映射。"""

    source_set_sha256: str
    """source_files 重算得到的来源集合摘要。"""

    large_value_count: int
    """large_values 中完整值的数量。"""


@dataclass(frozen=True, slots=True)
class WorkbookArtifact:
    """原子发布成功后的工作簿路径、摘要和回读结果。"""

    path: Path
    """最终发布路径。"""

    sha256: str
    """规范化 XLSX 二进制的 SHA-256。"""

    source_set_sha256: str
    """输入目录全部文件的稳定来源集合摘要。"""

    verification: WorkbookVerification
    """发布前完成的独立回读验证。"""


@dataclass(slots=True)
class _Partition:
    """一个逻辑 sheet 的物理分片及其写出统计。"""

    logical_sheet: str
    """所属逻辑 sheet 名称。"""

    physical_sheet: str
    """XLSX 内的实际 sheet 名称。"""

    partition_index: int
    """从一开始的稳定分片编号。"""

    row_count: int
    """不含表头的事实行数。"""

    column_count: int
    """表头列数量。"""

    header_sha256: str
    """稳定表头 JSON 的 SHA-256。"""


class _LargeValues:
    """集中管理超出 Excel 限制的规范化值和分片记录。"""

    def __init__(self, max_cell_chars: int) -> None:
        """初始化分片阈值与去重记录。"""

        if max_cell_chars < 1:
            raise ValueError("max_cell_chars 必须为正整数")
        self.max_cell_chars = max_cell_chars
        """每个 inline 或 chunk 单元格允许的最大字符数。"""

        self._rows: dict[tuple[str, str, int, str], list[dict[str, Any]]] = {}
        """按来源定位键保存的大值分片。"""

    def store(
        self,
        source_table: str,
        source_file: str,
        source_line: int,
        json_path: str,
        text: str,
    ) -> tuple[str, str, str]:
        """内联短值；长值写入分片表并返回标记、存储位置和摘要。"""

        digest = hashlib.sha256(text.encode("utf-8")).hexdigest()
        if len(text) <= self.max_cell_chars:
            return text, "inline", digest
        key = (source_table, source_file, source_line, json_path)
        existing = self._rows.get(key)
        if existing is None:
            chunks = [text[index : index + self.max_cell_chars] for index in range(0, len(text), self.max_cell_chars)]
            self._rows[key] = [
                {
                    "source_table": source_table,
                    "source_file": source_file,
                    "source_line": source_line,
                    "json_path": json_path,
                    "chunk_index": index + 1,
                    "value_sha256": digest,
                    "char_count": len(text),
                    "byte_count": len(text.encode("utf-8")),
                    "chunk_text": chunk,
                }
                for index, chunk in enumerate(chunks)
            ]
        elif existing[0]["value_sha256"] != digest:
            raise WorkbookValidationError(f"同一大值来源定位重复且内容不同：{key}")
        return f"{LARGE_VALUE_MARKER}{digest}", "large_values", digest

    def rows(self) -> Iterator[Mapping[str, Any]]:
        """按来源定位和分片编号稳定返回全部 large_values 行。"""

        for key in sorted(self._rows):
            yield from self._rows[key]


def write_task_workbook(
    task: TaskDataset | str | Path,
    output_path: str | Path,
    *,
    code_version: str = "unknown",
    max_data_rows: int = EXCEL_MAX_DATA_ROWS,
    max_cell_chars: int = EXCEL_MAX_CELL_CHARS,
) -> WorkbookArtifact:
    """QC 通过后原子写出一个完整 workbook-v2，并在替换前独立回读验证。"""

    if max_data_rows < 1:
        raise ValueError("max_data_rows 必须为正整数")
    dataset = task if isinstance(task, TaskDataset) else read_task(task)
    report = run_task_qc(dataset)
    if not report.passed:
        raise WorkbookValidationError("Stage 1 QC 失败，禁止发布 workbook。")

    destination = Path(output_path).expanduser()
    if destination.suffix.lower() != ".xlsx":
        raise ValueError("工作簿输出必须使用 .xlsx 扩展名")
    root = dataset.root.resolve()
    if destination.resolve().is_relative_to(root):
        raise WorkbookValidationError("禁止在只读 task 目录内发布工作簿。")
    destination.parent.mkdir(parents=True, exist_ok=True)

    source_files = collect_source_files(dataset.root)
    source_digest = source_set_sha256(source_files)
    large_values = _LargeValues(max_cell_chars)
    factory = _RowFactory(dataset, report, source_files, source_digest, code_version, large_values)
    token = uuid.uuid4().hex
    raw_path = destination.parent / f".{destination.name}.{token}.raw.xlsx"
    normalized_path = destination.parent / f".{destination.name}.{token}.tmp"
    try:
        _write_workbook(raw_path, factory, max_data_rows)
        _normalize_xlsx_archive(raw_path, normalized_path)
        _unlink_temporary_file(raw_path)
        verification = verify_task_workbook(normalized_path, max_data_rows=max_data_rows, max_cell_chars=max_cell_chars)
        current_source_files = collect_source_files(dataset.root)
        if current_source_files != source_files or source_set_sha256(current_source_files) != source_digest:
            raise WorkbookValidationError("来源文件在 QC/hash 与发布之间发生变化，禁止发布混合版本工作簿。")
        _fsync_file(normalized_path)
        _replace_file(normalized_path, destination)
        verification = replace(verification, path=destination)
        return WorkbookArtifact(destination, file_sha256(destination), source_digest, verification)
    except Exception:
        _unlink_temporary_file(raw_path)
        _unlink_temporary_file(normalized_path)
        raise


def verify_task_workbook(
    workbook_path: str | Path,
    *,
    max_data_rows: int = EXCEL_MAX_DATA_ROWS,
    max_cell_chars: int = EXCEL_MAX_CELL_CHARS,
) -> WorkbookVerification:
    """只读取 XLSX 本身，验证分片、表头、键、类型、摘要和大值引用。"""

    if max_data_rows < 1 or max_cell_chars < 1:
        raise ValueError("验证容量参数必须为正整数")
    path = Path(workbook_path).expanduser()
    if not path.is_file():
        raise FileNotFoundError(f"工作簿不存在：{path}")
    stream = path.open("rb")
    workbook = load_workbook(stream, read_only=True, data_only=False)
    try:
        if "sheet_index" not in workbook.sheetnames:
            raise WorkbookValidationError("工作簿缺少 sheet_index。")
        index_rows = _read_sheet_rows(workbook["sheet_index"])
        contracts = {contract.name: contract for contract in SHEET_CONTRACTS}
        physical_sheets, indexed_row_counts = _parse_sheet_index(index_rows, workbook.sheetnames, contracts)
        expected_names = set(SHEET_NAMES)
        if set(physical_sheets) != expected_names:
            raise WorkbookValidationError("sheet_index 的逻辑 sheet 集合与 workbook-v2 契约不一致。")
        logical_rows: dict[str, int] = {}
        primary_keys: dict[str, set[tuple[Any, ...]]] = {}
        deferred_foreign_rows: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
        audit_rows: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
        normalized_sheets = {"metadata_kv", "candidate_diag", "event_payload", "row_kv", "large_values"}
        for logical_name in SHEET_NAMES:
            contract = contracts[logical_name]
            keys: set[tuple[Any, ...]] = set()
            row_count = 0
            for row_count, row in enumerate(
                _iter_logical_rows(workbook, contract, physical_sheets[logical_name], max_cell_chars),
                start=1,
            ):
                key = _primary_key(contract, row)
                if key in keys:
                    raise WorkbookValidationError(f"{logical_name} 主键重复。")
                keys.add(key)
                if all(foreign_key.ref_sheet in primary_keys for foreign_key in contract.foreign_keys):
                    _verify_row_foreign_keys(logical_name, contract, row, primary_keys)
                elif contract.foreign_keys:
                    deferred_foreign_rows[logical_name].append(row)
                if logical_name in normalized_sheets or logical_name in {"source_files", "provenance", "qc_checks"}:
                    audit_rows[logical_name].append(row)
                elif any(
                    isinstance(value, str) and value.startswith(LARGE_VALUE_MARKER)
                    for value in row.values()
                ):
                    audit_row = dict(row)
                    audit_row["@logical_row_index"] = row_count
                    audit_rows[logical_name].append(audit_row)
            logical_rows[logical_name] = row_count
            if row_count != indexed_row_counts[logical_name]:
                raise WorkbookValidationError(f"{logical_name} 实际行数与 sheet_index 不一致。")
            primary_keys[logical_name] = keys
        _verify_foreign_keys(contracts, deferred_foreign_rows, primary_keys)
        _verify_large_values(audit_rows["large_values"], audit_rows, max_cell_chars)
        _verify_source_digest(audit_rows["source_files"], audit_rows["provenance"])
        _verify_fact_row_counts(audit_rows["source_files"], logical_rows)
        _verify_qc_rows(audit_rows["qc_checks"])
        provenance = audit_rows["provenance"]
        large_digests = {row["value_sha256"] for row in audit_rows["large_values"]}
        return WorkbookVerification(
            path=path,
            passed=True,
            logical_row_counts=logical_rows,
            physical_sheets=physical_sheets,
            source_set_sha256=str(provenance[0]["source_set_sha256"]),
            large_value_count=len(large_digests),
        )
    finally:
        workbook.close()
        stream.close()


def decode_workbook_text(value: str) -> str:
    """把 workbook 内的空文本或保留前缀编码还原为原始文本。"""

    if value == EMPTY_TEXT_MARKER:
        return ""
    if value.startswith(LITERAL_TEXT_MARKER):
        return value[len(LITERAL_TEXT_MARKER) :]
    return value


class _RowFactory:
    """按 workbook-v2 契约流式生成各逻辑 sheet 的行。"""

    def __init__(
        self,
        dataset: TaskDataset,
        report: StageOneQcReport,
        source_files: tuple[SourceFileRecord, ...],
        source_digest: str,
        code_version: str,
        large_values: _LargeValues,
    ) -> None:
        """保存只读输入、QC、来源摘要和大值存储。"""

        self.dataset = dataset
        """只读 schema-v2 task。"""

        self.report = report
        """已经通过的 Stage 1 QC 报告。"""

        self.source_files = source_files
        """递归来源文件清单。"""

        self.source_digest = source_digest
        """来源集合 SHA-256。"""

        self.code_version = code_version
        """写入 provenance 的代码版本。"""

        self.large_values = large_values
        """超长值集中分片存储。"""

        self.contracts = {contract.name: contract for contract in SHEET_CONTRACTS}
        """稳定逻辑 sheet 契约索引。"""

        self.file_hashes = {item.relative_path: item.sha256 for item in source_files}
        """相对路径到来源文件摘要的索引。"""

    def rows(self, logical_sheet: str) -> Iterable[Mapping[str, Any]]:
        """按逻辑 sheet 名调用对应行生成器。"""

        method = getattr(self, f"_rows_{logical_sheet}", None)
        if method is None:
            raise WorkbookValidationError(f"没有实现 workbook sheet：{logical_sheet}")
        return method()

    def _rows_README(self) -> Iterable[Mapping[str, Any]]:
        """返回 workbook 使用说明和诚实边界。"""

        return (
            {"section": "contract", "content": "EgoAnchor Stage 1 workbook-v2；所有事实来自只读 schema-v2 task。"},
            {"section": "nulls", "content": "空单元格表示原始值不可用，不等于零、false 或空字符串。"},
            {"section": "units", "content": "位置使用米，时间使用毫秒，旋转角使用度；列级单位见 data_dictionary。"},
            {"section": "reference", "content": "平台控制器 pose 是同一 Quest 时间线的参考，不是外部光学真值。"},
            {"section": "lineage", "content": "source_files、来源行号和 SHA-256 用于重建输入 lineage。"},
        )

    def _rows_provenance(self) -> Iterable[Mapping[str, Any]]:
        """返回单行工作簿 provenance。"""

        manifest = self.dataset.manifest
        session_id = str(manifest["session_id"])
        return (
            {
                "workbook_id": stable_workbook_id(session_id, self.source_digest),
                "session_id": session_id,
                "source_directory": self.dataset.root.resolve().as_posix(),
                "schema_version": int(manifest["schema_version"]),
                "workbook_contract_version": 2,
                "config_hash": str(manifest["config_hash"]),
                "code_version": self.code_version,
                "generated_at_utc": reproducible_generated_at(manifest),
                "input_sha256": self.source_digest,
                "source_set_sha256": self.source_digest,
            },
        )

    def _rows_source_files(self) -> Iterable[Mapping[str, Any]]:
        """返回递归来源文件清单。"""

        return (item.to_dict() for item in self.source_files)

    def _rows_manifest(self) -> Iterable[Mapping[str, Any]]:
        """返回 manifest 的稳定 session 标量。"""

        raw = dict(self.dataset.manifest)
        raw.update({"source_file": "manifest.json", "source_row_sha256": self.file_hashes["manifest.json"]})
        return (self._project("manifest", raw),)

    def _rows_metadata_kv(self) -> Iterator[Mapping[str, Any]]:
        """无损规范化 manifest 与 python_session 的全部 JSON 叶节点。"""

        session_id = str(self.dataset.manifest["session_id"])
        for filename, document in (
            ("manifest.json", self.dataset.manifest),
            ("python_session.json", self.dataset.python_session),
        ):
            source_hash = self.file_hashes[filename]
            for value in flatten_json(document):
                yield {
                    "document": filename,
                    "session_id": session_id,
                    **self._normalized("metadata_kv", filename, 1, value.json_path, value.value_type, value.value),
                    "source_file": filename,
                    "source_line": 1,
                    "source_row_sha256": source_hash,
                }

    def _rows_variants(self) -> Iterator[Mapping[str, Any]]:
        """合并 manifest variant definition 与 config。"""

        manifest = self.dataset.manifest
        session_id = str(manifest["session_id"])
        configs = {str(item["label"]): item for item in manifest["variant_configs"]}
        source_hash = self.file_hashes["manifest.json"]
        for definition in manifest["variant_definitions"]:
            variant_id = str(definition["variant_id"])
            yield self._project(
                "variants",
                {
                    **configs[variant_id],
                    **definition,
                    "session_id": session_id,
                    "source_file": "manifest.json",
                    "source_row_sha256": source_hash,
                },
            )

    def _rows_trial_plan(self) -> Iterator[Mapping[str, Any]]:
        """返回一到五的冻结任务计划。"""

        manifest = self.dataset.manifest
        source_hash = self.file_hashes["manifest.json"]
        for index, item in enumerate(manifest["trial_plan"], start=1):
            scenario_id = str(item.get("scenario_id") or "")
            experiment_id = str(item.get("experiment_id") or "")
            yield self._project(
                "trial_plan",
                {
                    **item,
                    "session_id": manifest["session_id"],
                    "task_number": index,
                    "condition_id": item.get("condition_id") or f"{experiment_id}/{scenario_id}",
                    "task_label": item.get("task_label") or "",
                    "task_description": item.get("task_description") or "",
                    "source_file": "manifest.json",
                    "source_row_sha256": source_hash,
                },
            )

    def _rows_completed_trials(self) -> Iterator[Mapping[str, Any]]:
        """从合并 events 中的最终 trial_ended 行恢复完成 trial 与来源。"""

        wanted = {
            (str(item["experiment_id"]), str(item["scenario_id"]), str(item["trial_id"]))
            for item in self.dataset.manifest["completed_tasks"]
        }
        for source_row in self.dataset.iter_rows("events"):
            row = source_row.data
            key = (str(row.get("experiment_id") or ""), str(row.get("scenario_id") or ""), str(row.get("trial_id") or ""))
            if row.get("event") != "trial_ended" or key not in wanted:
                continue
            raw_payload = row.get("payload")
            payload: Mapping[str, Any] = raw_payload if isinstance(raw_payload, Mapping) else {}
            yield self._project(
                "completed_trials",
                {
                    **row,
                    "condition_id": payload.get("condition_id") or "",
                    "event_row_id": self._event_row_id(source_row),
                    **self._source_fields(source_row),
                },
            )

    def _rows_writer_stats(self) -> Iterator[Mapping[str, Any]]:
        """把两端停止态 writer 统计合并为统一文件粒度长表。"""

        session_id = str(self.dataset.manifest["session_id"])
        python_stats = self.dataset.python_session["log_writer_stats"]
        manifest_stats = self.dataset.manifest["log_writer_stats"]
        sources = (
            ("python_candidates.jsonl", python_stats["python_candidates.jsonl"], "python_session.json", True),
            ("python_events.jsonl", python_stats["python_events.jsonl"], "python_session.json", True),
            ("unity_reference.jsonl", manifest_stats["unity_reference.jsonl"], "manifest.json", False),
            ("unity_admission.jsonl", manifest_stats["unity_admission.jsonl"], "manifest.json", False),
            ("unity_render.jsonl", manifest_stats["unity_render.jsonl"], "manifest.json", False),
            ("unity_events.jsonl", manifest_stats["events.jsonl"]["unity"], "manifest.json", False),
        )
        for filename, stats, source_file, is_python in sources:
            yield self._project(
                "writer_stats",
                {
                    "session_id": session_id,
                    "file_name": filename,
                    "writer_state": "python_stopped" if is_python else "unity_stopped",
                    "rows_written": stats.get("rows_written"),
                    "dropped_rows": stats.get("dropped_rows"),
                    "log_write_failures": stats.get("log_write_failures", 0),
                    "stats_pending": False,
                    "stats_source": source_file,
                    "source_file": source_file,
                    "source_row_sha256": self.file_hashes[source_file],
                },
            )
        yield self._project(
            "writer_stats",
            {
                "session_id": session_id,
                "file_name": "events.jsonl",
                "writer_state": "deterministic_merge_verified",
                "rows_written": self.report.metrics["events.jsonl"],
                "dropped_rows": 0,
                "log_write_failures": 0,
                "stats_pending": False,
                "stats_source": "manifest.json+python_session.json",
                "source_file": "manifest.json",
                "source_row_sha256": self.file_hashes["manifest.json"],
            },
        )

    def _rows_python_candidates(self) -> Iterator[Mapping[str, Any]]:
        """流式标量化 Python candidate 事实行。"""

        yield from self._fact_rows("python_candidates", "python_candidates")

    def _rows_candidate_flags(self) -> Iterator[Mapping[str, Any]]:
        """把 reliability_flags 数组展开为有序子表。"""

        for source_row in self.dataset.iter_rows("python_candidates"):
            for index, flag in enumerate(source_row.data.get("reliability_flags") or []):
                yield self._project(
                    "candidate_flags",
                    {
                        "session_id": source_row.data["session_id"],
                        "candidate_id": source_row.data["candidate_id"],
                        "flag_index": index,
                        "flag": flag,
                        **self._source_fields(source_row),
                    },
                )

    def _rows_candidate_diag(self) -> Iterator[Mapping[str, Any]]:
        """规范化每条 candidate 的 render_diagnostics。"""

        for source_row in self.dataset.iter_rows("python_candidates"):
            for value in flatten_json(source_row.data.get("render_diagnostics") or {}, prefix="render_diagnostics"):
                yield {
                    "session_id": source_row.data["session_id"],
                    "candidate_id": source_row.data["candidate_id"],
                    **self._normalized(
                        "candidate_diag",
                        source_row.source_file,
                        source_row.source_line,
                        value.json_path,
                        value.value_type,
                        value.value,
                    ),
                    **self._source_fields(source_row),
                }

    def _rows_unity_reference(self) -> Iterator[Mapping[str, Any]]:
        """流式标量化平台 reference、HMD 和 camera pose。"""

        yield from self._fact_rows("unity_reference", "unity_reference")

    def _rows_unity_admission(self) -> Iterator[Mapping[str, Any]]:
        """流式标量化 candidate×variant admission 事实。"""

        yield from self._fact_rows("unity_admission", "unity_admission")

    def _rows_unity_render(self) -> Iterator[Mapping[str, Any]]:
        """流式标量化 render tick×variant 输出和显示事实。"""

        yield from self._fact_rows("unity_render", "unity_render")

    def _rows_python_events(self) -> Iterator[Mapping[str, Any]]:
        """返回完整 Python 事件标量事实。"""

        yield from self._event_rows("python_events", "python_events")

    def _rows_unity_events(self) -> Iterator[Mapping[str, Any]]:
        """返回完整 Unity 事件标量事实。"""

        yield from self._event_rows("unity_events", "unity_events")

    def _rows_events(self) -> Iterator[Mapping[str, Any]]:
        """返回确定性合并事件的完整标量事实。"""

        yield from self._event_rows("events", "events")

    def _rows_event_payload(self) -> Iterator[Mapping[str, Any]]:
        """只规范化 canonical events payload，避免两个分片重复。"""

        for source_row in self.dataset.iter_rows("events"):
            payload = source_row.data.get("payload") or {}
            event_role = payload.get("event_role") if isinstance(payload, Mapping) else ""
            for value in flatten_json(payload, prefix="payload"):
                yield {
                    "event_row_id": self._event_row_id(source_row),
                    **self._normalized(
                        "event_payload",
                        source_row.source_file,
                        source_row.source_line,
                        value.json_path,
                        value.value_type,
                        value.value,
                    ),
                    "event_role": event_role or "",
                    **self._source_fields(source_row),
                }

    def _rows_row_kv(self) -> Iterator[Mapping[str, Any]]:
        """保存 typed/child sheets 未声明的未来 JSONL 字段。"""

        child_fields = {
            "python_candidates": {"reliability_flags", "render_diagnostics"},
            "events": {"payload"},
        }
        for table in (
            "python_candidates",
            "unity_reference",
            "unity_admission",
            "unity_render",
            "python_events",
            "unity_events",
            "events",
        ):
            typed_sheet = table
            contract = self.contracts[typed_sheet]
            represented = {
                column.source_path.split("[", 1)[0]
                for column in contract.columns
                if column.source_path and column.source_path not in {"source_file", "source_line", "source_row_sha256"}
            }
            represented.update(child_fields.get(table, set()))
            for source_row in self.dataset.iter_rows(table):
                extras = {key: value for key, value in source_row.data.items() if key not in represented}
                if not extras:
                    continue
                for value in flatten_json(extras):
                    yield {
                        "session_id": source_row.data["session_id"],
                        **self._normalized(
                            "row_kv",
                            source_row.source_file,
                            source_row.source_line,
                            value.json_path,
                            value.value_type,
                            value.value,
                        ),
                        **self._source_fields(source_row),
                    }

    def _rows_large_values(self) -> Iterable[Mapping[str, Any]]:
        """返回此前各规范化 sheet 收集的超长值分片。"""

        return self.large_values.rows()

    def _rows_qc_checks(self) -> Iterator[Mapping[str, Any]]:
        """把总门禁、指标和 warning 写成稳定 QC 长表。"""

        yield {
            "check_id": "stage1_qc",
            "status": "pass",
            "severity": "error",
            "observed": "0 hard errors",
            "expected": "0 hard errors",
            "details": "schema-v2 Stage 1 全量硬 QC 通过。",
            "source_file": "",
            "source_line": None,
        }
        for key in sorted(self.report.metrics):
            yield {
                "check_id": f"metric.{key}",
                "status": "pass",
                "severity": "audit",
                "observed": _canonical_json(self.report.metrics[key]),
                "expected": "recorded",
                "details": "QC 审计指标。",
                "source_file": "",
                "source_line": None,
            }
        for index, warning in enumerate(self.report.warnings, start=1):
            yield {
                "check_id": f"warning.{warning.code}.{index:03d}",
                "status": "warning",
                "severity": "warning",
                "observed": warning.message,
                "expected": "reviewed",
                "details": warning.message,
                "source_file": warning.source_file,
                "source_line": warning.source_line,
            }

    def _rows_data_dictionary(self) -> Iterator[Mapping[str, Any]]:
        """直接从 workbook-v2 contract 发布列级数据字典。"""

        for contract in SHEET_CONTRACTS:
            for column in contract.columns:
                yield {
                    "sheet": contract.name,
                    "column": column.name,
                    "dtype": column.dtype,
                    "unit": column.unit,
                    "nullable": column.nullable,
                    "source_json_path": column.source_path,
                    "description": column.description,
                }

    def _fact_rows(self, table: str, logical_sheet: str) -> Iterator[Mapping[str, Any]]:
        """按 contract source_path 标量投影一个 schema-v2 事实表。"""

        for source_row in self.dataset.iter_rows(table):
            raw = {**source_row.data, **self._source_fields(source_row)}
            yield self._project(logical_sheet, raw)

    def _event_rows(self, table: str, logical_sheet: str) -> Iterator[Mapping[str, Any]]:
        """投影一个事件表并加入稳定 event_row_id。"""

        for source_row in self.dataset.iter_rows(table):
            raw = {
                **source_row.data,
                "event_row_id": self._event_row_id(source_row),
                **self._source_fields(source_row),
            }
            yield self._project(logical_sheet, raw)

    def _project(self, logical_sheet: str, raw: Mapping[str, Any]) -> dict[str, Any]:
        """按列 source_path 从原始映射取值，并保留显式派生列。"""

        contract = self.contracts[logical_sheet]
        projected: dict[str, Any] = {}
        for column in contract.columns:
            if column.name in raw:
                projected[column.name] = raw[column.name]
            else:
                projected[column.name] = _extract_source_value(raw, column.source_path)
        return projected

    def _normalized(
        self,
        source_table: str,
        source_file: str,
        source_line: int,
        json_path: str,
        value_type: str,
        value: Any,
    ) -> dict[str, Any]:
        """编码一个规范化 JSON 值，并在需要时转存 large_values。"""

        text = _canonical_json(value)
        stored, storage, digest = self.large_values.store(
            source_table,
            source_file,
            source_line,
            json_path,
            text,
        )
        return {
            "json_path": json_path,
            "value_type": value_type,
            "value_json": stored,
            "value_storage": storage,
            "value_sha256": digest,
        }

    @staticmethod
    def _source_fields(source_row: SourceRow) -> dict[str, Any]:
        """返回三个固定来源追踪字段。"""

        return {
            "source_file": source_row.source_file,
            "source_line": source_row.source_line,
            "source_row_sha256": source_row.source_row_sha256,
        }

    @staticmethod
    def _event_row_id(source_row: SourceRow) -> str:
        """用 source_file:source_line 构造稳定事件行 ID。"""

        return f"{source_row.source_file}:{source_row.source_line}"


def _extract_source_value(raw: Mapping[str, Any], source_path: str) -> Any:
    """读取 `field` 或 `field[index]` 形式的简单契约来源路径。"""

    if not source_path:
        return None
    if "[" not in source_path:
        return raw.get(source_path)
    field, suffix = source_path.split("[", 1)
    try:
        index = int(suffix.rstrip("]"))
    except ValueError:
        return None
    value = raw.get(field)
    if not isinstance(value, (list, tuple)) or index >= len(value):
        return None
    return value[index]


def _canonical_json(value: Any) -> str:
    """返回禁止 NaN、键有序且无多余空白的 JSON 文本。"""

    return json.dumps(value, ensure_ascii=False, allow_nan=False, sort_keys=True, separators=(",", ":"))


def _write_workbook(path: Path, factory: _RowFactory, max_data_rows: int) -> None:
    """用 write-only openpyxl 写出全部逻辑 sheet 和分片目录。"""

    workbook = Workbook(write_only=True)
    timestamp = reproducible_generated_at(factory.dataset.manifest)
    workbook.properties.created = timestamp
    workbook.properties.modified = timestamp
    partitions: list[_Partition] = []
    deferred = {"large_values", "sheet_index"}
    for contract in SHEET_CONTRACTS:
        if contract.name in deferred:
            continue
        partitions.extend(
            _write_logical_sheet(
                workbook,
                contract,
                factory.rows(contract.name),
                max_data_rows,
                factory.large_values,
            )
        )
    large_contract = factory.contracts["large_values"]
    partitions.extend(
        _write_logical_sheet(
            workbook,
            large_contract,
            factory.rows("large_values"),
            max_data_rows,
            factory.large_values,
        )
    )
    index_contract = factory.contracts["sheet_index"]
    index_header_hash = _header_sha256(index_contract)
    index_row_count = len(partitions) + 1
    index_rows = [
        {
            "logical_sheet": item.logical_sheet,
            "physical_sheet": item.physical_sheet,
            "partition_index": item.partition_index,
            "row_count": item.row_count,
            "column_count": item.column_count,
            "header_sha256": item.header_sha256,
        }
        for item in partitions
    ]
    index_rows.append(
        {
            "logical_sheet": "sheet_index",
            "physical_sheet": "sheet_index",
            "partition_index": 1,
            "row_count": index_row_count,
            "column_count": len(index_contract.columns),
            "header_sha256": index_header_hash,
        }
    )
    _write_logical_sheet(
        workbook,
        index_contract,
        index_rows,
        EXCEL_MAX_DATA_ROWS,
        factory.large_values,
    )
    workbook.save(path)
    workbook.close()


def _write_logical_sheet(
    workbook: Workbook,
    contract: SheetContract,
    rows: Iterable[Mapping[str, Any]],
    max_data_rows: int,
    large_values: _LargeValues,
) -> list[_Partition]:
    """写一个逻辑 sheet；超过上限时重命名首分片并稳定续写。"""

    header_hash = _header_sha256(contract)
    partition_index = 1
    worksheet = _create_worksheet(workbook, contract.name, contract)
    _append_header(worksheet, contract)
    partitions = [_Partition(contract.name, contract.name, 1, 0, len(contract.columns), header_hash)]
    for logical_row_index, row in enumerate(rows, start=1):
        if partitions[-1].row_count >= max_data_rows:
            if partition_index == 1:
                first_name = f"{contract.name}_001"
                worksheet.title = first_name
                partitions[0].physical_sheet = first_name
            partition_index += 1
            physical_name = f"{contract.name}_{partition_index:03d}"
            worksheet = _create_worksheet(workbook, physical_name, contract)
            _append_header(worksheet, contract)
            partitions.append(
                _Partition(contract.name, physical_name, partition_index, 0, len(contract.columns), header_hash)
            )
        cells = [
            _write_cell(
                worksheet,
                contract,
                column,
                row.get(column.name),
                row,
                logical_row_index,
                large_values,
            )
            for column in contract.columns
        ]
        worksheet.append(cells)
        partitions[-1].row_count += 1
    return partitions


def _create_worksheet(workbook: Workbook, name: str, contract: SheetContract) -> Any:
    """创建可滚动审计的物理 sheet，并写入冻结窗格与稳定列宽。"""

    worksheet = workbook.create_sheet(name)
    worksheet.freeze_panes = "A2"
    for index, column in enumerate(contract.columns, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = _column_width(column)
    return worksheet


def _column_width(column: Any) -> float:
    """按列名和逻辑类型返回稳定、可读的 Excel 列宽。"""

    name = str(column.name)
    if "sha256" in name or name.endswith("_hash"):
        preferred = 68
    elif name in {"content", "description", "details", "source_directory"}:
        preferred = 72
    elif "json_path" in name or name.endswith("_path") or name.endswith("_file"):
        preferred = 48
    elif name.endswith("_id"):
        preferred = 30
    elif column.dtype == "datetime":
        preferred = 24
    elif column.dtype in {"int", "float"}:
        preferred = 18
    elif column.dtype == "bool":
        preferred = 12
    else:
        preferred = 24
    return float(max(len(name) + 2, preferred))


def _append_header(worksheet: Any, contract: SheetContract) -> None:
    """写入加粗且强制为文本类型的稳定表头。"""

    cells = []
    for column in contract.columns:
        cell = WriteOnlyCell(worksheet, value=column.name)
        cell.data_type = "s"
        cell.number_format = "@"
        cell.font = Font(bold=True)
        cells.append(cell)
    worksheet.append(cells)


def _write_cell(
    worksheet: Any,
    contract: SheetContract,
    column: Any,
    value: Any,
    row: Mapping[str, Any],
    row_index: int,
    large_values: _LargeValues,
) -> WriteOnlyCell:
    """校验逻辑类型、阻止公式解析，并把超长文本转入 large_values。"""

    if value is None:
        if not column.nullable:
            raise WorkbookValidationError(f"{contract.name}.{column.name} 不允许为空。")
        return WriteOnlyCell(worksheet, value=None)
    if column.dtype in {"text", "json"}:
        text = value if isinstance(value, str) else (_canonical_json(value) if column.dtype == "json" else str(value))
        already_chunked = (
            column.name == "value_json"
            and row.get("value_storage") == "large_values"
            and text.startswith(LARGE_VALUE_MARKER)
        )
        if len(text) > large_values.max_cell_chars and contract.name != "large_values" and not already_chunked:
            source_file = str(row.get("source_file") or f"@workbook/{contract.name}")
            raw_line = row.get("source_line")
            source_line = raw_line if isinstance(raw_line, int) else row_index
            text, _, _ = large_values.store(
                contract.name,
                source_file,
                source_line,
                column.source_path,
                text,
            )
            already_chunked = True
        if column.dtype == "text" and not already_chunked:
            text = _encode_workbook_text(text)
        if len(text) > EXCEL_MAX_CELL_CHARS:
            raise WorkbookValidationError(f"{contract.name}.{column.name} 超过 Excel 单元格限制。")
        cell = WriteOnlyCell(worksheet, value=text)
        cell.data_type = "s"
        cell.number_format = "@"
        return cell
    if column.dtype == "bool":
        if type(value) is not bool:
            raise WorkbookValidationError(f"{contract.name}.{column.name} 必须是 bool。")
        return WriteOnlyCell(worksheet, value=value)
    if column.dtype == "int":
        if type(value) is not int:
            raise WorkbookValidationError(f"{contract.name}.{column.name} 必须是 int。")
        return WriteOnlyCell(worksheet, value=value)
    if column.dtype == "float":
        if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value)):
            raise WorkbookValidationError(f"{contract.name}.{column.name} 必须是有限数值。")
        return WriteOnlyCell(worksheet, value=float(value))
    if column.dtype == "datetime":
        if not isinstance(value, datetime):
            raise WorkbookValidationError(f"{contract.name}.{column.name} 必须是 datetime。")
        cell = WriteOnlyCell(worksheet, value=value)
        cell.number_format = "yyyy-mm-dd hh:mm:ss"
        return cell
    raise WorkbookValidationError(f"未知 XLSX 数据类型：{column.dtype}")


def _encode_workbook_text(value: str) -> str:
    """可逆编码空文本和内部保留前缀，避免 Excel 空值折叠。"""

    if value == "":
        return EMPTY_TEXT_MARKER
    if value == EMPTY_TEXT_MARKER or value.startswith((LARGE_VALUE_MARKER, LITERAL_TEXT_MARKER)):
        return f"{LITERAL_TEXT_MARKER}{value}"
    return value


def _header_sha256(contract: SheetContract) -> str:
    """计算逻辑 sheet 表头顺序的稳定 SHA-256。"""

    encoded = _canonical_json(list(contract.column_names())).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _unlink_temporary_file(path: Path) -> None:
    """删除临时 XLSX；只对 Windows 短暂共享锁做有界重试。"""

    for attempt in range(_WINDOWS_FILE_RETRY_ATTEMPTS):
        try:
            path.unlink(missing_ok=True)
            return
        except PermissionError:
            if attempt + 1 == _WINDOWS_FILE_RETRY_ATTEMPTS:
                raise
            time.sleep(_WINDOWS_FILE_RETRY_DELAY_SECONDS * (attempt + 1))


def _replace_file(source: Path, destination: Path) -> None:
    """原子替换正式 XLSX；Windows 短暂共享锁期间保留旧文件并有界重试。"""

    for attempt in range(_WINDOWS_FILE_RETRY_ATTEMPTS):
        try:
            os.replace(source, destination)
            return
        except PermissionError:
            if attempt + 1 == _WINDOWS_FILE_RETRY_ATTEMPTS:
                raise
            time.sleep(_WINDOWS_FILE_RETRY_DELAY_SECONDS * (attempt + 1))


def _normalize_xlsx_archive(source: Path, destination: Path) -> None:
    """固定 ZIP 条目顺序和时间戳，使相同输入产生相同 XLSX 字节。"""

    fixed_timestamp = (1980, 1, 1, 0, 0, 0)
    with zipfile.ZipFile(source, "r") as input_archive, zipfile.ZipFile(
        destination,
        "w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as output_archive:
        for original in sorted(input_archive.infolist(), key=lambda item: item.filename):
            normalized = zipfile.ZipInfo(original.filename, fixed_timestamp)
            normalized.compress_type = zipfile.ZIP_DEFLATED
            normalized.external_attr = original.external_attr
            normalized.create_system = original.create_system
            normalized.flag_bits = original.flag_bits
            data = input_archive.read(original.filename)
            if original.filename == "docProps/core.xml":
                data = _normalize_core_properties(data)
            output_archive.writestr(normalized, data)


def _normalize_core_properties(data: bytes) -> bytes:
    """固定 OpenPyXL 保存时自动改写的 core 创建和修改时间。"""

    namespaces = {
        "dcterms": "http://purl.org/dc/terms/",
        "xsi": "http://www.w3.org/2001/XMLSchema-instance",
    }
    ET.register_namespace("dcterms", namespaces["dcterms"])
    ET.register_namespace("xsi", namespaces["xsi"])
    root = ET.fromstring(data)
    for tag in ("created", "modified"):
        node = root.find(f"{{{namespaces['dcterms']}}}{tag}")
        if node is not None:
            node.text = "1980-01-01T00:00:00Z"
    return ET.tostring(root, encoding="utf-8", xml_declaration=False)


def _fsync_file(path: Path) -> None:
    """在原子替换前把临时工作簿刷新到存储设备。"""

    with path.open("r+b", buffering=0) as handle:
        handle.flush()
        os.fsync(handle.fileno())


def _read_sheet_rows(worksheet: Any) -> list[Mapping[str, Any]]:
    """读取一个物理 sheet 为表头映射行，并拒绝公式和重复表头。"""

    iterator = worksheet.iter_rows()
    try:
        header_cells = next(iterator)
    except StopIteration as exc:
        raise WorkbookValidationError(f"物理 sheet 没有表头：{worksheet.title}") from exc
    headers = [cell.value for cell in header_cells]
    if any(not isinstance(value, str) or not value for value in headers):
        raise WorkbookValidationError(f"物理 sheet 表头无效：{worksheet.title}")
    if len(headers) != len(set(headers)):
        raise WorkbookValidationError(f"物理 sheet 表头重复：{worksheet.title}")
    rows: list[Mapping[str, Any]] = []
    for cells in iterator:
        if any(cell.data_type == "f" for cell in cells):
            raise WorkbookValidationError(f"工作簿禁止公式单元格：{worksheet.title}")
        rows.append(dict(zip(headers, (cell.value for cell in cells), strict=True)))
    return rows


def _parse_sheet_index(
    rows: Iterable[Mapping[str, Any]],
    workbook_sheetnames: Iterable[str],
    contracts: Mapping[str, SheetContract],
) -> tuple[dict[str, tuple[str, ...]], dict[str, int]]:
    """校验 sheet_index 并返回按分片号排序的逻辑到物理映射。"""

    available = set(workbook_sheetnames)
    grouped: dict[str, list[tuple[int, str]]] = defaultdict(list)
    row_counts: dict[str, int] = defaultdict(int)
    seen_physical: set[str] = set()
    for row in rows:
        logical = str(row.get("logical_sheet") or "")
        physical = str(row.get("physical_sheet") or "")
        partition = row.get("partition_index")
        row_count = row.get("row_count")
        column_count = row.get("column_count")
        if not logical or not physical or type(partition) is not int or partition < 1:
            raise WorkbookValidationError("sheet_index 包含非法逻辑名、物理名或分片号。")
        contract = contracts.get(logical)
        if contract is None:
            raise WorkbookValidationError(f"sheet_index 使用未知逻辑 sheet：{logical}")
        if type(row_count) is not int or row_count < 0:
            raise WorkbookValidationError(f"sheet_index 行数无效：{physical}")
        if column_count != len(contract.columns) or row.get("header_sha256") != _header_sha256(contract):
            raise WorkbookValidationError(f"sheet_index 列数或表头摘要错误：{physical}")
        if physical not in available or physical in seen_physical:
            raise WorkbookValidationError(f"sheet_index 物理 sheet 无效或重复：{physical}")
        seen_physical.add(physical)
        grouped[logical].append((partition, physical))
        row_counts[logical] += row_count
    if seen_physical != available:
        raise WorkbookValidationError("sheet_index 没有覆盖全部物理 sheet。")
    result: dict[str, tuple[str, ...]] = {}
    for logical, items in grouped.items():
        ordered = sorted(items)
        if [index for index, _ in ordered] != list(range(1, len(ordered) + 1)):
            raise WorkbookValidationError(f"{logical} 的分片号不连续。")
        result[logical] = tuple(name for _, name in ordered)
    return result, dict(row_counts)


def _iter_logical_rows(
    workbook: Any,
    contract: SheetContract,
    physical_names: tuple[str, ...],
    max_cell_chars: int,
) -> Iterator[Mapping[str, Any]]:
    """按分片顺序读取逻辑 sheet，并执行表头和单元格类型检查。"""

    expected_headers = list(contract.column_names())
    for physical_name in physical_names:
        worksheet = workbook[physical_name]
        iterator = worksheet.iter_rows()
        try:
            header_cells = next(iterator)
        except StopIteration as exc:
            raise WorkbookValidationError(f"物理 sheet 没有表头：{physical_name}") from exc
        headers = [cell.value for cell in header_cells]
        if headers != expected_headers:
            raise WorkbookValidationError(f"{physical_name} 表头与 {contract.name} 契约不一致。")
        for cells in iterator:
            row: dict[str, Any] = {}
            literal_text_columns: set[str] = set()
            for index, column in enumerate(contract.columns):
                cell = cells[index] if index < len(cells) else None
                _verify_cell(contract.name, column, cell, max_cell_chars)
                value = cell.value if cell is not None else None
                if column.dtype == "text" and isinstance(value, str) and value.startswith(LITERAL_TEXT_MARKER):
                    literal_text_columns.add(column.name)
                row[column.name] = decode_workbook_text(value) if column.dtype == "text" and isinstance(value, str) else value
            if literal_text_columns:
                row[_LITERAL_TEXT_COLUMNS_KEY] = frozenset(literal_text_columns)
            yield row


def _verify_cell(logical_sheet: str, column: Any, cell: Any, max_cell_chars: int) -> None:
    """验证回读单元格没有公式，并符合逻辑 dtype 与容量约束。"""

    value = cell.value if cell is not None else None
    if cell is not None and cell.data_type == "f":
        raise WorkbookValidationError(f"工作簿禁止公式：{logical_sheet}.{column.name}")
    if value is None:
        if not column.nullable:
            raise WorkbookValidationError(f"主键或必填列为空：{logical_sheet}.{column.name}")
        return
    if column.dtype in {"text", "json"}:
        if not isinstance(value, str) or len(value) > EXCEL_MAX_CELL_CHARS:
            raise WorkbookValidationError(f"文本类型或容量错误：{logical_sheet}.{column.name}")
        if logical_sheet == "large_values" and column.name == "chunk_text" and len(value) > max_cell_chars:
            raise WorkbookValidationError("large_values chunk_text 超过声明分片阈值。")
        return
    if column.dtype == "bool" and type(value) is not bool:
        raise WorkbookValidationError(f"bool 类型错误：{logical_sheet}.{column.name}")
    if column.dtype == "int" and type(value) is not int:
        raise WorkbookValidationError(f"int 类型错误：{logical_sheet}.{column.name}")
    if column.dtype == "float" and (
        isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(float(value))
    ):
        raise WorkbookValidationError(f"float 类型错误：{logical_sheet}.{column.name}")
    if column.dtype == "datetime" and not isinstance(value, datetime):
        raise WorkbookValidationError(f"datetime 类型错误：{logical_sheet}.{column.name}")


def _primary_key(contract: SheetContract, row: Mapping[str, Any]) -> tuple[Any, ...]:
    """读取一行主键，并拒绝空值。"""

    key = tuple(row.get(column) for column in contract.primary_key)
    if any(value is None or value == "" for value in key):
        raise WorkbookValidationError(f"{contract.name} 主键包含空值：{key}")
    return key


def _verify_foreign_keys(
    contracts: Mapping[str, SheetContract],
    rows_by_sheet: Mapping[str, list[Mapping[str, Any]]],
    primary_keys: Mapping[str, set[tuple[Any, ...]]],
) -> None:
    """验证引用表在后方时暂存的少量跨 sheet 外键。"""

    for sheet_name, contract in contracts.items():
        for row in rows_by_sheet.get(sheet_name, []):
            _verify_row_foreign_keys(sheet_name, contract, row, primary_keys)


def _verify_row_foreign_keys(
    sheet_name: str,
    contract: SheetContract,
    row: Mapping[str, Any],
    primary_keys: Mapping[str, set[tuple[Any, ...]]],
) -> None:
    """验证一行的全部显式外键。"""

    for foreign_key in contract.foreign_keys:
        referenced = primary_keys[foreign_key.ref_sheet]
        key = tuple(row.get(column) for column in foreign_key.columns)
        if any(value is None or value == "" for value in key):
            raise WorkbookValidationError(f"{sheet_name} 外键为空：{foreign_key.columns}")
        if key not in referenced:
            raise WorkbookValidationError(
                f"{sheet_name} 外键不存在：{foreign_key.columns} -> {foreign_key.ref_sheet}: {key}"
            )


def _verify_large_values(
    large_rows: list[Mapping[str, Any]],
    rows_by_sheet: Mapping[str, list[Mapping[str, Any]]],
    max_cell_chars: int,
) -> None:
    """重组所有大值分片、校验摘要，并确保每个 marker 都可解析。"""

    grouped: dict[tuple[str, str, int, str], list[Mapping[str, Any]]] = defaultdict(list)
    for row in large_rows:
        key = (
            str(row["source_table"]),
            str(row["source_file"]),
            int(row["source_line"]),
            str(row["json_path"]),
        )
        grouped[key].append(row)
    group_digests: dict[tuple[str, str, int, str], str] = {}
    for key, rows in grouped.items():
        ordered = sorted(rows, key=lambda item: int(item["chunk_index"]))
        if [int(row["chunk_index"]) for row in ordered] != list(range(1, len(ordered) + 1)):
            raise WorkbookValidationError(f"large_values 分片号不连续：{key}")
        chunks = [str(row["chunk_text"]) for row in ordered]
        if any(len(chunk) > max_cell_chars for chunk in chunks):
            raise WorkbookValidationError(f"large_values 分片超限：{key}")
        text = "".join(chunks)
        digest = hashlib.sha256(text.encode("utf-8")).hexdigest()
        first = ordered[0]
        byte_count = len(text.encode("utf-8"))
        for row in ordered:
            if (
                row["value_sha256"] != digest
                or row["char_count"] != len(text)
                or row["byte_count"] != byte_count
            ):
                raise WorkbookValidationError(f"large_values 分片元数据不一致：{key}")
        group_digests[key] = digest

    normalized_sheets = ("metadata_kv", "candidate_diag", "event_payload", "row_kv")
    for sheet_name in normalized_sheets:
        for row in rows_by_sheet[sheet_name]:
            storage = row.get("value_storage")
            value = str(row.get("value_json") or "")
            digest = str(row.get("value_sha256") or "")
            if storage == "inline":
                if hashlib.sha256(value.encode("utf-8")).hexdigest() != digest:
                    raise WorkbookValidationError(f"{sheet_name} inline 值摘要不一致。")
                continue
            if storage != "large_values" or not value.startswith(LARGE_VALUE_MARKER):
                raise WorkbookValidationError(f"{sheet_name} 使用未知大值存储状态。")
            marker_digest = value[len(LARGE_VALUE_MARKER) :]
            key = (
                sheet_name,
                str(row["source_file"]),
                int(row["source_line"]),
                str(row["json_path"]),
            )
            if marker_digest != digest or group_digests.get(key) != digest:
                raise WorkbookValidationError(f"{sheet_name} 大值 marker 与来源分片不一致：{key}")

    contracts = {contract.name: contract for contract in SHEET_CONTRACTS}
    for sheet_name, rows in rows_by_sheet.items():
        if sheet_name in {*normalized_sheets, "large_values"}:
            continue
        contract = contracts[sheet_name]
        columns = {column.name: column for column in contract.columns}
        for fallback_row_index, row in enumerate(rows, start=1):
            raw_logical_index = row.get("@logical_row_index")
            row_index = raw_logical_index if type(raw_logical_index) is int else fallback_row_index
            literal_text_columns = row.get(_LITERAL_TEXT_COLUMNS_KEY, ())
            for column_name, value in row.items():
                if column_name not in columns:
                    continue
                if not isinstance(value, str) or not value.startswith(LARGE_VALUE_MARKER):
                    continue
                digest = value[len(LARGE_VALUE_MARKER) :]
                if len(digest) != 64 or any(character not in "0123456789abcdef" for character in digest):
                    continue
                if column_name in literal_text_columns:
                    continue
                source_file = str(row.get("source_file") or f"@workbook/{sheet_name}")
                raw_line = row.get("source_line")
                source_line = raw_line if type(raw_line) is int else row_index
                key = (sheet_name, source_file, source_line, columns[column_name].source_path)
                if group_digests.get(key) != digest:
                    raise WorkbookValidationError(f"{sheet_name} 大值 marker 与来源分片不一致：{key}")


def _verify_source_digest(
    source_rows: list[Mapping[str, Any]],
    provenance_rows: list[Mapping[str, Any]],
) -> None:
    """从 source_files 重算来源集合摘要并核对 provenance。"""

    if len(provenance_rows) != 1:
        raise WorkbookValidationError("provenance 必须恰有一行。")
    records = tuple(
        SourceFileRecord(
            relative_path=str(row["relative_path"]),
            source_kind=str(row["source_kind"]),
            exists=bool(row["exists"]),
            byte_count=int(row["byte_count"]),
            row_count=int(row["row_count"]),
            sha256=str(row["sha256"]),
        )
        for row in source_rows
    )
    digest = source_set_sha256(records)
    provenance = provenance_rows[0]
    if digest != provenance.get("source_set_sha256") or digest != provenance.get("input_sha256"):
        raise WorkbookValidationError("provenance 来源集合摘要与 source_files 不一致。")


def _verify_fact_row_counts(
    source_rows: list[Mapping[str, Any]],
    logical_row_counts: Mapping[str, int],
) -> None:
    """核对七个事实 sheet 的行数与对应 JSONL 来源文件完全一致。"""

    expected_by_file = {str(row["relative_path"]): int(row["row_count"]) for row in source_rows}
    mappings = {
        "python_candidates.jsonl": "python_candidates",
        "unity_reference.jsonl": "unity_reference",
        "unity_admission.jsonl": "unity_admission",
        "unity_render.jsonl": "unity_render",
        "python_events.jsonl": "python_events",
        "unity_events.jsonl": "unity_events",
        "events.jsonl": "events",
    }
    for filename, logical_sheet in mappings.items():
        if filename not in expected_by_file:
            raise WorkbookValidationError(f"source_files 缺少固定事实来源：{filename}")
        if logical_row_counts.get(logical_sheet) != expected_by_file[filename]:
            raise WorkbookValidationError(f"{logical_sheet} 行数与 {filename} 不一致。")


def _verify_qc_rows(rows: list[Mapping[str, Any]]) -> None:
    """要求 qc_checks 含通过的总门禁，且不得包含 fail。"""

    statuses = {str(row.get("check_id")): str(row.get("status")) for row in rows}
    if statuses.get("stage1_qc") != "pass" or "fail" in statuses.values():
        raise WorkbookValidationError("qc_checks 未证明 Stage 1 硬 QC 通过。")


__all__ = [
    "EMPTY_TEXT_MARKER",
    "EXCEL_MAX_CELL_CHARS",
    "EXCEL_MAX_DATA_ROWS",
    "WorkbookArtifact",
    "WorkbookValidationError",
    "WorkbookVerification",
    "decode_workbook_text",
    "verify_task_workbook",
    "write_task_workbook",
]
