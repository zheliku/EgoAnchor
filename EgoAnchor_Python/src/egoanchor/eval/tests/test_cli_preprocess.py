"""Task 5 Stage 1 preprocess CLI 的批处理发布测试。"""

from __future__ import annotations

import contextlib
import io
import json
import shutil
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from egoanchor.eval import verify_task_workbook
from egoanchor.eval import cli as eval_cli

from .test_reader_qc import _write_valid_task


class PreprocessCliTests(unittest.TestCase):
    """验证 preprocess 的批次 QC 门禁、固定命名和审计输出。"""

    def test_preprocess_publishes_a_verified_workbook_with_stable_task_name(self) -> None:
        """通过 QC 的 task 写为 task_N_complete.xlsx，并返回来源与文件摘要。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output_root = Path(tmp) / "complete"
            output = io.StringIO()

            with contextlib.redirect_stdout(output):
                exit_code = eval_cli.main(
                    [
                        "preprocess",
                        str(root),
                        "--out",
                        str(output_root),
                        "--code-version",
                        "test-version",
                    ]
                )

            expected_workbook = output_root / "task_1_complete.xlsx"
            self.assertEqual(exit_code, eval_cli.EXIT_OK)
            self.assertTrue(expected_workbook.is_file())
            self.assertTrue(verify_task_workbook(expected_workbook).passed)
            payload = json.loads(output.getvalue())
            self.assertTrue(payload["passed"])
            self.assertEqual(payload["tasks"][0]["output_workbook"], str(expected_workbook))
            self.assertEqual(len(payload["tasks"][0]["workbook_sha256"]), 64)
            self.assertEqual(len(payload["tasks"][0]["input_sha256"]), 64)
            workbook = load_workbook(expected_workbook, read_only=True, data_only=False)
            provenance_rows = list(workbook["provenance"].iter_rows(values_only=True))
            provenance = dict(zip(provenance_rows[0], provenance_rows[1], strict=True))
            workbook.close()
            self.assertEqual(provenance["code_version"], "test-version")

    def test_preprocess_qc_failure_stops_the_entire_batch_before_writing(self) -> None:
        """批次中任一 task 硬 QC 失败时不发布任何工作簿，并返回退出码二。"""

        with tempfile.TemporaryDirectory() as tmp:
            valid_root = _write_valid_task(Path(tmp))
            invalid_root = valid_root.with_name("task_2_s01_controller_right")
            shutil.copytree(valid_root, invalid_root)
            session_path = invalid_root / "python_session.json"
            session = json.loads(session_path.read_text(encoding="utf-8"))
            session["state"] = "python_running"
            session_path.write_text(json.dumps(session), encoding="utf-8")
            output_root = Path(tmp) / "complete"
            output = io.StringIO()

            with contextlib.redirect_stdout(output):
                exit_code = eval_cli.main(
                    ["preprocess", str(valid_root), str(invalid_root), "--out", str(output_root)]
                )

            self.assertEqual(exit_code, eval_cli.EXIT_DATA_ERROR)
            self.assertFalse(output_root.exists())
            payload = json.loads(output.getvalue())
            self.assertFalse(payload["passed"])
            self.assertTrue(payload["tasks"][0]["passed"])
            self.assertFalse(payload["tasks"][1]["passed"])

    def test_preprocess_missing_task_directory_returns_io_error(self) -> None:
        """不存在的 task 目录属于缺源错误，必须返回退出码一。"""

        with tempfile.TemporaryDirectory() as tmp:
            missing_root = Path(tmp) / "task_1_missing_controller_right"
            output_root = Path(tmp) / "complete"
            output = io.StringIO()

            with contextlib.redirect_stderr(output):
                exit_code = eval_cli.main(
                    ["preprocess", str(missing_root), "--out", str(output_root)]
                )

            self.assertEqual(exit_code, eval_cli.EXIT_IO_ERROR)
            self.assertFalse(output_root.exists())
            self.assertIn("task 目录不存在", output.getvalue())

    def test_preprocess_missing_required_file_returns_io_error(self) -> None:
        """固定输入文件缺失属于缺源错误，必须返回退出码一。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            (root / "events.jsonl").unlink()
            output_root = Path(tmp) / "complete"
            output = io.StringIO()

            with contextlib.redirect_stderr(output):
                exit_code = eval_cli.main(["preprocess", str(root), "--out", str(output_root)])

            self.assertEqual(exit_code, eval_cli.EXIT_IO_ERROR)
            self.assertFalse(output_root.exists())
            self.assertIn("缺少固定文件", output.getvalue())

    def test_preprocess_invalid_schema_returns_data_error(self) -> None:
        """文件齐全但 schema 非法时仍属于数据契约错误，必须返回退出码二。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            manifest_path = root / "manifest.json"
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["schema_version"] = 1
            manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
            output_root = Path(tmp) / "complete"
            output = io.StringIO()

            with contextlib.redirect_stdout(output):
                exit_code = eval_cli.main(["preprocess", str(root), "--out", str(output_root)])

            self.assertEqual(exit_code, eval_cli.EXIT_DATA_ERROR)
            self.assertFalse(output_root.exists())

    def test_preprocess_batch_derives_each_task_number_once(self) -> None:
        """两个不同任务号必须分别发布固定文件名，不能复用或覆盖输出。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            second_root = root.with_name("task_2_s01_controller_right")
            shutil.copytree(root, second_root)
            output_root = Path(tmp) / "complete"
            output = io.StringIO()

            with contextlib.redirect_stdout(output):
                exit_code = eval_cli.main(
                    ["preprocess", str(root), str(second_root), "--out", str(output_root)]
                )

            self.assertEqual(exit_code, eval_cli.EXIT_OK)
            self.assertTrue((output_root / "task_1_complete.xlsx").is_file())
            self.assertTrue((output_root / "task_2_complete.xlsx").is_file())

    def test_preprocess_rejects_duplicate_task_number_before_writing(self) -> None:
        """两个目录解析出相同任务号时必须返回退出码二，且不创建发布目录。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            duplicate_root = root.with_name("task_1_duplicate_controller_right")
            shutil.copytree(root, duplicate_root)
            output_root = Path(tmp) / "complete"
            output = io.StringIO()

            with contextlib.redirect_stderr(output):
                exit_code = eval_cli.main(
                    ["preprocess", str(root), str(duplicate_root), "--out", str(output_root)]
                )

            self.assertEqual(exit_code, eval_cli.EXIT_DATA_ERROR)
            self.assertFalse(output_root.exists())
            self.assertIn("重复 task 编号", output.getvalue())

    def test_preprocess_rejects_output_inside_read_only_task(self) -> None:
        """CLI 在任何写入前拒绝把发布目录放入原始 task 内部。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = io.StringIO()

            with contextlib.redirect_stderr(output):
                exit_code = eval_cli.main(["preprocess", str(root), "--out", str(root / "derived")])

            self.assertEqual(exit_code, eval_cli.EXIT_DATA_ERROR)
            self.assertFalse((root / "derived").exists())
            self.assertIn("禁止在只读 task 目录内发布工作簿", output.getvalue())


if __name__ == "__main__":
    unittest.main()
