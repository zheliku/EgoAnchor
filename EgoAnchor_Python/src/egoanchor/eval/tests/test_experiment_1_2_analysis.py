"""实验一/二论文分析核心与构建边界测试。"""

from __future__ import annotations

import json
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path
from typing import Any
from unittest.mock import patch

import numpy as np
from matplotlib.patches import PathPatch
from matplotlib.text import Text
from openpyxl import Workbook  # type: ignore[import-untyped]
from scipy.spatial.transform import Rotation, Slerp  # type: ignore[import-untyped]

from egoanchor.eval import cli as eval_cli
from egoanchor.eval.experiments.experiment_1_2.analysis import (
    build_analysis,
    build_exp1_behavior_figure,
    load_settings,
    settings_sha256,
    METHODS,
    PaperResults,
    PerformanceSamples,
    TaskResults,
    TEMPORAL_STRATEGY_VARIANTS,
    analyze_workbooks,
    build_temporal_strategy_panel,
    build_exp1_dynamic_table,
    build_exp1_static_table,
    build_exp1_transition_table,
    build_exp2_attribution_figure,
    build_exp2_attribution_table,
    cache_key,
    cache_path,
    eligible_trials,
    iter_rows,
    load_task_results,
    merge_task_results,
    paired_metric_matrix,
    risk_coverage_curve,
    rotation_lag_metrics,
    summarize_risk_coverage,
    translation_lag_metrics,
    write_analysis_artifacts,
    write_task_results,
)


class Experiment12AnalysisTests(unittest.TestCase):
    """冻结新管线只读取 Stage 1 XLSX 且不恢复旧阶段命令。"""

    def test_formal_cli_does_not_allow_parameter_overrides(self) -> None:
        """正式论文入口只能读取冻结的 paper.toml。"""

        arguments = eval_cli.build_parser().parse_args(["analyze", "exp1-2"])

        self.assertFalse(hasattr(arguments, "settings"))
        self.assertEqual(len(settings_sha256()), 64)

    def test_temporal_panel_uses_extrapolation_linear_and_hermite_order(self) -> None:
        """图 3(d) 固定比较无 StaticLock 的外推、默认 Linear/SLERP 与 Hermite。"""

        self.assertEqual(
            TEMPORAL_STRATEGY_VARIANTS,
            (
                "Smoothed KF Extrapolation",
                "EgoAnchor w/o StaticLock",
                "Hermite Interpolation",
            ),
        )

    def test_lag_metrics_recover_known_translation_and_rotation_delay(self) -> None:
        """已知延迟轨迹必须按正毫秒方向恢复平移和旋转时延。"""

        times = np.arange(0.0, 2000.0, 10.0)
        delay_ms = 120.0
        query_times = np.clip(times - delay_ms, times[0], times[-1])
        reference_positions = np.column_stack(
            (
                0.08 * np.sin(times / 170.0),
                0.05 * np.cos(times / 230.0),
                0.03 * np.sin(times / 310.0 + 0.4),
            )
        )
        display_positions = np.column_stack(
            [
                np.interp(query_times, times, reference_positions[:, axis])
                for axis in range(3)
            ]
        )
        reference_angles = 0.0015 * times + 0.18 * np.sin(times / 260.0)
        reference_rotations = Rotation.from_euler("z", reference_angles[:, None])
        display_rotations = Slerp(times, reference_rotations)(query_times)
        settings = replace(
            load_settings(),
            lag_maximum_ms=250.0,
            lag_step_ms=5.0,
            lag_minimum_samples=30,
        )

        translation_lag, translation_rmse, _ = translation_lag_metrics(
            times,
            display_positions,
            reference_positions,
            settings,
        )
        rotation_lag, rotation_rmse, _ = rotation_lag_metrics(
            times,
            display_rotations.as_quat(),
            reference_rotations.as_quat(),
            settings,
        )

        self.assertEqual(translation_lag, delay_ms)
        self.assertEqual(rotation_lag, delay_ms)
        self.assertLess(translation_rmse, 1e-9)
        self.assertLess(rotation_rmse, 1e-9)

    def test_main_tables_split_exp1_by_aspect_and_keep_audit_table(self) -> None:
        """实验一正文表按三个评价方面各自成表，实验二表继续用于审计。"""

        def rows(variants: tuple[str, ...], **metrics: float) -> dict[str, tuple[dict[str, object], ...]]:
            """构造两个身份完整且可严格配对的片段。"""

            return {
                variant: tuple(
                    {
                        "session_id": "session",
                        "trial_id": "trial",
                        "segment_id": f"segment_{index}",
                        **{
                            key: value + variant_index + index
                            for key, value in metrics.items()
                        },
                    }
                    for index in range(2)
                )
                for variant_index, variant in enumerate(variants)
            }

        static_variants = (*METHODS, "EgoAnchor w/o StaticLock")
        temporal_variants = (
            *METHODS,
            "EgoAnchor w/o StaticLock",
            "Smoothed KF Extrapolation",
            "Hermite Interpolation",
        )
        occlusion = rows(
            (*METHODS, "EgoAnchor w/o VCD"),
            translation_p95_mm=4.0,
            rotation_p95_deg=1.8,
        )
        results = PaperResults(
            workbook_sha256={},
            static_segments=rows(
                static_variants,
                centered_p95_mm=1.0,
                absolute_p95_mm=2.0,
                frame_increment_p95_mm=0.1,
                centered_rotation_p95_deg=0.5,
                absolute_rotation_p95_deg=1.5,
                frame_rotation_increment_p95_deg=0.05,
            ),
            translation_segments=rows(
                temporal_variants,
                effective_lag_ms=200.0,
                aligned_rmse_mm=5.0,
                current_time_rmse_mm=25.0,
                aligned_residual_increment_p95_mm=0.5,
            ),
            rotation_segments=rows(
                temporal_variants,
                effective_lag_ms=220.0,
                aligned_rmse_deg=2.0,
                current_time_rmse_deg=12.0,
                aligned_residual_increment_p95_deg=0.2,
            ),
            occlusion_episodes=occlusion,
            transition_segments=rows(METHODS, response_ms=150.0),
            stop_segments=rows(
                ("EgoAnchor w/o StaticLock", "Smoothed KF Extrapolation"),
                forward_overshoot_mm=1.0,
                settling_time_ms=300.0,
            ),
            correction_segments={},
            capture_alignment=(
                {"capture_p95_mm": 2.0, "arrival_p95_mm": 5.0},
                {"capture_p95_mm": 3.0, "arrival_p95_mm": 7.0},
            ),
            vcd_aurc_segments=(
                {"aurc_mm": 2.0, "full_coverage_risk_mm": 4.0, "risk_gain_mm": 2.0},
                {"aurc_mm": 3.0, "full_coverage_risk_mm": 5.0, "risk_gain_mm": 2.0},
            ),
            vcd_risk_coverage=(),
            performance={},
        )

        static_table = build_exp1_static_table(results)
        dynamic_table = build_exp1_dynamic_table(results)
        transition_table = build_exp1_transition_table(results)
        attribution_table = build_exp2_attribution_table(results)
        exp1_tables = (static_table, dynamic_table, transition_table)

        for table in (*exp1_tables, attribution_table):
            self.assertTrue(table.endswith("\n"))
            self.assertFalse(table.endswith("\n\n"))
        # 三个评价方面各自成表，均按自然宽度排在单栏内：不横向撑满，不改字号，
        # 也不动行距；唯一允许的宽度手段是列距。
        for index, table in enumerate(exp1_tables):
            self.assertIn(r"\begin{table}[t]", table)
            self.assertNotIn(r"\begin{table*}", table)
            self.assertIn(
                r"\begin{tabular}{@{}l" + "c" * (3, 4, 2)[index] + r"@{}}",
                table,
            )
            for forbidden in (
                r"\begin{tabular*}",
                r"\extracolsep",
                r"\small",
                r"\footnotesize",
                r"\renewcommand{\arraystretch}",
                r"\resizebox",
                r"\multirow",
                "$n=",
                "Mdn [",
                "[Q1",
            ):
                self.assertNotIn(forbidden, table)
            # 通道并入单元格后每个方法只占一行，表内不再出现通道行标签。
            for label in ("Arrival", "Capture", "One-Euro", "EgoAnchor"):
                self.assertIn(f"{label} &", table)
            self.assertNotIn("平移 &", table)
            self.assertNotIn("旋转 &", table)
        # 只有指标最多的动态保真度需要收紧列距，其余两张表保持模板列距；
        # 赋值写在 table 环境内，因此不外溢到后续表格。
        for table in (static_table, transition_table):
            self.assertNotIn(r"\setlength{\tabcolsep}", table)
        self.assertIn(
            "\n" + r"\setlength{\tabcolsep}{2pt}" + "\n" + r"\begin{tabular}",
            dynamic_table,
        )
        self.assertGreater(
            dynamic_table.index(r"\setlength{\tabcolsep}"),
            dynamic_table.index(r"\caption{"),
        )
        # 方面名与 §5.1 的评价指标及 §6.1 的三个小节逐一对应，且各自单一概念。
        for aspect, table in zip(
            ("静态保真度", "动态保真度", "转换响应"), exp1_tables, strict=True
        ):
            self.assertIn(aspect, table)
        for label, table in zip(
            ("tab:exp1-static", "tab:exp1-dynamic", "tab:exp1-transition"),
            exp1_tables,
            strict=True,
        ):
            self.assertIn(rf"\label{{{label}}}", table)
        # 每个指标列带一个方向箭头，箭头紧随指标名而非单位。
        for count, table in zip((3, 4, 2), exp1_tables, strict=True):
            self.assertEqual(table.count(r"$\downarrow$"), count)
            self.assertNotIn(r"mm/$^\circ$\,$\downarrow$", table)
        for header, table in (
            (r"头动泄漏\,$\downarrow$", static_table),
            (r"绝对注册\,$\downarrow$", static_table),
            (r"静止抖动\,$\downarrow$", static_table),
            (r"有效时延\,$\downarrow$", dynamic_table),
            (r"LA-RMSE\,$\downarrow$", dynamic_table),
            (r"CT-RMSE\,$\downarrow$", dynamic_table),
            (r"残余抖动\,$\downarrow$", dynamic_table),
            (r"遮挡误差\,$\downarrow$", transition_table),
            (r"起动转换\,$\downarrow$", transition_table),
        ):
            self.assertIn(header, table)
        # 单元格按平移/旋转顺序书写，两通道各自独立加粗最优中位数；通道分隔用
        # 裸斜杠，双通道单元格是全表最宽成分，两侧窄空会把列距逼到过挤的档位。
        self.assertIn(
            r"Arrival & \textbf{1.50}/\textbf{1.00} & "
            r"\textbf{2.50}/\textbf{2.00} & \textbf{0.60}/\textbf{0.55} \\",
            static_table,
        )
        self.assertIn(r"Capture & 2.50/2.00 & 3.50/3.00", static_table)
        for table in exp1_tables:
            self.assertNotIn(r"\,/\,", table)
        # 时延按 5~ms 网格的实际分辨率保留一位小数，其余指标保留两位。
        self.assertIn(
            r"Arrival & \textbf{200.5}/\textbf{220.5} & "
            r"\textbf{5.50}/\textbf{2.50} & \textbf{25.50}/\textbf{12.50} & "
            r"\textbf{1.00}/\textbf{0.70} \\",
            dynamic_table,
        )
        # 起动转换为跨通道标量，单元格内不再拼接通道。
        self.assertIn(
            r"Arrival & \textbf{4.50}/\textbf{2.30} & \textbf{150.5} \\",
            transition_table,
        )
        combined = "".join(exp1_tables) + attribution_table
        self.assertNotIn("Arrival-Hold &", combined)
        self.assertNotIn("Capture-Hold &", combined)
        self.assertNotIn("One-Euro Anchor &", combined)
        self.assertIn("受评设计 & 参照指标 & 启用 & 关闭", attribution_table)
        for row_label in ("采集时刻对齐", "StaticLock", "VCD 判别性", "时序策略"):
            self.assertIn(f"{row_label} &", attribution_table)
        for removed in ("VCD 接纳", "Hermite 补充", "停止护栏"):
            self.assertNotIn(removed, attribution_table)
        self.assertIn(r"$\times$", attribution_table)
        self.assertNotIn(">40", combined)
        self.assertNotIn("灾难性失效", combined)

    def test_paper_composites_use_single_axes_and_hide_hermite(self) -> None:
        """正文组合图固定为四面板，实验一无双轴且实验二不显示 Hermite。"""

        results = _presentation_results()
        self.assertIn("Hermite Interpolation", results.translation_segments)

        def left_title(axis: Any) -> Text:
            """通过 Matplotlib 公共 artist 查询返回左对齐标题。"""

            return next(
                artist
                for artist in axis.findobj(Text)
                if artist.get_text() == axis.get_title(loc="left")
            )

        experiment_one = build_exp1_behavior_figure(results)
        self.assertEqual(len(experiment_one.axes), 4)
        self.assertTrue(all(axis.get_yscale() == "linear" for axis in experiment_one.axes))
        self.assertEqual(
            [axis.get_title(loc="left") for axis in experiment_one.axes],
            [
                "(a) Static translation",
                "(b) Static rotation",
                "(c) Dynamic translation",
                "(d) Dynamic rotation",
            ],
        )
        self.assertLess(experiment_one.subplotpars.wspace, 0.4)
        for axis in experiment_one.axes:
            title = left_title(axis)
            self.assertAlmostEqual(float(title.get_fontsize()), 7.2)
            self.assertEqual(title.get_fontweight(), "bold")
            self.assertAlmostEqual(axis.yaxis.label.get_fontsize(), 7.4)
            boxes = [patch for patch in axis.patches if isinstance(patch, PathPatch)]
            self.assertEqual(len(boxes), 8)
            self.assertTrue(
                all(
                    float(np.asarray(patch.get_facecolor(), dtype=float)[-1]) == 0.0
                    for patch in boxes
                )
            )
            styles = [patch.get_linestyle() for patch in boxes]
            self.assertEqual(styles[0::2], ["-"] * 4)
            self.assertEqual(styles[1::2], ["--"] * 4)
            means = [
                line
                for line in axis.lines
                if line.get_marker() == "o"
                and abs(line.get_markersize() - 3.6) < 1.0e-12
            ]
            self.assertEqual(len(means), 8)
        self.assertEqual(
            [text.get_text() for text in experiment_one.legends[0].get_texts()],
            ["Error / lag-aligned RMSE", "Residual jitter P95", "Mean"],
        )

        experiment_two = build_exp2_attribution_figure(results)
        self.assertEqual(len(experiment_two.axes), 4)
        temporal_legend = [
            text.get_text() for text in experiment_two.axes[3].get_legend().get_texts()
        ]
        self.assertEqual(temporal_legend, ["Smoothed KF", "Linear/SLERP"])
        self.assertNotIn("Hermite", " ".join(temporal_legend))
        self.assertFalse(experiment_two.axes[3].get_legend().get_visible())
        self.assertEqual(
            [text.get_text() for text in experiment_two.legends[0].get_texts()],
            ["IQR", "Median", "Smoothed KF", "Linear/SLERP"],
        )
        self.assertEqual(
            [axis.get_title(loc="left") for axis in experiment_two.axes],
            [
                "(a) Capture-time alignment",
                "(b) StaticLock",
                "(c) VCD risk-coverage",
                "(d) Temporal strategy",
            ],
        )
        self.assertLess(experiment_two.subplotpars.wspace, 0.4)
        for axis in experiment_two.axes:
            title = left_title(axis)
            self.assertAlmostEqual(float(title.get_fontsize()), 7.2)
            self.assertEqual(title.get_fontweight(), "bold")
            self.assertAlmostEqual(axis.yaxis.label.get_fontsize(), 7.4)

    def test_paper_vcd_axis_matches_aurc_step_intervals(self) -> None:
        """正文 VCD 阶梯在零覆盖起步，并把完整同分组风险画在左侧区间。"""

        figure = build_exp2_attribution_figure(_presentation_results())
        axis = figure.axes[2]
        event_lines = tuple(line for line in axis.lines if line.get_label() != "Median")
        median_line = next(line for line in axis.lines if line.get_label() == "Median")

        self.assertEqual(len(event_lines), 2)
        for line in (*event_lines, median_line):
            self.assertEqual(line.get_drawstyle(), "steps-pre")
            self.assertEqual(float(line.get_xdata()[0]), 0.0)
        np.testing.assert_allclose(event_lines[0].get_xdata(), (0.0, 0.5, 1.0))
        np.testing.assert_allclose(event_lines[0].get_ydata(), (3.0, 3.0, 5.0))
        coverage_index = int(np.searchsorted(event_lines[0].get_xdata(), 0.75, side="left"))
        self.assertEqual(float(event_lines[0].get_ydata()[coverage_index]), 5.0)

    def test_paper_attribution_rejects_duplicate_pair_identities(self) -> None:
        """正文组件配对不得静默覆盖 StaticLock 或采集对齐的重复身份。"""

        results = _presentation_results()
        full_variant = METHODS[-1]
        static_segments = dict(results.static_segments)
        static_segments[full_variant] = (
            *static_segments[full_variant],
            dict(static_segments[full_variant][0]),
        )
        with self.subTest(path="StaticLock"), self.assertRaisesRegex(
            ValueError,
            "片段键重复",
        ):
            build_exp2_attribution_figure(
                replace(results, static_segments=static_segments)
            )

        capture_alignment = (
            *results.capture_alignment,
            dict(results.capture_alignment[0]),
        )
        with self.subTest(path="capture alignment"), self.assertRaisesRegex(
            ValueError,
            "片段键重复",
        ):
            build_exp2_attribution_figure(
                replace(results, capture_alignment=capture_alignment)
            )

    def test_vcd_risk_coverage_keeps_score_ties_indivisible(self) -> None:
        """同分候选必须整组进入曲线，AURC 不得依赖候选行顺序。"""

        curve, aurc = risk_coverage_curve((0.9, 0.9, 0.2), (1.0, 9.0, 2.0))
        reordered, reordered_aurc = risk_coverage_curve((0.9, 0.2, 0.9), (9.0, 2.0, 1.0))

        self.assertEqual([row["coverage"] for row in curve], [2.0 / 3.0, 1.0])
        self.assertEqual([row["score_tie_count"] for row in curve], [2, 1])
        self.assertAlmostEqual(float(curve[0]["selective_risk_mm"]), 5.0)
        self.assertAlmostEqual(aurc, 14.0 / 3.0)
        self.assertEqual(curve, reordered)
        self.assertAlmostEqual(aurc, reordered_aurc)

    def test_vcd_risk_coverage_all_equal_scores_reduce_to_full_risk(self) -> None:
        """所有分数相同时只有全覆盖点，AURC 等于候选平均风险。"""

        curve, aurc = risk_coverage_curve((0.5, 0.5, 0.5), (1.0, 2.0, 6.0))

        self.assertEqual(len(curve), 1)
        self.assertEqual(curve[0]["coverage"], 1.0)
        self.assertAlmostEqual(aurc, 3.0)

    def test_vcd_risk_coverage_rejects_invalid_scores(self) -> None:
        """缺失、非有限或越界分数不得静默进入论文指标。"""

        for scores in ((), (float("nan"),), (-0.01,), (1.01,)):
            with self.subTest(scores=scores), self.assertRaises(ValueError):
                risk_coverage_curve(scores, (1.0,) if scores else ())

    def test_vcd_plot_summary_uses_whole_tie_group_at_target_coverage(self) -> None:
        """固定 coverage 汇总不得在 event 的首个完整同分组内部插值。"""

        rows = (
            {"session_id": "s", "trial_id": "t", "segment_id": "a", "coverage": 0.6, "selective_risk_mm": 2.0},
            {"session_id": "s", "trial_id": "t", "segment_id": "a", "coverage": 1.0, "selective_risk_mm": 4.0},
            {"session_id": "s", "trial_id": "t", "segment_id": "b", "coverage": 0.4, "selective_risk_mm": 6.0},
            {"session_id": "s", "trial_id": "t", "segment_id": "b", "coverage": 1.0, "selective_risk_mm": 8.0},
        )

        summary = summarize_risk_coverage(rows)

        self.assertEqual(len(summary), 20)
        self.assertAlmostEqual(float(summary[0]["coverage"]), 0.05)
        self.assertAlmostEqual(float(summary[0]["selective_risk_median_mm"]), 4.0)
        self.assertAlmostEqual(float(summary[-1]["selective_risk_median_mm"]), 6.0)

    def test_xlsx_reader_streams_selected_columns_from_stage_one_sheet(self) -> None:
        """新分析 reader 直接消费 Stage 1 sheet，不改写原始 workbook。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "task_1_complete.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "unity_render"
            sheet.append(("session_id", "variant_id", "render_mono_ms", "unused"))
            sheet.append(("session", "EgoAnchor", 123.5, "ignored"))
            workbook.save(path)
            before = path.read_bytes()

            rows = list(iter_rows(path, "unity_render", ("variant_id", "render_mono_ms")))

            self.assertEqual(
                rows,
                [{"variant_id": "EgoAnchor", "render_mono_ms": 123.5}],
            )
            self.assertEqual(path.read_bytes(), before)

    def test_build_analysis_rejects_non_xlsx_input_before_writing(self) -> None:
        """分析入口拒绝 JSON/CSV，保持初始 XLSX 是唯一分析桥梁。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "task_1.json"
            source.write_text("{}", encoding="utf-8")
            output = root / "output"

            with self.assertRaisesRegex(ValueError, "五本 Stage 1 XLSX"):
                build_analysis(
                    (source,),
                    output,
                    root / "cache",
                    "batch_test",
                    {},
                    settings=load_settings(),
                    config_sha256=settings_sha256(),
                )
            self.assertFalse(output.exists())

    def test_analysis_artifacts_remove_only_retired_figure_tex(self) -> None:
        """重建时删除旧拼图 TeX，但保留同目录内未托管文件。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            output_root = Path(temporary_directory) / "analysis"
            legacy_root = output_root / "tex" / "figures"
            legacy_root.mkdir(parents=True)
            retired = (
                legacy_root / "figure2_experiment1.tex",
                legacy_root / "figure3_experiment2.tex",
            )
            for path in retired:
                path.write_text("retired", encoding="utf-8")
            unrelated = legacy_root / "researcher_note.tex"
            unrelated.write_text("keep", encoding="utf-8")

            artifacts = write_analysis_artifacts(
                _presentation_results(),
                output_root,
            )

            self.assertTrue(all(not path.exists() for path in retired))
            self.assertEqual(unrelated.read_text(encoding="utf-8"), "keep")
            self.assertNotIn("figure2_tex", artifacts)
            self.assertNotIn("figure3_tex", artifacts)

    def test_task_cache_round_trips_non_finite_metrics_as_strict_json(self) -> None:
        """task 缓存必须显式编码 NaN，不能写出非标准 JSON 数字。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook = root / "task_1_complete.xlsx"
            workbook.touch()
            destination = cache_path(root / "cache", workbook)
            key = cache_key("a" * 64, "b" * 64)
            results = self._task_results(workbook, "a" * 64, nan_metric=True)

            write_task_results(destination, key, results)
            restored = load_task_results(destination, key, workbook)

            self.assertIsNotNone(restored)
            assert restored is not None
            self.assertNotIn("NaN", destination.read_text(encoding="utf-8"))
            self.assertTrue(
                np.isnan(float(restored.static_segments["EgoAnchor"][0]["value"]))
            )

    def test_task_cache_result_tampering_is_a_cache_miss(self) -> None:
        """指标缓存正文被修改后必须重算，而不是只依赖未变化的 key。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook = root / "task_1_complete.xlsx"
            workbook.touch()
            destination = cache_path(root / "cache", workbook)
            key = cache_key("a" * 64, "b" * 64)
            write_task_results(destination, key, self._task_results(workbook, "a" * 64))

            document = json.loads(destination.read_text(encoding="utf-8"))
            document["result"]["workbook_sha256"] = "c" * 64
            destination.write_text(json.dumps(document), encoding="utf-8")

            self.assertIsNone(load_task_results(destination, key, workbook))

    def test_build_analysis_uses_five_task_caches_without_hashing_workbooks(self) -> None:
        """五项缓存全部命中时不得重新读取或计算 Stage 1 workbook。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook_root = root / "workbooks"
            cache_root = root / "cache"
            output_root = root / "analysis"
            workbook_root.mkdir()
            workbooks = tuple(
                workbook_root / f"task_{number}_complete.xlsx" for number in range(1, 6)
            )
            expected = {}
            parameter_digest = settings_sha256()
            for number, workbook in enumerate(workbooks, start=1):
                workbook.touch()
                digest = str(number) * 64
                expected[str(workbook.resolve())] = digest
                results = self._task_results(
                    workbook,
                    digest,
                    include_vcd=number == 5,
                )
                write_task_results(
                    cache_path(cache_root, workbook),
                    cache_key(digest, parameter_digest),
                    results,
                )
            progress: list[str] = []

            with (
                patch(
                    "egoanchor.eval.experiments.experiment_1_2.analysis.pipeline.calculate_workbook_sha256",
                    side_effect=AssertionError("缓存命中不应哈希 workbook"),
                ),
                patch(
                    "egoanchor.eval.experiments.experiment_1_2.analysis.pipeline.analyze_task_workbook",
                    side_effect=AssertionError("缓存命中不应扫描 workbook"),
                ),
                patch(
                    "egoanchor.eval.experiments.experiment_1_2.analysis.pipeline.publish_figures",
                    return_value={},
                ),
                patch(
                    "egoanchor.eval.experiments.experiment_1_2.analysis.pipeline.write_analysis_artifacts",
                    return_value={},
                ),
            ):
                payload = build_analysis(
                    workbooks,
                    output_root,
                    cache_root,
                    "batch_test",
                    expected,
                    settings=load_settings(),
                    config_sha256=parameter_digest,
                    progress=progress.append,
                )

            self.assertTrue(payload["passed"])
            self.assertEqual(set(payload["task_cache"].values()), {"hit"})
            self.assertEqual(
                [message for message in progress if "使用指标缓存" in message],
                [f"Task {number}: 使用指标缓存" for number in range(1, 6)],
            )

            changed_workbook = workbooks[2]
            changed_digest = "a" * 64
            expected[str(changed_workbook.resolve())] = changed_digest
            rebuilt = self._task_results(changed_workbook, changed_digest)
            with (
                patch(
                    "egoanchor.eval.experiments.experiment_1_2.analysis.pipeline.calculate_workbook_sha256",
                    return_value=changed_digest,
                ) as hash_workbook,
                patch(
                    "egoanchor.eval.experiments.experiment_1_2.analysis.pipeline.analyze_task_workbook",
                    return_value=rebuilt,
                ) as analyze_workbook,
                patch(
                    "egoanchor.eval.experiments.experiment_1_2.analysis.pipeline.publish_figures",
                    return_value={},
                ),
                patch(
                    "egoanchor.eval.experiments.experiment_1_2.analysis.pipeline.write_analysis_artifacts",
                    return_value={},
                ),
            ):
                replaced_payload = build_analysis(
                    workbooks,
                    output_root,
                    cache_root,
                    "batch_replaced_task_3",
                    expected,
                    settings=load_settings(),
                    config_sha256=parameter_digest,
                )

            hash_workbook.assert_called_once_with(changed_workbook.resolve())
            self.assertEqual(analyze_workbook.call_count, 1)
            self.assertEqual(replaced_payload["task_cache"][changed_workbook.name], "rebuilt")
            self.assertEqual(
                sum(state == "hit" for state in replaced_payload["task_cache"].values()),
                4,
            )

    def test_task_merge_summarizes_raw_performance_samples(self) -> None:
        """性能统计必须合并原始样本，不能对各 task 中位数再次汇总。"""

        first = self._task_results(Path("task_1_complete.xlsx"), "1" * 64)
        second = self._task_results(
            Path("task_5_complete.xlsx"),
            "5" * 64,
            include_vcd=True,
        )
        first = replace(
            first,
            performance_samples=PerformanceSamples((0.0, 10.0), (1.0,), (50.0,)),
        )
        second = replace(
            second,
            performance_samples=PerformanceSamples((100.0,), (3.0,), (150.0,)),
        )

        merged = merge_task_results((first, second))

        self.assertEqual(merged.performance["track_total_ms_median"], 10.0)
        self.assertEqual(merged.performance["register_total_ms_median"], 2.0)
        self.assertEqual(merged.performance["pose_publish_interval_ms_median"], 100.0)

    @staticmethod
    def _task_results(
        workbook: Path,
        digest: str,
        *,
        include_vcd: bool = False,
        nan_metric: bool = False,
    ) -> TaskResults:
        """构造只覆盖缓存与合并契约的最小 task 结果。"""

        static = (
            {"session_id": "s", "trial_id": "t", "segment_id": "e", "value": np.nan},
        ) if nan_metric else ()
        vcd_curve = (
            {
                "session_id": "s5",
                "trial_id": "t",
                "segment_id": "e",
                "coverage": 1.0,
                "selective_risk_mm": 1.0,
            },
        ) if include_vcd else ()
        vcd_aurc = (
            {
                "session_id": "s5",
                "trial_id": "t",
                "segment_id": "e",
                "aurc_mm": 1.0,
            },
        ) if include_vcd else ()
        return TaskResults(
            workbook_path=str(workbook.resolve()),
            workbook_sha256=digest,
            static_segments={"EgoAnchor": static} if static else {},
            translation_segments={},
            rotation_segments={},
            occlusion_episodes={},
            transition_segments={},
            stop_segments={},
            correction_segments={},
            capture_alignment=(),
            vcd_risk_coverage=vcd_curve,
            vcd_aurc_segments=vcd_aurc,
            performance_samples=PerformanceSamples((1.0,), (2.0,), (100.0,)),
        )

    def test_v3_stage_one_workbook_is_rejected_before_analysis(self) -> None:
        """旧 linear_v2 工作簿不得被新时序策略矩阵的论文入口接受。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "task_1_complete.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "metadata_kv"
            sheet.append(("document", "json_path", "value_json"))
            sheet.append(("manifest.json", "variant_matrix_id", '"exp12_9_linear_v2"'))
            workbook.save(path)

            with self.assertRaisesRegex(ValueError, "exp12_9_smoothed_hermite_v4"):
                analyze_workbooks((path,))

    def test_eligible_trials_exclude_rejected_retries(self) -> None:
        """论文指标只接受最终结束且没有后续作废的 trial。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "task_1_complete.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "events"
            sheet.append(
                (
                    "session_id",
                    "trial_id",
                    "event",
                    "event_type",
                    "mono_ms",
                    "created_unix_ms",
                    "event_row_id",
                )
            )
            sheet.append(("s", "old", "trial_ended", "trial_ended", 10.0, 10.0, "e1"))
            sheet.append(("s", "old", "trial_rejected", "trial_rejected", 11.0, 11.0, "e2"))
            sheet.append(("s", "new", "trial_ended", "trial_ended", 20.0, 20.0, "e3"))
            workbook.save(path)

            self.assertEqual(eligible_trials((path,)), frozenset({("s", "new")}))

    def test_figure_pairing_uses_event_identity_instead_of_row_position(self) -> None:
        """实验一折线必须按稳定片段键配对，即使各方法行顺序不同。"""

        rows = {}
        for method_index, method in enumerate(METHODS):
            method_rows = [
                {"session_id": "s", "trial_id": "t", "segment_id": "b", "value": method_index + 10.0},
                {"session_id": "s", "trial_id": "t", "segment_id": "a", "value": method_index + 1.0},
            ]
            if method_index % 2:
                method_rows.reverse()
            rows[method] = tuple(method_rows)

        matrix = paired_metric_matrix(rows, METHODS, ("value",))

        self.assertEqual(matrix[:, :, 0].tolist(), [[1.0, 2.0, 3.0, 4.0], [10.0, 11.0, 12.0, 13.0]])

    def test_figure_pairing_rejects_missing_method_episode(self) -> None:
        """任一方法缺失 episode 时禁止生成看似配对的趋势线。"""

        rows: dict[str, tuple[dict[str, object], ...]] = {
            method: ({"session_id": "s", "trial_id": "t", "segment_id": "a", "value": 1.0},)
            for method in METHODS
        }
        rows[METHODS[-1]] = ()

        with self.assertRaisesRegex(ValueError, "配对不完整"):
            paired_metric_matrix(rows, METHODS, ("value",))

    def test_temporal_panel_keeps_all_points_and_expands_axis(self) -> None:
        """时序面板必须显示全部片段，且固定画布的纵向坐标轴齐平。"""

        points = np.asarray(
            [
                ((260.0, 58.0), (350.0, 8.0), (360.0, 9.0)),
                ((280.0, 45.0), (365.0, 12.0), (375.0, 14.0)),
            ],
            dtype=float,
        )

        figure = build_temporal_strategy_panel(points)
        axis = figure.axes[0]
        labeled_collections = {
            collection.get_label(): collection
            for collection in axis.collections
            if collection.get_label() in {"Smoothed KF", "Linear/SLERP", "Hermite"}
        }

        self.assertEqual(tuple(figure.get_size_inches()), (2.8, 2.18))
        self.assertEqual(set(labeled_collections), {"Smoothed KF", "Linear/SLERP", "Hermite"})
        self.assertTrue(
            all(len(collection.get_offsets()) == 2 for collection in labeled_collections.values())
        )
        self.assertGreater(axis.get_ylim()[1], float(np.max(points[:, :, 1])))
        self.assertAlmostEqual(axis.get_position().y0, 0.25)
        self.assertAlmostEqual(axis.get_position().y1, 0.97)
        self.assertEqual(
            [text.get_text() for text in axis.get_legend().get_texts()],
            ["Smoothed KF", "Linear/SLERP", "Hermite"],
        )
        self.assertFalse(axis.get_legend().get_frame_on())
        self.assertEqual(axis.get_legend().borderaxespad, 0.0)

def _presentation_results() -> PaperResults:
    """构造覆盖两张正文组合图全部输入契约的最小结果。"""

    def rows(
        variants: tuple[str, ...],
        **metrics: float,
    ) -> dict[str, tuple[dict[str, object], ...]]:
        """为每个配置构造三个身份严格匹配的正值片段。"""

        return {
            variant: tuple(
                {
                    "session_id": "session",
                    "trial_id": "trial",
                    "segment_id": f"segment_{index}",
                    **{
                        key: value + 0.2 * variant_index + 0.1 * index
                        for key, value in metrics.items()
                    },
                }
                for index in range(3)
            )
            for variant_index, variant in enumerate(variants)
        }

    temporal_variants = (
        *METHODS,
        "EgoAnchor w/o StaticLock",
        "Smoothed KF Extrapolation",
        "Hermite Interpolation",
    )
    risk_rows = tuple(
        {
            "session_id": "session",
            "trial_id": "trial",
            "segment_id": f"event_{event}",
            "coverage": coverage,
            "selective_risk_mm": risk + event,
        }
        for event in range(2)
        for coverage, risk in ((0.5, 3.0), (1.0, 5.0))
    )
    return PaperResults(
        workbook_sha256={},
        static_segments=rows(
            temporal_variants,
            centered_p95_mm=1.0,
            absolute_p95_mm=2.0,
            frame_increment_p95_mm=0.1,
            centered_rotation_p95_deg=0.5,
            absolute_rotation_p95_deg=1.5,
            frame_rotation_increment_p95_deg=0.05,
        ),
        translation_segments=rows(
            temporal_variants,
            effective_lag_ms=200.0,
            aligned_rmse_mm=5.0,
            current_time_rmse_mm=25.0,
            aligned_residual_increment_p95_mm=0.5,
        ),
        rotation_segments=rows(
            temporal_variants,
            effective_lag_ms=220.0,
            aligned_rmse_deg=2.0,
            current_time_rmse_deg=12.0,
            aligned_residual_increment_p95_deg=0.2,
        ),
        occlusion_episodes=rows(
            temporal_variants,
            translation_p95_mm=4.0,
            translation_max_mm=6.0,
            rotation_p95_deg=1.8,
            rotation_max_deg=2.7,
            catastrophic_gt40=0.0,
        ),
        transition_segments=rows(temporal_variants, response_ms=150.0),
        stop_segments=rows(
            temporal_variants,
            forward_overshoot_mm=3.0,
            reverse_return_mm=1.0,
            settling_time_ms=240.0,
        ),
        correction_segments=rows(
            temporal_variants,
            position_step_p95_mm=2.0,
            rotation_step_p95_deg=1.0,
        ),
        capture_alignment=tuple(
            {
                "session_id": "session",
                "trial_id": "trial",
                "segment_id": f"segment_{index}",
                "capture_p95_mm": 2.0 + index,
                "arrival_p95_mm": 5.0 + index,
            }
            for index in range(3)
        ),
        vcd_aurc_segments=tuple(
            {
                "session_id": "session",
                "trial_id": "trial",
                "segment_id": f"event_{index}",
                "aurc_mm": 3.0 + index,
                "full_coverage_risk_mm": 5.0 + index,
                "risk_gain_mm": 2.0,
            }
            for index in range(3)
        ),
        vcd_risk_coverage=risk_rows,
        performance={},
    )


if __name__ == "__main__":
    unittest.main()
