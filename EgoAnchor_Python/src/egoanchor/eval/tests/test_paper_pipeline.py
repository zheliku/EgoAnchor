"""论文分析入口的边界与最小计算测试。"""

from __future__ import annotations

import json
import tempfile
import unittest
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import numpy as np
from openpyxl import Workbook  # type: ignore[import-untyped]

from egoanchor.eval import cli as eval_cli
from egoanchor.eval.paper_analysis import (
    METHODS,
    PaperResults,
    PerformanceSamples,
    TaskResults,
    TEMPORAL_STRATEGY_VARIANTS,
    analyze_workbooks,
    build_analysis,
    build_exp1_table,
    build_exp2_table,
    build_point_panel,
    build_translation_panel,
    cache_key,
    cache_path,
    eligible_trials,
    iter_rows,
    load_task_results,
    merge_task_results,
    paired_metric_matrix,
    risk_coverage_curve,
    settings_sha256,
    summarize_risk_coverage,
    write_task_results,
)


class PaperPipelineTests(unittest.TestCase):
    """冻结新管线只读取 Stage 1 XLSX 且不恢复旧阶段命令。"""

    def test_formal_cli_does_not_allow_parameter_overrides(self) -> None:
        """正式论文入口只能读取冻结的 paper.toml。"""

        arguments = eval_cli.build_parser().parse_args(["analyze"])

        self.assertFalse(hasattr(arguments, "settings"))
        self.assertEqual(len(settings_sha256()), 64)

    def test_temporal_panel_uses_extrapolation_linear_and_hermite_order(self) -> None:
        """图 3(d) 固定比较无 StaticLock 的外推、默认 Linear/SLERP 与 Hermite。"""

        self.assertEqual(
            TEMPORAL_STRATEGY_VARIANTS,
            (
                "Smoothed KF Extrapolation",
                "EgoAnchor w/o StaticLock",
                "Hermite Interpolation",
            ),
        )

    def test_main_tables_publish_occlusion_translation_p95(self) -> None:
        """两张主表使用两位小数的连续遮挡误差，阈值次数只留在审计产物。"""

        def rows(variants: tuple[str, ...], **metrics: float) -> dict[str, tuple[dict[str, object], ...]]:
            """构造两个身份完整且可严格配对的片段。"""

            return {
                variant: tuple(
                    {
                        "session_id": "session",
                        "trial_id": "trial",
                        "segment_id": f"segment_{index}",
                        **{
                            key: value + variant_index + index
                            for key, value in metrics.items()
                        },
                    }
                    for index in range(2)
                )
                for variant_index, variant in enumerate(variants)
            }

        static_variants = (*METHODS, "EgoAnchor w/o StaticLock")
        temporal_variants = (
            *METHODS,
            "EgoAnchor w/o StaticLock",
            "Smoothed KF Extrapolation",
            "Hermite Interpolation",
        )
        occlusion = rows((*METHODS, "EgoAnchor w/o VCD"), translation_p95_mm=4.0)
        results = PaperResults(
            workbook_sha256={},
            static_segments=rows(
                static_variants,
                centered_p95_mm=1.0,
                absolute_p95_mm=2.0,
                frame_increment_p95_mm=0.1,
            ),
            translation_segments=rows(
                temporal_variants,
                effective_lag_ms=200.0,
                aligned_rmse_mm=5.0,
            ),
            rotation_segments=rows(
                temporal_variants,
                effective_lag_ms=220.0,
                aligned_rmse_deg=2.0,
            ),
            occlusion_episodes=occlusion,
            transition_segments=rows(METHODS, response_ms=150.0),
            stop_segments=rows(
                ("EgoAnchor w/o StaticLock", "Smoothed KF Extrapolation"),
                forward_overshoot_mm=1.0,
                settling_time_ms=300.0,
            ),
            correction_segments={},
            capture_alignment=(
                {"capture_p95_mm": 2.0, "arrival_p95_mm": 5.0},
                {"capture_p95_mm": 3.0, "arrival_p95_mm": 7.0},
            ),
            vcd_aurc_segments=(
                {"aurc_mm": 2.0, "full_coverage_risk_mm": 4.0, "risk_gain_mm": 2.0},
                {"aurc_mm": 3.0, "full_coverage_risk_mm": 5.0, "risk_gain_mm": 2.0},
            ),
            vcd_risk_coverage=(),
            performance={},
        )

        experiment_one = build_exp1_table(results)
        experiment_two = build_exp2_table(results)

        self.assertIn("遮挡期平移误差 P95", experiment_one)
        self.assertIn(r"\begin{tabular}{lccccccc}", experiment_one)
        self.assertIn("Capture & 2.50 [2.25, 2.75]", experiment_one)
        for label in ("Arrival", "Capture", "One-Euro", "EgoAnchor"):
            self.assertIn(f"{label} &", experiment_one)
        self.assertNotIn("Arrival-Hold &", experiment_one)
        self.assertNotIn("Capture-Hold &", experiment_one)
        self.assertNotIn("One-Euro Anchor &", experiment_one)
        self.assertIn("遮挡期平移误差 P95", experiment_two)
        self.assertIn("7.50 [7.25, 7.75]", experiment_two)
        self.assertIn("个 episode 对照更高", experiment_two)
        self.assertNotIn(">40", experiment_one + experiment_two)
        self.assertNotIn("灾难性失效", experiment_one + experiment_two)

    def test_vcd_risk_coverage_keeps_score_ties_indivisible(self) -> None:
        """同分候选必须整组进入曲线，AURC 不得依赖候选行顺序。"""

        curve, aurc = risk_coverage_curve((0.9, 0.9, 0.2), (1.0, 9.0, 2.0))
        reordered, reordered_aurc = risk_coverage_curve((0.9, 0.2, 0.9), (9.0, 2.0, 1.0))

        self.assertEqual([row["coverage"] for row in curve], [2.0 / 3.0, 1.0])
        self.assertEqual([row["score_tie_count"] for row in curve], [2, 1])
        self.assertAlmostEqual(float(curve[0]["selective_risk_mm"]), 5.0)
        self.assertAlmostEqual(aurc, 14.0 / 3.0)
        self.assertEqual(curve, reordered)
        self.assertAlmostEqual(aurc, reordered_aurc)

    def test_vcd_risk_coverage_all_equal_scores_reduce_to_full_risk(self) -> None:
        """所有分数相同时只有全覆盖点，AURC 等于候选平均风险。"""

        curve, aurc = risk_coverage_curve((0.5, 0.5, 0.5), (1.0, 2.0, 6.0))

        self.assertEqual(len(curve), 1)
        self.assertEqual(curve[0]["coverage"], 1.0)
        self.assertAlmostEqual(aurc, 3.0)

    def test_vcd_risk_coverage_rejects_invalid_scores(self) -> None:
        """缺失、非有限或越界分数不得静默进入论文指标。"""

        for scores in ((), (float("nan"),), (-0.01,), (1.01,)):
            with self.subTest(scores=scores), self.assertRaises(ValueError):
                risk_coverage_curve(scores, (1.0,) if scores else ())

    def test_vcd_plot_summary_uses_whole_tie_group_at_target_coverage(self) -> None:
        """固定 coverage 汇总不得在 event 的首个完整同分组内部插值。"""

        rows = (
            {"session_id": "s", "trial_id": "t", "segment_id": "a", "coverage": 0.6, "selective_risk_mm": 2.0},
            {"session_id": "s", "trial_id": "t", "segment_id": "a", "coverage": 1.0, "selective_risk_mm": 4.0},
            {"session_id": "s", "trial_id": "t", "segment_id": "b", "coverage": 0.4, "selective_risk_mm": 6.0},
            {"session_id": "s", "trial_id": "t", "segment_id": "b", "coverage": 1.0, "selective_risk_mm": 8.0},
        )

        summary = summarize_risk_coverage(rows)

        self.assertEqual(len(summary), 20)
        self.assertAlmostEqual(float(summary[0]["coverage"]), 0.05)
        self.assertAlmostEqual(float(summary[0]["selective_risk_median_mm"]), 4.0)
        self.assertAlmostEqual(float(summary[-1]["selective_risk_median_mm"]), 6.0)

    def test_xlsx_reader_streams_selected_columns_from_stage_one_sheet(self) -> None:
        """新分析 reader 直接消费 Stage 1 sheet，不改写原始 workbook。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "task_1_complete.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "unity_render"
            sheet.append(("session_id", "variant_id", "render_mono_ms", "unused"))
            sheet.append(("session", "EgoAnchor", 123.5, "ignored"))
            workbook.save(path)
            before = path.read_bytes()

            rows = list(iter_rows(path, "unity_render", ("variant_id", "render_mono_ms")))

            self.assertEqual(
                rows,
                [{"variant_id": "EgoAnchor", "render_mono_ms": 123.5}],
            )
            self.assertEqual(path.read_bytes(), before)

    def test_build_analysis_rejects_non_xlsx_input_before_writing(self) -> None:
        """分析入口拒绝 JSON/CSV，保持初始 XLSX 是唯一分析桥梁。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "task_1.json"
            source.write_text("{}", encoding="utf-8")
            output = root / "output"

            with self.assertRaisesRegex(ValueError, "五本 Stage 1 XLSX"):
                build_analysis(
                    (source,),
                    output,
                    "figures/panels",
                    root / "cache",
                    "batch_test",
                    {},
                )
            self.assertFalse(output.exists())

    def test_task_cache_round_trips_non_finite_metrics_as_strict_json(self) -> None:
        """task 缓存必须显式编码 NaN，不能写出非标准 JSON 数字。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook = root / "task_1_complete.xlsx"
            workbook.touch()
            destination = cache_path(root / "cache", workbook)
            key = cache_key("a" * 64, "b" * 64)
            results = self._task_results(workbook, "a" * 64, nan_metric=True)

            write_task_results(destination, key, results)
            restored = load_task_results(destination, key, workbook)

            self.assertIsNotNone(restored)
            self.assertNotIn("NaN", destination.read_text(encoding="utf-8"))
            self.assertTrue(
                np.isnan(float(restored.static_segments["EgoAnchor"][0]["value"]))
            )

    def test_task_cache_result_tampering_is_a_cache_miss(self) -> None:
        """指标缓存正文被修改后必须重算，而不是只依赖未变化的 key。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook = root / "task_1_complete.xlsx"
            workbook.touch()
            destination = cache_path(root / "cache", workbook)
            key = cache_key("a" * 64, "b" * 64)
            write_task_results(destination, key, self._task_results(workbook, "a" * 64))

            document = json.loads(destination.read_text(encoding="utf-8"))
            document["result"]["workbook_sha256"] = "c" * 64
            destination.write_text(json.dumps(document), encoding="utf-8")

            self.assertIsNone(load_task_results(destination, key, workbook))

    def test_build_analysis_uses_five_task_caches_without_hashing_workbooks(self) -> None:
        """五项缓存全部命中时不得重新读取或计算 Stage 1 workbook。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            workbook_root = root / "workbooks"
            cache_root = root / "cache"
            output_root = root / "analysis"
            workbook_root.mkdir()
            workbooks = tuple(
                workbook_root / f"task_{number}_complete.xlsx" for number in range(1, 6)
            )
            expected = {}
            parameter_digest = settings_sha256()
            for number, workbook in enumerate(workbooks, start=1):
                workbook.touch()
                digest = str(number) * 64
                expected[str(workbook.resolve())] = digest
                results = self._task_results(
                    workbook,
                    digest,
                    include_vcd=number == 5,
                )
                write_task_results(
                    cache_path(cache_root, workbook),
                    cache_key(digest, parameter_digest),
                    results,
                )
            progress: list[str] = []

            with (
                patch(
                    "egoanchor.eval.paper_analysis.pipeline.calculate_workbook_sha256",
                    side_effect=AssertionError("缓存命中不应哈希 workbook"),
                ),
                patch(
                    "egoanchor.eval.paper_analysis.pipeline.analyze_task_workbook",
                    side_effect=AssertionError("缓存命中不应扫描 workbook"),
                ),
                patch("egoanchor.eval.paper_analysis.pipeline.publish_figures", return_value={}),
                patch(
                    "egoanchor.eval.paper_analysis.pipeline.write_analysis_artifacts",
                    return_value={},
                ),
            ):
                payload = build_analysis(
                    workbooks,
                    output_root,
                    "figures/panels",
                    cache_root,
                    "batch_test",
                    expected,
                    progress.append,
                )

            self.assertTrue(payload["passed"])
            self.assertEqual(set(payload["task_cache"].values()), {"hit"})
            self.assertEqual(
                [message for message in progress if "使用指标缓存" in message],
                [f"Task {number}: 使用指标缓存" for number in range(1, 6)],
            )

            changed_workbook = workbooks[2]
            changed_digest = "a" * 64
            expected[str(changed_workbook.resolve())] = changed_digest
            rebuilt = self._task_results(changed_workbook, changed_digest)
            with (
                patch(
                    "egoanchor.eval.paper_analysis.pipeline.calculate_workbook_sha256",
                    return_value=changed_digest,
                ) as hash_workbook,
                patch(
                    "egoanchor.eval.paper_analysis.pipeline.analyze_task_workbook",
                    return_value=rebuilt,
                ) as analyze_workbook,
                patch("egoanchor.eval.paper_analysis.pipeline.publish_figures", return_value={}),
                patch(
                    "egoanchor.eval.paper_analysis.pipeline.write_analysis_artifacts",
                    return_value={},
                ),
            ):
                replaced_payload = build_analysis(
                    workbooks,
                    output_root,
                    "figures/panels",
                    cache_root,
                    "batch_replaced_task_3",
                    expected,
                )

            hash_workbook.assert_called_once_with(changed_workbook.resolve())
            self.assertEqual(analyze_workbook.call_count, 1)
            self.assertEqual(replaced_payload["task_cache"][changed_workbook.name], "rebuilt")
            self.assertEqual(
                sum(state == "hit" for state in replaced_payload["task_cache"].values()),
                4,
            )

    def test_task_merge_summarizes_raw_performance_samples(self) -> None:
        """性能统计必须合并原始样本，不能对各 task 中位数再次汇总。"""

        first = self._task_results(Path("task_1_complete.xlsx"), "1" * 64)
        second = self._task_results(
            Path("task_5_complete.xlsx"),
            "5" * 64,
            include_vcd=True,
        )
        first = replace(
            first,
            performance_samples=PerformanceSamples((0.0, 10.0), (1.0,), (50.0,)),
        )
        second = replace(
            second,
            performance_samples=PerformanceSamples((100.0,), (3.0,), (150.0,)),
        )

        merged = merge_task_results((first, second))

        self.assertEqual(merged.performance["track_total_ms_median"], 10.0)
        self.assertEqual(merged.performance["register_total_ms_median"], 2.0)
        self.assertEqual(merged.performance["pose_publish_interval_ms_median"], 100.0)

    @staticmethod
    def _task_results(
        workbook: Path,
        digest: str,
        *,
        include_vcd: bool = False,
        nan_metric: bool = False,
    ) -> TaskResults:
        """构造只覆盖缓存与合并契约的最小 task 结果。"""

        static = (
            {"session_id": "s", "trial_id": "t", "segment_id": "e", "value": np.nan},
        ) if nan_metric else ()
        vcd_curve = (
            {
                "session_id": "s5",
                "trial_id": "t",
                "segment_id": "e",
                "coverage": 1.0,
                "selective_risk_mm": 1.0,
            },
        ) if include_vcd else ()
        vcd_aurc = (
            {
                "session_id": "s5",
                "trial_id": "t",
                "segment_id": "e",
                "aurc_mm": 1.0,
            },
        ) if include_vcd else ()
        return TaskResults(
            workbook_path=str(workbook.resolve()),
            workbook_sha256=digest,
            static_segments={"EgoAnchor": static} if static else {},
            translation_segments={},
            rotation_segments={},
            occlusion_episodes={},
            transition_segments={},
            stop_segments={},
            correction_segments={},
            capture_alignment=(),
            vcd_risk_coverage=vcd_curve,
            vcd_aurc_segments=vcd_aurc,
            performance_samples=PerformanceSamples((1.0,), (2.0,), (100.0,)),
        )

    def test_v3_stage_one_workbook_is_rejected_before_analysis(self) -> None:
        """旧 linear_v2 工作簿不得被新时序策略矩阵的论文入口接受。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "task_1_complete.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "metadata_kv"
            sheet.append(("document", "json_path", "value_json"))
            sheet.append(("manifest.json", "variant_matrix_id", '"exp12_9_linear_v2"'))
            workbook.save(path)

            with self.assertRaisesRegex(ValueError, "exp12_9_smoothed_hermite_v4"):
                analyze_workbooks((path,))

    def test_eligible_trials_exclude_rejected_retries(self) -> None:
        """论文指标只接受最终结束且没有后续作废的 trial。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "task_1_complete.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "events"
            sheet.append(
                (
                    "session_id",
                    "trial_id",
                    "event",
                    "event_type",
                    "mono_ms",
                    "created_unix_ms",
                    "event_row_id",
                )
            )
            sheet.append(("s", "old", "trial_ended", "trial_ended", 10.0, 10.0, "e1"))
            sheet.append(("s", "old", "trial_rejected", "trial_rejected", 11.0, 11.0, "e2"))
            sheet.append(("s", "new", "trial_ended", "trial_ended", 20.0, 20.0, "e3"))
            workbook.save(path)

            self.assertEqual(eligible_trials((path,)), frozenset({("s", "new")}))

    def test_figure_pairing_uses_event_identity_instead_of_row_position(self) -> None:
        """实验一折线必须按稳定片段键配对，即使各方法行顺序不同。"""

        rows = {}
        for method_index, method in enumerate(METHODS):
            method_rows = [
                {"session_id": "s", "trial_id": "t", "segment_id": "b", "value": method_index + 10.0},
                {"session_id": "s", "trial_id": "t", "segment_id": "a", "value": method_index + 1.0},
            ]
            if method_index % 2:
                method_rows.reverse()
            rows[method] = tuple(method_rows)

        matrix = paired_metric_matrix(rows, METHODS, ("value",))

        self.assertEqual(matrix[:, :, 0].tolist(), [[1.0, 2.0, 3.0, 4.0], [10.0, 11.0, 12.0, 13.0]])

    def test_figure_pairing_rejects_missing_method_episode(self) -> None:
        """任一方法缺失 episode 时禁止生成看似配对的趋势线。"""

        rows = {
            method: ({"session_id": "s", "trial_id": "t", "segment_id": "a", "value": 1.0},)
            for method in METHODS
        }
        rows[METHODS[-1]] = ()

        with self.assertRaisesRegex(ValueError, "配对不完整"):
            paired_metric_matrix(rows, METHODS, ("value",))

    def test_experiment_one_panels_use_consistent_method_legend(self) -> None:
        """图二三个面板均保留一致的四方法图例，独立阅读时不依赖相邻子图。"""

        values = {method: np.asarray((1.0, 2.0)) for method in METHODS}
        point_figure = build_point_panel(
            values,
            "Metric",
            np.asarray(((1.0, 1.1, 1.2, 1.3), (2.0, 2.1, 2.2, 2.3))),
        )
        rows = {
            method: (
                {
                    "session_id": "s",
                    "trial_id": "t",
                    "segment_id": "a",
                    "effective_lag_ms": 200.0 + index,
                    "aligned_rmse_mm": 4.0 + index,
                },
                {
                    "session_id": "s",
                    "trial_id": "t",
                    "segment_id": "b",
                    "effective_lag_ms": 220.0 + index,
                    "aligned_rmse_mm": 5.0 + index,
                },
            )
            for index, method in enumerate(METHODS)
        }
        translation_figure = build_translation_panel(
            SimpleNamespace(translation_segments=rows)
        )

        point_legend = point_figure.axes[0].get_legend()
        translation_legend = translation_figure.axes[0].get_legend()
        self.assertIsNotNone(point_legend)
        self.assertIsNotNone(translation_legend)
        self.assertEqual(
            [text.get_text() for text in point_legend.get_texts()],
            ["Arrival", "Capture", "One-Euro", "EgoAnchor"],
        )
        self.assertEqual(
            [text.get_text() for text in translation_legend.get_texts()],
            ["Arrival", "Capture", "One-Euro", "EgoAnchor"],
        )

    def test_translation_panel_does_not_connect_methods(self) -> None:
        """图二(b) 不连接跨方法散点，避免方法顺序造成视觉混淆。"""

        rows = {
            method: (
                {
                    "session_id": "s",
                    "trial_id": "t",
                    "segment_id": "a",
                    "effective_lag_ms": 200.0 + index,
                    "aligned_rmse_mm": 4.0 + index,
                },
            )
            for index, method in enumerate(METHODS)
        }
        figure = build_translation_panel(SimpleNamespace(translation_segments=rows))

        paired_lines = [line for line in figure.axes[0].lines if line.get_alpha() == 0.20]
        self.assertEqual(paired_lines, [])

    def test_panel_titles_are_owned_by_tex_fragment(self) -> None:
        """绘图代码不重复写入由手工引入 TeX 子标题承担的小标题。"""

        values = {method: np.asarray((1.0, 2.0)) for method in METHODS}
        figure = build_point_panel(
            values,
            "Metric",
            np.asarray(((1.0, 1.1, 1.2, 1.3), (2.0, 2.1, 2.2, 2.3))),
        )

        self.assertEqual(figure.axes[0].get_title(), "")


if __name__ == "__main__":
    unittest.main()
