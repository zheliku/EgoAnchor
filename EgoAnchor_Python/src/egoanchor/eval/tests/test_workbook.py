"""Task 4 Stage 1 workbook writer 与回读验证测试。"""

from __future__ import annotations

import hashlib
import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from egoanchor.eval import (
    WorkbookValidationError,
    collect_source_files,
    decode_workbook_text,
    source_set_sha256,
    verify_task_workbook,
    write_task_workbook,
)

from .test_reader_qc import _read_jsonl, _write_jsonl, _write_valid_task
from ..preprocess import workbook as workbook_module


class WorkbookWriterTests(unittest.TestCase):
    """验证工作簿无损、原子、可分片且能够独立回读。"""

    def test_source_files_include_nested_audit_files_and_stable_digest(self) -> None:
        """来源目录中的嵌套审计文件必须进入文件清单和集合摘要。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            audit_file = root / "audit_samples" / "nested" / "sample.bin"
            audit_file.parent.mkdir(parents=True)
            audit_file.write_bytes(b"audit-sample")

            files = collect_source_files(root)
            first_digest = source_set_sha256(files)
            second_digest = source_set_sha256(collect_source_files(root))

            self.assertIn("audit_samples/nested/sample.bin", {item.relative_path for item in files})
            self.assertEqual(first_digest, second_digest)
            audit_file.write_bytes(b"changed")
            self.assertNotEqual(first_digest, source_set_sha256(collect_source_files(root)))

    def test_empty_and_reserved_text_are_encoded_reversibly(self) -> None:
        """空文本与保留前缀不得在 Excel 中变成 null 或内部 marker。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            candidate_path = root / "python_candidates.jsonl"
            candidates = _read_jsonl(candidate_path)
            literal_large_marker = f"@large:{'a' * 64}"
            candidates[0]["failure_reason"] = ""
            candidates[0]["pose_source"] = literal_large_marker
            _write_jsonl(candidate_path, candidates)
            output = Path(tmp) / "text-roundtrip.xlsx"
            write_task_workbook(root, output, code_version="test-version")
            workbook = load_workbook(output, read_only=True, data_only=False)
            sheet = workbook["python_candidates"]
            header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            row = next(sheet.iter_rows(min_row=2, max_row=2))
            encoded_empty = row[header.index("failure_reason")].value
            encoded_reserved = row[header.index("pose_source")].value
            workbook.close()

            self.assertIsInstance(encoded_empty, str)
            self.assertEqual(decode_workbook_text(encoded_empty), "")
            self.assertEqual(decode_workbook_text(encoded_reserved), literal_large_marker)

    def test_workbook_roundtrip_preserves_rows_types_hashes_and_is_deterministic(self) -> None:
        """合法 fixture 写出后通过独立回读，文本类型和二进制摘要保持稳定。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            candidate_path = root / "python_candidates.jsonl"
            candidates = _read_jsonl(candidate_path)
            candidates[0]["failure_reason"] = "=HYPERLINK(\"https://invalid\")"
            _write_jsonl(candidate_path, candidates)
            session = json.loads((root / "python_session.json").read_text(encoding="utf-8"))
            session["log_writer_stats"]["python_candidates.jsonl"]["rows_written"] = 1
            (root / "python_session.json").write_text(json.dumps(session), encoding="utf-8")
            first = Path(tmp) / "first.xlsx"
            second = Path(tmp) / "second.xlsx"

            first_result = write_task_workbook(root, first, code_version="test-version")
            second_result = write_task_workbook(root, second, code_version="test-version")
            verification = verify_task_workbook(first)

            self.assertTrue(verification.passed)
            self.assertEqual(first_result.verification.path, first)
            self.assertTrue(first_result.verification.path.is_file())
            self.assertEqual(first_result.sha256, second_result.sha256)
            self.assertEqual(first_result.sha256, hashlib.sha256(first.read_bytes()).hexdigest())
            self.assertEqual(verification.logical_row_counts["python_candidates"], 1)
            self.assertEqual(verification.logical_row_counts["unity_admission"], 8)
            self.assertEqual(verification.logical_row_counts["unity_render"], 8)

            workbook = load_workbook(first, read_only=True, data_only=False)
            sheet = workbook["python_candidates"]
            header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            row = next(sheet.iter_rows(min_row=2, max_row=2))
            failure_cell = row[header.index("failure_reason")]
            candidate_cell = row[header.index("candidate_id")]
            self.assertEqual(failure_cell.data_type, "s")
            self.assertEqual(candidate_cell.data_type, "s")
            self.assertFalse(any(cell.data_type == "f" for worksheet in workbook for cells in worksheet for cell in cells))
            workbook.close()

    def test_workbook_sheets_freeze_headers_and_define_readable_column_widths(self) -> None:
        """每个物理 sheet 都必须冻结表头，并为所有契约列写入稳定列宽。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "formatted.xlsx"
            write_task_workbook(root, output, code_version="test-version")

            workbook = load_workbook(output, read_only=False, data_only=False)
            for worksheet in workbook.worksheets:
                self.assertEqual(worksheet.freeze_panes, "A2", worksheet.title)
                for cell in worksheet[1]:
                    column_letter = cell.column_letter
                    self.assertIn(column_letter, worksheet.column_dimensions, worksheet.title)
                    width = worksheet.column_dimensions[column_letter].width
                    self.assertGreaterEqual(width, len(str(cell.value)) + 2, worksheet.title)
            workbook.close()

    def test_qc_failure_never_replaces_existing_output(self) -> None:
        """硬 QC 失败时不得留下临时文件，也不得替换已有正式文件。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            session_path = root / "python_session.json"
            session = json.loads(session_path.read_text(encoding="utf-8"))
            session["state"] = "python_running"
            session_path.write_text(json.dumps(session), encoding="utf-8")
            output = Path(tmp) / "existing.xlsx"
            output.write_bytes(b"existing-output")

            with self.assertRaises(WorkbookValidationError):
                write_task_workbook(root, output, code_version="test-version")

            self.assertEqual(output.read_bytes(), b"existing-output")
            self.assertFalse(any(output.parent.glob(f".{output.name}.*.tmp")))

    def test_output_inside_raw_task_is_rejected_without_creating_directory(self) -> None:
        """禁止输出到 raw task 内部，并且拒绝前不得创建目标父目录。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = root / "derived" / "forbidden.xlsx"

            with self.assertRaises(WorkbookValidationError):
                write_task_workbook(root, output, code_version="test-version")

            self.assertFalse(output.parent.exists())

    def test_large_values_are_chunked_without_truncation(self) -> None:
        """超过单元格限制的规范化值必须分片，并能由回读验证完整重建。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            candidate_path = root / "python_candidates.jsonl"
            candidates = _read_jsonl(candidate_path)
            candidates[0]["render_diagnostics"]["long_text"] = "长" * 200
            _write_jsonl(candidate_path, candidates)
            output = Path(tmp) / "chunked.xlsx"

            write_task_workbook(
                root,
                output,
                code_version="test-version",
                max_cell_chars=64,
            )
            verification = verify_task_workbook(output, max_cell_chars=64)

            self.assertTrue(verification.passed)
            self.assertGreater(verification.large_value_count, 0)
            workbook = load_workbook(output, read_only=True, data_only=False)
            chunks = list(workbook["large_values"].iter_rows(min_row=2, values_only=True))
            self.assertTrue(chunks)
            self.assertTrue(all(len(str(row[-1])) <= 64 for row in chunks))
            workbook.close()

    def test_fact_sheet_splits_with_stable_partition_names(self) -> None:
        """事实行超过注入上限时使用从一开始的稳定三位分片名。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "split.xlsx"

            write_task_workbook(
                root,
                output,
                code_version="test-version",
                max_data_rows=2,
            )
            verification = verify_task_workbook(output, max_data_rows=2)

            self.assertEqual(verification.logical_row_counts["unity_admission"], 8)
            self.assertEqual(
                verification.physical_sheets["unity_admission"],
                (
                    "unity_admission_001",
                    "unity_admission_002",
                    "unity_admission_003",
                    "unity_admission_004",
                ),
            )

    def test_readback_rejects_tampered_foreign_key(self) -> None:
        """独立回读必须拒绝指向未知 candidate 的 admission 外键。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "tampered.xlsx"
            write_task_workbook(root, output, code_version="test-version")
            workbook = load_workbook(output)
            sheet = workbook["unity_admission"]
            header = [cell.value for cell in sheet[1]]
            sheet.cell(row=2, column=header.index("candidate_id") + 1, value="unknown-candidate")
            workbook.save(output)
            workbook.close()

            with self.assertRaises(WorkbookValidationError):
                verify_task_workbook(output)

    def test_unknown_nested_fields_are_preserved_in_row_kv(self) -> None:
        """未来新增的嵌套 JSONL 字段必须无损进入通用 overflow 子表。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            candidate_path = root / "python_candidates.jsonl"
            candidates = _read_jsonl(candidate_path)
            candidates[0]["future_extension"] = {"values": ["001", 2], "empty": []}
            _write_jsonl(candidate_path, candidates)
            output = Path(tmp) / "overflow.xlsx"

            write_task_workbook(root, output, code_version="test-version")
            workbook = load_workbook(output, read_only=True, data_only=False)
            rows = list(workbook["row_kv"].iter_rows(min_row=2, values_only=True))
            header = [cell.value for cell in workbook["row_kv"][1]]
            path_index = header.index("json_path")
            value_index = header.index("value_json")
            values = {str(row[path_index]): row[value_index] for row in rows}
            workbook.close()

            self.assertEqual(values["future_extension.values[0]"], '"001"')
            self.assertEqual(values["future_extension.values[1]"], "2")
            self.assertEqual(values["future_extension.empty"], "[]")
            source_index = header.index("source_file")
            self.assertTrue(
                any(row[source_index] == "python_events.jsonl" and row[path_index] == "payload.condition_id" for row in rows)
            )
            self.assertTrue(
                any(row[source_index] == "unity_events.jsonl" and row[path_index] == "payload.condition_id" for row in rows)
            )

    def test_readback_rejects_tampered_sheet_index_count(self) -> None:
        """sheet_index 声明的物理行数与实际行数不一致时必须硬失败。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "tampered-index.xlsx"
            write_task_workbook(root, output, code_version="test-version")
            workbook = load_workbook(output)
            sheet = workbook["sheet_index"]
            header = [cell.value for cell in sheet[1]]
            sheet.cell(row=2, column=header.index("row_count") + 1, value=999999)
            workbook.save(output)
            workbook.close()

            with self.assertRaises(WorkbookValidationError):
                verify_task_workbook(output)

    def test_readback_rejects_fact_count_not_matching_source_file(self) -> None:
        """事实 sheet 与对应 JSONL 行数不一致时，即使同步篡改 index 也必须失败。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "tampered-fact-count.xlsx"
            write_task_workbook(root, output, code_version="test-version")
            workbook = load_workbook(output)
            workbook["unity_reference"].delete_rows(2, 1)
            index_sheet = workbook["sheet_index"]
            header = [cell.value for cell in index_sheet[1]]
            logical_column = header.index("logical_sheet") + 1
            count_column = header.index("row_count") + 1
            for row_index in range(2, index_sheet.max_row + 1):
                if index_sheet.cell(row=row_index, column=logical_column).value == "unity_reference":
                    index_sheet.cell(row=row_index, column=count_column, value=0)
                    break
            workbook.save(output)
            workbook.close()

            with self.assertRaises(WorkbookValidationError):
                verify_task_workbook(output)

    def test_large_value_marker_is_bound_to_exact_source(self) -> None:
        """大值 marker 改指另一条合法摘要时仍必须被来源绑定检查拒绝。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            candidate_path = root / "python_candidates.jsonl"
            candidates = _read_jsonl(candidate_path)
            candidates[0]["render_diagnostics"].update({"long_a": "a" * 200, "long_b": "b" * 200})
            _write_jsonl(candidate_path, candidates)
            output = Path(tmp) / "tampered-large.xlsx"
            write_task_workbook(root, output, code_version="test-version", max_cell_chars=64)
            workbook = load_workbook(output)
            sheet = workbook["candidate_diag"]
            header = [cell.value for cell in sheet[1]]
            storage_column = header.index("value_storage") + 1
            value_column = header.index("value_json") + 1
            large_rows = [
                row_index
                for row_index in range(2, sheet.max_row + 1)
                if sheet.cell(row=row_index, column=storage_column).value == "large_values"
            ]
            first_value = sheet.cell(row=large_rows[0], column=value_column).value
            second_value = sheet.cell(row=large_rows[1], column=value_column).value
            self.assertNotEqual(first_value, second_value)
            sheet.cell(row=large_rows[0], column=value_column, value=second_value)
            workbook.save(output)
            workbook.close()

            with self.assertRaises(WorkbookValidationError):
                verify_task_workbook(output, max_cell_chars=64)

    def test_typed_large_marker_requires_exact_source_group(self) -> None:
        """typed sheet 的内部 marker 缺少精确来源分片时必须硬失败。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            candidate_path = root / "python_candidates.jsonl"
            candidates = _read_jsonl(candidate_path)
            candidates[0]["failure_reason"] = "x" * 200
            _write_jsonl(candidate_path, candidates)
            output = Path(tmp) / "missing-large-source.xlsx"
            write_task_workbook(root, output, code_version="test-version", max_cell_chars=64)

            workbook = load_workbook(output)
            large_sheet = workbook["large_values"]
            large_header = [cell.value for cell in large_sheet[1]]
            table_column = large_header.index("source_table") + 1
            path_column = large_header.index("json_path") + 1
            deleted_rows = [
                row_index
                for row_index in range(2, large_sheet.max_row + 1)
                if large_sheet.cell(row=row_index, column=table_column).value == "python_candidates"
                and large_sheet.cell(row=row_index, column=path_column).value == "failure_reason"
            ]
            self.assertTrue(deleted_rows)
            for row_index in reversed(deleted_rows):
                large_sheet.delete_rows(row_index, 1)

            index_sheet = workbook["sheet_index"]
            index_header = [cell.value for cell in index_sheet[1]]
            physical_column = index_header.index("physical_sheet") + 1
            count_column = index_header.index("row_count") + 1
            for row_index in range(2, index_sheet.max_row + 1):
                if index_sheet.cell(row=row_index, column=physical_column).value == "large_values":
                    old_count = int(index_sheet.cell(row=row_index, column=count_column).value)
                    index_sheet.cell(row=row_index, column=count_column, value=old_count - len(deleted_rows))
                    break
            workbook.save(output)
            workbook.close()

            with self.assertRaises(WorkbookValidationError):
                verify_task_workbook(output, max_cell_chars=64)

    def test_source_mutation_before_publish_is_rejected(self) -> None:
        """QC/hash 后来源文件变化时不得发布混合版本工作簿。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "mutated-source.xlsx"
            candidate_path = root / "python_candidates.jsonl"
            original_write = workbook_module._write_workbook

            def mutate_after_write(*args: object, **kwargs: object) -> None:
                """先完成临时 XLSX，再修改一个来源文件模拟同步竞态。"""

                original_write(*args, **kwargs)
                candidate_path.write_bytes(candidate_path.read_bytes() + b"\n")

            with patch.object(workbook_module, "_write_workbook", side_effect=mutate_after_write):
                with self.assertRaises(WorkbookValidationError):
                    write_task_workbook(root, output, code_version="test-version")

            self.assertFalse(output.exists())

    def test_temporary_raw_archive_delete_retries_after_windows_file_lock(self) -> None:
        """刚关闭的临时 ZIP 首次被 Windows 锁定时，writer 应短暂重试并发布。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "retry.xlsx"
            original_unlink = Path.unlink
            raw_delete_attempts = 0

            def intermittent_unlink(path: Path, *, missing_ok: bool = False) -> None:
                """仅让 raw XLSX 的第一次删除模拟短暂的系统文件锁。"""

                nonlocal raw_delete_attempts
                if path.name.endswith(".raw.xlsx"):
                    raw_delete_attempts += 1
                    if raw_delete_attempts == 1:
                        raise PermissionError(32, "sharing violation", str(path))
                original_unlink(path, missing_ok=missing_ok)

            with patch.object(Path, "unlink", new=intermittent_unlink):
                artifact = write_task_workbook(root, output, code_version="test-version")

            self.assertTrue(artifact.path.is_file())
            self.assertGreaterEqual(raw_delete_attempts, 2)

    def test_readme_freezes_header_and_declares_column_widths(self) -> None:
        """每个新工作簿的 README 都冻结首行，并携带稳定的列宽定义。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "readme-layout.xlsx"
            write_task_workbook(root, output, code_version="test-version")

            with zipfile.ZipFile(output) as archive:
                readme_xml = archive.read("xl/worksheets/sheet1.xml")

            self.assertIn(b'topLeftCell="A2"', readme_xml)
            self.assertIn(b"<cols>", readme_xml)

    def test_atomic_replace_retries_after_windows_file_lock(self) -> None:
        """正式 XLSX 首次被 Windows 锁定时，writer 应重试原子替换。"""

        with tempfile.TemporaryDirectory() as tmp:
            root = _write_valid_task(Path(tmp))
            output = Path(tmp) / "replace-retry.xlsx"
            output.write_bytes(b"previous-workbook")
            original_replace = workbook_module.os.replace
            replace_attempts = 0

            def intermittent_replace(source: Path, destination: Path) -> None:
                """仅让正式目标的第一次替换模拟短暂的系统文件锁。"""

                nonlocal replace_attempts
                replace_attempts += 1
                if replace_attempts == 1:
                    raise PermissionError(5, "access denied", str(destination))
                original_replace(source, destination)

            with patch.object(workbook_module.os, "replace", side_effect=intermittent_replace):
                artifact = write_task_workbook(root, output, code_version="test-version")

            self.assertTrue(artifact.path.is_file())
            self.assertGreaterEqual(replace_attempts, 2)
            self.assertTrue(verify_task_workbook(output).passed)


if __name__ == "__main__":
    unittest.main()
