"""GPT v4 论文分析入口的边界与最小计算测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

import numpy as np
from openpyxl import Workbook  # type: ignore[import-untyped]

from egoanchor.eval import cli as eval_cli
from egoanchor.eval.gpt_v4 import (
    METHODS,
    build_point_panel,
    build_translation_panel,
    iter_rows,
    paired_metric_matrix,
)


class GptV4PipelineTests(unittest.TestCase):
    """冻结新管线只读取 Stage 1 XLSX 且不恢复旧阶段命令。"""

    def test_cli_replaces_old_analysis_stages_with_one_paper_build(self) -> None:
        """旧 analyze/publish/materialize 命令不得作为兼容层保留。"""

        self.assertEqual(eval_cli.STAGE_COMMANDS, ("qc", "preprocess", "build-paper"))

    def test_xlsx_reader_streams_selected_columns_from_stage_one_sheet(self) -> None:
        """新分析 reader 直接消费 Stage 1 sheet，不改写原始 workbook。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "task_1_complete.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "unity_render"
            sheet.append(("session_id", "variant_id", "render_mono_ms", "unused"))
            sheet.append(("session", "EgoAnchor", 123.5, "ignored"))
            workbook.save(path)
            before = path.read_bytes()

            rows = list(iter_rows(path, "unity_render", ("variant_id", "render_mono_ms")))

            self.assertEqual(
                rows,
                [{"variant_id": "EgoAnchor", "render_mono_ms": 123.5}],
            )
            self.assertEqual(path.read_bytes(), before)

    def test_build_paper_rejects_non_xlsx_input_before_writing(self) -> None:
        """论文入口拒绝 JSON/CSV，保持初始 XLSX 是唯一分析桥梁。"""

        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            source = root / "task_1.json"
            source.write_text("{}", encoding="utf-8")
            output = root / "output"

            code = eval_cli.main(
                [
                    "build-paper",
                    str(source),
                    "--out",
                    str(output),
                    "--paper-root",
                    str(root / "paper"),
                ]
            )

            self.assertEqual(code, eval_cli.EXIT_DATA_ERROR)
            self.assertFalse(output.exists())

    def test_figure_pairing_uses_event_identity_instead_of_row_position(self) -> None:
        """实验一折线必须按稳定片段键配对，即使各方法行顺序不同。"""

        rows = {}
        for method_index, method in enumerate(METHODS):
            method_rows = [
                {"session_id": "s", "trial_id": "t", "segment_id": "b", "value": method_index + 10.0},
                {"session_id": "s", "trial_id": "t", "segment_id": "a", "value": method_index + 1.0},
            ]
            if method_index % 2:
                method_rows.reverse()
            rows[method] = tuple(method_rows)

        matrix = paired_metric_matrix(rows, METHODS, ("value",))

        self.assertEqual(matrix[:, :, 0].tolist(), [[1.0, 2.0, 3.0, 4.0], [10.0, 11.0, 12.0, 13.0]])

    def test_figure_pairing_rejects_missing_method_episode(self) -> None:
        """任一方法缺失 episode 时禁止生成看似配对的趋势线。"""

        rows = {
            method: ({"session_id": "s", "trial_id": "t", "segment_id": "a", "value": 1.0},)
            for method in METHODS
        }
        rows[METHODS[-1]] = ()

        with self.assertRaisesRegex(ValueError, "配对不完整"):
            paired_metric_matrix(rows, METHODS, ("value",))

    def test_experiment_one_panels_have_local_legends(self) -> None:
        """图二的三个独立子图都必须自带方法图例。"""

        values = {method: np.asarray((1.0, 2.0)) for method in METHODS}
        point_figure = build_point_panel(
            values,
            "(a) Test",
            "Metric",
            np.asarray(((1.0, 1.1, 1.2, 1.3), (2.0, 2.1, 2.2, 2.3))),
        )
        rows = {
            method: (
                {
                    "session_id": "s",
                    "trial_id": "t",
                    "segment_id": "a",
                    "effective_lag_ms": 200.0 + index,
                    "aligned_rmse_mm": 4.0 + index,
                },
                {
                    "session_id": "s",
                    "trial_id": "t",
                    "segment_id": "b",
                    "effective_lag_ms": 220.0 + index,
                    "aligned_rmse_mm": 5.0 + index,
                },
            )
            for index, method in enumerate(METHODS)
        }
        translation_figure = build_translation_panel(
            SimpleNamespace(translation_segments=rows)
        )

        self.assertIsNotNone(point_figure.axes[0].get_legend())
        self.assertIsNotNone(translation_figure.axes[0].get_legend())

    def test_translation_panel_connects_paired_method_points(self) -> None:
        """图二(b) 用轻量折线显示同一片段在四种方法间的变化方向。"""

        rows = {
            method: (
                {
                    "session_id": "s",
                    "trial_id": "t",
                    "segment_id": "a",
                    "effective_lag_ms": 200.0 + index,
                    "aligned_rmse_mm": 4.0 + index,
                },
            )
            for index, method in enumerate(METHODS)
        }
        figure = build_translation_panel(SimpleNamespace(translation_segments=rows))

        paired_lines = [line for line in figure.axes[0].lines if line.get_alpha() == 0.20]
        self.assertEqual(len(paired_lines), 1)


if __name__ == "__main__":
    unittest.main()
