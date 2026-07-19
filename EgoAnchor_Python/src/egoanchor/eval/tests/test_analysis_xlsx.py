"""Stage 2 实验分析 XLSX 的同源、确定性和原子发布测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook

from egoanchor.eval import publish_analysis_outputs


def _metric_row() -> dict[str, object]:
    """创建一条实验一 event 指标行。"""

    return {
        "session_id": "session",
        "experiment_id": "exp1_system_characterization",
        "scenario_id": "static_head_motion",
        "trial_id": "trial",
        "event_id": "event",
        "condition_id": "exp1_system_characterization/static_head_motion",
        "variant_id": "EgoAnchor",
        "metric_key": "translation_event_pninetyfive_mm",
        "metric_value": 3.5,
        "metric_unit": "mm",
        "aggregation_level": "event",
        "input_workbook_sha256": "a" * 64,
    }


class AnalysisWorkbookTests(unittest.TestCase):
    """验证实验一审阅工作簿不成为新计算层。"""

    def test_exp1_workbook_is_typed_filtered_formula_free_and_deterministic(self) -> None:
        """同源表必须保持类型、实验过滤、公式安全和稳定二进制 hash。"""

        input_workbook = SimpleNamespace(
            path=Path("task_1_complete.xlsx"),
            sha256="a" * 64,
            session_id="session",
            row_count=10,
        )
        tables = {
            "exp1/event_metrics": [_metric_row()],
            "analysis_qc": [
                {
                    "check_id": "formula_text",
                    "status": "passed",
                    "observed": "=not-a-formula",
                    "expected": "text",
                    "details": "审阅工作簿必须防公式注入",
                }
            ],
            "numbers": [
                {
                    "experiment": "exp1_system_characterization",
                    "macro_name": "SessionCount",
                    "value": 1,
                    "source_csv": "common/trial_windows.csv",
                    "source_sha256": "",
                },
                {
                    "experiment": "exp2_design_attribution",
                    "macro_name": "SessionCount",
                    "value": 1,
                    "source_csv": "common/trial_windows.csv",
                    "source_sha256": "",
                },
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "results"
            first = publish_analysis_outputs(
                output,
                tables,
                input_workbooks=(input_workbook,),
                code_version="test-version",
                parameter_set_id="b" * 64,
            )
            first_hash = first.workbook_sha256["exp1_analysis.xlsx"]
            workbook_path = output / "exp1_analysis.xlsx"
            workbook = load_workbook(workbook_path, read_only=False, data_only=False)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    [
                        "workbook_info",
                        "sheet_index",
                        "inputs",
                        "analysis_qc",
                        "metric_catalog",
                        "trial_windows",
                        "event_metrics",
                        "trial_metrics",
                        "session_metrics",
                        "scenario_summary",
                        "head_motion_trace",
                        "start_stop_trace",
                        "lag_tradeoff",
                        "occlusion_trace",
                        "paper_numbers",
                        "paper_tables",
                        "lineage",
                    ],
                )
                event_sheet = workbook["event_metrics"]
                header = [cell.value for cell in event_sheet[1]]
                metric_column = header.index("metric_value") + 1
                self.assertEqual(event_sheet.cell(2, metric_column).value, 3.5)
                self.assertEqual(event_sheet.freeze_panes, "A2")
                self.assertEqual(event_sheet.auto_filter.ref, event_sheet.dimensions)
                observed_column = [cell.value for cell in workbook["analysis_qc"][1]].index("observed") + 1
                self.assertEqual(workbook["analysis_qc"].cell(2, observed_column).data_type, "s")
                self.assertEqual(
                    workbook["analysis_qc"].cell(2, observed_column).value,
                    "=not-a-formula",
                )
                numbers = list(workbook["paper_numbers"].iter_rows(min_row=2, values_only=True))
                self.assertEqual(len(numbers), 1)
                self.assertEqual(numbers[0][0], "exp1_system_characterization")
                lineage_header = [cell.value for cell in workbook["lineage"][1]]
                output_column = lineage_header.index("output_path") + 1
                self.assertTrue(
                    any(
                        row[output_column - 1] == "paper/numbers.csv"
                        for row in workbook["lineage"].iter_rows(min_row=2, values_only=True)
                    )
                )
                for sheet in workbook.worksheets:
                    for row in sheet.iter_rows():
                        self.assertFalse(any(cell.data_type == "f" for cell in row))
            finally:
                workbook.close()

            second = publish_analysis_outputs(
                output,
                tables,
                input_workbooks=(input_workbook,),
                code_version="test-version",
                parameter_set_id="b" * 64,
            )
            self.assertEqual(second.workbook_sha256["exp1_analysis.xlsx"], first_hash)

    def test_workbook_failure_preserves_existing_output(self) -> None:
        """XLSX 构建失败时不得提交同批 CSV 或破坏旧结果目录。"""

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "results"
            output.mkdir()
            sentinel = output / "sentinel.txt"
            sentinel.write_text("keep", encoding="utf-8")
            with patch(
                "egoanchor.eval.analysis.csv_output.write_analysis_workbooks",
                side_effect=ValueError("simulated workbook failure"),
            ):
                with self.assertRaisesRegex(ValueError, "simulated workbook failure"):
                    publish_analysis_outputs(output, {"analysis_qc": []})
            self.assertEqual(sentinel.read_text(encoding="utf-8"), "keep")

    def test_exp2_workbook_is_published_when_exp2_rows_exist(self) -> None:
        """实验二有数据时必须同时发布独立审阅工作簿。"""

        input_workbook = SimpleNamespace(
            path=Path("task_1_complete.xlsx"),
            sha256="a" * 64,
            session_id="session",
            row_count=10,
        )
        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "results"
            publish_analysis_outputs(
                output,
                {"exp2/event_metrics": [_metric_row()]},
                input_workbooks=(input_workbook,),
                code_version="test-version",
                parameter_set_id="b" * 64,
            )
            workbook = load_workbook(output / "exp2_analysis.xlsx", read_only=True, data_only=False)
            try:
                self.assertIn("paired_summary", workbook.sheetnames)
                self.assertIn("mechanism_attribution", workbook.sheetnames)
                self.assertIn("paper_tables", workbook.sheetnames)
            finally:
                workbook.close()

    def test_raw_workbook_cleanup_retries_windows_file_lock(self) -> None:
        """OpenPyXL 临时文件被短暂占用时应重试删除并完成发布。"""

        real_unlink = Path.unlink
        failures = 0

        def intermittent_unlink(path: Path, *args, **kwargs) -> None:
            """只让实验一 raw XLSX 的第一次删除模拟共享锁。"""

            nonlocal failures
            if path.name == ".exp1_analysis.raw.xlsx" and failures == 0:
                failures += 1
                raise PermissionError("simulated Windows workbook lock")
            real_unlink(path, *args, **kwargs)

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "results"
            with patch("pathlib.Path.unlink", new=intermittent_unlink):
                publish_analysis_outputs(output, {"analysis_qc": []})
            self.assertEqual(failures, 1)
            self.assertTrue((output / "exp1_analysis.xlsx").is_file())
            self.assertFalse((output / ".exp1_analysis.raw.xlsx").exists())


if __name__ == "__main__":
    unittest.main()
