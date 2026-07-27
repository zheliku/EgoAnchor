"""实验三正式模板、统计与结果产物测试。"""

from __future__ import annotations

import tempfile
import unittest
from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from egoanchor.eval import cli as eval_cli
from egoanchor.eval.paper_analysis.experiment_3 import (
    analyze_scores,
    build_raw_template,
    derive_scores,
    load_settings,
    publish_figures,
    read_workbook,
    signed_rank_test,
    validate_for_analysis,
    write_results_workbook,
)


REPOSITORY_ROOT = Path(__file__).resolve().parents[5]
"""包含论文目录与 Python 工程的仓库根目录。"""

R2_WORKBOOK = (
    REPOSITORY_ROOT
    / "2026-EgoAnchor"
    / "material"
    / "EgoAnchor_Experiment3_Simulated_Claude-Opus-5-1M_R2_v5_1_24P.xlsx"
)
"""只用于测试分析契约的冻结结构模拟工作簿。"""


class Experiment3Tests(unittest.TestCase):
    """验证空白模板、实时公式边界和离线结果链。"""

    def test_raw_template_is_empty_and_keeps_live_formulas_outside_raw_regions(self) -> None:
        """模板只保留设计映射，原始值区为空且公式仅位于派生表。"""

        settings = load_settings()
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "experiment3_raw.xlsx"
            build_raw_template(settings, output)
            data = read_workbook(output)
            self.assertEqual(data.source_kind, "formal")
            self.assertEqual(len(data.blocks), 144)
            workbook = load_workbook(output, data_only=False)
            try:
                self.assertEqual(
                    [workbook["Participants"].cell(3, column).value for column in range(12, 25)],
                    [None] * 13,
                )
                self.assertEqual(
                    [workbook["Records"].cell(5, column).value for column in range(11, 42)],
                    [None] * 31,
                )
                self.assertEqual(workbook["Records"]["X151"].value, "A/B归属回忆确认")
                self.assertEqual(workbook["Records"]["Y151"].value, "方法级记录有效")
                self.assertTrue(str(workbook["Derived"]["E5"].value).startswith("=IF(AND(COUNTIFS"))
                self.assertIn("QUARTILE.INC", str(workbook["Analysis"]["D16"].value))
                self.assertTrue(workbook.calculation.fullCalcOnLoad)
                self.assertTrue(workbook.calculation.forceFullCalc)
                self.assertEqual(workbook.properties.identifier, "EgoAnchor.Experiment3.RawData.v5.1")
                self.assertEqual(workbook.properties.category, "formal-participant-data")
                self.assertFalse(any("78" in str(item) for item in workbook["Analysis"].merged_cells.ranges))
                for sheet_name in ("Participants", "Records"):
                    self.assertFalse(
                        any(
                            cell.data_type == "f"
                            for row in workbook[sheet_name].iter_rows()
                            for cell in row
                        )
                    )
            finally:
                workbook.close()
            with self.assertRaises(FileExistsError):
                build_raw_template(settings, output)

    def test_reduced_aq_mode_updates_excel_and_python_contract_together(self) -> None:
        """AQ 缩减模式同时移除 EQ3/IQ1，并把有效区块门槛改为 11 项。"""

        settings = replace(load_settings(), aq_mode="reduced")
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "experiment3_reduced.xlsx"
            build_raw_template(settings, output)
            workbook = load_workbook(output, data_only=False)
            try:
                eq_formula = str(workbook["Derived"]["M5"].value)
                iq_formula = str(workbook["Derived"]["N5"].value)
                valid_formula = str(workbook["Derived"]["E5"].value)
                self.assertIn("Records!S5,Records!T5", eq_formula)
                self.assertNotIn("Records!U5", eq_formula)
                self.assertIn("Records!O5,Records!P5", iq_formula)
                self.assertNotIn("Records!V5", iq_formula)
                self.assertIn("=11", valid_formula)
            finally:
                workbook.close()

    def test_r2_simulation_reproduces_frozen_family_directions(self) -> None:
        """模拟输入只用于验证计分、配对方向和家族校正实现。"""

        settings = replace(load_settings(), bootstrap_iterations=1000, clmm_enabled=False)
        data = read_workbook(R2_WORKBOOK)
        validation = validate_for_analysis(
            data,
            minimum_participants=settings.minimum_participants,
            aq_mode=settings.aq_mode,
            q10_enabled=settings.q10_enabled,
            allow_synthetic=True,
        )
        scores = derive_scores(data, settings)
        tables = analyze_scores(data, scores, settings)
        primary = tables.primary.set_index("Outcome")
        scales = tables.scales.set_index("Outcome")
        self.assertEqual(validation["included_count"], 24)
        self.assertEqual(int(primary.loc["Q1", "N"]), 24)
        self.assertLess(float(primary.loc["Q1", "p_Holm"]), 0.01)
        self.assertGreater(float(primary.loc["Q2", "p_Holm"]), 0.05)
        self.assertLess(float(scales.loc["STIAS", "p_Holm"]), 0.01)
        workbook = load_workbook(R2_WORKBOOK, data_only=True, read_only=True)
        try:
            analysis = workbook["Analysis"]
            for row_number, outcome in enumerate(
                ("Q1", "Q8", "Q2", "Q9", "Q3", "Q6", "Q7"),
                start=5,
            ):
                offline = primary.loc[outcome]
                self.assertAlmostEqual(
                    float(analysis.cell(row_number, 5).value),
                    float(offline["Difference_Median"]),
                )
                self.assertAlmostEqual(
                    float(analysis.cell(row_number, 6).value),
                    float(offline["Difference_Mean"]),
                )
                self.assertAlmostEqual(
                    float(analysis.cell(row_number, 7).value),
                    float(offline["Difference_SD"]),
                )
                self.assertAlmostEqual(
                    float(analysis.cell(row_number, 8).value),
                    float(offline["dz"]),
                )
            for row_number, outcome in enumerate(
                ("AQ_EQ", "AQ_IQ", "TIA_RC", "TIA_UP", "STIAS"),
                start=15,
            ):
                offline = scales.loc[outcome]
                self.assertAlmostEqual(
                    float(analysis.cell(row_number, 4).value),
                    float(offline["Difference_Median"]),
                )
                self.assertAlmostEqual(
                    float(analysis.cell(row_number, 5).value),
                    float(offline["Difference_SD"]),
                )
                self.assertAlmostEqual(
                    float(analysis.cell(row_number, 6).value),
                    float(offline["dz"]),
                )
        finally:
            workbook.close()

    def test_results_workbook_and_figures_share_one_analysis_object(self) -> None:
        """结果 XLSX 可回读绘图，并同时生成非空 PNG/PDF。"""

        settings = replace(load_settings(), bootstrap_iterations=1000, clmm_enabled=False)
        data = read_workbook(R2_WORKBOOK)
        validation = validate_for_analysis(
            data,
            minimum_participants=settings.minimum_participants,
            aq_mode=settings.aq_mode,
            q10_enabled=settings.q10_enabled,
            allow_synthetic=True,
        )
        scores = derive_scores(data, settings)
        tables = analyze_scores(data, scores, settings)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            results = write_results_workbook(
                root / "experiment3_analysis.xlsx",
                data=data,
                scores=scores,
                tables=tables,
                clmm_coefficients=tables.primary.iloc[0:0],
                clmm_contrasts=tables.primary.iloc[0:0],
                settings=settings,
                settings_sha256="0" * 64,
                validation=validation,
            )
            figures = publish_figures(results, root, settings)
            self.assertGreater(results.stat().st_size, 50_000)
            self.assertEqual(set(figures), {"paired_png", "paired_pdf", "scales_png", "scales_pdf"})
            for path in figures.values():
                self.assertGreater(path.stat().st_size, 10_000)

    def test_signed_rank_exact_sign_dp_handles_zeros_and_ties(self) -> None:
        """精确秩检验删除零差，并以平均秩保留并列。"""

        result = signed_rank_test([0.0, 1.0, 1.0, -2.0])
        self.assertEqual(result["n_nonzero"], 3)
        self.assertAlmostEqual(float(result["w"]), 3.0)
        self.assertAlmostEqual(float(result["rank_biserial"]), 0.0)
        self.assertEqual(float(result["p_value"]), 1.0)

    def test_cli_exposes_nested_experiment3_workflow(self) -> None:
        """统一 CLI 暴露固定路径的模板、验证、分析和绘图命令。"""

        parser = eval_cli.build_parser()
        args = parser.parse_args(["experiment3", "analyze"])
        self.assertEqual(args.experiment3_command, "analyze")
        self.assertIs(args.handler, eval_cli._run_experiment3_analyze)


if __name__ == "__main__":
    unittest.main()
