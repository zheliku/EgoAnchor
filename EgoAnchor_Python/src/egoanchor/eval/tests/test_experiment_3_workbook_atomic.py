"""实验三结果工作簿原子发布与失败保护测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]

from egoanchor.eval.experiments.experiment_3.analysis import (
    MAIN_FAMILY,
    PRIMARY_OUTCOMES,
    SCALE_FAMILY,
    SCALE_OUTCOMES,
    AnalysisTables,
    Exp3Data,
    load_settings,
    write_results_workbook,
)


EXPECTED_SHEETS = (
    "说明",
    "样本与质控",
    "主结果",
    "分物体描述",
    "量表信度",
    "选择结果",
)
"""结果工作簿 v2 契约的固定六页顺序。"""

EXPECTED_OUTCOMES = (
    "Q1 静止稳定",
    "Q8 位置正确",
    "Q2 运动附着",
    "Q9 姿态一致",
    "Q3 恢复一致",
    "Q6 依赖意愿",
    "Q7 稳定—响应平衡",
    "AQ 嵌入质量",
    "AQ 交互质量",
    "TiA 可靠性/能力",
    "TiA 理解/可预测性",
    "S-TIAS 信任",
)
"""主结果页必须呈现的十二项冻结结局。"""


class Experiment3WorkbookAtomicTests(unittest.TestCase):
    """验证回读失败不破坏旧结果，成功时才原子替换。"""

    def test_readback_failure_preserves_existing_workbook_and_removes_temporary_file(self) -> None:
        """冻结结局顺序错误时，旧结果字节不变且临时 XLSX 被清理。"""

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            output = root / "experiment3_analysis.xlsx"
            _write_old_workbook(output)
            original_bytes = output.read_bytes()
            invalid_tables = _tables(tuple(reversed((*PRIMARY_OUTCOMES, *SCALE_OUTCOMES))))

            with self.assertRaisesRegex(ValueError, "主结果必须按冻结顺序"):
                _write(output, invalid_tables)

            self.assertEqual(output.read_bytes(), original_bytes)
            self.assertFalse(tuple(root.glob(f".{output.stem}.*.tmp.xlsx")))

    def test_success_replaces_existing_workbook_after_strict_readback(self) -> None:
        """合法六页和十二结局通过回读后，正式结果才替换旧文件。"""

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            output = root / "experiment3_analysis.xlsx"
            _write_old_workbook(output)
            original_bytes = output.read_bytes()

            result = _write(output, _tables((*PRIMARY_OUTCOMES, *SCALE_OUTCOMES)))

            self.assertEqual(result, output.resolve())
            self.assertNotEqual(output.read_bytes(), original_bytes)
            self.assertFalse(tuple(root.glob(f".{output.stem}.*.tmp.xlsx")))
            workbook = load_workbook(output, read_only=True, data_only=True)
            try:
                self.assertEqual(tuple(workbook.sheetnames), EXPECTED_SHEETS)
                self.assertEqual(
                    tuple(workbook["主结果"].cell(row, 2).value for row in range(5, 17)),
                    EXPECTED_OUTCOMES,
                )
                self.assertEqual(
                    sum(
                        workbook["分物体描述"].cell(row, 1).value is not None
                        for row in range(5, 26)
                    ),
                    21,
                )
            finally:
                workbook.close()


def _write(output: Path, tables: AnalysisTables) -> Path:
    """使用不依赖真实来源门禁的最小契约写入一个结果工作簿。"""

    data = Exp3Data(
        participants=pd.DataFrame(),
        blocks=pd.DataFrame(),
        methods=pd.DataFrame(),
        finals=pd.DataFrame(),
        source_kind="synthetic",
        source_path="memory://experiment-3-atomic-test",
        source_sha256="a" * 64,
    )
    validation = {
        "included_participants": ("P001", "P002"),
        "included_count": 2,
        "paper_eligible": False,
        "response_fingerprint": "test-only",
        "warnings": ("仅用于原子写入测试",),
    }
    return write_results_workbook(
        output,
        data=data,
        tables=tables,
        settings=load_settings(),
        settings_sha256="b" * 64,
        batch_config_path=output.parent / "batch.toml",
        paper_config_path=output.parent / "paper.toml",
        validation=validation,
    )


def _tables(outcome_order: tuple[str, ...]) -> AnalysisTables:
    """构造只覆盖工作簿写入契约的最小六表。"""

    results = pd.DataFrame(
        _result_row(outcome)
        for outcome in outcome_order
    )
    objects = pd.DataFrame(
        _object_row(outcome, object_key)
        for outcome in PRIMARY_OUTCOMES
        for object_key in ("blue_mouse", "stapler", "gamepad")
    )
    return AnalysisTables(
        sample=pd.DataFrame(columns=("Section", "Variable", "Category")),
        results=results,
        objects=objects,
        reliability=pd.DataFrame(),
        manipulation=pd.DataFrame(columns=("Type",)),
        choices=pd.DataFrame(columns=("Measure", "Category")),
    )


def _result_row(outcome: str) -> dict[str, object]:
    """返回一个数值完整、可稳定写入的主结果测试行。"""

    return {
        "Family": MAIN_FAMILY if outcome in PRIMARY_OUTCOMES else SCALE_FAMILY,
        "Outcome": outcome,
        "N": 2,
        "N_Nonzero": 2,
        "OneEuro_Q1": 3.0,
        "OneEuro_Median": 3.5,
        "OneEuro_Q3": 4.0,
        "EgoAnchor_Q1": 5.0,
        "EgoAnchor_Median": 5.5,
        "EgoAnchor_Q3": 6.0,
        "Difference_Q1": 1.0,
        "Difference_Median": 2.0,
        "Difference_Q3": 3.0,
        "W": 0.0,
        "p_Holm": 0.04,
        "r_rb": 1.0,
        "r_rb_CI_Low": 1.0,
        "r_rb_CI_High": 1.0,
        "r_rb_CI_Status": "degenerate_at_bound",
    }


def _object_row(outcome: str, object_key: str) -> dict[str, object]:
    """返回一个逐物体描述测试行，不包含任何推断字段。"""

    return {
        "Outcome": outcome,
        "Object_Key": object_key,
        "N": 2,
        "OneEuro_Q1": 3.0,
        "OneEuro_Median": 3.5,
        "OneEuro_Q3": 4.0,
        "EgoAnchor_Q1": 5.0,
        "EgoAnchor_Median": 5.5,
        "EgoAnchor_Q3": 6.0,
        "Difference_Q1": 1.0,
        "Difference_Median": 2.0,
        "Difference_Q3": 3.0,
        "Direction": "EgoAnchor_higher",
    }


def _write_old_workbook(path: Path) -> None:
    """写入一个合法但内容不同的旧工作簿，用于逐字节保护断言。"""

    workbook = Workbook()
    workbook.active.title = "旧结果"
    workbook.active["A1"] = "必须保留到新结果完成回读"
    try:
        workbook.save(path)
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
