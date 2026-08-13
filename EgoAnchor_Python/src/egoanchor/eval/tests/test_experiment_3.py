"""实验三五表模板、统计和完整构建回归测试。"""

from __future__ import annotations

import tempfile
import unittest
from dataclasses import replace
from pathlib import Path
from shutil import copyfile
from unittest import mock

from openpyxl import load_workbook  # type: ignore[import-untyped]
from PIL import Image

from egoanchor.eval import cli as eval_cli
from egoanchor.eval.experiments.experiment_3 import (
    ExperimentPaths,
    SettingsSnapshot,
    build_raw_template,
    load_settings,
    load_settings_snapshot,
    plan_assets,
    validate_workflow,
)
from egoanchor.eval.experiments.experiment_3 import workflow as experiment_workflow
from egoanchor.eval.experiments.experiment_3.analysis import (
    EXP3_ARTIFACTS,
    PRIMARY_OUTCOMES,
    WORKBOOK_CONTRACT_ID,
    analyze_scores,
    build_analysis,
    derive_scores,
    publish_figures,
    read_workbook,
    validate_complete_pair_counts,
    validate_for_analysis,
    write_results_workbook,
)


REPOSITORY_ROOT = Path(__file__).resolve().parents[5]
"""包含论文目录与 Python 工程的仓库根目录。"""

PYTHON_ROOT = REPOSITORY_ROOT / "EgoAnchor_Python"
"""包含 ``pixi.toml`` 的 Python 工程根目录。"""

RAW_WORKBOOK = (
    REPOSITORY_ROOT
    / "2026-EgoAnchor"
    / "material"
    / "EgoAnchor_Experiment3_RawData_Template_v5_3.xlsx"
)
"""实验三当前唯一的五表正式分析源。"""

PAPER_CONFIG = (
    PYTHON_ROOT / "src" / "egoanchor" / "eval" / "config" / "paper.toml"
)
"""受版本控制的实验三科学参数配置。"""


class Experiment3Tests(unittest.TestCase):
    """验证五表输入、直接统计和确定性分析产物。"""

    def test_settings_and_outcome_contract(self) -> None:
        """参数版本、样本阈值和连续 Q1--Q7 顺序不可漂移。"""

        settings = load_settings()
        self.assertEqual(settings.contract_version, 5)
        self.assertEqual(settings.template_version, "v5.3")
        self.assertEqual(settings.alpha, 0.05)
        self.assertEqual(settings.confidence_level, 0.95)
        self.assertEqual(settings.target_participants, 24)
        self.assertEqual(settings.minimum_participants, 18)
        self.assertEqual(PRIMARY_OUTCOMES, ("Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7"))
        self.assertFalse(hasattr(settings, "q10_enabled"))
        self.assertEqual(EXP3_ARTIFACTS.version, 9)

        source = PAPER_CONFIG.read_text(encoding="utf-8")
        cases = (
            ("minimum", "minimum_participants = 18", "minimum_participants = 18.5", "必须是 int"),
            ("contract", "version = 5", "version = true", "必须是 int"),
            ("alpha", "alpha = 0.05", "alpha = 1", "必须是 float"),
            ("width", "width_inches = 7.15", "width_inches = nan", "必须是有限浮点数"),
        )
        with tempfile.TemporaryDirectory() as directory:
            for name, original, replacement, error in cases:
                with self.subTest(parameter=name):
                    config = Path(directory) / f"{name}.toml"
                    config.write_text(source.replace(original, replacement, 1), encoding="utf-8")
                    with self.assertRaisesRegex(ValueError, error):
                        load_settings(config)

    def test_settings_snapshot_binds_values_and_digest(self) -> None:
        """配置快照固定同一次读取的值和摘要。"""

        source = PAPER_CONFIG.read_text(encoding="utf-8")
        with tempfile.TemporaryDirectory() as directory:
            config = Path(directory) / "paper.toml"
            config.write_text(source, encoding="utf-8")
            first = load_settings_snapshot(config)
            config.write_text(
                source.replace("bootstrap_iterations = 10000", "bootstrap_iterations = 10001", 1),
                encoding="utf-8",
            )
            second = load_settings_snapshot(config)
        self.assertEqual(first.settings.bootstrap_iterations, 10000)
        self.assertEqual(second.settings.bootstrap_iterations, 10001)
        self.assertNotEqual(first.sha256, second.sha256)

    def test_raw_template_keeps_design_and_clears_five_input_tables(self) -> None:
        """空白模板保留 24 单元设计，只清空参与者回答。"""

        settings = load_settings()
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "experiment3_blank.xlsx"
            build_raw_template(settings, output, source_template=RAW_WORKBOOK)
            workbook = load_workbook(output, data_only=False)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["Questionnaire", "Participants", "Block", "Method", "Final"],
                )
                self.assertEqual(
                    workbook.properties.identifier,
                    WORKBOOK_CONTRACT_ID,
                )
                questionnaire_items = tuple(
                    workbook["Questionnaire"].cell(row, 2).value
                    for row in range(12, 25)
                )
                self.assertEqual(
                    questionnaire_items,
                    (
                        "Q1", "Q2", "Q3", "Q4", "Q5", "Q6", "Q7",
                        "AQ_EQ1", "AQ_EQ2", "AQ_EQ3", "AQ_IQ1", "AQ_IQ2", "AQ_IQ3",
                    ),
                )
                self.assertNotIn("Q10_OPT", questionnaire_items)
                self.assertEqual(
                    workbook["Questionnaire"]["D20"].value,
                    "虚拟内容看起来真实、自然地融入了真实物体及其周围环境。",
                )
                b5_validations = tuple(
                    validation.formula1
                    for validation in workbook["Participants"].data_validations.dataValidation
                    if "P2:P25" in str(validation.sqref)
                )
                self.assertEqual(
                    b5_validations,
                    ('"从未,1–5 次,6–20 次,21 次及以上"',),
                )
                for sheet_name, first_input_column, last_input_column, last_row in (
                    ("Participants", 12, 22, 25),
                    ("Block", 11, 27, 145),
                    ("Method", 5, 20, 49),
                    ("Final", 2, 11, 25),
                ):
                    with self.subTest(sheet=sheet_name):
                        worksheet = workbook[sheet_name]
                        self.assertTrue(
                            all(
                                worksheet.cell(row, column).value is None
                                for row in range(2, last_row + 1)
                                for column in range(first_input_column, last_input_column + 1)
                            )
                        )
                self.assertEqual(workbook["Participants"]["A2"].value, "P001")
                self.assertEqual(workbook["Block"]["A2"].value, "P001")
                self.assertEqual(workbook["Method"]["A2"].value, "P001")
                self.assertEqual(workbook["Final"]["A2"].value, "P001")
                self.assertEqual(workbook["Participants"].freeze_panes, "L2")
                self.assertEqual(workbook["Block"].freeze_panes, "K2")
                self.assertTrue(workbook["Participants"].data_validations.count)
                self.assertTrue(workbook["Block"].data_validations.count)
            finally:
                workbook.close()
            described = read_workbook(output)
            self.assertEqual(len(described.participants), 24)
            self.assertEqual(len(described.blocks), 144)

    def test_reader_uses_new_numbering_and_rejects_contract_drift(self) -> None:
        """五表 reader 保留评分语义，并拒绝旧题号或错误契约。"""

        data = read_workbook(RAW_WORKBOOK)
        row = data.blocks.loc[
            (data.blocks["Participant_ID"] == "P002")
            & (data.blocks["Block_Index"] == 1)
        ].iloc[0]
        self.assertEqual(
            tuple(row[item] for item in PRIMARY_OUTCOMES),
            (5, 7, 4, 6, 5, 4, 6),
        )

        with tempfile.TemporaryDirectory() as directory:
            bad_contract = Path(directory) / "bad_contract.xlsx"
            copyfile(RAW_WORKBOOK, bad_contract)
            workbook = load_workbook(bad_contract)
            try:
                workbook.properties.identifier = "EgoAnchor.Experiment3.RawData.v5.1"
                workbook.save(bad_contract)
            finally:
                workbook.close()
            with self.assertRaisesRegex(ValueError, "契约标识不匹配"):
                read_workbook(bad_contract)

            bad_fallback = Path(directory) / "bad_fallback.xlsx"
            copyfile(RAW_WORKBOOK, bad_fallback)
            workbook = load_workbook(bad_fallback)
            try:
                workbook.properties.identifier = None
                workbook.properties.description = "EgoAnchor 实验三 v5.2 五表原始数据。"
                workbook.save(bad_fallback)
            finally:
                workbook.close()
            with self.assertRaisesRegex(ValueError, "v5.3 文档属性不匹配"):
                read_workbook(bad_fallback)

            bad_order = Path(directory) / "bad_order.xlsx"
            copyfile(RAW_WORKBOOK, bad_order)
            workbook = load_workbook(bad_order)
            try:
                workbook["Questionnaire"]["B14"] = "Q9"
                workbook.save(bad_order)
            finally:
                workbook.close()
            with self.assertRaisesRegex(ValueError, "1--13 页固定排列"):
                read_workbook(bad_order)

    def test_validate_and_derive_complete_pairs(self) -> None:
        """新数据直接形成 24 人的十二项参与者内配对。"""

        settings = replace(load_settings(), bootstrap_iterations=1000)
        data = read_workbook(RAW_WORKBOOK)
        readiness = validate_for_analysis(data, aq_mode=settings.aq_mode)
        scores = derive_scores(data, settings)
        pair_counts = validate_complete_pair_counts(
            scores.paired_scores,
            settings.minimum_participants,
        )
        tables = analyze_scores(data, scores, settings)
        self.assertEqual(readiness["included_count"], 24)
        self.assertEqual(set(pair_counts), set((*PRIMARY_OUTCOMES, "AQ_EQ", "AQ_IQ", "TIA_RC", "TIA_UP", "STIAS")))
        self.assertTrue(all(count == 24 for count in pair_counts.values()))
        self.assertEqual(tuple(tables.results["Outcome"][:7]), PRIMARY_OUTCOMES)
        age = tables.sample[tables.sample["Variable"] == "Age"].iloc[0]
        self.assertEqual(int(age["N"]), 24)
        self.assertEqual(int(age["Denominator"]), 24)
        self.assertTrue(
            (tables.sample[tables.sample["Section"] == "Design_Balance"]["Status"] == "balanced").all()
        )

    def test_validation_rejects_score_range_and_design_drift(self) -> None:
        """直接分析仍拒绝非法评分和错位的区块顺序。"""

        settings = load_settings()
        with tempfile.TemporaryDirectory() as directory:
            invalid_score = Path(directory) / "invalid_score.xlsx"
            copyfile(RAW_WORKBOOK, invalid_score)
            workbook = load_workbook(invalid_score)
            try:
                workbook["Block"]["K2"] = 8
                workbook.save(invalid_score)
            finally:
                workbook.close()
            with self.assertRaisesRegex(ValueError, "必须是 1--7 整数"):
                validate_for_analysis(read_workbook(invalid_score), aq_mode=settings.aq_mode)

            invalid_order = Path(directory) / "invalid_order.xlsx"
            copyfile(RAW_WORKBOOK, invalid_order)
            workbook = load_workbook(invalid_order)
            try:
                workbook["Block"]["B2"] = 2
                workbook["Block"]["B3"] = 1
                workbook.save(invalid_order)
            finally:
                workbook.close()
            with self.assertRaisesRegex(ValueError, "实际区块顺序与 Participants 计划不一致"):
                read_workbook(invalid_order)

    def test_build_requires_minimum_pairs_for_every_outcome(self) -> None:
        """任一冻结结局配对不足时拒绝发布。"""

        settings = replace(load_settings(), bootstrap_iterations=1000)
        data = read_workbook(RAW_WORKBOOK)
        methods = data.methods.copy()
        participant_ids = sorted(methods["Participant_ID"].astype(str).unique())
        mask = methods["Participant_ID"].astype(str).isin(participant_ids[1:])
        rc_items = ("TIA_RC1", "TIA_RC2", "TIA_RC3_REV", "TIA_RC4", "TIA_RC5_REV", "TIA_RC6")
        for item in rc_items:
            methods[item] = methods[item].astype(object)
        methods.loc[mask, list(rc_items)] = "无法回答"
        incomplete = replace(data, methods=methods)
        with mock.patch(
            "egoanchor.eval.experiments.experiment_3.workflow.read_workbook",
            return_value=incomplete,
        ), self.assertRaisesRegex(ValueError, r"TIA_RC N=1"):
            paths = _temporary_paths(RAW_WORKBOOK, Path(tempfile.gettempdir()) / "unused")
            with (
                mock.patch(
                    "egoanchor.eval.experiments.experiment_3.workflow.load_paths",
                    return_value=paths,
                ),
                mock.patch(
                    "egoanchor.eval.experiments.experiment_3.workflow.load_settings_snapshot",
                    return_value=SettingsSnapshot(settings=settings, sha256="0" * 64),
                ),
            ):
                validate_workflow()

    def test_failed_staged_publish_preserves_previous_tree(self) -> None:
        """产物写出失败时保留上一轮分析目录。"""

        settings = replace(load_settings(), bootstrap_iterations=1000)
        data = read_workbook(RAW_WORKBOOK)
        with tempfile.TemporaryDirectory(dir=PYTHON_ROOT / "data") as directory:
            analysis_root = Path(directory) / "analysis"
            analysis_root.mkdir()
            marker = analysis_root / "previous_complete.txt"
            marker.write_text("previous", encoding="utf-8")
            with (
                mock.patch(
                    "egoanchor.eval.experiments.experiment_3.analysis.pipeline.read_workbook",
                    return_value=data,
                ),
                mock.patch(
                    "egoanchor.eval.experiments.experiment_3.analysis.pipeline.write_subjective_table",
                    side_effect=RuntimeError("simulated TeX failure"),
                ),
                self.assertRaisesRegex(RuntimeError, "simulated TeX failure"),
            ):
                build_analysis(
                    settings,
                    input_workbook=RAW_WORKBOOK,
                    output_root=analysis_root,
                    project_root=PYTHON_ROOT,
                    config_sha256="0" * 64,
                    batch_config_path=Path(directory) / "batch.toml",
                    paper_config_path=Path(directory) / "paper.toml",
                )
            self.assertEqual(marker.read_text(encoding="utf-8"), "previous")
            self.assertFalse(tuple(analysis_root.parent.glob(".analysis.build-*")))

    def test_interrupted_staged_build_cleans_staging(self) -> None:
        """生成阶段收到 Ctrl+C 时保留旧构建并清理本轮暂存目录。"""

        settings = replace(load_settings(), bootstrap_iterations=1000)
        data = read_workbook(RAW_WORKBOOK)
        with tempfile.TemporaryDirectory(dir=PYTHON_ROOT / "data") as directory:
            analysis_root = Path(directory) / "analysis"
            analysis_root.mkdir()
            marker = analysis_root / "previous_complete.txt"
            marker.write_text("previous", encoding="utf-8")
            with (
                mock.patch(
                    "egoanchor.eval.experiments.experiment_3.analysis.pipeline.read_workbook",
                    return_value=data,
                ),
                mock.patch(
                    "egoanchor.eval.experiments.experiment_3.analysis.pipeline.write_subjective_table",
                    side_effect=KeyboardInterrupt(),
                ),
                self.assertRaises(KeyboardInterrupt),
            ):
                build_analysis(
                    settings,
                    input_workbook=RAW_WORKBOOK,
                    output_root=analysis_root,
                    project_root=PYTHON_ROOT,
                    config_sha256="0" * 64,
                    batch_config_path=Path(directory) / "batch.toml",
                    paper_config_path=Path(directory) / "paper.toml",
                )
            self.assertEqual(marker.read_text(encoding="utf-8"), "previous")
            self.assertFalse(tuple(analysis_root.parent.glob(".analysis.build-*")))

    def test_complete_build_writes_six_results_pages_and_one_figure_pair(self) -> None:
        """完整构建写出 XLSX、TeX 和同一张复合图的 PNG/PDF。"""

        settings = replace(load_settings(), bootstrap_iterations=1000)
        with tempfile.TemporaryDirectory(dir=PYTHON_ROOT / "data") as directory:
            root = Path(directory)
            analysis_root = root / "analysis"
            progress: list[str] = []
            payload = build_analysis(
                settings,
                input_workbook=RAW_WORKBOOK,
                output_root=analysis_root,
                project_root=PYTHON_ROOT,
                config_sha256="0" * 64,
                batch_config_path=root / "batch.toml",
                paper_config_path=root / "paper.toml",
                progress=progress.append,
            )
            self.assertEqual(payload["build"]["details"]["included_count"], 24)
            self.assertEqual(payload["build"]["details"]["publish_mode"], "managed_files")
            self.assertEqual(len(progress), 6)
            self.assertEqual(progress[-1], "提交完整分析产物")
            results = analysis_root / "results" / "experiment3_analysis.xlsx"
            workbook = load_workbook(results, read_only=True, data_only=True)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["说明", "样本与质控", "主结果", "分物体描述", "量表信度", "选择结果"],
                )
            finally:
                workbook.close()
            figures = {
                "png": analysis_root / "figures" / "figure4_exp3_subjective_outcomes.png",
                "pdf": analysis_root / "figures" / "figure4_exp3_subjective_outcomes.pdf",
            }
            self.assertTrue(all(path.stat().st_size > 10_000 for path in figures.values()))
            with Image.open(figures["png"]) as image:
                expected = tuple(round(value * settings.figure_dpi) for value in settings.figure_size)
                self.assertEqual(image.size, expected)

            paths = _temporary_paths(RAW_WORKBOOK, analysis_root)
            with (
                mock.patch(
                    "egoanchor.eval.experiments.experiment_3.workflow.load_paths",
                    return_value=paths,
                ),
                mock.patch(
                    "egoanchor.eval.experiments.experiment_3.workflow.load_settings_snapshot",
                    return_value=SettingsSnapshot(settings=settings, sha256="0" * 64),
                ),
            ):
                plan = plan_assets()
            self.assertEqual(
                {asset.key for asset in plan.assets},
                {"figure4_png", "figure4_pdf", "subjective_table"},
            )

    def test_results_and_figure_reject_inconsistent_pair_difference(self) -> None:
        """结果簿和图共同消费同一批配对分。"""

        settings = replace(load_settings(), bootstrap_iterations=1000)
        data = read_workbook(RAW_WORKBOOK)
        validation = validate_for_analysis(data, aq_mode=settings.aq_mode)
        scores = derive_scores(data, settings)
        tables = analyze_scores(data, scores, settings)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            results = write_results_workbook(
                root / "experiment3_analysis.xlsx",
                data=data,
                tables=tables,
                settings=settings,
                settings_sha256="0" * 64,
                batch_config_path=root / "batch.toml",
                paper_config_path=root / "paper.toml",
                validation=validation,
            )
            result_workbook = load_workbook(results, read_only=True, data_only=False)
            try:
                self.assertEqual(
                    result_workbook.sheetnames,
                    ["说明", "样本与质控", "主结果", "分物体描述", "量表信度", "选择结果"],
                )
                self.assertFalse(
                    any(
                        cell.value == "Q10"
                        for sheet in result_workbook.worksheets
                        for row in sheet.iter_rows()
                        for cell in row
                    )
                )
            finally:
                result_workbook.close()
            publish_figures(scores, tables, root, settings)
            bad = scores.paired_scores.copy()
            bad.at[bad.index[0], "Difference"] += 0.25
            with self.assertRaisesRegex(ValueError, "配对差不等于"):
                publish_figures(
                    replace(scores, paired_scores=bad),
                    tables,
                    root / "bad",
                    settings,
                )

    def test_cli_exposes_experiment3_as_analyze_target(self) -> None:
        """实验三继续使用统一 analyze 入口。"""

        parser = eval_cli.build_parser()
        args = parser.parse_args(["analyze", "exp3"])
        self.assertEqual(args.target, "exp3")
        self.assertIs(args.handler, eval_cli._run_analyze)

    def test_cli_progress_total_matches_reported_stages(self) -> None:
        """实验三上报全部阶段后进度条必须恰好到达 100%。"""

        parser = eval_cli.build_parser()
        args = parser.parse_args(["analyze", "exp3"])
        bar = mock.MagicMock()

        def report_all_stages(
            target: str,
            *,
            rebuild_experiment_1_2: bool = False,
            experiment_3_progress: object | None = None,
        ) -> dict[str, bool]:
            """模拟工作区按真实阶段数依次回调。"""

            self.assertEqual(target, "exp3")
            self.assertFalse(rebuild_experiment_1_2)
            assert callable(experiment_3_progress)
            for number in range(6):
                experiment_3_progress(f"stage {number}")
            return {"passed": True}

        with (
            mock.patch.object(eval_cli, "tqdm") as progress,
            mock.patch.object(
                eval_cli,
                "analyze_workspace",
                side_effect=report_all_stages,
            ),
        ):
            progress.return_value.__enter__.return_value = bar
            result = eval_cli._run_analyze(args)

        self.assertTrue(result["passed"])
        progress.assert_called_once_with(
            total=6,
            desc="analyze exp3",
            unit="stage",
            leave=False,
        )
        self.assertEqual(bar.update.call_count, 6)


def _temporary_paths(source: Path, analysis_root: Path) -> ExperimentPaths:
    """为 workflow 和论文资源计划构造仓库内测试路径。"""

    sandbox = analysis_root.parent
    return ExperimentPaths(
        project_root=PYTHON_ROOT,
        source_template=source,
        input_workbook=source,
        analysis_root=analysis_root,
        paper_root=sandbox / "paper",
        figure_destination=sandbox / "paper" / "figures",
        table_destination=sandbox / "paper" / "tables" / "exp3_subjective.tex",
        batch_config_path=sandbox / "batch.toml",
    )


if __name__ == "__main__":
    unittest.main()
