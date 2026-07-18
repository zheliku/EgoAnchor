"""Task 9 Stage 2 XLSX loader 的边界测试。"""

from __future__ import annotations

import tempfile
import unittest
from unittest.mock import patch
from pathlib import Path

from openpyxl import Workbook

from egoanchor.eval import LoadedBatch, WorkbookInput, load_workbook_batch


class LoaderTests(unittest.TestCase):
    """验证 loader 只接受 Stage 1 workbook，并拒绝不完整批次。"""

    def test_loader_rejects_non_xlsx_input_before_opening(self) -> None:
        """JSONL 或目录不能伪装成 Stage 2 输入。"""

        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "raw.jsonl"
            source.write_text("{}\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "XLSX"):
                load_workbook_batch((source,))

    def test_loader_rejects_workbook_without_required_sheet(self) -> None:
        """缺少 variants 等 Stage 1 sheet 时必须硬失败。"""

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "task_1_complete.xlsx"
            workbook = Workbook()
            workbook.active.title = "manifest"
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "缺少必需 sheet"):
                load_workbook_batch((path,))

    def test_loader_rejects_duplicate_session_across_workbooks(self) -> None:
        """同一 session 不能在同一分析批次中出现两次。"""

        source = WorkbookInput(
            Path("one.xlsx"),
            "a" * 64,
            "same-session",
            1,
            "object",
            "formal",
            "v1",
            "config",
            "params",
            "model",
        )
        loaded = LoadedBatch((source,), (), (), (), (), (), ())
        with patch("egoanchor.eval.analysis.loader.load_workbook", return_value=(source, loaded)):
            with self.assertRaisesRegex(ValueError, "重复 session_id"):
                load_workbook_batch((Path("one.xlsx"), Path("two.xlsx")))

    def test_loader_orders_same_workbook_set_independently_of_argument_order(self) -> None:
        """相同 workbook 集合的加载和输出顺序不得受命令行参数顺序影响。"""

        def loaded(path: Path) -> tuple[WorkbookInput, LoadedBatch]:
            """为路径构造具有一致批次签名的最小加载结果。"""

            session_id = path.stem
            source = WorkbookInput(
                path,
                ("a" if session_id == "a" else "b") * 64,
                session_id,
                1,
                "object",
                "formal",
                "v1",
                "config",
                "params",
                "model",
            )
            return source, LoadedBatch((source,), (), (), (), (), (), ())

        with patch("egoanchor.eval.analysis.loader.load_workbook", side_effect=loaded):
            forward = load_workbook_batch((Path("a.xlsx"), Path("b.xlsx")))
            reverse = load_workbook_batch((Path("b.xlsx"), Path("a.xlsx")))

        self.assertEqual(
            tuple(item.session_id for item in forward.inputs),
            tuple(item.session_id for item in reverse.inputs),
        )
        self.assertEqual(tuple(item.session_id for item in forward.inputs), ("a", "b"))


if __name__ == "__main__":
    unittest.main()
