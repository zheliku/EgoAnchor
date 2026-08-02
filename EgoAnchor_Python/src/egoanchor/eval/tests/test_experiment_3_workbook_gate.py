"""实验三结果工作簿来源门禁与回读防篡改测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook  # type: ignore[import-untyped]

from egoanchor.eval.experiments.experiment_3.analysis import (
    MAIN_FAMILY,
    OBJECTS,
    PRIMARY_OUTCOMES,
    SCALE_FAMILY,
    SCALE_OUTCOMES,
    AnalysisTables,
    Exp3Data,
    load_settings,
    write_results_workbook,
)
from egoanchor.eval.experiments.experiment_3.analysis import workbook as workbook_module


_INPUT_SHA = "1" * 64
_FINGERPRINT = "2" * 64
_SETTINGS_SHA = "3" * 64
_INCLUDED = tuple(f"P{index:03d}" for index in range(1, 19))
"""单元测试使用的稳定摘要与最低完整配对参与者集合。"""


class Experiment3WorkbookGateTests(unittest.TestCase):
    """验证结果簿不能脱离输入来源结论，也不能在写出后静默漂移。"""

    def test_writer_rejects_source_gate_status_incompatible_with_input_kind(self) -> None:
        """formal/nonformal 状态必须与同一次 ``Exp3Data`` 快照兼容。"""

        cases = (
            ("formal", "nonformal", False, "formal 输入不得标记为 nonformal"),
            ("synthetic", "approved", True, "只适用于 formal 输入"),
            ("unknown", "unapproved_formal", False, "只适用于 formal 输入"),
        )
        with tempfile.TemporaryDirectory() as directory:
            for source_kind, status, eligible, message in cases:
                with self.subTest(source_kind=source_kind, status=status):
                    validation = _validation(source_kind, status, eligible)
                    with self.assertRaisesRegex(ValueError, message):
                        _write(
                            Path(directory) / f"{source_kind}_{status}.xlsx",
                            _data(source_kind),
                            validation,
                        )

    def test_writer_rejects_detached_source_kind_and_invalid_fingerprint(self) -> None:
        """读取器元数据必须绑定当前输入且响应指纹必须为规范 SHA-256。"""

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "result.xlsx"
            detached = _validation("synthetic", "known_synthetic", False)
            with self.assertRaisesRegex(ValueError, "未绑定当前输入的 source_kind"):
                _write(output, _data("formal"), detached)

            invalid = _validation("formal", "approved", True)
            invalid["response_fingerprint"] = "A" * 64
            with self.assertRaisesRegex(ValueError, "核心响应指纹"):
                _write(output, _data("formal"), invalid)

    def test_known_synthetic_status_can_override_a_formal_self_marker(self) -> None:
        """已知合成指纹优先拒绝时，formal 自报标记不能恢复论文资格。"""

        with tempfile.TemporaryDirectory() as directory:
            output = _write(
                Path(directory) / "known_synthetic.xlsx",
                _data("formal"),
                _validation("formal", "known_synthetic", False),
            )
            facts = _read_info_facts(output)
            self.assertEqual(facts["来源门禁状态"], "known_synthetic")
            self.assertEqual(facts["可用于论文"], "否")

    def test_readback_rejects_tampered_gate_and_digest_facts(self) -> None:
        """说明页任一来源或配置事实被改写后，回读必须失败。"""

        tampered_facts = {
            "来源门禁状态": "known_synthetic",
            "可用于论文": "否",
            "输入 SHA-256": "4" * 64,
            "参数 SHA-256": "5" * 64,
            "核心响应指纹": "6" * 64,
        }
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            for label, replacement in tampered_facts.items():
                with self.subTest(label=label):
                    output = _write(
                        root / f"tampered_{len(label)}.xlsx",
                        _data("formal"),
                        _validation("formal", "approved", True),
                    )
                    _replace_info_fact(output, label, replacement)
                    with self.assertRaisesRegex(ValueError, label):
                        _verify(output)

    def test_readback_rejects_replaced_outcome_object_key(self) -> None:
        """21 行数量不变时，重复或替换任一物体键仍必须失败。"""

        with tempfile.TemporaryDirectory() as directory:
            output = _write(
                Path(directory) / "tampered_object.xlsx",
                _data("formal"),
                _validation("formal", "approved", True),
            )
            workbook = load_workbook(output)
            try:
                worksheet = workbook["分物体描述"]
                worksheet.cell(5, 2, "错误物体")
                workbook.save(output)
            finally:
                workbook.close()
            with self.assertRaisesRegex(ValueError, "七个主条目乘三个对象"):
                _verify(output)


def _data(source_kind: str) -> Exp3Data:
    """构造只供结果簿契约测试使用的最小输入快照。"""

    return Exp3Data(
        participants=pd.DataFrame(),
        blocks=pd.DataFrame(),
        methods=pd.DataFrame(),
        finals=pd.DataFrame(),
        source_kind=source_kind,
        source_path="P:/test/source.xlsx",
        source_sha256=_INPUT_SHA,
    )


def _validation(source_kind: str, status: str, paper_eligible: bool) -> dict[str, Any]:
    """构造与读取器正式返回字段一致的门禁摘要。"""

    return {
        "included_participants": _INCLUDED,
        "included_count": len(_INCLUDED),
        "warnings": (),
        "source_kind": source_kind,
        "response_fingerprint": _FINGERPRINT,
        "paper_eligible": paper_eligible,
        "source_gate_status": status,
        "source_gate_reason": "测试门禁原因",
    }


def _tables() -> AnalysisTables:
    """构造覆盖六页写入路径且冻结键完整的最小结果表。"""

    sample = pd.DataFrame(
        columns=(
            "Section",
            "Variable",
            "Category",
            "N",
            "Denominator",
            "Proportion",
            "Mean",
            "SD",
            "Q1",
            "Median",
            "Q3",
            "Min",
            "Max",
            "Expected_At_Actual_N",
            "Deviation_From_Actual_Balance",
            "Status",
        )
    )
    results = pd.DataFrame(
        _result_row(outcome, MAIN_FAMILY)
        for outcome in PRIMARY_OUTCOMES
    )
    scale_results = pd.DataFrame(
        _result_row(outcome, SCALE_FAMILY)
        for outcome in SCALE_OUTCOMES
    )
    results = pd.concat((results, scale_results), ignore_index=True)
    objects = pd.DataFrame(
        _object_row(outcome, object_key)
        for outcome in PRIMARY_OUTCOMES
        for object_key in OBJECTS
    )
    return AnalysisTables(
        sample=sample,
        results=results,
        objects=objects,
        reliability=pd.DataFrame(),
        manipulation=pd.DataFrame(columns=("Type",)),
        choices=pd.DataFrame(columns=("Measure", "Category")),
    )


def _result_row(outcome: str, family: str) -> dict[str, Any]:
    """返回一行可写入主结果页的中性推断结果。"""

    return {
        "Family": family,
        "Outcome": outcome,
        "N": len(_INCLUDED),
        "N_Nonzero": len(_INCLUDED),
        "OneEuro_Q1": 3.0,
        "OneEuro_Median": 4.0,
        "OneEuro_Q3": 5.0,
        "EgoAnchor_Q1": 3.0,
        "EgoAnchor_Median": 4.0,
        "EgoAnchor_Q3": 5.0,
        "Difference_Q1": -1.0,
        "Difference_Median": 0.0,
        "Difference_Q3": 1.0,
        "W": 40.0,
        "p_Holm": 1.0,
        "r_rb": 0.0,
        "r_rb_CI_Low": -0.2,
        "r_rb_CI_High": 0.2,
        "r_rb_CI_Status": "estimated",
    }


def _object_row(outcome: str, object_key: str) -> dict[str, Any]:
    """返回一行可写入分物体描述页的中性结果。"""

    row = _result_row(outcome, MAIN_FAMILY)
    row.update({"Object_Key": object_key, "Direction": "median_tie"})
    return row


def _write(destination: Path, data: Exp3Data, validation: dict[str, Any]) -> Path:
    """通过公开入口写出并完成第一次回读验证。"""

    return write_results_workbook(
        destination,
        data=data,
        tables=_tables(),
        settings=load_settings(),
        settings_sha256=_SETTINGS_SHA,
        batch_config_path=destination.parent / "batch.toml",
        paper_config_path=destination.parent / "paper.toml",
        validation=validation,
    )


def _verify(path: Path) -> None:
    """按最初写入事实重新执行内部回读器。"""

    workbook_module._verify_results_workbook(
        path,
        input_sha256=_INPUT_SHA,
        settings_sha256=_SETTINGS_SHA,
        included_count=len(_INCLUDED),
        source_gate_status="approved",
        response_fingerprint=_FINGERPRINT,
    )


def _read_info_facts(path: Path) -> dict[str, Any]:
    """读取说明页两列事实，供成功路径断言。"""

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = workbook["说明"]
        return {
            str(worksheet.cell(row, 1).value): worksheet.cell(row, 2).value
            for row in range(1, worksheet.max_row + 1)
            if worksheet.cell(row, 1).value is not None
        }
    finally:
        workbook.close()


def _replace_info_fact(path: Path, label: str, replacement: str) -> None:
    """在保留其余结构的前提下篡改单个说明页事实。"""

    workbook = load_workbook(path)
    try:
        worksheet = workbook["说明"]
        for row in range(1, worksheet.max_row + 1):
            if worksheet.cell(row, 1).value == label:
                worksheet.cell(row, 2, replacement)
                workbook.save(path)
                return
        raise AssertionError(f"测试工作簿缺少说明页事实：{label}")
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
