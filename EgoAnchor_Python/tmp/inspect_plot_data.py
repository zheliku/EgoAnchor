"""临时诊断：读出 figure_plot_data.xlsx 的所有 sheet 与关键统计，核对表 2 的平移通道口径。"""

from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import load_workbook

PATH = Path("P:/VSCode-Project/EgoAnchor/EgoAnchor_Python/data/experiments/experiment_1_2/analysis/plots/figure_plot_data.xlsx")


def main() -> None:
    wb = load_workbook(PATH, read_only=True, data_only=True)
    print("sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = rows[0]
        body = [r for r in rows[1:] if r and any(v is not None for v in r)]
        print("\n" + "=" * 74)
        print(f"sheet={name}  rows={len(body)}")
        print("header:", header)
        for r in body[:4]:
            print("  ", r)
        # 若含 method 与数值列，按 method 汇总中位数与 n
        if "method" in [str(h) for h in header]:
            mi = [str(h) for h in header].index("method")
            numeric_cols = [
                (i, str(h))
                for i, h in enumerate(header)
                if h is not None and i != mi
            ]
            groups: dict[str, list[tuple]] = {}
            for r in body:
                groups.setdefault(str(r[mi]), []).append(r)
            for method, rs in groups.items():
                parts = []
                for i, h in numeric_cols:
                    vals = [float(x) for x in (row[i] for row in rs) if isinstance(x, (int, float))]
                    if vals:
                        parts.append(f"{h}: med={np.median(vals):.2f} n={len(vals)}")
                if parts:
                    print(f"  [{method}] " + " | ".join(parts))


if __name__ == "__main__":
    main()
