"""临时诊断：按 panel/series 汇总 figure_plot_data.xlsx，核对 n、中位数与自举 95% CI。"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

PATH = Path("P:/VSCode-Project/EgoAnchor/EgoAnchor_Python/data/experiments/experiment_1_2/analysis/plots/figure_plot_data.xlsx")


def boot_ci(values, iterations=20000, seed=20260826):
    a = np.asarray(values, dtype=float)
    a = a[np.isfinite(a)]
    if a.size == 0:
        return float("nan"), float("nan"), float("nan")
    rng = np.random.default_rng(seed)
    draws = rng.integers(0, a.size, size=(iterations, a.size))
    med = np.median(a[draws], axis=1)
    return float(np.median(a)), float(np.quantile(med, 0.025)), float(np.quantile(med, 0.975))


def main() -> None:
    wb = load_workbook(PATH, read_only=True, data_only=True)
    for sheet in ("Figure2", "Figure3"):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h) for h in rows[0]]
        idx = {name: i for i, name in enumerate(header)}
        buckets = defaultdict(list)
        for r in rows[1:]:
            if not r or r[idx["panel"]] is None:
                continue
            y = r[idx["y_value"]]
            if not isinstance(y, (int, float)):
                continue
            key = (r[idx["panel"]], r[idx["series"]], r[idx["y_metric"]])
            buckets[key].append(float(y))
        print("\n" + "#" * 78)
        print(f"# {sheet}")
        print("#" * 78)
        for (panel, series, metric), vals in buckets.items():
            med, lo, hi = boot_ci(vals)
            print(f"{panel:32s} | {str(series):46s} | {metric:38s} "
                  f"n={len(vals):3d} med={med:9.2f} CI=[{lo:.2f}, {hi:.2f}]")


if __name__ == "__main__":
    main()
