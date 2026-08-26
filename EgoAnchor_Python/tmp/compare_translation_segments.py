"""临时诊断：核对已发表图数据与当前重算在平移通道上的片段身份差异。"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

from egoanchor.eval.experiments.experiment_1_2.analysis import (
    analyze_workbooks,
    load_settings,
)

ROOT = Path("P:/VSCode-Project/EgoAnchor/EgoAnchor_Python")
PLOT = ROOT / "data/experiments/experiment_1_2/analysis/plots/figure_plot_data.xlsx"
WORKBOOKS = tuple(sorted((ROOT / "data/experiments/experiment_1_2/workbooks").glob("task_*_complete.xlsx")))


def published_translation_segments() -> dict[tuple[str, str, str], float]:
    wb = load_workbook(PLOT, read_only=True, data_only=True)
    ws = wb["Figure2"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h) for h in rows[0]]
    i = {n: k for k, n in enumerate(header)}
    out: dict[tuple[str, str, str], float] = {}
    for r in rows[1:]:
        if not r or r[i["panel"]] != "(c) Dynamic translation":
            continue
        if r[i["y_metric"]] != "aligned_rmse_mm" or r[i["variant_id"]] != "EgoAnchor":
            continue
        key = (str(r[i["session_id"]]), str(r[i["trial_id"]]), str(r[i["segment_id"]]))
        out[key] = float(r[i["y_value"]])
    return out


def main() -> None:
    published = published_translation_segments()
    print(f"published EgoAnchor translation segments: n={len(published)}")
    by_session = defaultdict(list)
    for (s, t, e) in published:
        by_session[s].append((t, e))
    for s in sorted(by_session):
        print(f"  {s}: {len(by_session[s])}")

    results = analyze_workbooks(WORKBOOKS, load_settings(ROOT / "src/egoanchor/eval/config/paper.toml"))
    fresh = {
        (str(r["session_id"]), str(r["trial_id"]), str(r["segment_id"])): float(r["aligned_rmse_mm"])
        for r in results.translation_segments["EgoAnchor"]
    }
    print(f"\nfresh EgoAnchor translation segments: n={len(fresh)}")
    fby = defaultdict(list)
    for (s, t, e) in fresh:
        fby[s].append((t, e))
    for s in sorted(fby):
        print(f"  {s}: {len(fby[s])}")

    print("\nonly in fresh (not published):")
    for k in sorted(set(fresh) - set(published)):
        print(f"  {k}  rmse={fresh[k]:.2f}")
    print("\nonly in published (not fresh):")
    for k in sorted(set(published) - set(fresh)):
        print(f"  {k}  rmse={published[k]:.2f}")

    common = sorted(set(fresh) & set(published))
    print(f"\ncommon n={len(common)}; value agreement:")
    diffs = [abs(fresh[k] - published[k]) for k in common]
    if diffs:
        print(f"  max |diff| = {max(diffs):.6f}")
    pub_vals = list(published.values())
    print(f"\npublished median = {np.median(pub_vals):.2f} (paper table says 9.12)")
    print(f"fresh median     = {np.median(list(fresh.values())):.2f}")


if __name__ == "__main__":
    main()
