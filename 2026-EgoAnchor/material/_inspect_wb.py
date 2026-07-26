"""临时检查脚本：打印实验三各工作簿的结构（表名、尺寸、表头区域）。"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

MAT = Path(__file__).resolve().parent


def dump(path: Path, max_rows: int = 14, max_cols: int = 14) -> None:
    print("=" * 100)
    print(f"FILE: {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    for ws in wb.worksheets:
        print(f"  -- sheet {ws.title!r}: dims={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
    wb.close()


def dump_sheet(path: Path, sheet: str, rows: int, cols: int, start_row: int = 1) -> None:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    print("=" * 100)
    print(f"FILE {path.name} SHEET {sheet} dims={ws.dimensions}")
    for r in range(start_row, min(ws.max_row, start_row + rows - 1) + 1):
        vals = []
        for c in range(1, min(ws.max_column, cols) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                v = ""
            s = str(v).replace("\n", "\\n")
            if len(s) > 34:
                s = s[:31] + "..."
            vals.append(s)
        print(f"  r{r:>4}: " + " | ".join(vals))
    wb.close()


if __name__ == "__main__":
    if len(sys.argv) == 1:
        for name in [
            "EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx",
            "EgoAnchor_Experiment3_Simulated_24P_v5_1_Claude-Fable-5.xlsx",
            "EgoAnchor_Experiment3_Simulated_Claude-Opus-5-1M_v5_1_24P.xlsx",
            "GPT-5.6_Thinking_Experiment3_Synthetic_VSCodeSafe.xlsx",
        ]:
            dump(MAT / name)
    else:
        f, sheet, rows, cols = sys.argv[1], sys.argv[2], int(sys.argv[3]), int(sys.argv[4])
        start = int(sys.argv[5]) if len(sys.argv) > 5 else 1
        dump_sheet(MAT / f, sheet, rows, cols, start)
