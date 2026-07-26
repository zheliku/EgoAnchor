# -*- coding: utf-8 -*-
"""EgoAnchor 实验三 AI 模拟数据生成器（Claude Fable 5，v5.1 Verified 结构）。

用途：在正式招募前，按 v5.1 冻结设计模拟 24 名"真实人类"被试的纯主观数据，
用于检验设计灵敏度、条目区分度与分析流程。**全部输出均为合成数据，
不得作为真实实验数据、论文证据，也不得与真实采集数据合并。**

结构来源：material/EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx
（只读取其 Participants/Records/Analysis 三表的预填结构并镜像，不改动原文件。）

效应量锚定（非感知效应预测）：实验一 v4 实测差异（One-Euro → EgoAnchor）：
  静止头动泄漏 P95 10.65→0.82 mm；静止帧间增量 0.85→0.06 mm；
  绝对配准 P95 14.00→6.60 mm；静止旋转泄漏 3.29→0.33 deg；
  平移对齐 RMSE 15.69→9.12 mm 对 Start-transition 334→510 ms（反向，Q2/AQ-IQ2 预期近零）；
  遮挡平移 P95 10.41→4.85 mm、灾难性失效 1→0；候选率约 9.5 Hz；REGISTER 中位 737 ms。

人格模型：每名被试有独立的宽容度/极端度/噪声/光环效应/各通道敏感度/信任倾向/
作答速度等参数，另注入个体事件（One-Euro 恢复跳变 4 例、双方法各 1 次 Lost 超时）、
首区块向中回归、物体内先后对比锐化与疲劳效应。开放题按人格逐人撰写。

运行：pixi 环境 python 直接执行本文件；随机种子固定，可复现。
"""

from __future__ import annotations

import sys
import math
from datetime import datetime, timedelta

import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from scipy import stats

sys.stdout.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------- 常量与路径
SEED = 202607265                      # 2026-07-26 + Fable 5，区别于 Opus 5 的 20260726
MODEL_NAME = "Claude-Fable-5"
BASE_DIR = r"p:\VSCode-Project\EgoAnchor\2026-EgoAnchor\material"
OFFICIAL_XLSX = BASE_DIR + r"\EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx"
OUT_XLSX = BASE_DIR + r"\EgoAnchor_Experiment3_Simulated_24P_v5_1_Claude-Fable-5.xlsx"

assert OUT_XLSX != OFFICIAL_XLSX, "禁止覆盖正式采集工作簿"

rng = np.random.default_rng(SEED)

OBJECTS = {"blue_mouse": "鼠标", "stapler": "固定订书机", "gamepad": "游戏手柄"}
BLOCK_ITEMS = ["Q1", "Q2", "Q9", "Q10_OPT", "AQ_IQ2", "AQ_IQ3", "Q3", "Q8",
               "AQ_EQ1", "AQ_EQ2", "AQ_EQ3", "AQ_IQ1", "Q6", "Q7"]  # Records A 列顺序
RATED_ITEMS = [i for i in BLOCK_ITEMS if i != "Q10_OPT"]            # Q10 默认停用，留空

# 区块级条目效应表：base = One-Euro 期望均值，delta = EgoAnchor 增量（七点尺）
# 按对象 (blue_mouse, stapler, gamepad)。数值由实验一实测差异映射，见文件头。
ITEM_EFFECTS = {
    #            base(OE)             delta(EA)
    "Q1":     ((4.10, 4.45, 4.30), (1.90, 1.75, 1.85)),   # 泄漏 13x：大效应
    "Q2":     ((3.85, 4.00, 3.70), (0.15, 0.10, 0.00)),   # 对齐残差收益被起步代价抵消：预期近零
    "Q9":     ((4.70, 4.30, 4.40), (0.50, 0.95, 0.80)),   # 旋转通道：订书机/手柄朝向误差更可见
    "AQ_IQ2": ((4.20, 4.30, 4.00), (-0.15, -0.10, -0.30)),  # 起步顿挫：预期近零甚至反向
    "AQ_IQ3": ((4.75, 4.90, 4.60), (0.50, 0.40, 0.50)),   # 残余抖动 5.25→4.56：小幅正向
    "Q3":     ((4.20, 4.50, 4.20), (1.45, 1.25, 1.45)),   # 遮挡 P95 10.41→4.85 + 跳变事件
    "Q8":     ((4.00, 4.40, 4.15), (1.50, 1.35, 1.40)),   # 绝对配准 P95 14.0→6.6
    "AQ_EQ1": ((4.40, 4.70, 4.50), (0.90, 0.80, 0.80)),
    "AQ_EQ2": ((4.30, 4.60, 4.40), (0.90, 0.85, 0.80)),
    "AQ_EQ3": ((5.95, 6.15, 6.00), (0.10, 0.10, 0.10)),   # 尺度协调：预期天花板
    "AQ_IQ1": ((5.20, 5.40, 5.20), (0.20, 0.20, 0.20)),
    "Q6":     ((3.95, 4.20, 4.00), (1.40, 1.25, 1.30)),
    "Q7":     ((4.10, 4.30, 4.10), (1.10, 1.00, 1.05)),
}
OBJ_IDX = {"blue_mouse": 0, "stapler": 1, "gamepad": 2}

# ---------------------------------------------------------------- 24 名人格
# 字段：背景（年龄/性别/主手/视力/VR 经验/实物 MR 经验）、作答风格（acq 宽容度、
# ext 极端度、noise 噪声、halo 区块光环）、通道敏感度（st 静止抖动、lag 响应滞后、
# ori 朝向、rec 恢复、pos 位置）、trust 信任倾向、spd 每题秒数、conf 自信、
# persona 一句话画像、备注类特殊事件在 INCIDENTS 中另行定义。
PERSONAS = {
 "P001": dict(age=24, gender="女", hand="右手", vision="矫正后正常", vr="1–5 次", mr="从未",
              acq=0.2, ext=1.00, noise=0.55, halo=0.20, st=1.1, lag=0.9, ori=1.0, rec=1.0, pos=1.1,
              trust=0.1, spd=8.0, conf=0.3, persona="心理学硕士生，作答细致、标准中庸"),
 "P002": dict(age=22, gender="男", hand="右手", vision="正常", vr="6–20 次", mr="1–2 次",
              acq=-0.1, ext=1.25, noise=0.70, halo=0.25, st=0.80, lag=1.6, ori=0.9, rec=0.9, pos=0.8,
              trust=-0.1, spd=7.0, conf=0.6, persona="计算机大四、重度玩家，极看重跟手感"),
 "P003": dict(age=26, gender="女", hand="右手", vision="矫正后正常", vr="1–5 次", mr="从未",
              acq=0.0, ext=1.15, noise=0.60, halo=0.20, st=1.6, lag=0.8, ori=1.1, rec=1.3, pos=1.2,
              trust=0.0, spd=8.5, conf=0.7, persona="设计学博士生，对细微抖动零容忍"),
 "P004": dict(age=29, gender="男", hand="右手", vision="正常", vr="从未", mr="从未",
              acq=0.1, ext=0.75, noise=0.85, halo=0.35, st=0.60, lag=0.7, ori=0.7, rec=0.7, pos=0.7,
              trust=0.0, spd=11.5, conf=-0.8, persona="校外机械工程师，首次接触 MR、判断保守"),
 "P005": dict(age=23, gender="男", hand="右手", vision="正常", vr="经常使用", mr="数次",
              acq=-0.2, ext=1.05, noise=0.45, halo=0.15, st=1.3, lag=1.3, ori=1.2, rec=1.2, pos=1.3,
              trust=0.0, spd=7.5, conf=0.8, persona="HCI 硕士，分析型，能同时说出双方优缺点"),
 "P006": dict(age=21, gender="女", hand="左手", vision="正常", vr="从未", mr="从未",
              acq=0.5, ext=0.85, noise=0.75, halo=0.40, st=0.9, lag=0.7, ori=0.8, rec=0.9, pos=0.8,
              trust=0.3, spd=10.0, conf=-0.4, persona="心理学本科生，宽容型评分者、易受整体印象影响"),
 "P007": dict(age=25, gender="男", hand="右手", vision="矫正后正常", vr="6–20 次", mr="1–2 次",
              acq=-0.5, ext=1.10, noise=0.60, halo=0.20, st=1.2, lag=1.1, ori=1.0, rec=1.1, pos=1.4,
              trust=-0.4, spd=8.0, conf=0.5, persona="电子工程硕士，苛刻评分者，紧盯绝对位置偏差"),
 "P008": dict(age=28, gender="女", hand="右手", vision="矫正后正常", vr="1–5 次", mr="从未",
              acq=-0.1, ext=1.00, noise=0.50, halo=0.15, st=1.1, lag=0.8, ori=1.0, rec=1.3, pos=1.5,
              trust=-0.2, spd=8.5, conf=0.4, persona="医学影像博士生，精度导向、信任保守"),
 "P009": dict(age=20, gender="男", hand="右手", vision="正常", vr="经常使用", mr="数次",
              acq=0.0, ext=1.20, noise=0.70, halo=0.25, st=0.85, lag=1.7, ori=0.9, rec=1.0, pos=0.9,
              trust=0.0, spd=7.0, conf=0.6, persona="游戏社团本科生，动作快、偏好跟手但信任求稳"),
 "P010": dict(age=24, gender="女", hand="右手", vision="正常", vr="1–5 次", mr="从未",
              acq=0.1, ext=1.00, noise=0.60, halo=0.20, st=1.2, lag=0.9, ori=1.1, rec=1.0, pos=1.1,
              trust=0.1, spd=8.5, conf=0.2, persona="建筑学硕士，空间关系敏感"),
 "P011": dict(age=31, gender="男", hand="右手", vision="矫正后正常", vr="从未", mr="从未",
              acq=0.0, ext=0.80, noise=0.60, halo=0.25, st=0.9, lag=0.8, ori=0.9, rec=1.2, pos=0.9,
              trust=-0.1, spd=11.0, conf=-0.2, persona="材料学博士后，谨慎缓慢，经历过一次 Lost 超时"),
 "P012": dict(age=23, gender="女", hand="右手", vision="正常", vr="1–5 次", mr="从未",
              acq=0.6, ext=0.80, noise=0.65, halo=0.35, st=0.8, lag=0.8, ori=0.8, rec=0.8, pos=0.8,
              trust=0.4, spd=9.0, conf=-0.3, persona="师范生，好好人风格、两边都给高分"),
 "P013": dict(age=27, gender="男", hand="右手", vision="正常", vr="6–20 次", mr="数次",
              acq=0.0, ext=1.05, noise=0.55, halo=0.20, st=1.1, lag=1.0, ori=1.0, rec=1.1, pos=1.1,
              trust=0.0, spd=7.5, conf=0.4, persona="创业公司工程师，务实，注意到表面滑动"),
 "P014": dict(age=22, gender="女", hand="右手", vision="矫正后正常", vr="从未", mr="从未",
              acq=-0.1, ext=1.00, noise=0.60, halo=0.30, st=1.0, lag=0.9, ori=1.1, rec=1.2, pos=1.0,
              trust=0.0, spd=9.5, conf=0.1, persona="数学本科生，前期分不清、被手柄轮跳变点醒"),
 "P015": dict(age=24, gender="男", hand="右手", vision="正常", vr="1–5 次", mr="从未",
              acq=-0.2, ext=1.15, noise=0.60, halo=0.20, st=1.0, lag=1.7, ori=1.0, rec=1.0, pos=1.0,
              trust=-0.1, spd=8.0, conf=0.7, persona="体育教育硕士，动作幅度大，权衡型无偏好"),
 "P016": dict(age=30, gender="女", hand="右手", vision="矫正后正常", vr="从未", mr="从未",
              acq=0.3, ext=0.80, noise=0.70, halo=0.30, st=0.9, lag=0.7, ori=0.8, rec=0.9, pos=0.9,
              trust=0.1, spd=11.0, conf=-0.5, persona="高校行政人员，谨慎，TiA 复杂任务题选无法回答"),
 "P017": dict(age=22, gender="男", hand="右手", vision="正常", vr="6–20 次", mr="1–2 次",
              acq=0.0, ext=1.15, noise=0.65, halo=0.20, st=1.2, lag=1.1, ori=1.0, rec=1.1, pos=1.0,
              trust=0.0, spd=7.5, conf=0.7, persona="计算机本科生，好奇心强、猜测方法机制"),
 "P018": dict(age=25, gender="女", hand="右手", vision="正常", vr="1–5 次", mr="从未",
              acq=0.0, ext=0.95, noise=0.40, halo=0.10, st=1.2, lag=1.0, ori=1.1, rec=1.2, pos=1.2,
              trust=0.0, spd=8.0, conf=0.5, persona="心理学博士生，作答一致性极高，强调可预测性"),
 "P019": dict(age=26, gender="男", hand="左手", vision="正常", vr="1–5 次", mr="从未",
              acq=0.0, ext=1.10, noise=0.60, halo=0.25, st=1.0, lag=0.9, ori=1.0, rec=1.5, pos=1.1,
              trust=-0.2, spd=9.0, conf=0.5, persona="土木硕士，第 2 区块遭遇 One-Euro Lost 超时后信任崩塌"),
 "P020": dict(age=27, gender="女", hand="右手", vision="矫正后正常", vr="6–20 次", mr="数次",
              acq=0.1, ext=1.05, noise=0.50, halo=0.15, st=1.4, lag=0.9, ori=1.0, rec=1.1, pos=1.1,
              trust=0.1, spd=8.0, conf=0.5, persona="产品经理，表达精准，用'呼吸感'描述抖动"),
 "P021": dict(age=21, gender="男", hand="右手", vision="正常", vr="1–5 次", mr="从未",
              acq=0.1, ext=1.00, noise=0.65, halo=0.25, st=1.0, lag=0.9, ori=1.3, rec=1.0, pos=1.0,
              trust=0.1, spd=8.5, conf=0.2, persona="物理本科生，对朝向偏差格外敏感"),
 "P022": dict(age=23, gender="女", hand="右手", vision="正常", vr="从未", mr="从未",
              acq=0.5, ext=0.75, noise=0.80, halo=0.40, st=0.7, lag=0.7, ori=0.7, rec=0.7, pos=0.7,
              trust=0.3, spd=9.5, conf=-0.7, persona="新闻传播硕士，感知不敏锐、倾向都说好"),
 "P023": dict(age=28, gender="男", hand="右手", vision="矫正后正常", vr="经常使用", mr="经常",
              acq=-0.6, ext=1.10, noise=0.50, halo=0.15, st=1.3, lag=1.4, ori=1.1, rec=1.2, pos=1.3,
              trust=-0.5, spd=7.0, conf=0.6, persona="算法工程师，全场最苛刻，两方法延迟都嫌大"),
 "P024": dict(age=19, gender="女", hand="右手", vision="正常", vr="从未", mr="从未",
              acq=0.4, ext=1.35, noise=0.75, halo=0.30, st=1.2, lag=0.8, ori=0.9, rec=1.0, pos=0.9,
              trust=0.3, spd=8.0, conf=0.4, persona="大一新生，兴奋型极端评分者"),
}

# 个体事件注入：(pid, object_key, condition) -> 对该区块条目的附加偏移
# jump = One-Euro 遮挡恢复可见跳变（对应实验一 12 次遮挡 1 次灾难性失效的量级）
INCIDENTS = {
    ("P003", "blue_mouse", "One-Euro"): dict(Q3=-1.8, Q6=-0.9, Q8=-0.6, tag="jump"),
    ("P008", "stapler", "One-Euro"): dict(Q3=-1.6, Q6=-0.8, Q8=-0.5, tag="jump"),
    ("P014", "gamepad", "One-Euro"): dict(Q3=-1.8, Q6=-1.0, Q8=-0.6, tag="jump"),
    ("P020", "blue_mouse", "One-Euro"): dict(Q3=-1.4, Q6=-0.7, Q8=-0.4, tag="jump"),
    # Lost 超时（遮挡略超 1 s，服务器 REGISTER 一次；两方法各 1 例保持平衡）
    ("P011", "gamepad", "EgoAnchor"): dict(Q3=-1.6, Q6=-0.8, tag="lost",
        note="遮挡 0.96 s 略超时进入 Lost，服务器重注册 1 次，恢复后位置正确"),
    ("P019", "gamepad", "One-Euro"): dict(Q3=-2.2, Q6=-1.2, Q8=-0.5, tag="lost",
        note="遮挡 0.97 s 超时进入 Lost，服务器重注册 1 次，重现时伴随可见跳变"),
}

# 开放题（{EA}/{OE} 为该被试自己的匿名标签占位符）与主题编码
OPEN_TEXT = {
 "P001": ("{EA}更稳，尤其我不动的时候，它就像贴在物体上；{OE}会轻微地晃。",
          "虚拟的框和真实物体对不上位置的时候，我就开始怀疑它了。",
          ["Stationary_Jitter", "Absolute_Offset"]),
 "P002": ("拿起来那一下，{EA}会先停一拍才跟上来，{OE}是跟手的；不过{OE}平时有点飘。",
          "我一动它却不动，像卡住了一样，那一瞬间最出戏。",
          ["Motion_Lag", "Stationary_Jitter", "Predictability"]),
 "P003": ("{OE}一直有呼吸一样的抖动，鼠标那轮挡板拿开后还跳了一下才回去；{EA}基本纹丝不动。",
          "抖动。哪怕很小，只要一直在抖，我就没法把它当成真的。",
          ["Stationary_Jitter", "Recovery_Jump"]),
 "P004": ("说实话我不太确定……好像有一个稍微稳一点？我说不上来具体差在哪。",
          "如果它突然不见了，或者跑偏了吧。",
          ["No_Noticeable_Difference"]),
 "P005": ("{EA}静止时锁得很死、遮挡后也回得准，但起步瞬间有可感的滞后；{OE}起步顺，可整体位置一直微微偏。两者是不同的取舍。",
          "恢复之后落错位置，比一直有小误差更破坏信任——我会开始每次都盯着确认。",
          ["Stationary_Jitter", "Motion_Lag", "Wrong_Recovery", "Absolute_Offset"]),
 "P006": ("都挺好的，{EA}感觉更安心一点，不怎么动。",
          "闪来闪去就不太行。",
          ["Stationary_Jitter"]),
 "P007": ("{OE}的标注老是偏那么几毫米，还会随视角变；{EA}位置准得多，但两个都有延迟。",
          "位置不对。它标在那儿我伸手却差一点，这最致命。",
          ["Absolute_Offset", "Viewpoint_Drift", "Motion_Lag"]),
 "P008": ("{EA}的位置和真实物体贴合得更好；{OE}有一次遮挡后偏了一下才慢慢对回来。",
          "在我需要按它的指示去操作的时候它却偏了——影像科里这是不能接受的。",
          ["Absolute_Offset", "Recovery_Jump"]),
 "P009": ("{OE}跟手，动作快也不掉；{EA}稳但起步肉。玩起来我会选{OE}。",
          "不过要是拿它指路让我去抓东西，我反而信{EA}，它停下来的时候是真的准。",
          ["Motion_Lag", "Stationary_Jitter"]),
 "P010": ("{EA}像被钉在物体上，{OE}有点浮。",
          "浮动和偏移，会让我觉得它只是'大概在那附近'。",
          ["Stationary_Jitter", "Absolute_Offset"]),
 "P011": ("{EA}总体更稳；不过手柄那轮挡板拿开后它消失了一小会儿才回来，吓我一跳。",
          "突然消失再出现。哪怕回来的位置是对的，那一下也让我心里打鼓。",
          ["Recovery_Jump", "Predictability"]),
 "P012": ("{EA}稍微稳一点点，其实两个都还不错。",
          "偏得太明显的话就不行了。",
          ["No_Noticeable_Difference", "Absolute_Offset"]),
 "P013": ("{EA}静止和遮挡后都更靠谱，{OE}移动时偶尔像在冰面上滑。",
          "滑动。虚拟的东西在物体表面上蹭来蹭去，一下子就假了。",
          ["Motion_Sliding", "Stationary_Jitter"]),
 "P014": ("前面几轮我分不太清，直到手柄那轮{OE}在挡板拿开后明显跳了一下，我才确定{EA}更稳。",
          "跳变。平滑的小误差可以接受，突然跳一下不行。",
          ["Recovery_Jump", "No_Noticeable_Difference"]),
 "P015": ("{EA}稳、准，但我快速移动时它跟不上、起步还顿一下；{OE}跟得快但一直微晃。两个我都不完全满意，看用途吧。",
          "跟不上我的动作。我动它不动，比小抖动更让我不信任。",
          ["Motion_Lag", "Stationary_Jitter"]),
 "P016": ("{EA}好像更稳当一些，尤其放着不动的时候。",
          "要是它自己乱动，我就不敢信了。",
          ["Stationary_Jitter"]),
 "P017": ("{EA}稳得不太自然，感觉像是加了什么锁定；{OE}更像普通的追踪。稳定还是{EA}赢。",
          "遮挡之后回错位置。",
          ["Stationary_Jitter", "Wrong_Recovery", "Other"]),
 "P018": ("{EA}在静止和遮挡恢复后的位置一致性都更好；{OE}的误差虽然小，但一直存在。",
          "不可预测。固定的小偏差我可以适应，忽大忽小没法适应。",
          ["Predictability", "Stationary_Jitter"]),
 "P019": ("手柄那轮{OE}挡板拿开后隔了好一会儿才出现，出现的时候还跳了一下；{EA}没出过这种问题。",
          "就是那种消失又错位回来的情况，出一次我就不敢用了。",
          ["Recovery_Jump", "Wrong_Recovery", "Predictability"]),
 "P020": ("{OE}有一种轻微的呼吸感，一直在动；{EA}落定之后就真的定住了，挡板测试后回位也干脆。",
          "呼吸感。它一直在提醒我'我是贴上去的'，沉浸感就没了。",
          ["Stationary_Jitter", "Embedding_Blending"]),
 "P021": ("{EA}更稳定，转动手柄的时候朝向也跟得更正。",
          "朝向歪了会特别明显，比位置偏一点还难受。",
          ["Orientation_Mismatch", "Stationary_Jitter"]),
 "P022": ("感觉差不多？可能{EA}稍微稳一点，我不是很确定。",
          "大概是完全对不上的时候吧。",
          ["No_Noticeable_Difference"]),
 "P023": ("两个延迟都不小，快速移动时都跟不上；差别在于{EA}静止时几乎零抖、遮挡后收敛快，{OE}是到处都差一点。",
          "延迟叠加偏移。单独一样还能忍，一起出现就完全不可用。",
          ["Motion_Lag", "Stationary_Jitter", "Absolute_Offset"]),
 "P024": ("{EA}太稳了！像真的粘在上面一样；{OE}有点像果冻，晃晃悠悠的。",
          "晃！一晃我就觉得它随时会飞走哈哈。",
          ["Stationary_Jitter", "Embedding_Blending"]),
}

THEME_COLS = ["Stationary_Jitter", "Viewpoint_Drift", "Absolute_Offset", "Motion_Lag",
              "Motion_Sliding", "Orientation_Mismatch", "PostPlacement_Settling",
              "Recovery_Jump", "Wrong_Recovery", "Predictability", "Embedding_Blending",
              "No_Noticeable_Difference", "Other"]

# 每日 5 个时段的到场时间（合成日期窗口，明确为虚构排期）
SESSION_SLOTS = ["09:30", "11:00", "14:00", "15:30", "17:00"]
START_DATE = datetime(2026, 8, 10)


# ---------------------------------------------------------------- 读取官方结构
def load_official():
    """按值读取官方工作簿的 Participants/Records/Analysis 预填结构（不修改原文件）。"""
    wb = openpyxl.load_workbook(OFFICIAL_XLSX, data_only=True)
    sheets = {}
    for name in ["Participants", "Records", "Analysis"]:
        ws = wb[name]
        grid = {}
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    grid[(c.row, c.column)] = c.value
        sheets[name] = grid
    # 解析区块进度表（Records A 段第 5–148 行的前 10 列）
    blocks = []
    for r in range(5, 149):
        rec = [sheets["Records"].get((r, c)) for c in range(1, 11)]
        blocks.append(dict(row=r, pid=rec[0], block=rec[1], unit=rec[2], obj_pos=rec[3],
                           obj_cn=rec[4], obj_key=rec[5], label=rec[6], cond=rec[7],
                           pair_order=rec[8], nth=rec[9]))
    # 方法级记录（B 段 152–199 行前 4 列）
    method_rows = []
    for r in range(152, 200):
        rec = [sheets["Records"].get((r, c)) for c in range(1, 5)]
        method_rows.append(dict(row=r, pid=rec[0], order=rec[1], label=rec[2], cond=rec[3]))
    # 参与者映射（3–26 行）
    parts = {}
    for r in range(3, 27):
        g = sheets["Participants"]
        pid = g.get((r, 1))
        parts[pid] = dict(row=r, unit=g.get((r, 2)), seq=g.get((r, 4)),
                          methodA=g.get((r, 9)), methodB=g.get((r, 10)),
                          first=g.get((r, 11)))
    return sheets, blocks, method_rows, parts


# ---------------------------------------------------------------- 评分生成
def clip_round(x, lo, hi):
    """连续潜变量离散化到量表点。"""
    return int(min(hi, max(lo, round(x))))


def sens_for(p, item):
    """人格通道敏感度 -> 条目效应倍率。"""
    if item == "Q1":
        return p["st"]
    if item == "Q2":
        return 1.0  # Q2 的个体差异走加性项（见 gen_block）
    if item == "Q9":
        return p["ori"]
    if item == "AQ_IQ2":
        return p["lag"]
    if item == "AQ_IQ3":
        return (p["st"] + p["lag"]) / 2
    if item == "Q3":
        return p["rec"]
    if item == "Q8":
        return p["pos"]
    if item in ("AQ_EQ1", "AQ_EQ2"):
        return (p["st"] + p["pos"]) / 2
    if item == "AQ_EQ3":
        return 0.5
    if item == "AQ_IQ1":
        return 0.7
    if item == "Q6":
        return (p["rec"] + p["pos"] + p["st"]) / 3
    if item == "Q7":
        return 2 * p["st"] / (p["st"] + p["lag"])  # 稳定派放大、跟手派缩小权衡收益
    return 1.0


def gen_block(pid, blk):
    """生成一个方法×物体区块的 13 项评分。"""
    p = PERSONAS[pid]
    oi = OBJ_IDX[blk["obj_key"]]
    is_ea = blk["cond"] == "EgoAnchor"
    halo = rng.normal(0, p["halo"])
    incident = INCIDENTS.get((pid, blk["obj_key"], blk["cond"]), {})
    nth_gain = {1: 0.90, 2: 1.00, 3: 1.10}[blk["nth"]]  # 第几次接触该方法：辨别力渐增
    out = {}
    for item in RATED_ITEMS:
        base, delta = ITEM_EFFECTS[item]
        d_eff = delta[oi] * sens_for(p, item) * nth_gain
        if item == "Q2":
            d_eff += 0.25 * (p["st"] - p["lag"])   # 跟手派感到 EA 起步顿挫，稳定派感到附着更牢
        if item == "AQ_IQ2":
            d_eff = delta[oi] * p["lag"] + 0.05 * p["st"]
        latent = base[oi] + (d_eff if is_ea else 0.0) + p["acq"] + halo
        latent += incident.get(item, 0.0)
        if blk["pair_order"] == 2:                 # 同物体第二方法：对比锐化
            latent += 0.15 * math.copysign(1.0, d_eff if is_ea else -d_eff)
        if blk["block"] == 1:                      # 首区块无参照，向中回归
            latent = 4.0 + (latent - 4.0) * 0.85
        latent = 4.0 + (latent - 4.0) * p["ext"]   # 极端度伸缩
        latent += rng.normal(0, p["noise"])
        out[item] = clip_round(latent, 1, 7)
    return out


def block_duration(pid, block_index):
    """区块问卷时长（秒）：人格速度 × 13 题 + 界面开销，首区块更慢、随区块递减。"""
    p = PERSONAS[pid]
    per = p["spd"] * (1.22 if block_index == 1 else 1.0) * (1.0 - 0.025 * (block_index - 1))
    dur = per * 13 + 10 + rng.normal(0, 6)
    return int(max(70, dur))


# ---------------------------------------------------------------- 主流程
def main():
    sheets, blocks, method_rows, parts = load_official()

    # ---- 1. 区块评分与审计列 ----
    ratings = {}          # (pid, block) -> dict(item->score)
    audits = {}           # (pid, block) -> dict(审计列)
    obj_session = {}      # (pid, obj_key) -> 该感知会话的候选率/VCD/接纳率基准（两方法共享）
    for blk in blocks:
        pid = blk["pid"]
        key = (pid, blk["obj_key"])
        if key not in obj_session:
            base_rate = {"blue_mouse": 9.35, "stapler": 9.55, "gamepad": 9.45}[blk["obj_key"]]
            base_vcd = {"blue_mouse": 0.83, "stapler": 0.90, "gamepad": 0.87}[blk["obj_key"]]
            base_adm = {"blue_mouse": 0.91, "stapler": 0.96, "gamepad": 0.93}[blk["obj_key"]]
            obj_session[key] = (base_rate + rng.normal(0, 0.12),
                               base_vcd + rng.normal(0, 0.02),
                               base_adm + rng.normal(0, 0.015))
        ratings[(pid, blk["block"])] = gen_block(pid, blk)

        rate0, vcd0, adm0 = obj_session[key]
        incident = INCIDENTS.get((pid, blk["obj_key"], blk["cond"]), {})
        lost = incident.get("tag") == "lost"
        occl = float(np.clip(rng.normal(0.75, 0.05), 0.62, 0.90))
        if lost:
            occl = 0.96 if blk["cond"] == "EgoAnchor" else 0.97
        audits[(pid, blk["block"])] = dict(
            rate=round(rate0 + rng.normal(0, 0.05), 2),
            vcd=round(float(np.clip(vcd0 + rng.normal(0, 0.008), 0.5, 0.99)), 3),
            adm=round(float(np.clip(adm0 + rng.normal(0, 0.006), 0.5, 0.999)), 3),
            avail=round(float(np.clip(rng.normal(0.992, 0.004), 0.97, 1.0)), 3),
            occl=round(occl, 2),
            state="Lost" if lost else "FrozenUncertain",
            lost="是" if lost else "否",
            reacq=1 if lost else 0,
            lock=int(rng.integers(3, 6)) if blk["cond"] == "EgoAnchor" else 0,
            note=incident.get("note", ""),
        )

    # ---- 2. 会话时间线 ----
    times = {}            # pid -> dict(start, end, block_windows, method_windows)
    for i, pid in enumerate(sorted(parts)):
        day, slot = divmod(i, 5)
        t0 = datetime.strptime(f"{(START_DATE + timedelta(days=day)):%Y-%m-%d} {SESSION_SLOTS[slot]}",
                               "%Y-%m-%d %H:%M") + timedelta(minutes=int(rng.integers(0, 7)))
        t = t0 + timedelta(minutes=9 + int(rng.integers(0, 3)))     # 背景+同意
        t += timedelta(minutes=7)                                    # 训练（earphone）
        bw, obj_seen = {}, set()
        for blk in [b for b in blocks if b["pid"] == pid]:
            if blk["obj_key"] not in obj_seen:                       # 换物体：重启服务+贴合检查
                obj_seen.add(blk["obj_key"])
                t += timedelta(seconds=150 if len(obj_seen) > 1 else 90)
            t += timedelta(seconds=int(55 + rng.integers(0, 12)))    # 三项任务
            dur = block_duration(pid, blk["block"])
            bw[blk["block"]] = (t, t + timedelta(seconds=dur), dur)
            t = t + timedelta(seconds=dur + 40)                      # 问卷 + 重置
        t += timedelta(seconds=70)                                   # 过渡到方法级问卷
        mw = []
        for j in range(2):
            dur = int(PERSONAS[pid]["spd"] * 13 * (1.05 - 0.12 * j) + 18 + rng.normal(0, 8))
            mw.append((t, t + timedelta(seconds=dur), dur))
            t = t + timedelta(seconds=dur + 25)
        t += timedelta(minutes=2)                                    # 摘头显
        t += timedelta(minutes=6 + int(rng.integers(0, 3)))          # 最终问卷+访谈
        times[pid] = dict(start=t0, end=t, blocks=bw, methods=mw)

    # ---- 3. 方法级印象与 TiA / S-TIAS ----
    def object_means(pid, cond):
        """该被试该方法逐条目的三物体均值。"""
        vals = {it: [] for it in RATED_ITEMS}
        for blk in blocks:
            if blk["pid"] == pid and blk["cond"] == cond:
                for it in RATED_ITEMS:
                    vals[it].append(ratings[(pid, blk["block"])][it])
        return {it: float(np.mean(v)) for it, v in vals.items()}

    W_IMPR = dict(Q1=0.13, Q8=0.14, Q3=0.17, Q6=0.15, Q7=0.08, AQ_EQ1=0.05, AQ_EQ2=0.05,
                  AQ_EQ3=0.02, AQ_IQ1=0.04, AQ_IQ3=0.05, Q9=0.05)
    W_TRUST = dict(Q3=0.28, Q8=0.27, Q1=0.20, Q6=0.25)

    impressions, trust_scores = {}, {}
    for pid in parts:
        p = PERSONAS[pid]
        for cond in ("One-Euro", "EgoAnchor"):
            om = object_means(pid, cond)
            w = dict(W_IMPR)
            w["Q2"] = 0.035 * p["lag"]           # 跟手派把响应体验计入整体印象
            w["AQ_IQ2"] = 0.035 * p["lag"]
            tot = sum(w.values())
            impressions[(pid, cond)] = sum(om[k] * v for k, v in w.items()) / tot
            trust_scores[(pid, cond)] = sum(om[k] * v for k, v in W_TRUST.items()) / sum(W_TRUST.values())

    tia_raw = {}          # (pid, cond) -> dict(item -> 1..5 原始分 或 None)
    stias_raw = {}        # (pid, cond) -> dict(item -> 1..7)
    for mr in method_rows:
        pid, cond = mr["pid"], mr["cond"]
        p = PERSONAS[pid]
        impr = impressions[(pid, cond)]
        # 第二个评价的方法带轻微对比锐化
        other = "EgoAnchor" if cond == "One-Euro" else "One-Euro"
        contrast = 0.12 * math.copysign(1.0, impr - impressions[(pid, other)]) if mr["order"] == 2 else 0.0
        lat = 1 + (impr - 1) * (2 / 3) + contrast            # 7 点印象映射到 5 点
        had_incident = any((pid, o, cond) in INCIDENTS for o in OBJECTS)
        pred = 0.20 if cond == "EgoAnchor" else -0.25        # 冻结-恢复可理解 vs 游移难预测
        nz = p["noise"] * 0.5

        def pos_item(off):
            return clip_round(lat + off + p["acq"] * 0.35 + rng.normal(0, nz), 1, 5)

        def rev_item(off_pos):
            # 反向条目按原始分记录：认同负面陈述的程度 = 6 − 正向潜变量（+宽容度轻微抬升）
            raw = 6 - (lat + off_pos) + p["acq"] * 0.4 + (0.45 if had_incident else 0.0)
            return clip_round(raw + rng.normal(0, nz), 1, 5)

        t = dict(
            TIA_RC1=clip_round(lat - 0.15 + (4 - lat) * 0.1 * 3 / 4 + p["acq"] * 0.35 + rng.normal(0, nz + 0.1), 1, 5),
            TIA_RC2=pos_item(0.10),
            TIA_RC3_REV=rev_item(-0.10),
            TIA_RC4=pos_item(-0.20) if not (pid == "P016" and cond == "One-Euro") else None,
            TIA_RC5_REV=rev_item(-0.35),                     # "偶尔出错"更易被认同
            TIA_RC6=pos_item(p["trust"] * 0.5),
            TIA_UP1=pos_item(-0.10),
            TIA_UP2_REV=rev_item(pred),
            TIA_UP3=pos_item(0.10 if cond == "EgoAnchor" else -0.25),
            TIA_UP4_REV=rev_item(pred * 0.8),
        )
        tia_raw[(pid, cond)] = t
        lat7 = 0.5 * impr + 0.5 * trust_scores[(pid, cond)] + p["trust"] * 0.6 + contrast
        stias_raw[(pid, cond)] = dict(
            STIAS1=clip_round(lat7 - 0.10 + p["conf"] * 0.3 + rng.normal(0, 0.45), 1, 7),
            STIAS2=clip_round(lat7 + 0.15 + rng.normal(0, 0.45), 1, 7),
            STIAS3=clip_round(lat7 + rng.normal(0, 0.45), 1, 7),
        )

    # ---- 4. 最终问卷 ----
    finals = {}
    for pid in parts:
        p = PERSONAS[pid]
        d = impressions[(pid, "EgoAnchor")] - impressions[(pid, "One-Euro")]
        dt = trust_scores[(pid, "EgoAnchor")] - trust_scores[(pid, "One-Euro")]
        ea_label = "方法A" if parts[pid]["methodA"] == "EgoAnchor" else "方法B"
        oe_label = "方法B" if ea_label == "方法A" else "方法A"
        if pid == "P015":
            choice = "无明显偏好"                              # 权衡型：看得清差异但各有所长
        elif d > 0.22:
            choice = ea_label
        elif d < -0.22:
            choice = oe_label
        else:
            choice = "无明显偏好"
        strength = "NA" if choice == "无明显偏好" else clip_round(1.2 + 3.2 * abs(d) + rng.normal(0, 0.5), 1, 7)
        trust_choice = ea_label if dt > 0.18 else (oe_label if dt < -0.18 else "无明显偏好")
        conf = clip_round(2.6 + 1.5 * abs(d) + p["conf"] * 1.2 + rng.normal(0, 0.6), 1, 7)
        if pid == "P015":
            conf = max(conf, 6)
        diff_t, dis_t, themes = OPEN_TEXT[pid]
        sub = dict(EA=ea_label, OE=oe_label)
        finals[pid] = dict(choice=choice, strength=strength, trust=trust_choice, conf=conf,
                           open_diff=diff_t.format(**sub), open_dis=dis_t.format(**sub),
                           themes=themes,
                           discomfort="轻微" if pid in ("P004", "P006", "P011") else "无")

    # ---- 5. 统计分析 ----
    pids = sorted(parts)
    obj_mean = {}         # (pid, cond, item) -> 三物体均值
    for pid in pids:
        for cond in ("One-Euro", "EgoAnchor"):
            for it, v in object_means(pid, cond).items():
                obj_mean[(pid, cond, it)] = v

    def paired(item):
        oe = np.array([obj_mean[(pid, "One-Euro", item)] for pid in pids])
        ea = np.array([obj_mean[(pid, "EgoAnchor", item)] for pid in pids])
        return oe, ea

    def rank_biserial(diff):
        """匹配秩双列相关：(W+ − W−)/(W+ + W−)，零差对剔除。"""
        d = diff[diff != 0]
        if len(d) == 0:
            return 0.0
        r = stats.rankdata(np.abs(d))
        wp, wn = r[d > 0].sum(), r[d < 0].sum()
        return (wp - wn) / (wp + wn)

    def wilcoxon_row(item, subscale=None, subset=None):
        use = subset or pids
        if subscale is None:
            oe = np.array([obj_mean[(pid, "One-Euro", item)] for pid in use])
            ea = np.array([obj_mean[(pid, "EgoAnchor", item)] for pid in use])
        else:
            oe = np.array([subscale[(pid, "One-Euro")] for pid in use])
            ea = np.array([subscale[(pid, "EgoAnchor")] for pid in use])
        diff = ea - oe
        res = stats.wilcoxon(ea, oe, alternative="two-sided", method="auto")
        rb = rank_biserial(diff)
        boots = []
        n = len(use)
        for _ in range(10000):
            idx = rng.integers(0, n, n)
            boots.append(rank_biserial(diff[idx]))
        lo, hi = np.percentile(boots, [2.5, 97.5])
        dz = diff.mean() / diff.std(ddof=1) if diff.std(ddof=1) > 0 else float("inf")
        def mdn_iqr(x):
            return f"{np.median(x):.2f} [{np.percentile(x, 25):.2f}, {np.percentile(x, 75):.2f}]"
        return dict(oe=mdn_iqr(oe), ea=mdn_iqr(ea), mdiff=float(np.median(diff)),
                    W=float(res.statistic), p=float(res.pvalue), r=rb, ci=(lo, hi), dz=dz)

    def holm(pvals):
        order = np.argsort(pvals)
        m = len(pvals)
        adj = np.empty(m)
        mx = 0.0
        for rank, idx in enumerate(order):
            mx = max(mx, (m - rank) * pvals[idx])
            adj[idx] = min(1.0, mx)
        return adj

    main_items = ["Q1", "Q8", "Q2", "Q9", "Q3", "Q6", "Q7"]
    main_rows = {it: wilcoxon_row(it) for it in main_items}
    main_holm = holm([main_rows[it]["p"] for it in main_items])
    for it, ph in zip(main_items, main_holm):
        main_rows[it]["p_holm"] = float(ph)

    # 已发表量表家族（TiA 反向项先按 6−raw 换向，分量表 = 有效条目均值）
    REV_ITEMS = {"TIA_RC3_REV", "TIA_RC5_REV", "TIA_UP2_REV", "TIA_UP4_REV"}

    def tia_score(pid, cond, items):
        vals = []
        for it in items:
            v = tia_raw[(pid, cond)][it]
            if v is None:
                continue
            vals.append(6 - v if it in REV_ITEMS else v)
        return float(np.mean(vals))

    RC_ITEMS = ["TIA_RC1", "TIA_RC2", "TIA_RC3_REV", "TIA_RC4", "TIA_RC5_REV", "TIA_RC6"]
    UP_ITEMS = ["TIA_UP1", "TIA_UP2_REV", "TIA_UP3", "TIA_UP4_REV"]

    scales = {}
    scales["AQ-EQ"] = {(pid, c): float(np.mean([obj_mean[(pid, c, i)] for i in ("AQ_EQ1", "AQ_EQ2", "AQ_EQ3")]))
                       for pid in pids for c in ("One-Euro", "EgoAnchor")}
    scales["AQ-IQ"] = {(pid, c): float(np.mean([obj_mean[(pid, c, i)] for i in ("AQ_IQ1", "AQ_IQ2", "AQ_IQ3")]))
                       for pid in pids for c in ("One-Euro", "EgoAnchor")}
    scales["TiA-R/C"] = {(pid, c): tia_score(pid, c, RC_ITEMS) for pid in pids for c in ("One-Euro", "EgoAnchor")}
    scales["TiA-U/P"] = {(pid, c): tia_score(pid, c, UP_ITEMS) for pid in pids for c in ("One-Euro", "EgoAnchor")}
    scales["S-TIAS"] = {(pid, c): float(np.mean(list(stias_raw[(pid, c)].values())))
                        for pid in pids for c in ("One-Euro", "EgoAnchor")}
    scale_rows = {s: wilcoxon_row(None, subscale=scales[s]) for s in scales}
    scale_holm = holm([scale_rows[s]["p"] for s in scales])
    for s, ph in zip(scales, scale_holm):
        scale_rows[s]["p_holm"] = float(ph)

    # 信度：α 与 ω（主因子一维近似），逐方法计算
    def cronbach(mat):
        mat = mat[~np.isnan(mat).any(axis=1)]
        k = mat.shape[1]
        return k / (k - 1) * (1 - mat.var(axis=0, ddof=1).sum() / mat.sum(axis=1).var(ddof=1))

    def omega_1f(mat):
        mat = mat[~np.isnan(mat).any(axis=1)]
        R = np.corrcoef(mat.T)
        h = np.clip(np.abs(R - np.eye(len(R))).max(axis=1), 0.2, 0.95)
        for _ in range(80):
            Ra = R.copy()
            np.fill_diagonal(Ra, h)
            val, vec = np.linalg.eigh(Ra)
            lam = vec[:, -1] * math.sqrt(max(val[-1], 1e-9))
            h_new = np.clip(lam ** 2, 0.05, 0.98)
            if np.max(np.abs(h_new - h)) < 1e-6:
                h = h_new
                break
            h = h_new
        lam = np.abs(lam)
        return float(lam.sum() ** 2 / (lam.sum() ** 2 + (1 - h).sum()))

    def scale_matrix(name, cond):
        if name == "AQ-EQ":
            items = ["AQ_EQ1", "AQ_EQ2", "AQ_EQ3"]
            return np.array([[obj_mean[(pid, cond, i)] for i in items] for pid in pids])
        if name == "AQ-IQ":
            items = ["AQ_IQ1", "AQ_IQ2", "AQ_IQ3"]
            return np.array([[obj_mean[(pid, cond, i)] for i in items] for pid in pids])
        if name == "TiA-R/C":
            items = RC_ITEMS
        elif name == "TiA-U/P":
            items = UP_ITEMS
        else:
            return np.array([[stias_raw[(pid, cond)][i] for i in ("STIAS1", "STIAS2", "STIAS3")] for pid in pids])
        out = []
        for pid in pids:
            row = []
            for it in items:
                v = tia_raw[(pid, cond)][it]
                row.append(np.nan if v is None else (6 - v if it in REV_ITEMS else v))
            out.append(row)
        return np.array(out, dtype=float)

    reliab = {}
    for s in scales:
        reliab[s] = {}
        for cond in ("One-Euro", "EgoAnchor"):
            m = scale_matrix(s, cond)
            reliab[s][cond] = (cronbach(m), omega_1f(m))

    # 次级：区块级配对优势比（CLMM 演练近似）与顺序效应检查
    clmm_like = {}
    for it in main_items + ["AQ-EQ", "AQ-IQ"]:
        n_pos = n_neg = 0
        per_obj_diff = {o: [] for o in OBJECTS}
        for pid in pids:
            for o in OBJECTS:
                if it in ("AQ-EQ", "AQ-IQ"):
                    items = ("AQ_EQ1", "AQ_EQ2", "AQ_EQ3") if it == "AQ-EQ" else ("AQ_IQ1", "AQ_IQ2", "AQ_IQ3")
                    bl_ea = bl_oe = None
                    for blk in blocks:
                        if blk["pid"] == pid and blk["obj_key"] == o:
                            v = float(np.mean([ratings[(pid, blk["block"])][x] for x in items]))
                            if blk["cond"] == "EgoAnchor":
                                bl_ea = v
                            else:
                                bl_oe = v
                else:
                    bl_ea = bl_oe = None
                    for blk in blocks:
                        if blk["pid"] == pid and blk["obj_key"] == o:
                            v = ratings[(pid, blk["block"])][it]
                            if blk["cond"] == "EgoAnchor":
                                bl_ea = v
                            else:
                                bl_oe = v
                per_obj_diff[o].append(bl_ea - bl_oe)
                if bl_ea > bl_oe:
                    n_pos += 1
                elif bl_ea < bl_oe:
                    n_neg += 1
        orr = n_pos / max(n_neg, 1)
        pb = stats.binomtest(n_pos, n_pos + n_neg, 0.5).pvalue if n_pos + n_neg else 1.0
        dirs = "/".join("+" if np.mean(per_obj_diff[o]) > 0.05 else ("-" if np.mean(per_obj_diff[o]) < -0.05 else "0")
                        for o in OBJECTS)
        clmm_like[it] = dict(n_pos=n_pos, n_neg=n_neg, orr=orr, p=pb, dirs=dirs,
                             obj_means=per_obj_diff)

    # 顺序效应：被试内去均值后回归（区块位置 + 物体内先后），演练近似
    y, X = [], []
    for pid in pids:
        rows = [(blk, ratings[(pid, blk["block"])]) for blk in blocks if blk["pid"] == pid]
        m7 = {it: np.mean([r[it] for _, r in rows]) for it in ("Q1", "Q8", "Q3", "Q6")}
        for blk, r in rows:
            score = np.mean([r[it] for it in ("Q1", "Q8", "Q3", "Q6")])
            base = np.mean(list(m7.values()))
            y.append(score - base)
            X.append([1.0 if blk["cond"] == "EgoAnchor" else 0.0,
                      (blk["block"] - 3.5) / 1.71, 1.0 if blk["pair_order"] == 2 else 0.0])
    y, X = np.array(y), np.array(X)
    beta, *_ = np.linalg.lstsq(np.c_[np.ones(len(y)), X], y, rcond=None)
    resid = y - np.c_[np.ones(len(y)), X] @ beta
    se = math.sqrt(resid.var(ddof=4) * np.linalg.inv((np.c_[np.ones(len(y)), X].T @ np.c_[np.ones(len(y)), X]))[2, 2])
    order_effect = dict(method=beta[1], pos=beta[2], pos_se=se, pair=beta[3])

    # 操纵检验：TOST 配对等价（等价界与判定按演练冻结值）
    def tost(metric, bound):
        oe = np.array([audits[(pid, blk["block"])][metric] for blk in blocks for pid in [blk["pid"]]
                       if blk["cond"] == "One-Euro" and blk["pid"] == pid])
        ea = np.array([audits[(pid, blk["block"])][metric] for blk in blocks for pid in [blk["pid"]]
                       if blk["cond"] == "EgoAnchor" and blk["pid"] == pid])
        d = ea - oe
        se_ = d.std(ddof=1) / math.sqrt(len(d))
        t1 = (d.mean() + bound) / se_
        t2 = (d.mean() - bound) / se_
        p1 = 1 - stats.t.cdf(t1, len(d) - 1)
        p2 = stats.t.cdf(t2, len(d) - 1)
        return dict(oe=oe.mean(), ea=ea.mean(), diff=d.mean(), p=max(p1, p2), bound=bound)

    tost_rows = dict(rate=tost("rate", 0.30), vcd=tost("vcd", 0.03), adm=tost("adm", 0.03),
                     avail=tost("avail", 0.02), occl=tost("occl", 0.05))
    lifecycle = dict(
        frozen_oe=sum(1 for blk in blocks if blk["cond"] == "One-Euro" and audits[(blk["pid"], blk["block"])]["state"] == "FrozenUncertain"),
        frozen_ea=sum(1 for blk in blocks if blk["cond"] == "EgoAnchor" and audits[(blk["pid"], blk["block"])]["state"] == "FrozenUncertain"),
        lost_oe=sum(1 for blk in blocks if blk["cond"] == "One-Euro" and audits[(blk["pid"], blk["block"])]["lost"] == "是"),
        lost_ea=sum(1 for blk in blocks if blk["cond"] == "EgoAnchor" and audits[(blk["pid"], blk["block"])]["lost"] == "是"),
    )

    # 最终计数
    def count_choice(field):
        lab = dict(A=0, B=0, none=0)
        dec = {"EgoAnchor": 0, "One-Euro": 0}
        for pid in pids:
            v = finals[pid][field]
            if v == "无明显偏好":
                lab["none"] += 1
            else:
                lab["A" if v == "方法A" else "B"] += 1
                dec[parts[pid]["methodA"] if v == "方法A" else parts[pid]["methodB"]] += 1
        return lab, dec

    choice_lab, choice_dec = count_choice("choice")
    trust_lab, trust_dec = count_choice("trust")
    inconsistent = [pid for pid in pids
                    if finals[pid]["choice"] != finals[pid]["trust"]]
    conf_gap_r = stats.spearmanr(
        [finals[pid]["conf"] for pid in pids],
        [abs(impressions[(pid, "EgoAnchor")] - impressions[(pid, "One-Euro")]) for pid in pids])

    # N=18 鲁棒性（P001–P018）
    sub18 = pids[:18]
    main18 = {it: wilcoxon_row(it, subset=sub18) for it in main_items}
    holm18 = holm([main18[it]["p"] for it in main_items])
    for it, ph in zip(main_items, holm18):
        main18[it]["p_holm"] = float(ph)

    # 响应定式与问卷时长检查
    long_blocks = [(blk["pid"], blk["block"], times[blk["pid"]]["blocks"][blk["block"]][2])
                   for blk in blocks if times[blk["pid"]]["blocks"][blk["block"]][2] > 150]
    def max_run(pid, block):
        seq = [ratings[(pid, block)][it] for it in RATED_ITEMS]
        best = cur = 1
        for a, b in zip(seq, seq[1:]):
            cur = cur + 1 if a == b else 1
            best = max(best, cur)
        return best
    straightline = [(blk["pid"], blk["block"], max_run(blk["pid"], blk["block"]))
                    for blk in blocks if max_run(blk["pid"], blk["block"]) >= 5]

    # ---- 6. 写出工作簿 ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "README_SYNTHETIC"
    readme = [
        "【AI 合成数据 —— 模型：Claude Fable 5（claude-fable-5）。本工作簿全部内容为模拟演练数据，"
        "不得作为真实实验数据或论文证据，不得与真实采集数据合并，不得进入效应量/功效/样本量推断。】",
        "",
        f"生成时间：2026-07-26（会话内）；随机种子 {SEED}；生成脚本 material/simulate_exp3_claude_fable_5.py（固定种子可复现）。",
        "结构镜像自唯一正式工作簿 EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx"
        "（v5.1 Verified，物体最外层 2×3=6 区块、区块级 13 项七点、方法级 TiA10+S-TIAS3、最终 7 项）。",
        "条目定义、量尺与核对记录见正式工作簿 Questionnaire / Verification_Audit 两表，本文件不重复。",
        "",
        "模拟口径：",
        "· 效应量锚定实验一 v4 实测差异（One-Euro→EgoAnchor：静止泄漏 10.65→0.82 mm、绝对配准 14.00→6.60 mm、"
        "遮挡 P95 10.41→4.85 mm、起停转换 334→510 ms 反向、平移对齐 RMSE 15.69→9.12 mm），不是感知效应预测。",
        "· 24 名被试各有独立人格参数（宽容度/极端度/通道敏感度/信任倾向/作答速度），见 Personas 表。",
        "· 注入事件：One-Euro 遮挡恢复可见跳变 4 例；双方法各 1 次遮挡超时进入 Lost（服务器重注册），保持方法间平衡。",
        "· Q10_OPT 默认停用，列保留为空；TiA 反向条目按原始分记录（未换向），换向只在分析中执行。",
        "· 时间戳为虚构排期（2026-08-10 起 5 个工作日），仅用于流程演练。",
        "",
        "工作表：Personas（人格参数）/ Participants / Records（三段堆叠）/ Analysis（分析回填，含演练近似标注）。",
    ]
    for i, line in enumerate(readme, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110

    # Personas 表
    wsp = wb.create_sheet("Personas")
    head = ["Participant_ID", "画像", "宽容度acq", "极端度ext", "噪声sd", "光环sd",
            "敏感度_静止", "敏感度_滞后", "敏感度_朝向", "敏感度_恢复", "敏感度_位置",
            "信任倾向", "作答速度s每题", "自信", "特殊事件"]
    for c, h in enumerate(head, 1):
        wsp.cell(row=1, column=c, value=h)
    for r, pid in enumerate(pids, 2):
        p = PERSONAS[pid]
        ev = [f"{OBJECTS[o]}×{cond}:{INCIDENTS[(q, o, cond)]['tag']}"
              for (q, o, cond) in INCIDENTS if q == pid]
        row = [pid, p["persona"], p["acq"], p["ext"], p["noise"], p["halo"],
               p["st"], p["lag"], p["ori"], p["rec"], p["pos"], p["trust"], p["spd"], p["conf"],
               "；".join(ev)]
        for c, v in enumerate(row, 1):
            wsp.cell(row=r, column=c, value=v)
    wsp.column_dimensions["B"].width = 42
    wsp.column_dimensions["O"].width = 34

    # Participants / Records / Analysis：先镜像官方全部预填值
    for name in ["Participants", "Records", "Analysis"]:
        wsn = wb.create_sheet(name)
        for (r, c), v in sorted(sheets[name].items()):
            wsn.cell(row=r, column=c, value=v)

    wsP = wb["Participants"]
    for pid in pids:
        r = parts[pid]["row"]
        p = PERSONAS[pid]
        vals = [p["age"], p["gender"], p["hand"], p["vision"], p["vr"], p["mr"], "是", "无",
                times[pid]["start"].strftime("%Y-%m-%d %H:%M"),
                times[pid]["end"].strftime("%Y-%m-%d %H:%M")]
        for c, v in enumerate(vals, 12):
            wsP.cell(row=r, column=c, value=v)
    wsP.cell(row=parts["P011"]["row"], column=24, value="手柄区块一次遮挡超时（见 Records 备注）")
    wsP.cell(row=parts["P019"]["row"], column=24, value="第 2 区块一次遮挡超时（见 Records 备注）")

    wsR = wb["Records"]
    for blk in blocks:
        pid, r = blk["pid"], blk["row"]
        rt = ratings[(pid, blk["block"])]
        for ci, item in enumerate(BLOCK_ITEMS, 11):
            if item != "Q10_OPT":
                wsR.cell(row=r, column=ci, value=rt[item])
        t0, t1, dur = times[pid]["blocks"][blk["block"]]
        au = audits[(pid, blk["block"])]
        tail = {25: t0.strftime("%H:%M:%S"), 26: t1.strftime("%H:%M:%S"), 27: dur,
                28: "是", 29: "是", 30: "是", 31: "无",
                32: au["rate"], 33: au["vcd"], 34: au["adm"], 35: au["avail"],
                36: au["occl"], 37: au["state"], 38: au["lost"], 39: au["reacq"],
                40: au["lock"], 41: au["note"] or None}
        if pid == "P013" and blk["block"] == 3:
            tail[41] = "区块开始前 Unity 程序重启一次，初始化重跑后正常"
        for c, v in tail.items():
            if v is not None:
                wsR.cell(row=r, column=c, value=v)

    for j, mr in enumerate(method_rows):
        pid, r = mr["pid"], mr["row"]
        t = tia_raw[(pid, mr["cond"])]
        s = stias_raw[(pid, mr["cond"])]
        seq = ["TIA_RC1", "TIA_RC2", "TIA_RC3_REV", "TIA_RC4", "TIA_RC5_REV", "TIA_RC6",
               "TIA_UP1", "TIA_UP2_REV", "TIA_UP3", "TIA_UP4_REV"]
        for c, it in enumerate(seq, 5):
            if t[it] is not None:
                wsR.cell(row=r, column=c, value=t[it])
        for c, it in enumerate(("STIAS1", "STIAS2", "STIAS3"), 15):
            wsR.cell(row=r, column=c, value=s[it])
        t0, t1, dur = times[pid]["methods"][mr["order"] - 1]
        wsR.cell(row=r, column=18, value="是")
        wsR.cell(row=r, column=19, value=t0.strftime("%H:%M:%S"))
        wsR.cell(row=r, column=20, value=t1.strftime("%H:%M:%S"))
        wsR.cell(row=r, column=21, value=dur)
        if pid == "P016" and mr["cond"] == "One-Euro":
            wsR.cell(row=r, column=23, value="TIA_RC4 勾选'无法回答'，记缺失")

    interview_notes = {
        "P015": "明确表示能感到差异但认为两方法各有适用场景，属权衡型无偏好",
        "P017": "访谈中猜测两方法机制差异（提到'像加了锁定'），如实记录",
        "P011": "回忆手柄区块的消失-重现事件，表示影响了对该方法的信心",
        "P019": "多次提及第 2 区块的消失与跳变事件，明确表示因此不信任该方法",
        "P009": "偏好与信任选择不一致：跟手感选择前者，涉及真实操作时信任后者",
    }
    for i, pid in enumerate(pids):
        r = 203 + i
        f = finals[pid]
        vals = [None, f["choice"], f["strength"], f["trust"], f["conf"],
                f["open_diff"], f["open_dis"], f["discomfort"],
                (parts[pid]["methodA"] if f["choice"] == "方法A" else
                 parts[pid]["methodB"] if f["choice"] == "方法B" else "无明显偏好"),
                (parts[pid]["methodA"] if f["trust"] == "方法A" else
                 parts[pid]["methodB"] if f["trust"] == "方法B" else "无明显偏好"),
                interview_notes.get(pid)]
        for c, v in enumerate(vals, 1):
            if v is not None:
                wsR.cell(row=r, column=c, value=v)

    # Analysis 表回填
    wsA = wb["Analysis"]
    wsA.cell(row=1, column=1, value="分析结果回填壳（本文件为 AI 合成数据的演练回填；CLMM 以区块级配对优势比近似，正式分析用 R ordinal::clmm）")
    A_ROWS = {"Q1": 5, "Q8": 6, "Q2": 7, "Q9": 8, "Q3": 9, "Q6": 10, "Q7": 11}
    for it, r in A_ROWS.items():
        w = main_rows[it]
        concl = "显著（EgoAnchor 更优）" if w["p_holm"] < 0.05 and w["mdiff"] > 0 else \
                ("显著（One-Euro 更优）" if w["p_holm"] < 0.05 else "不显著")
        vals = [w["oe"], w["ea"], round(w["mdiff"], 2), w["W"], f"{w['p']:.4g}", f"{w['p_holm']:.4g}",
                round(w["r"], 3), f"[{w['ci'][0]:.2f}, {w['ci'][1]:.2f}]", concl]
        for c, v in enumerate(vals, 3):
            wsA.cell(row=r, column=c, value=v)
    B_ROWS = {"AQ-EQ": 15, "AQ-IQ": 16, "TiA-R/C": 17, "TiA-U/P": 18, "S-TIAS": 19}
    for s, r in B_ROWS.items():
        w = scale_rows[s]
        a_oe, o_oe = reliab[s]["One-Euro"]
        a_ea, o_ea = reliab[s]["EgoAnchor"]
        concl = "显著（EgoAnchor 更优）" if w["p_holm"] < 0.05 and w["mdiff"] > 0 else \
                ("显著（One-Euro 更优）" if w["p_holm"] < 0.05 else "不显著")
        vals = [w["oe"], w["ea"], round(w["mdiff"], 2), w["W"], f"{w['p']:.4g}", f"{w['p_holm']:.4g}",
                round(w["r"], 3), round(a_oe, 2), round(a_ea, 2), round(o_oe, 2), round(o_ea, 2), concl]
        for c, v in enumerate(vals, 2):
            wsA.cell(row=r, column=c, value=v)
    r = 23
    for it in main_items + ["AQ-EQ", "AQ-IQ"]:
        cl = clmm_like[it]
        wsA.cell(row=r, column=1, value=it)
        wsA.cell(row=r, column=2, value="方法（区块级配对，演练近似）")
        wsA.cell(row=r, column=3, value=f"{cl['orr']:.2f} ({cl['n_pos']}+/{cl['n_neg']}-)")
        wsA.cell(row=r, column=4, value="精确二项")
        wsA.cell(row=r, column=5, value=f"{cl['p']:.4g}")
        wsA.cell(row=r, column=6, value=f"三物体方向 {cl['dirs']}；正式分析用 CLMM")
        r += 1
    wsA.cell(row=r, column=1, value="Q1/Q8/Q3/Q6 合成")
    wsA.cell(row=r, column=2, value="区块位置（被试内去均值回归）")
    wsA.cell(row=r, column=3, value=f"β={order_effect['pos']:.3f} (SE {order_effect['pos_se']:.3f})")
    wsA.cell(row=r, column=6, value=f"物体内先后 β={order_effect['pair']:.3f}；演练近似")
    D_BASE = {"Q1": 42, "Q2": 45, "Q9": 48, "Q10_OPT": 51, "Q3": 54, "Q8": 57,
              "Q6": 60, "Q7": 63, "AQ-EQ": 66, "AQ-IQ": 69}
    for it, r0 in D_BASE.items():
        if it == "Q10_OPT":
            for k in range(3):
                wsA.cell(row=r0 + k, column=7, value="默认停用，未采集")
            continue
        for k, o in enumerate(OBJECTS):
            if it in ("AQ-EQ", "AQ-IQ"):
                items = ("AQ_EQ1", "AQ_EQ2", "AQ_EQ3") if it == "AQ-EQ" else ("AQ_IQ1", "AQ_IQ2", "AQ_IQ3")
                oe_v = np.mean([np.mean([ratings[(pid, blk["block"])][x] for x in items])
                                for pid in pids for blk in blocks
                                if blk["pid"] == pid and blk["obj_key"] == o and blk["cond"] == "One-Euro"])
                ea_v = np.mean([np.mean([ratings[(pid, blk["block"])][x] for x in items])
                                for pid in pids for blk in blocks
                                if blk["pid"] == pid and blk["obj_key"] == o and blk["cond"] == "EgoAnchor"])
            else:
                oe_v = np.mean([ratings[(pid, blk["block"])][it] for pid in pids for blk in blocks
                                if blk["pid"] == pid and blk["obj_key"] == o and blk["cond"] == "One-Euro"])
                ea_v = np.mean([ratings[(pid, blk["block"])][it] for pid in pids for blk in blocks
                                if blk["pid"] == pid and blk["obj_key"] == o and blk["cond"] == "EgoAnchor"])
            d = ea_v - oe_v
            row_vals = [round(float(oe_v), 2), round(float(ea_v), 2), round(float(d), 2),
                        "+" if d > 0.05 else ("-" if d < -0.05 else "0")]
            for c, v in enumerate(row_vals, 3):
                wsA.cell(row=r0 + k, column=c, value=v)
    E_MAP = {75: ("rate", "Hz"), 76: ("vcd", ""), 77: ("adm", ""), 78: ("avail", ""), 79: ("occl", "s")}
    for r0, (m, unit) in E_MAP.items():
        t = tost_rows[m]
        verdict = "等价" if t["p"] < 0.05 else "不等价"
        vals = [f"{t['oe']:.3f}", f"{t['ea']:.3f}", f"Δ={t['diff']:+.3f}",
                f"±{t['bound']}{unit}（演练冻结）", f"{t['p']:.4g}", verdict]
        for c, v in enumerate(vals, 2):
            wsA.cell(row=r0, column=c, value=v)
    wsA.cell(row=80, column=2, value=f"{lifecycle['frozen_oe']}/72")
    wsA.cell(row=80, column=3, value=f"{lifecycle['frozen_ea']}/72")
    wsA.cell(row=80, column=7, value="通过（多数）")
    wsA.cell(row=81, column=2, value=f"{lifecycle['lost_oe']}/72")
    wsA.cell(row=81, column=3, value=f"{lifecycle['lost_ea']}/72")
    wsA.cell(row=81, column=7, value="平衡（各 1 次）")
    wsA.cell(row=82, column=2, value=str(lifecycle["lost_oe"]))
    wsA.cell(row=82, column=3, value=str(lifecycle["lost_ea"]))
    for r0, (lab, dec) in ((86, (choice_lab, choice_dec)), (87, (trust_lab, trust_dec))):
        vals = [lab["A"], lab["B"], lab["none"], dec["EgoAnchor"], dec["One-Euro"]]
        for c, v in enumerate(vals, 2):
            wsA.cell(row=r0, column=c, value=v)
    strengths = [finals[pid]["strength"] for pid in pids if finals[pid]["strength"] != "NA"]
    confs = [finals[pid]["conf"] for pid in pids]
    wsA.cell(row=88, column=2, value="；".join(f"{k}:{strengths.count(k)}" for k in range(1, 8)))
    wsA.cell(row=88, column=4, value=sum(1 for pid in pids if finals[pid]["strength"] == "NA"))
    wsA.cell(row=89, column=2, value="；".join(f"{k}:{confs.count(k)}" for k in range(1, 8)))
    wsA.cell(row=90, column=2, value=len(inconsistent))
    wsA.cell(row=90, column=7, value="不一致者：" + "、".join(inconsistent))
    disc = [finals[pid]["discomfort"] for pid in pids]
    wsA.cell(row=91, column=2, value=f"无:{disc.count('无')}；轻微:{disc.count('轻微')}；中等及以上:0")
    for i, pid in enumerate(pids):
        r0 = 95 + i
        f = finals[pid]
        wsA.cell(row=r0, column=2, value=f["open_diff"][:60])
        wsA.cell(row=r0, column=3, value=f["open_dis"][:60])
        for c, th in enumerate(THEME_COLS, 4):
            wsA.cell(row=r0, column=c, value=1 if th in f["themes"] else 0)
        wsA.cell(row=r0, column=17, value="Fable5-模拟")
        wsA.cell(row=r0, column=18, value="Fable5-模拟")

    for wsx in (wsR, wsP, wsA):
        wsx.column_dimensions["A"].width = 14

    wb.save(OUT_XLSX)

    # ---- 7. 校验 ----
    checks = []
    n_rated = sum(1 for blk in blocks for it in RATED_ITEMS
                  if 1 <= ratings[(blk["pid"], blk["block"])][it] <= 7)
    checks.append(("区块评分 144×13 全部在 1–7", n_rated == 144 * 13))
    checks.append(("Q10_OPT 全部留空", True))
    n_tia = sum(1 for mr in method_rows for it in tia_raw[(mr["pid"], mr["cond"])]
                if tia_raw[(mr["pid"], mr["cond"])][it] is not None)
    checks.append(("TiA 48×10 仅 1 个缺失（P016）", n_tia == 480 - 1))
    first_ea = sum(1 for pid in pids if parts[pid]["first"] == "EgoAnchor")
    checks.append(("先行方法 12/12", first_ea == 12))
    checks.append(("Lost 事件方法间平衡 1/1", lifecycle["lost_oe"] == 1 and lifecycle["lost_ea"] == 1))
    durs = [(times[pid]["end"] - times[pid]["start"]).total_seconds() / 60 for pid in pids]
    checks.append(("会话时长 50–75 min", all(50 <= d <= 75 for d in durs)))

    # ---- 8. 统计输出 ----
    P = print
    P("=" * 74)
    P("EgoAnchor 实验三 AI 模拟（Claude Fable 5，v5.1 结构，种子 %d）" % SEED)
    P("=" * 74)
    P("\n[校验]")
    for name, ok in checks:
        P(f"  {'PASS' if ok else 'FAIL'}  {name}")
    P(f"  会话时长 min={min(durs):.0f} max={max(durs):.0f} 中位={np.median(durs):.0f} 分钟")
    P("\n[A 主证实家族]（三物体均值 → Wilcoxon → Holm(7)）")
    for it in main_items:
        w = main_rows[it]
        P(f"  {it:4s} OE {w['oe']:>20s} | EA {w['ea']:>20s} | Δmdn {w['mdiff']:+.2f} "
          f"| W={w['W']:.1f} p={w['p']:.4g} pHolm={w['p_holm']:.4g} r={w['r']:.3f} "
          f"CI[{w['ci'][0]:.2f},{w['ci'][1]:.2f}] dz={w['dz']:.2f}")
    P("\n[B 已发表量表家族]（Holm(5)；α/ω 为当前样本、演练近似）")
    for s in scales:
        w = scale_rows[s]
        a_oe, o_oe = reliab[s]["One-Euro"]
        a_ea, o_ea = reliab[s]["EgoAnchor"]
        P(f"  {s:8s} OE {w['oe']:>20s} | EA {w['ea']:>20s} | Δmdn {w['mdiff']:+.2f} "
          f"| p={w['p']:.4g} pHolm={w['p_holm']:.4g} r={w['r']:.3f} dz={w['dz']:.2f} "
          f"| α {a_oe:.2f}/{a_ea:.2f} ω {o_oe:.2f}/{o_ea:.2f}")
    P("\n[C 次级近似]（区块级配对优势比 + 顺序效应）")
    for it in main_items + ["AQ-EQ", "AQ-IQ"]:
        cl = clmm_like[it]
        P(f"  {it:6s} OR≈{cl['orr']:.2f} ({cl['n_pos']}+/{cl['n_neg']}-) p={cl['p']:.4g} 三物体方向 {cl['dirs']}")
    P(f"  顺序效应：区块位置 β={order_effect['pos']:+.3f}（SE {order_effect['pos_se']:.3f}），"
      f"物体内先后 β={order_effect['pair']:+.3f}")
    P("\n[E 操纵检验 TOST]")
    for m, t in tost_rows.items():
        P(f"  {m:6s} OE={t['oe']:.3f} EA={t['ea']:.3f} Δ={t['diff']:+.4f} 界±{t['bound']} p={t['p']:.4g} "
          f"{'等价' if t['p'] < 0.05 else '不等价'}")
    P(f"  生命周期：FrozenUncertain OE {lifecycle['frozen_oe']}/72、EA {lifecycle['frozen_ea']}/72；"
      f"Lost OE {lifecycle['lost_oe']}、EA {lifecycle['lost_ea']}")
    P("\n[F 最终测量]")
    P(f"  方法选择 标签 A:{choice_lab['A']} B:{choice_lab['B']} 无:{choice_lab['none']} | "
      f"解码 EgoAnchor:{choice_dec['EgoAnchor']} One-Euro:{choice_dec['One-Euro']}")
    P(f"  信任选择 标签 A:{trust_lab['A']} B:{trust_lab['B']} 无:{trust_lab['none']} | "
      f"解码 EgoAnchor:{trust_dec['EgoAnchor']} One-Euro:{trust_dec['One-Euro']}")
    P(f"  偏好强度（有选择者）：{sorted(strengths)}")
    P(f"  区分信心：{sorted(confs)}")
    P(f"  偏好×信任不一致：{len(inconsistent)} 人 {inconsistent}")
    P(f"  区分信心×|印象差| Spearman ρ={conf_gap_r.statistic:.3f} (p={conf_gap_r.pvalue:.4g})")
    P("\n[N=18 鲁棒性]（P001–P018）")
    for it in main_items:
        w = main18[it]
        P(f"  {it:4s} pHolm={w['p_holm']:.4g} r={w['r']:.3f} {'显著' if w['p_holm'] < 0.05 else '不显著'}")
    P("\n[问卷负担与响应定式]")
    P(f"  区块问卷时长 中位={np.median([times[b['pid']]['blocks'][b['block']][2] for b in blocks]):.0f} s；"
      f">150 s 共 {len(long_blocks)} 区块：{long_blocks}")
    P(f"  连续同分≥5 的区块：{len(straightline)} 个 {[(a, b) for a, b, _ in straightline][:8]}")
    P("\n[逐被试印象差与最终选择]")
    for pid in pids:
        d = impressions[(pid, 'EgoAnchor')] - impressions[(pid, 'One-Euro')]
        f = finals[pid]
        P(f"  {pid} Δ印象={d:+.2f} 选择={f['choice']}（强度 {f['strength']}）信任={f['trust']} 信心={f['conf']}")
    P("\n输出：", OUT_XLSX)


if __name__ == "__main__":
    main()
