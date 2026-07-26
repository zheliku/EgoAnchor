# -*- coding: utf-8 -*-
"""EgoAnchor 实验三（v5.1 冻结结构）AI 模拟演练数据生成器。

模型：Claude Opus 5（1M context），模型 ID ``claude-opus-5[1m]``。

**本脚本产出的全部数值都是 AI 合成数据，不是真实被试数据。**
用途只有一个：在正式招募前检验 v5.1 冻结设计的灵敏度、条目区分度与分析链路是否可用。
禁止作为论文证据、禁止与真实采集数据合并、禁止写入正式结果表。

运行方式（在 ``EgoAnchor_Python`` 目录下）::

    pixi run python ..\\2026-EgoAnchor\\material\\simulate_exp3_claude_opus_5_1m.py

脚本以官方采集工作簿
``EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx``
为结构母版：复制一份后只写入应答列与审计列，不改动预填的平衡设计列，
以保证模拟结果可以与正式采集表逐单元格对照。种子固定，重复运行全部单元格取值完全一致
（xlsx 文件哈希会因 openpyxl 写入的创建时间元数据而不同，与数据无关）。
"""

from __future__ import annotations

import json
import math
import shutil
from dataclasses import dataclass, field
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.formatting.formatting import ConditionalFormattingList
from scipy import stats

# ---------------------------------------------------------------------------
# 路径与常量
# ---------------------------------------------------------------------------

MATERIAL_DIR = Path(__file__).resolve().parent
SOURCE_WORKBOOK = (
    MATERIAL_DIR
    / "EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx"
)
OUTPUT_WORKBOOK = MATERIAL_DIR / "EgoAnchor_Experiment3_Simulated_Claude-Opus-5-1M_v5_1_24P.xlsx"
SUMMARY_JSON = MATERIAL_DIR / "_sim_opus_5_1m_summary.json"  # 供报告撰写核对，非交付物

SEED = 20260727  # 固定随机种子；与上一轮 Opus 5 演练（20260726）区分
MODEL_NAME = "Claude Opus 5 (1M context) / claude-opus-5[1m]"

# Records A 段（区块记录）在母版中的行范围与列号
BLOCK_FIRST_ROW, BLOCK_LAST_ROW = 5, 148
# Records B 段（方法级记录）
METHOD_FIRST_ROW, METHOD_LAST_ROW = 152, 199
# Records C 段（最终问卷）
FINAL_FIRST_ROW, FINAL_LAST_ROW = 203, 226
# Participants 数据行
PART_FIRST_ROW, PART_LAST_ROW = 3, 26

# 区块级 14 列（K..X），顺序与母版表头一致：页序 1,2,3,3',4..13
# 每项给出 (item_id, 简称, 两方法均值基线, 条件效应 δ = EgoAnchor − One-Euro)
BLOCK_ITEMS: list[tuple[str, str, float, float]] = [
    ("Q1", "静止稳定", 4.90, +1.10),
    ("Q2", "运动附着", 5.04, +0.18),
    ("Q9", "姿态一致", 5.06, +0.52),
    ("Q10_OPT", "放置后稳定", 0.0, 0.0),  # 默认停用，不生成数据
    ("AQ_IQ2", "响应及时", 5.06, -0.18),
    ("AQ_IQ3", "运动平滑", 4.98, +0.85),
    ("Q3", "恢复一致", 4.76, +0.92),
    ("Q8", "位置正确", 4.81, +0.86),
    ("AQ_EQ1", "融合", 4.89, +0.58),
    ("AQ_EQ2", "嵌入可信", 4.83, +0.66),
    ("AQ_EQ3", "尺度协调", 5.63, +0.06),
    ("AQ_IQ1", "交互直观", 5.24, +0.28),
    ("Q6", "依赖意愿", 4.74, +0.88),
    ("Q7", "稳定-响应平衡", 4.76, +0.62),
]
BLOCK_ITEM_COL = {item[0]: 11 + idx for idx, item in enumerate(BLOCK_ITEMS)}  # K = 11

# 已发表量表的分量表归属：同一分量表内的条目共享一个区块内潜在因子，
# 这是量表具备内部一致性的来源；自制条目不属于任何分量表，不进入 α/ω。
BLOCK_SUBSCALE = {
    "AQ_EQ1": "AQ_EQ",
    "AQ_EQ2": "AQ_EQ",
    "AQ_EQ3": "AQ_EQ",
    "AQ_IQ1": "AQ_IQ",
    "AQ_IQ2": "AQ_IQ",
    "AQ_IQ3": "AQ_IQ",
}
SUBSCALE_FACTOR_SD = 0.46  # 区块内分量表共享因子的标准差

# 方法级 TiA 10 项（五点），列 E..N；反向项在采集表中只存原始分
# (item_id, 分量表, 是否反向, 构念基线（两方法均值）, 构念效应 δ)
TIA_ITEMS: list[tuple[str, str, bool, float, float]] = [
    ("TIA_RC1", "RC", False, 3.55, +0.32),
    ("TIA_RC2", "RC", False, 3.50, +0.55),
    ("TIA_RC3_REV", "RC", True, 3.45, +0.50),
    ("TIA_RC4", "RC", False, 3.40, +0.30),
    ("TIA_RC5_REV", "RC", True, 3.30, +0.45),
    ("TIA_RC6", "RC", False, 3.52, +0.58),
    ("TIA_UP1", "UP", False, 3.42, +0.25),
    ("TIA_UP2_REV", "UP", True, 3.48, +0.40),
    ("TIA_UP3", "UP", False, 3.30, +0.15),
    ("TIA_UP4_REV", "UP", True, 3.40, +0.32),
]
TIA_COL = {item[0]: 5 + idx for idx, item in enumerate(TIA_ITEMS)}  # E = 5

# S-TIAS 3 项（七点程度尺），列 O..Q
STIAS_ITEMS: list[tuple[str, float, float]] = [
    ("STIAS1", 4.65, +0.75),
    ("STIAS2", 4.60, +0.85),
    ("STIAS3", 4.55, +0.80),
]
STIAS_COL = {item[0]: 15 + idx for idx, item in enumerate(STIAS_ITEMS)}  # O = 15

# 物体调制：增益放大/缩小条件效应；偏移平移两方法的共同质量水平
OBJECT_TUNING = {
    "blue_mouse": {"gain": 1.20, "shift": -0.22, "cand": 9.18, "vcd": 0.862},
    "stapler": {"gain": 0.82, "shift": +0.30, "cand": 9.61, "vcd": 0.912},
    "gamepad": {"gain": 1.02, "shift": +0.04, "cand": 9.29, "vcd": 0.879},
}

# 遮挡时长冻结值（0.6–0.9 s 区间内，避开 0.45 / 1.0 s 边界）
OCCLUSION_TARGET_S = 0.75

# 主观单项评分的噪声尺度：真实七点单项的重测波动通常在 0.8–1.2 分量级，
# 人格表里的 noise/halo 是相对值，这两个系数把它放大到该量级。
ITEM_NOISE_SCALE = 1.30  # 条目层随机波动
HALO_SCALE = 1.15  # 区块整体印象波动（同区块 13 项共享，不在配对差中抵消）

# 人×条目随机斜率的标准差：persona 的 acuity 已承担人层敏感度差异
# （0.30–1.34，含两名明确的非辨别者），本项只负责条目特异的残差权重
# ——例如某人重视平滑却不在意恢复一致性——因此取值小于 acuity 的离散度。
# 缺少该成分时，三物体求均值会把噪声抹平，配对符号完全由恒正的 δ 决定，
# r_rb 被结构性推到 1.0；真实主观量表不会出现全员同向。
PERSON_ITEM_SLOPE_SD = 0.38

# 操纵检验等价界（采集前冻结）
EQUIV_BOUNDS = {
    "候选到达率 Hz": 0.50,
    "VCD 分数中位": 0.030,
    "VCD 接纳率": 0.050,
    "输出可用率": 0.010,
    "遮挡时长 s": 0.050,
}


# ---------------------------------------------------------------------------
# 虚拟参与者人格设定
# ---------------------------------------------------------------------------


@dataclass
class Persona:
    """单个虚拟参与者的稳定个体差异参数。

    acuity        感知敏锐度：直接缩放条件效应，0.3 表示几乎分辨不出两方法。
    leniency      宽严倾向：整体评分平移。
    style_gain    量表使用风格：>1 用两端，<1 挤在中间。
    noise         作答噪声标准差。
    halo          区块整体印象波动（同区块内 13 项共享）。
    fatigue       后两个区块的额外压缩强度。
    item_mult     条目特异的效应乘子（人格化偏好，可为负=反向）。
    decisiveness  最终强制选择的判定阈值，越小越容易做出选择。
    tag           人格标签，写入报告。
    """

    tag: str
    acuity: float
    leniency: float
    style_gain: float
    noise: float
    halo: float = 0.30
    fatigue: float = 0.05
    item_mult: dict[str, float] = field(default_factory=dict)
    object_mult: dict[str, float] = field(default_factory=dict)
    decisiveness: float = 0.35
    background: tuple = ()
    note: str = ""


# 24 名虚拟参与者。人格分布刻意覆盖真实样本中必然出现的几类人：
# 非辨别者、响应优先者、高敏感者、过稳怀疑者、极端/中庸/宽容/严苛评分者、疲劳型。
PERSONAS: dict[str, Persona] = {
    "P001": Persona(
        "高敏感 VR 老手", 1.34, -0.14, 1.16, 0.42, halo=0.26, decisiveness=0.24,
        background=(26, "男", "右手", "矫正后正常", "超过 20 次", "数次"),
        note="全程主动做左右头动对比，评分偏严但区分度高",
    ),
    "P002": Persona(
        "中庸稳健者", 0.96, +0.05, 0.64, 0.40, decisiveness=0.42,
        background=(23, "女", "右手", "矫正后正常", "6–20 次", "1–2 次"),
        note="几乎只用 3–5 分，自述“不想给满分也不想给低分”",
    ),
    "P003": Persona(
        "非辨别者 A", 0.30, +0.18, 0.86, 0.62, halo=0.40, decisiveness=0.55,
        background=(21, "女", "右手", "正常", "1–5 次", "从未"),
        note="多次表示两种方法“看着差不多”；TIA_RC4 选择无法回答",
    ),
    "P004": Persona(
        "响应优先者 A", 1.06, -0.05, 1.04, 0.44, decisiveness=0.30,
        item_mult={"AQ_IQ2": 2.6, "Q2": 0.15, "Q7": 0.30, "AQ_IQ1": 0.4},
        background=(29, "男", "右手", "矫正后正常", "超过 20 次", "经常"),
        note="把完整系统起动瞬间的迟疑明确算作缺点，反复提到“慢半拍”",
    ),
    "P005": Persona(
        "宽容乐观者", 0.86, +0.72, 0.90, 0.46, decisiveness=0.40,
        background=(22, "女", "右手", "正常", "1–5 次", "1–2 次"),
        note="整体给分偏高，认为“两个都挺好”",
    ),
    "P006": Persona(
        "严苛评分者", 1.12, -0.70, 1.06, 0.44, decisiveness=0.28,
        background=(31, "男", "左手", "矫正后正常", "6–20 次", "数次"),
        note="口头标准很高，认为叠加物“只要动一下就算错”",
    ),
    "P007": Persona(
        "极端评分者", 1.02, +0.10, 1.46, 0.48, halo=0.34, decisiveness=0.32,
        background=(20, "男", "右手", "正常", "1–5 次", "从未"),
        note="只用 1/2 与 6/7，中间刻度几乎不用",
    ),
    "P008": Persona(
        "高敏感 VR 老手 B", 1.30, -0.08, 1.10, 0.40, halo=0.24, decisiveness=0.24,
        background=(27, "女", "右手", "矫正后正常", "经常使用", "经常"),
        note="能自发区分“持续微抖”与“单次跳变”",
    ),
    "P009": Persona(
        "过稳怀疑者", 1.20, -0.02, 1.02, 0.44, decisiveness=0.30,
        item_mult={"AQ_EQ1": -0.85, "AQ_EQ2": -1.30, "Q6": -0.70, "AQ_IQ1": -0.3},
        background=(24, "男", "右手", "正常", "6–20 次", "1–2 次"),
        note="认为纹丝不动的叠加物“像贴纸”，在嵌入可信与依赖意愿上反向",
    ),
    "P010": Persona(
        "一般用户 A", 0.98, +0.02, 0.94, 0.48, decisiveness=0.36,
        background=(25, "女", "右手", "矫正后正常", "1–5 次", "1–2 次"),
        note="第 3 区块遇到感知服务崩溃，该区块作废",
    ),
    "P011": Persona(
        "非辨别者 B", 0.34, -0.10, 0.80, 0.66, halo=0.42, decisiveness=0.60,
        background=(33, "男", "右手", "矫正后正常", "从未", "从未"),
        note="首次使用 VR；TIA_RC1 两次都选无法回答",
    ),
    "P012": Persona(
        "一般用户 B", 1.00, +0.08, 0.98, 0.46, decisiveness=0.34,
        background=(22, "女", "右手", "正常", "1–5 次", "从未"),
    ),
    "P013": Persona(
        "谨慎中庸者", 0.82, -0.15, 0.70, 0.42, decisiveness=0.46,
        background=(28, "女", "右手", "矫正后正常", "6–20 次", "数次"),
        note="第 5 区块 VCD 连续拒绝导致一次 Lost；TIA_UP3 一次无法回答",
    ),
    "P014": Persona(
        "响应优先者 B", 1.08, +0.02, 1.00, 0.44, decisiveness=0.32,
        item_mult={"AQ_IQ2": 2.0, "Q2": 0.35, "Q7": 0.55},
        background=(24, "男", "右手", "正常", "超过 20 次", "数次"),
        note="把起动迟疑当作决定性缺点；承认稳定性与遮挡恢复更差仍选 One-Euro",
    ),
    "P015": Persona(
        "一般用户 C", 1.02, -0.05, 1.00, 0.50, decisiveness=0.34,
        background=(23, "男", "右手", "矫正后正常", "1–5 次", "1–2 次"),
    ),
    "P016": Persona(
        "手柄敏感玩家", 1.22, -0.06, 1.08, 0.44, decisiveness=0.26,
        object_mult={"gamepad": 1.34},
        background=(21, "男", "右手", "正常", "超过 20 次", "数次"),
        note="长期使用手柄，对手柄上的偏差特别敏感",
    ),
    "P017": Persona(
        "宽容乐观者 B", 0.90, +0.58, 0.92, 0.48, decisiveness=0.42,
        background=(30, "女", "右手", "矫正后正常", "1–5 次", "1–2 次"),
    ),
    "P018": Persona(
        "严苛工程背景者", 1.26, -0.52, 1.08, 0.38, halo=0.24, decisiveness=0.22,
        background=(27, "男", "右手", "矫正后正常", "经常使用", "经常"),
        note="有 AR 开发经验，在开放题中直接推测了机制差异（需求特征风险样例）",
    ),
    "P019": Persona(
        "一般用户 D", 0.94, +0.12, 0.96, 0.52, decisiveness=0.36,
        background=(19, "女", "右手", "正常", "从未", "从未"),
        note="第 2 区块单手搬移时手掌遮挡过多",
    ),
    "P020": Persona(
        "疲劳型", 1.04, +0.00, 0.98, 0.50, fatigue=0.26, decisiveness=0.38,
        background=(34, "男", "右手", "矫正后正常", "1–5 次", "1–2 次"),
        note="后两个区块明显加快作答，评分向中间收缩",
    ),
    "P021": Persona(
        "一般用户 E", 1.00, -0.02, 1.02, 0.48, decisiveness=0.34,
        background=(26, "女", "左手", "矫正后正常", "6–20 次", "1–2 次"),
        note="第 6 区块挡板移除时轻微碰到物体",
    ),
    "P022": Persona(
        "首因锚定型", 0.92, +0.06, 0.94, 0.46, decisiveness=0.40,
        background=(24, "男", "右手", "正常", "1–5 次", "从未"),
        note="第一个区块定基调，后续评分变化幅度偏小",
    ),
    "P023": Persona(
        "一般用户 F", 1.04, +0.04, 1.00, 0.48, decisiveness=0.34,
        background=(22, "女", "右手", "矫正后正常", "1–5 次", "1–2 次"),
    ),
    "P024": Persona(
        "高敏感 D（轻微晕动）", 1.28, -0.10, 1.12, 0.44, fatigue=0.18, decisiveness=0.26,
        background=(25, "男", "右手", "矫正后正常", "6–20 次", "数次"),
        note="结束时报告中等程度不适，与方法无关",
    ),
}

# 预设技术事件：(参与者, 区块序号, 是否整块作废, 说明, 受影响条目衰减)
TECH_EVENTS = {
    ("P006", 4): (False, "挡板移除偏慢，遮挡时长超出冻结窗口", {"Q3": -1.5, "Q6": -0.6}),
    ("P010", 3): (True, "感知服务崩溃，需人工重初始化；该区块按冻结规则作废", {}),
    ("P013", 5): (False, "VCD 连续拒绝，锚点一度进入 Lost 并等待服务器重注册",
                  {"Q3": -2.2, "Q6": -1.1, "Q8": -0.7}),
    ("P019", 2): (False, "单手搬移时手掌遮挡过多，候选到达率下降",
                  {"Q2": -1.2, "Q8": -0.8, "AQ_IQ2": -0.7}),
    ("P021", 6): (False, "挡板移除时轻微碰到物体，物体位置发生小幅改变",
                  {"Q3": -1.6, "Q8": -0.6}),
}

# TiA 无法回答（记缺失）的预设单元格：(参与者, 条目) → 两次施测都缺失或只缺一次
TIA_MISSING = {
    ("P003", "TIA_RC4"): "both",
    ("P011", "TIA_RC1"): "both",
    ("P013", "TIA_UP3"): "first",
}


# ---------------------------------------------------------------------------
# 开放题回答（逐人手写，与该人的人格与实际评分模式对应）
# ---------------------------------------------------------------------------

OPEN_DIFFERENCE = {
    "P001": "一个在我左右晃头的时候会跟着一起游，另一个基本钉死在那儿。差别最明显的不是拿起来的时候，是我头一动、它就露馅。",
    "P002": "感觉一个稍微稳一点吧，不过没有特别夸张的区别，我可能没看那么细。",
    "P003": "说实话我没太看出来。可能有一点点不一样，但让我具体说是哪儿不一样，我说不上来。",
    "P004": "稳定性上后面那个确实好，但它起步慢半拍——我手都抬起来了它还在原地愣一下，这个我挺在意的。",
    "P005": "两个都挺好用的，非要说的话有一个在放下之后安定得更快一些。",
    "P006": "一个静止的时候一直在微微地抖，抖得不大但一直在抖；另一个是真的不动。挡板拿开之后的表现也差一截。",
    "P007": "差太多了。一个完全就是贴死在鼠标上，另一个像是浮在上面，随时会飘。",
    "P008": "我把它们分成两种毛病：一种是一直在小幅度地呼吸，另一种是起步会愣一下。呼吸那个一直提醒我它是假的，愣一下那个又让我等它。两种我都能立刻看出来，但要说哪个整体更好，我真选不出来。",
    "P009": "一个稳得有点过头，纹丝不动，反而像一张图片糊在物体上；另一个会有很轻微的活动，看着更像真的在那儿。位置准不准是另一回事。",
    "P010": "遮住再拿开的时候差别最大，一个回来就在原位，另一个要挪一下才对上。",
    "P011": "我第一次用这种设备，不太确定该看什么。两个给我的感觉差不多，最后让我必须挑一个，我就凭最后那轮的印象挑了，说不上有多大把握。",
    "P012": "静止看的时候差别最明显，一个会跟着我的视线飘一点点。搬动的时候两个我都觉得还行。",
    "P013": "有一次挡板拿开之后它半天没回来，那次印象很深。其他时候我觉得区别不算很大。",
    "P014": "稳的那个确实更稳，挡板拿开之后也回得更准。但它起步太拖了——我手已经开始动了，它还在原地待着，等它追上来我这一下都做完了。真让我选，我选跟手的那个，虽然我知道它没那么稳。",
    "P015": "一个在我靠近看的时候会有点偏，另一个位置更实。转动物体的时候两个都还可以。",
    "P016": "手柄上差别最大，因为手柄大、我能看到整个轮廓。一个的轮廓边缘一直在动，另一个是死死咬住的。鼠标上我反而没那么确定。",
    "P017": "都还不错，一个在遮挡之后恢复得更利索一点。",
    "P018": "一个明显只跟着我的视角走，我头一转它就跟着漂，说明它没有真正在世界坐标里固定住；另一个应该是加了某种静止锁定，静止的时候完全不动，但代价是起步的响应慢了一点。",
    "P019": "拿起来的时候有一次我的手挡住了大半，那次两个都不太行。平时看的话有一个更贴一些。",
    "P020": "前面几轮我觉得区别挺清楚的，一个更稳；后面几轮说实话有点累了，看得没那么仔细。",
    "P021": "位置准确度上有区别，一个会整体偏出去一点点，虽然它自己很稳。",
    "P022": "第一次戴上看到的那个印象最深，后面都在跟它比。总的来说有一个更贴合，另一个会飘。",
    "P023": "静止的时候一个在慢慢挪，另一个不挪。挡板的测试里也是同一个表现更好。",
    "P024": "转动的时候两个差不多，但静止和遮挡恢复上区别明显。我后面有点晕，不过跟哪个方法没关系。",
}

OPEN_DISTRUST = {
    "P001": "它自己纠正位置的那一下。我一看见它在往回挪，就等于亲眼看见它刚才是错的，之后我就不敢信它了。",
    "P002": "一直轻轻地抖，看久了会让人怀疑它到底知不知道东西在哪儿。",
    "P003": "如果它整个跑到别的地方去了，那肯定不能信。轻微的偏我觉得还好。",
    "P004": "反应跟不上。我已经把东西放好了它还在追，这种滞后让我觉得它其实不知道现在发生了什么。",
    "P005": "突然消失或者跳到很远的地方吧，这个最吓人。",
    "P006": "持续的微抖比一次大偏移更糟。偏移我知道它偏多少，抖动我根本不知道下一秒它在哪儿。",
    "P007": "只要它离开物体一次，我就不信了，没有第二次机会。",
    "P008": "误差忽大忽小。固定偏五毫米我能自己在心里补偿，忽大忽小我补偿不了。",
    "P009": "太完美也会让我怀疑。如果它一动不动得不像真的，我会觉得它其实没在跟踪，只是把画面冻住了。",
    "P010": "遮挡之后回来的位置不对，还要自己再挪一次才对上。",
    "P011": "整个不见了应该最不能接受吧，其他的我不太说得清。",
    "P012": "我换个角度看它就偏了。这说明它只知道我看到的样子，不知道东西真正在哪儿。",
    "P013": "长时间不恢复。那次挡板拿开之后它空了好几秒，那几秒我完全不知道该不该继续等它。",
    "P014": "跟不上我的动作。它慢那半拍的时候，我看到的其实是我半秒前的世界，那一刻我完全不知道它现在到底知不知道东西在哪儿——这个比它抖一下更让我不敢用。",
    "P015": "位置整体偏出去。哪怕它很稳，偏了就是错的，稳只会让错误看起来更理直气壮。",
    "P016": "边缘和实物对不齐，尤其是手柄这种我很熟悉形状的东西，差一点点我立刻能看出来。",
    "P017": "闪烁或者忽然变位置。",
    "P018": "视角相关的漂移。这不是精度问题，是它根本没有把物体放在房间坐标系里，我一旦意识到这一点，就不会在需要精度的场合用它。",
    "P019": "被手挡住之后就乱掉。日常拿东西肯定会挡住一部分，如果一挡就不行，那没法用。",
    "P021": "物体被碰动之后它没跟上。那一下之后我就一直在确认它对不对，没法专心做别的。",
    "P020": "反复地小幅修正。每修正一次我就要重新判断一次它现在准不准，很累。",
    "P022": "第一眼就没对齐。第一印象错了，后面它再准我也会持怀疑态度。",
    "P023": "慢慢地往一个方向滑走。这种最阴险，因为你不盯着看不出来，等发现的时候已经偏很多了。",
    "P024": "遮挡之后跳回来的那一下。跳动本身让我不舒服，也让我觉得它刚才是瞎猜的。",
}

INTERVIEW_NOTES = {
    "P003": "A/B 归属回忆检查通过（第二次提示后）；自述“两边差不多”，区分信心低",
    "P009": "自发提出“太稳反而不真实”，与其嵌入可信评分方向一致",
    "P011": "A/B 归属回忆检查首次答错，重新说明后通过；TIA_RC1 两次均选无法回答",
    "P013": "第 5 区块技术事件已如实记录，参与者主动提及该次异常",
    "P014": "唯一在最终二选一中选择 One-Euro 者；明确承认稳定性与遮挡恢复更差，仍以“跟手”为决定性依据，且信任题选“无明显偏好”——偏好与信任分离个例",
    "P018": "自发推测机制（世界坐标固定 / 静止锁定），需求特征风险个例，如实记录",
    "P020": "后两个区块作答明显加快，疲劳可见",
    "P024": "结束不适为中等，休息 10 分钟后缓解；与方法无关",
}


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------


def clip_round(value: float, lo: int, hi: int) -> int:
    """把连续潜变量转成整数刻度分，并夹在量表范围内。"""
    return int(min(hi, max(lo, round(value))))


def fmt_clock(seconds: float) -> str:
    """把从当日 0 点起算的秒数格式化成 HH:MM:SS 字符串（保持 VSCodeSafe 纯文本）。"""
    total = int(round(seconds))
    return f"{total // 3600:02d}:{(total % 3600) // 60:02d}:{total % 60:02d}"


def median_iqr(values: np.ndarray, digits: int = 2) -> str:
    """返回 ``Mdn [Q1, Q3]`` 字符串。"""
    q1, med, q3 = np.percentile(values, [25, 50, 75])
    return f"{med:.{digits}f} [{q1:.{digits}f}, {q3:.{digits}f}]"


def rank_biserial(diff: np.ndarray) -> float:
    """匹配秩双列相关：正负符号秩之和的标准化差。"""
    nonzero = diff[diff != 0]
    if nonzero.size == 0:
        return 0.0
    ranks = stats.rankdata(np.abs(nonzero))
    pos = ranks[nonzero > 0].sum()
    neg = ranks[nonzero < 0].sum()
    return float((pos - neg) / (pos + neg))


def bootstrap_rb_ci(diff: np.ndarray, rng: np.random.Generator, n_boot: int = 10000) -> tuple[float, float]:
    """对匹配秩双列相关做参与者层面的自举置信区间。"""
    n = diff.size
    draws = np.empty(n_boot)
    for i in range(n_boot):
        sample = diff[rng.integers(0, n, n)]
        draws[i] = rank_biserial(sample)
    return float(np.percentile(draws, 2.5)), float(np.percentile(draws, 97.5))


def wilcoxon_pair(oe: np.ndarray, ea: np.ndarray) -> tuple[float, float]:
    """配对 Wilcoxon 符号秩检验，返回 (W, p)。全为结平时返回 (0, 1)。"""
    diff = ea - oe
    if np.all(diff == 0):
        return 0.0, 1.0
    result = stats.wilcoxon(ea, oe, zero_method="wilcox", alternative="two-sided")
    return float(result.statistic), float(result.pvalue)


def holm(pvals: list[float]) -> list[float]:
    """Holm 逐步降低法校正，返回与输入同序的校正后 p 值。"""
    order = np.argsort(pvals)
    m = len(pvals)
    adjusted = [0.0] * m
    running = 0.0
    for rank, idx in enumerate(order):
        value = (m - rank) * pvals[idx]
        running = max(running, value)
        adjusted[idx] = min(1.0, running)
    return adjusted


def cronbach_alpha(matrix: np.ndarray) -> float:
    """Cronbach's α；输入为 参与者 × 条目 矩阵（需完整数据）。"""
    k = matrix.shape[1]
    item_var = matrix.var(axis=0, ddof=1).sum()
    total_var = matrix.sum(axis=1).var(ddof=1)
    if total_var <= 0:
        return float("nan")
    return float(k / (k - 1) * (1 - item_var / total_var))


def mcdonald_omega(matrix: np.ndarray, iterations: int = 60) -> float:
    """McDonald's ω_total，用主轴因子法（PAF）拟合单因子模型近似求解。

    ω = (Σλ)² / [(Σλ)² + Σψ]。PAF 用平方多重相关初始化共同度并迭代，
    与 ML 解略有差异，仅用于模拟演练的量级参考。
    """
    corr = np.corrcoef(matrix, rowvar=False)
    k = corr.shape[0]
    inv = np.linalg.pinv(corr)
    communality = 1.0 - 1.0 / np.diag(inv)
    reduced = corr.copy()
    loadings = np.zeros(k)
    for _ in range(iterations):
        np.fill_diagonal(reduced, communality)
        eigenvalues, eigenvectors = np.linalg.eigh(reduced)
        top = int(np.argmax(eigenvalues))
        lam = max(eigenvalues[top], 1e-9)
        loadings = eigenvectors[:, top] * math.sqrt(lam)
        if loadings.sum() < 0:
            loadings = -loadings
        communality = np.clip(loadings**2, 0.0, 0.995)
    sd = matrix.std(axis=0, ddof=1)
    lam_unstd = loadings * sd
    uniqueness = np.clip(matrix.var(axis=0, ddof=1) - lam_unstd**2, 1e-9, None)
    total = lam_unstd.sum() ** 2
    return float(total / (total + uniqueness.sum()))


def paired_tost(oe: np.ndarray, ea: np.ndarray, bound: float) -> tuple[float, float, float]:
    """配对 TOST 等价检验，返回 (差值均值, 差异检验 p, TOST p)。"""
    diff = ea - oe
    n = diff.size
    mean = float(diff.mean())
    sd = float(diff.std(ddof=1))
    se = sd / math.sqrt(n) if sd > 0 else 1e-12
    df = n - 1
    p_diff = float(stats.ttest_rel(ea, oe).pvalue)
    t_lower = (mean + bound) / se
    t_upper = (mean - bound) / se
    p_lower = float(stats.t.sf(t_lower, df))
    p_upper = float(stats.t.cdf(t_upper, df))
    return mean, p_diff, max(p_lower, p_upper)


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------


def main() -> None:
    rng = np.random.default_rng(SEED)
    shutil.copyfile(SOURCE_WORKBOOK, OUTPUT_WORKBOOK)
    wb = openpyxl.load_workbook(OUTPUT_WORKBOOK)
    ws_part = wb["Participants"]
    ws_rec = wb["Records"]
    ws_ana = wb["Analysis"]

    # --- 读取母版预填的平衡设计（只读，不改写） ---------------------------------
    participants: list[str] = []
    design: dict[str, dict] = {}
    for row in range(PART_FIRST_ROW, PART_LAST_ROW + 1):
        pid = ws_part.cell(row, 1).value
        participants.append(pid)
        design[pid] = {
            "row": row,
            "unit": ws_part.cell(row, 2).value,
            "seq": ws_part.cell(row, 4).value,
            "label_A": ws_part.cell(row, 9).value,
            "label_B": ws_part.cell(row, 10).value,
            "first_method": ws_part.cell(row, 11).value,
            "objects": [ws_part.cell(row, c).value for c in (5, 6, 7)],
        }

    blocks: list[dict] = []
    for row in range(BLOCK_FIRST_ROW, BLOCK_LAST_ROW + 1):
        blocks.append(
            {
                "row": row,
                "pid": ws_rec.cell(row, 1).value,
                "block_index": int(ws_rec.cell(row, 2).value),
                "obj_pos": int(ws_rec.cell(row, 4).value),
                "object_key": ws_rec.cell(row, 6).value,
                "label": ws_rec.cell(row, 7).value,
                "method": ws_rec.cell(row, 8).value,
                "within_order": int(ws_rec.cell(row, 9).value),
            }
        )

    method_rows: list[dict] = []
    for row in range(METHOD_FIRST_ROW, METHOD_LAST_ROW + 1):
        method_rows.append(
            {
                "row": row,
                "pid": ws_rec.cell(row, 1).value,
                "order": int(ws_rec.cell(row, 2).value),
                "label": ws_rec.cell(row, 3).value,
                "method": ws_rec.cell(row, 4).value,
            }
        )

    # --- 逐参与者、逐物体生成区块级评分 --------------------------------------
    # ratings[(pid, method, object_key)][item_id] = 整数分；作废区块记 None
    ratings: dict[tuple[str, str, str], dict[str, int | None]] = {}
    block_meta: dict[tuple[str, str, str], dict] = {}

    for pid in participants:
        persona = PERSONAS[pid]
        # 方法级整体印象（同一方法在三个物体上共享的稳定偏移）
        method_impression = {
            "One-Euro": float(rng.normal(0, 0.20)),
            "EgoAnchor": float(rng.normal(0, 0.20)),
        }
        person_item_bias = {
            item_id: float(rng.normal(0, 0.34)) for item_id, *_ in BLOCK_ITEMS
        }
        # 人×条目随机斜率：除 persona 已写明的辨别力与条目偏好之外，
        # 每人对每个线索的权重仍有残差差异，少数人会把某一条目读成反向。
        # 这是主观量表不会出现全员同向的根本原因，缺少它 r_rb 会被结构性推到 1.0。
        person_item_slope = {
            item_id: float(rng.normal(1.0, PERSON_ITEM_SLOPE_SD))
            for item_id, *_ in BLOCK_ITEMS
        }
        my_blocks = [b for b in blocks if b["pid"] == pid]
        for obj_pos in (1, 2, 3):
            pair = sorted(
                (b for b in my_blocks if b["obj_pos"] == obj_pos),
                key=lambda b: b["within_order"],
            )
            object_key = pair[0]["object_key"]
            tuning = OBJECT_TUNING[object_key]
            # 对象位置带来的辨别力增长：越到后面越知道要看什么
            position_gain = {1: 0.88, 2: 1.00, 3: 1.08}[obj_pos]
            object_gain = tuning["gain"] * persona.object_mult.get(object_key, 1.0)

            latents: dict[str, dict[str, float]] = {}
            for blk in pair:
                method = blk["method"]
                sign = +0.5 if method == "EgoAnchor" else -0.5
                halo = float(rng.normal(0, persona.halo * HALO_SCALE))
                # 分量表共享因子：AQ-EQ / AQ-IQ 各自的条目在本区块内同涨同落
                subscale_factor = {
                    name: float(rng.normal(0, SUBSCALE_FACTOR_SD))
                    for name in ("AQ_EQ", "AQ_IQ")
                }
                # 疲劳只作用于第 5、6 区块
                fatigue = persona.fatigue if blk["block_index"] >= 5 else 0.0
                per_item = {}
                for item_id, _short, base, delta in BLOCK_ITEMS:
                    if item_id == "Q10_OPT":
                        continue
                    mult = persona.item_mult.get(item_id, 1.0)
                    delta_eff = (
                        delta
                        * persona.acuity
                        * object_gain
                        * position_gain
                        * mult
                        * person_item_slope[item_id]
                    )
                    group = BLOCK_SUBSCALE.get(item_id)
                    latent = (
                        base
                        + tuning["shift"]
                        + person_item_bias[item_id]
                        + method_impression[method]
                        + halo
                        + (subscale_factor[group] if group else 0.0)
                        + sign * delta_eff
                        + float(rng.normal(0, persona.noise * ITEM_NOISE_SCALE))
                    )
                    per_item[item_id] = latent
                latents[method] = per_item
                blk["_halo"] = halo
                blk["_fatigue"] = fatigue

            # 同物体内第二个区块存在对比效应：相对第一个方法作判断，差异被放大
            first, second = pair[0], pair[1]
            for item_id in latents[first["method"]]:
                gap = latents[second["method"]][item_id] - latents[first["method"]][item_id]
                latents[second["method"]][item_id] += 0.14 * gap

            for blk in pair:
                method = blk["method"]
                key = (pid, method, object_key)
                event = TECH_EVENTS.get((pid, blk["block_index"]))
                voided = bool(event and event[0])
                penalties = event[2] if event else {}
                style = persona.style_gain * (1.0 - blk["_fatigue"])
                item_scores: dict[str, int | None] = {}
                for item_id, _short, _base, _delta in BLOCK_ITEMS:
                    if item_id == "Q10_OPT":
                        item_scores[item_id] = None  # 预实验默认停用
                        continue
                    if voided:
                        item_scores[item_id] = None
                        continue
                    latent = latents[method][item_id] + penalties.get(item_id, 0.0)
                    shaped = 4.0 + style * (latent - 4.0) + persona.leniency
                    item_scores[item_id] = clip_round(shaped, 1, 7)
                ratings[key] = item_scores
                block_meta[key] = {
                    "block": blk,
                    "voided": voided,
                    "event": event,
                }

    # --- 写入 Records A 段：评分、时长、审计 ------------------------------------
    session_clock: dict[str, float] = {}
    block_audit: dict[tuple[str, str, str], dict] = {}
    for pid in participants:
        session_clock[pid] = 13 * 3600 + 30 * 60 + int(rng.integers(0, 40)) * 60

    for blk in blocks:
        pid = blk["pid"]
        persona = PERSONAS[pid]
        method = blk["method"]
        object_key = blk["object_key"]
        key = (pid, method, object_key)
        meta = block_meta[key]
        scores = ratings[key]
        row = blk["row"]

        # 任务 45–60 s + 区块问卷；问卷时长由人格与区块位置决定
        per_item_s = float(rng.normal(9.0, 1.4)) * (1.35 if persona.style_gain < 0.8 else 1.0)
        per_item_s *= 1.18 if blk["block_index"] == 1 else 1.0
        per_item_s *= 0.86 if blk["block_index"] >= 5 and persona.fatigue > 0.2 else 1.0
        duration = max(70.0, per_item_s * 13 + float(rng.normal(12, 6)))

        task_seconds = float(rng.normal(52, 5))
        start = session_clock[pid] + task_seconds
        end = start + duration
        session_clock[pid] = end + (120 if blk["block_index"] % 2 == 0 else 35)

        if not meta["voided"]:
            for item_id, _short, _base, _delta in BLOCK_ITEMS:
                value = scores[item_id]
                if value is not None:
                    ws_rec.cell(row, BLOCK_ITEM_COL[item_id], value)
            ws_rec.cell(row, 25, fmt_clock(start))
            ws_rec.cell(row, 26, fmt_clock(end))
            ws_rec.cell(row, 27, round(duration, 1))
            ws_rec.cell(row, 28, "是")
            ws_rec.cell(row, 29, "是")
            ws_rec.cell(row, 30, "是")
        else:
            ws_rec.cell(row, 28, "否")
            ws_rec.cell(row, 29, "否")
            ws_rec.cell(row, 30, "否")

        event = meta["event"]
        ws_rec.cell(row, 31, event[1] if event else "无")

        # 运行时审计：两方法共享同一候选流，逐区块独立时间窗口造成微小差异
        tuning = OBJECT_TUNING[object_key]
        degrade = 0.55 if event and "手掌遮挡" in event[1] else 1.0
        cand = float(rng.normal(tuning["cand"], 0.16)) - (0.9 if degrade < 1 else 0.0)
        vcd_med = float(np.clip(rng.normal(tuning["vcd"], 0.018), 0.5, 0.99))
        admit = float(np.clip(rng.normal(0.812, 0.024), 0.5, 0.98))
        out_avail = float(np.clip(rng.normal(0.9957, 0.0016), 0.95, 1.0))

        occl = float(rng.normal(OCCLUSION_TARGET_S, 0.042))
        if event and "挡板移除偏慢" in event[1]:
            occl += 0.30
        if event and "VCD 连续拒绝" in event[1]:
            occl += 0.24
        occl = float(np.clip(occl, 0.45, 1.60))
        if occl >= 1.0:
            state, lost = "Lost", "是"
        elif occl > 0.45:
            state, lost = "FrozenUncertain", "否"
        else:
            state, lost = "Coasting", "否"

        reacquire = 1 if lost == "是" else 0
        static_lock = int(rng.integers(2, 7)) if method == "EgoAnchor" else 0

        if not meta["voided"]:
            ws_rec.cell(row, 32, round(cand, 2))
            ws_rec.cell(row, 33, round(vcd_med, 3))
            ws_rec.cell(row, 34, round(admit, 3))
            ws_rec.cell(row, 35, round(out_avail, 4))
            ws_rec.cell(row, 36, round(occl, 3))
            ws_rec.cell(row, 37, state)
            ws_rec.cell(row, 38, lost)
            ws_rec.cell(row, 39, reacquire)
            ws_rec.cell(row, 40, static_lock)
        ws_rec.cell(
            row,
            41,
            "作废区块，不进入分析" if meta["voided"] else None,
        )

        block_audit[key] = {
            "cand": cand,
            "vcd": vcd_med,
            "admit": admit,
            "avail": out_avail,
            "occl": occl,
            "state": state,
            "lost": lost == "是",
            "duration": duration,
            "voided": meta["voided"],
        }

    # --- 方法级 TiA / S-TIAS ----------------------------------------------------
    # 方法级判断由三个物体的累积体验驱动：先算该方法的区块级综合印象
    block_composite: dict[tuple[str, str], float] = {}
    for pid in participants:
        for method in ("One-Euro", "EgoAnchor"):
            vals = []
            for obj in OBJECT_TUNING:
                key = (pid, method, obj)
                if key in ratings and ratings[key]["Q1"] is not None:
                    vals.extend(
                        v for iid, v in ratings[key].items() if iid != "Q10_OPT" and v is not None
                    )
            block_composite[(pid, method)] = float(np.mean(vals)) if vals else 4.5

    method_scores: dict[tuple[str, str], dict[str, int | None]] = {}
    for mrow in method_rows:
        pid, method, row = mrow["pid"], mrow["method"], mrow["row"]
        persona = PERSONAS[pid]
        carry = block_composite[(pid, method)] - 4.85  # 区块体验对方法级判断的传导
        # 方法级"总体印象"噪声：六个区块结束后凭记忆整体作答，同一方法下各条目共享
        # 该成分。它不随条目数增加而被平均掉，因此分量表均值仍保留场合噪声——
        # 若缺少它，6 项均值的配对差会几乎无残差，产出不真实的 r_rb≈1。
        impression = float(rng.normal(0, 0.46))
        scores: dict[str, int | None] = {}

        for item_id, _sub, reverse, base, delta in TIA_ITEMS:
            missing_mode = TIA_MISSING.get((pid, item_id))
            if missing_mode == "both" or (missing_mode == "first" and mrow["order"] == 1):
                scores[item_id] = None
                continue
            sign = +0.5 if method == "EgoAnchor" else -0.5
            construct = (
                base
                + sign * delta * persona.acuity * persona.item_mult.get(item_id, 1.0)
                + 0.30 * carry
                + impression
                + persona.leniency * 0.55
                + float(rng.normal(0, persona.noise * 0.85))
            )
            style5 = 1.0 + (persona.style_gain - 1.0) * 0.7
            shaped = 3.0 + style5 * (construct - 3.0)
            raw = 6.0 - shaped if reverse else shaped  # 反向项按原始措辞作答
            scores[item_id] = clip_round(raw, 1, 5)

        # S-TIAS 只有 3 项，共享印象项对其均值的影响更大，单独抽取
        stias_impression = float(rng.normal(0, 0.42))
        for item_id, base, delta in STIAS_ITEMS:
            sign = +0.5 if method == "EgoAnchor" else -0.5
            latent = (
                base
                + sign * delta * persona.acuity
                + 0.38 * carry
                + impression
                + stias_impression
                + persona.leniency * 0.7
                + float(rng.normal(0, persona.noise))
            )
            shaped = 4.0 + persona.style_gain * (latent - 4.0)
            scores[item_id] = clip_round(shaped, 1, 7)

        method_scores[(pid, method)] = scores

        for item_id in TIA_COL:
            value = scores[item_id]
            if value is not None:
                ws_rec.cell(row, TIA_COL[item_id], value)
        for item_id in STIAS_COL:
            ws_rec.cell(row, STIAS_COL[item_id], scores[item_id])
        ws_rec.cell(row, 18, "是")

        m_duration = float(np.clip(rng.normal(172, 26), 110, 280))
        m_start = session_clock[pid] + (90 if mrow["order"] == 1 else 25)
        ws_rec.cell(row, 19, fmt_clock(m_start))
        ws_rec.cell(row, 20, fmt_clock(m_start + m_duration))
        ws_rec.cell(row, 21, round(m_duration, 1))
        session_clock[pid] = m_start + m_duration
        ws_rec.cell(row, 22, "无")
        note = None
        if (pid, "TIA_RC1") in TIA_MISSING or (pid, "TIA_RC4") in TIA_MISSING:
            note = "存在“无法回答”条目，按缺失处理"
        elif (pid, "TIA_UP3") in TIA_MISSING and mrow["order"] == 1:
            note = "存在“无法回答”条目，按缺失处理"
        ws_rec.cell(row, 23, note)

    # --- 最终问卷 ---------------------------------------------------------------
    def obj_mean(pid: str, method: str, item_id: str) -> float:
        """某人某方法在三个物体上的条目均值（作废区块自动跳过）。"""
        vals = [
            ratings[(pid, method, obj)][item_id]
            for obj in OBJECT_TUNING
            if ratings[(pid, method, obj)][item_id] is not None
        ]
        return float(np.mean(vals))

    final_records: dict[str, dict] = {}
    for pid in participants:
        persona = PERSONAS[pid]
        d = design[pid]
        stias_diff = np.mean(
            [method_scores[(pid, "EgoAnchor")][i] for i, *_ in STIAS_ITEMS]
        ) - np.mean([method_scores[(pid, "One-Euro")][i] for i, *_ in STIAS_ITEMS])

        def diff(item_id: str) -> float:
            return obj_mean(pid, "EgoAnchor", item_id) - obj_mean(pid, "One-Euro", item_id)

        # 最终二选一由被试自己在意的线索驱动，而非 13 项等权平均。
        # persona.item_mult 表示"这个人多看重这条线索"，取绝对值作权重
        # （符号已经体现在 diff 里），于是响应优先者按响应及时性判断——
        # AQ-IQ2 的真实 δ 为负——少数人因此真的会选 One-Euro。
        def weighted(cues: dict[str, float], extra_w: float, extra_v: float,
                     noise_sd: float) -> float:
            num = den = 0.0
            for item_id, w in cues.items():
                weight = w * abs(persona.item_mult.get(item_id, 1.0))
                num += weight * diff(item_id)
                den += weight
            num += extra_w * extra_v
            den += extra_w
            return (num + float(rng.normal(0, noise_sd))) / max(den, 1e-6)

        pref_score = weighted(
            {"Q1": 0.28, "Q3": 0.28, "Q8": 0.28, "Q6": 1.05, "Q7": 1.05, "AQ_IQ2": 0.80},
            0.55, float(stias_diff), 0.22,
        )
        trust_score = weighted(
            {"Q6": 1.15, "Q3": 0.95, "Q8": 0.75, "AQ_IQ2": 0.45},
            0.85, float(stias_diff), 0.30,
        )

        def to_choice(score: float) -> str:
            if score > persona.decisiveness:
                return "EgoAnchor"
            if score < -persona.decisiveness:
                return "One-Euro"
            return "无明显偏好"

        method_choice = to_choice(pref_score)
        trust_choice = to_choice(trust_score)
        label_of = {d["label_A"]: "方法A", d["label_B"]: "方法B"}

        strength = (
            "NA"
            if method_choice == "无明显偏好"
            else clip_round(2.0 + 2.6 * abs(pref_score) + rng.normal(0, 0.5), 1, 7)
        )
        confidence = clip_round(
            2.2 + 2.4 * persona.acuity + 0.7 * abs(pref_score) + rng.normal(0, 0.55), 1, 7
        )
        discomfort = "中等" if pid == "P024" else ("轻微" if rng.random() < 0.25 else "无")

        row = FINAL_FIRST_ROW + participants.index(pid)
        ws_rec.cell(row, 2, label_of.get(method_choice, "无明显偏好"))
        ws_rec.cell(row, 3, strength)
        ws_rec.cell(row, 4, label_of.get(trust_choice, "无明显偏好"))
        ws_rec.cell(row, 5, confidence)
        ws_rec.cell(row, 6, OPEN_DIFFERENCE[pid])
        ws_rec.cell(row, 7, OPEN_DISTRUST[pid])
        ws_rec.cell(row, 8, discomfort)
        ws_rec.cell(row, 9, method_choice)
        ws_rec.cell(row, 10, trust_choice)
        ws_rec.cell(row, 11, INTERVIEW_NOTES.get(pid))

        final_records[pid] = {
            "method_choice": method_choice,
            "trust_choice": trust_choice,
            "strength": strength,
            "confidence": confidence,
            "discomfort": discomfort,
            "pref_score": float(pref_score),
        }

    # --- Participants 背景与时间 -------------------------------------------------
    for pid in participants:
        persona = PERSONAS[pid]
        row = design[pid]["row"]
        age, gender, hand, vision, vr_exp, mr_exp = persona.background
        ws_part.cell(row, 12, age)
        ws_part.cell(row, 13, gender)
        ws_part.cell(row, 14, hand)
        ws_part.cell(row, 15, vision)
        ws_part.cell(row, 16, vr_exp)
        ws_part.cell(row, 17, mr_exp)
        ws_part.cell(row, 18, "是")
        ws_part.cell(row, 19, "无")
        start = 13 * 3600 + 30 * 60
        ws_part.cell(row, 20, fmt_clock(start))
        ws_part.cell(row, 21, fmt_clock(session_clock[pid] + 420))
        ws_part.cell(row, 22, "是")
        ws_part.cell(
            row,
            23,
            "第 3 区块感知服务崩溃，该区块作废" if pid == "P010" else "无",
        )
        ws_part.cell(row, 24, f"{persona.tag}｜{persona.note}" if persona.note else persona.tag)

    # =======================================================================
    # 分析
    # =======================================================================
    boot_rng = np.random.default_rng(SEED + 1)

    def paired_item(item_id: str) -> tuple[np.ndarray, np.ndarray]:
        oe = np.array([obj_mean(p, "One-Euro", item_id) for p in participants])
        ea = np.array([obj_mean(p, "EgoAnchor", item_id) for p in participants])
        return oe, ea

    # --- A 段：主证实家族 -------------------------------------------------------
    primary = [("Q1", "静止稳定"), ("Q8", "位置正确"), ("Q2", "运动附着"),
               ("Q9", "姿态一致"), ("Q3", "恢复一致"), ("Q6", "依赖意愿"),
               ("Q7", "稳定-响应平衡")]
    primary_stats = []
    for item_id, short in primary:
        oe, ea = paired_item(item_id)
        w, p = wilcoxon_pair(oe, ea)
        diff = ea - oe
        r_rb = rank_biserial(diff)
        lo, hi = bootstrap_rb_ci(diff, boot_rng)
        dz = float(diff.mean() / diff.std(ddof=1)) if diff.std(ddof=1) > 0 else 0.0
        primary_stats.append(
            {"item": item_id, "short": short, "oe": oe, "ea": ea, "W": w, "p": p,
             "r": r_rb, "ci": (lo, hi), "dz": dz,
             "pos": int((diff > 0).sum()), "neg": int((diff < 0).sum()),
             "tie": int((diff == 0).sum())}
        )
    for record, p_adj in zip(primary_stats, holm([s["p"] for s in primary_stats])):
        record["p_holm"] = p_adj

    for offset, record in enumerate(primary_stats):
        row = 5 + offset
        ws_ana.cell(row, 3, median_iqr(record["oe"]))
        ws_ana.cell(row, 4, median_iqr(record["ea"]))
        ws_ana.cell(row, 5, f"{np.median(record['ea'] - record['oe']):+.2f}")
        ws_ana.cell(row, 6, round(record["W"], 1))
        ws_ana.cell(row, 7, f"{record['p']:.2e}" if record["p"] < 1e-3 else round(record["p"], 4))
        ws_ana.cell(row, 8, f"{record['p_holm']:.2e}" if record["p_holm"] < 1e-3 else round(record["p_holm"], 4))
        ws_ana.cell(row, 9, round(record["r"], 3))
        ws_ana.cell(row, 10, f"[{record['ci'][0]:.2f}, {record['ci'][1]:.2f}]")
        verdict = "显著（Holm 后）" if record["p_holm"] < 0.05 else "不显著"
        ws_ana.cell(
            row, 11,
            f"{verdict}｜dz={record['dz']:.2f}｜正/负/平={record['pos']}/{record['neg']}/{record['tie']}",
        )

    # --- B 段：已发表量表家族 ----------------------------------------------------
    def subscale_block(pid: str, method: str, items: list[str]) -> float:
        """AQ 子量表：先在区块内取 3 项均值，再在三个物体上取均值。"""
        vals = []
        for obj in OBJECT_TUNING:
            block = ratings[(pid, method, obj)]
            if block[items[0]] is None:
                continue
            vals.append(float(np.mean([block[i] for i in items])))
        return float(np.mean(vals))

    def tia_subscale(pid: str, method: str, sub: str) -> tuple[float, list[float]]:
        """TiA 分量表：反向项换向为 6−原始分后取有效条目均值。"""
        vals = []
        for item_id, subname, reverse, *_ in TIA_ITEMS:
            if subname != sub:
                continue
            raw = method_scores[(pid, method)][item_id]
            if raw is None:
                continue
            vals.append(6 - raw if reverse else raw)
        return float(np.mean(vals)), vals

    scale_defs = [
        ("AQ-EQ（3 项均值）", lambda p, m: subscale_block(p, m, ["AQ_EQ1", "AQ_EQ2", "AQ_EQ3"])),
        ("AQ-IQ（3 项均值）", lambda p, m: subscale_block(p, m, ["AQ_IQ1", "AQ_IQ2", "AQ_IQ3"])),
        ("TiA-R/C（6 项均值）", lambda p, m: tia_subscale(p, m, "RC")[0]),
        ("TiA-U/P（4 项均值）", lambda p, m: tia_subscale(p, m, "UP")[0]),
        ("S-TIAS（3 项均值）",
         lambda p, m: float(np.mean([method_scores[(p, m)][i] for i, *_ in STIAS_ITEMS]))),
    ]
    scale_stats = []
    for name, fn in scale_defs:
        oe = np.array([fn(p, "One-Euro") for p in participants])
        ea = np.array([fn(p, "EgoAnchor") for p in participants])
        w, p = wilcoxon_pair(oe, ea)
        diff = ea - oe
        dz = float(diff.mean() / diff.std(ddof=1)) if diff.std(ddof=1) > 0 else 0.0
        scale_stats.append({"name": name, "oe": oe, "ea": ea, "W": w, "p": p,
                            "r": rank_biserial(diff), "dz": dz})
    for record, p_adj in zip(scale_stats, holm([s["p"] for s in scale_stats])):
        record["p_holm"] = p_adj

    # 当前样本信度（α / ω）：AQ 用三物体均值后的条目分，TiA 用换向后的原始条目
    def reliability(item_ids: list[str], method: str, level: str) -> tuple[float, float]:
        rows = []
        for pid in participants:
            if level == "block":
                vals = []
                for item_id in item_ids:
                    per_obj = [
                        ratings[(pid, method, obj)][item_id]
                        for obj in OBJECT_TUNING
                        if ratings[(pid, method, obj)][item_id] is not None
                    ]
                    vals.append(float(np.mean(per_obj)))
                rows.append(vals)
            else:
                vals = []
                ok = True
                for item_id in item_ids:
                    raw = method_scores[(pid, method)][item_id]
                    if raw is None:
                        ok = False
                        break
                    reverse = next(r for i, _s, r, *_ in TIA_ITEMS if i == item_id) if item_id.startswith("TIA") else False
                    vals.append(6 - raw if reverse else raw)
                if ok:
                    rows.append(vals)
        matrix = np.array(rows, dtype=float)
        return cronbach_alpha(matrix), mcdonald_omega(matrix)

    rel_specs = [
        (["AQ_EQ1", "AQ_EQ2", "AQ_EQ3"], "block"),
        (["AQ_IQ1", "AQ_IQ2", "AQ_IQ3"], "block"),
        ([i for i, s, *_ in TIA_ITEMS if s == "RC"], "method"),
        ([i for i, s, *_ in TIA_ITEMS if s == "UP"], "method"),
        ([i for i, *_ in STIAS_ITEMS], "method"),
    ]
    for offset, (record, (items, level)) in enumerate(zip(scale_stats, rel_specs)):
        row = 15 + offset
        a_oe, w_oe = reliability(items, "One-Euro", level)
        a_ea, w_ea = reliability(items, "EgoAnchor", level)
        ws_ana.cell(row, 2, median_iqr(record["oe"]))
        ws_ana.cell(row, 3, median_iqr(record["ea"]))
        ws_ana.cell(row, 4, f"{np.median(record['ea'] - record['oe']):+.2f}")
        ws_ana.cell(row, 5, round(record["W"], 1))
        ws_ana.cell(row, 6, f"{record['p']:.2e}" if record["p"] < 1e-3 else round(record["p"], 4))
        ws_ana.cell(row, 7, f"{record['p_holm']:.2e}" if record["p_holm"] < 1e-3 else round(record["p_holm"], 4))
        ws_ana.cell(row, 8, round(record["r"], 3))
        ws_ana.cell(row, 9, round(a_oe, 3))
        ws_ana.cell(row, 10, round(a_ea, 3))
        ws_ana.cell(row, 11, round(w_oe, 3))
        ws_ana.cell(row, 12, round(w_ea, 3))
        ws_ana.cell(
            row, 13,
            ("显著（Holm 后）" if record["p_holm"] < 0.05 else "不显著") + f"｜dz={record['dz']:.2f}",
        )
        record["alpha"] = (a_oe, a_ea)
        record["omega"] = (w_oe, w_ea)

    # --- C 段：次级与稳健性（本次模拟不拟合 CLMM，改报顺序与标签泄漏诊断） -------
    ws_ana.cell(21, 1,
                "C. 次级与稳健性｜本次 AI 演练环境无序数混合模型依赖（statsmodels/ordinal 缺失），"
                "未拟合 CLMM；改以下列顺序效应与标签泄漏诊断替代，正式分析仍须按计划书 6.3 拟合 CLMM")
    ws_ana.cell(22, 1, "诊断项")
    ws_ana.cell(22, 2, "分组 1")
    ws_ana.cell(22, 3, "分组 2")
    ws_ana.cell(22, 4, "统计量")
    ws_ana.cell(22, 5, "p")
    ws_ana.cell(22, 6, "备注")

    diag_rows = []
    # 先行方法效应：以 Q1/Q3/Q6/Q7 的配对差为结局，比较两组先行方法
    key_items = ["Q1", "Q3", "Q6", "Q7"]
    comp_diff = np.array(
        [np.mean([obj_mean(p, "EgoAnchor", i) - obj_mean(p, "One-Euro", i) for i in key_items])
         for p in participants]
    )
    first_ea = np.array([design[p]["first_method"] == "EgoAnchor" for p in participants])
    u, pu = stats.mannwhitneyu(comp_diff[first_ea], comp_diff[~first_ea])
    diag_rows.append(("先行方法效应（核心 4 项配对差）",
                      f"先 EgoAnchor n=12, M={comp_diff[first_ea].mean():.2f}",
                      f"先 One-Euro n=12, M={comp_diff[~first_ea].mean():.2f}",
                      f"U={u:.1f}", f"{pu:.3f}",
                      "p>.05 表示先行方法未系统改变条件差"))
    # 标签映射泄漏：A=EgoAnchor 与 B=EgoAnchor 两组配对差是否不同
    a_is_ea = np.array([design[p]["label_A"] == "EgoAnchor" for p in participants])
    u2, pu2 = stats.mannwhitneyu(comp_diff[a_is_ea], comp_diff[~a_is_ea])
    diag_rows.append(("标签映射泄漏检查",
                      f"A=EgoAnchor n=12, M={comp_diff[a_is_ea].mean():.2f}",
                      f"B=EgoAnchor n=12, M={comp_diff[~a_is_ea].mean():.2f}",
                      f"U={u2:.1f}", f"{pu2:.3f}",
                      "p>.05 表示匿名标签本身未带来偏好"))
    # 对象内先后：同一方法排第 1 与排第 2 时的评分差
    order_first, order_second = [], []
    for blk in blocks:
        key = (blk["pid"], blk["method"], blk["object_key"])
        if ratings[key]["Q1"] is None:
            continue
        composite = float(np.mean([ratings[key][i] for i in key_items]))
        (order_first if blk["within_order"] == 1 else order_second).append(composite)
    u3, pu3 = stats.mannwhitneyu(order_first, order_second)
    diag_rows.append(("对象内先后效应（核心 4 项均值）",
                      f"排第 1 n={len(order_first)}, M={np.mean(order_first):.2f}",
                      f"排第 2 n={len(order_second)}, M={np.mean(order_second):.2f}",
                      f"U={u3:.1f}", f"{pu3:.3f}",
                      "对比效应会略微放大第二个区块的评分离散"))
    # 区块位置趋势
    idx = [blk["block_index"] for blk in blocks
           if ratings[(blk["pid"], blk["method"], blk["object_key"])]["Q1"] is not None]
    comp = [float(np.mean([ratings[(blk["pid"], blk["method"], blk["object_key"])][i] for i in key_items]))
            for blk in blocks
            if ratings[(blk["pid"], blk["method"], blk["object_key"])]["Q1"] is not None]
    rho, prho = stats.spearmanr(idx, comp)
    diag_rows.append(("区块位置趋势（1→6）", f"n={len(idx)}", "—",
                      f"rho={rho:.3f}", f"{prho:.3f}", "检查疲劳或练习导致的整体漂移"))
    # 问卷时长与 AQ 缩减规则
    durations = np.array([block_audit[k]["duration"] for k in block_audit if not block_audit[k]["voided"]])
    over = int((durations > 150).sum())
    diag_rows.append(("区块问卷时长", f"Mdn={np.median(durations):.0f} s",
                      f"P95={np.percentile(durations, 95):.0f} s",
                      f">150 s: {over}/{durations.size}",
                      f"{over / durations.size:.1%}",
                      "AQ 缩减规则触发判据（>150 s 或响应定式）"))

    for offset, values in enumerate(diag_rows):
        for col, value in enumerate(values, start=1):
            ws_ana.cell(23 + offset, col, value)

    # --- D 段：逐物体描述 --------------------------------------------------------
    d_items = [("Q1", None), ("Q2", None), ("Q9", None), ("Q10_OPT", None), ("Q3", None),
               ("Q8", None), ("Q6", None), ("Q7", None), ("AQ-EQ", None), ("AQ-IQ", None)]
    obj_labels = [("鼠标", "blue_mouse"), ("固定订书机", "stapler"), ("游戏手柄", "gamepad")]
    per_object = {}
    row = 42
    for item_id, _ in d_items:
        for label, obj_key in obj_labels:
            if item_id == "Q10_OPT":
                ws_ana.cell(row, 3, "—")
                ws_ana.cell(row, 4, "—")
                ws_ana.cell(row, 7, "预实验默认停用，本次未采集")
                row += 1
                continue
            oe_vals, ea_vals = [], []
            for pid in participants:
                if item_id == "AQ-EQ":
                    block_oe = ratings[(pid, "One-Euro", obj_key)]
                    block_ea = ratings[(pid, "EgoAnchor", obj_key)]
                    if block_oe["AQ_EQ1"] is None or block_ea["AQ_EQ1"] is None:
                        continue
                    oe_vals.append(np.mean([block_oe[i] for i in ("AQ_EQ1", "AQ_EQ2", "AQ_EQ3")]))
                    ea_vals.append(np.mean([block_ea[i] for i in ("AQ_EQ1", "AQ_EQ2", "AQ_EQ3")]))
                elif item_id == "AQ-IQ":
                    block_oe = ratings[(pid, "One-Euro", obj_key)]
                    block_ea = ratings[(pid, "EgoAnchor", obj_key)]
                    if block_oe["AQ_IQ1"] is None or block_ea["AQ_IQ1"] is None:
                        continue
                    oe_vals.append(np.mean([block_oe[i] for i in ("AQ_IQ1", "AQ_IQ2", "AQ_IQ3")]))
                    ea_vals.append(np.mean([block_ea[i] for i in ("AQ_IQ1", "AQ_IQ2", "AQ_IQ3")]))
                else:
                    a = ratings[(pid, "One-Euro", obj_key)][item_id]
                    b = ratings[(pid, "EgoAnchor", obj_key)][item_id]
                    if a is None or b is None:
                        continue
                    oe_vals.append(a)
                    ea_vals.append(b)
            oe_m, ea_m = float(np.mean(oe_vals)), float(np.mean(ea_vals))
            ws_ana.cell(row, 3, round(oe_m, 2))
            ws_ana.cell(row, 4, round(ea_m, 2))
            ws_ana.cell(row, 5, round(ea_m - oe_m, 2))
            ws_ana.cell(row, 6, "EgoAnchor 更高" if ea_m > oe_m else ("One-Euro 更高" if ea_m < oe_m else "持平"))
            ws_ana.cell(row, 7, f"n={len(oe_vals)}")
            per_object[(item_id, label)] = ea_m - oe_m
            row += 1

    # --- E 段：操纵检验（TOST） --------------------------------------------------
    ws_ana.unmerge_cells("A78:S78")  # 母版美化时把“输出可用率”数据行误合并为节标题带

    def audit_by_method(field: str) -> tuple[np.ndarray, np.ndarray]:
        oe, ea = [], []
        for pid in participants:
            for target, bucket in (("One-Euro", oe), ("EgoAnchor", ea)):
                vals = [
                    block_audit[(pid, target, obj)][field]
                    for obj in OBJECT_TUNING
                    if not block_audit[(pid, target, obj)]["voided"]
                ]
                bucket.append(float(np.mean(vals)))
        return np.array(oe), np.array(ea)

    manip_specs = [
        (75, "候选到达率 Hz", "cand", 3),
        (76, "VCD 分数中位", "vcd", 3),
        (77, "VCD 接纳率", "admit", 3),
        (78, "输出可用率", "avail", 4),
        (79, "遮挡时长 s", "occl", 3),
    ]
    manip_results = []
    for row_idx, name, field, digits in manip_specs:
        oe, ea = audit_by_method(field)
        bound = EQUIV_BOUNDS[name]
        mean_diff, p_diff, p_tost = paired_tost(oe, ea, bound)
        ws_ana.cell(row_idx, 1, name)
        ws_ana.cell(row_idx, 2, round(float(oe.mean()), digits))
        ws_ana.cell(row_idx, 3, round(float(ea.mean()), digits))
        ws_ana.cell(row_idx, 4, f"{mean_diff:+.{digits}f}（差异检验 p={p_diff:.3f}）")
        ws_ana.cell(row_idx, 5, f"±{bound}")
        ws_ana.cell(row_idx, 6, f"{p_tost:.2e}" if p_tost < 1e-3 else round(p_tost, 4))
        ws_ana.cell(row_idx, 7, "等价" if p_tost < 0.05 else "未达等价")
        if row_idx == 78:
            ws_ana.cell(row_idx, 8, "母版此行被误合并为节标题带，已在本演练副本中取消合并")
        manip_results.append((name, float(oe.mean()), float(ea.mean()), mean_diff, p_diff, p_tost))

    # 生命周期分布
    lifecycle = {"One-Euro": {}, "EgoAnchor": {}}
    for (pid, method, obj), audit in block_audit.items():
        if audit["voided"]:
            continue
        lifecycle[method][audit["state"]] = lifecycle[method].get(audit["state"], 0) + 1
    for row_idx, state in ((80, "FrozenUncertain"), (81, "Lost")):
        oe_n = lifecycle["One-Euro"].get(state, 0)
        ea_n = lifecycle["EgoAnchor"].get(state, 0)
        oe_total = sum(lifecycle["One-Euro"].values())
        ea_total = sum(lifecycle["EgoAnchor"].values())
        ws_ana.cell(row_idx, 2, f"{oe_n}/{oe_total} = {oe_n / oe_total:.1%}")
        ws_ana.cell(row_idx, 3, f"{ea_n}/{ea_total} = {ea_n / ea_total:.1%}")
        ws_ana.cell(row_idx, 4, f"差 {ea_n - oe_n} 次")
        ws_ana.cell(row_idx, 7, "成立" if (state == "FrozenUncertain" and min(oe_n / oe_total, ea_n / ea_total) > 0.85)
                    or (state == "Lost" and abs(oe_n - ea_n) <= 3) else "需复核")
    reacq = {"One-Euro": 0, "EgoAnchor": 0}
    for (pid, method, obj), audit in block_audit.items():
        if not audit["voided"] and audit["lost"]:
            reacq[method] += 1
    ws_ana.cell(82, 2, reacq["One-Euro"])
    ws_ana.cell(82, 3, reacq["EgoAnchor"])
    ws_ana.cell(82, 4, f"差 {reacq['EgoAnchor'] - reacq['One-Euro']} 次")
    ws_ana.cell(82, 7, "描述性")

    # --- F 段：最终测量计数 ------------------------------------------------------
    def count_choice(field: str) -> dict[str, int]:
        out = {"方法A": 0, "方法B": 0, "无明显偏好": 0, "EgoAnchor": 0, "One-Euro": 0}
        for pid in participants:
            decoded = final_records[pid][field]
            out[decoded] = out.get(decoded, 0) + 1
            if decoded != "无明显偏好":
                label = "方法A" if design[pid]["label_A"] == decoded else "方法B"
                out[label] += 1
        return out

    mc = count_choice("method_choice")
    tc = count_choice("trust_choice")
    ws_ana.cell(86, 2, mc["方法A"])
    ws_ana.cell(86, 3, mc["方法B"])
    ws_ana.cell(86, 4, mc["无明显偏好"])
    ws_ana.cell(86, 5, mc["EgoAnchor"])
    ws_ana.cell(86, 6, mc["One-Euro"])
    ws_ana.cell(86, 7, "标签计数已按参与者映射解码")
    ws_ana.cell(87, 2, tc["方法A"])
    ws_ana.cell(87, 3, tc["方法B"])
    ws_ana.cell(87, 4, tc["无明显偏好"])
    ws_ana.cell(87, 5, tc["EgoAnchor"])
    ws_ana.cell(87, 6, tc["One-Euro"])
    ws_ana.cell(87, 7, "与偏好选择独立作答")

    strengths = [final_records[p]["strength"] for p in participants if final_records[p]["strength"] != "NA"]
    confid = [final_records[p]["confidence"] for p in participants]
    ws_ana.cell(88, 2, f"n={len(strengths)}（做出方法选择者）")
    ws_ana.cell(88, 3, f"Mdn={np.median(strengths):.1f}")
    ws_ana.cell(88, 4, f"分布={dict(sorted({int(v): strengths.count(v) for v in set(strengths)}.items()))}")
    ws_ana.cell(88, 7, "无明显偏好者记 NA，不折算为 1")
    ws_ana.cell(89, 2, f"n={len(confid)}")
    ws_ana.cell(89, 3, f"Mdn={np.median(confid):.1f}")
    ws_ana.cell(89, 4, f"分布={dict(sorted({int(v): confid.count(v) for v in set(confid)}.items()))}")
    ws_ana.cell(89, 7, "低信心提示确实无差异，高信心提示权衡型无偏好")
    mismatch = [p for p in participants
                if final_records[p]["method_choice"] != final_records[p]["trust_choice"]]
    ws_ana.cell(90, 2, len(mismatch))
    ws_ana.cell(90, 4, "、".join(mismatch) if mismatch else "无")
    ws_ana.cell(90, 7, "偏好与信任分离本身是有信息量的结果")
    disc = {}
    for pid in participants:
        disc[final_records[pid]["discomfort"]] = disc.get(final_records[pid]["discomfort"], 0) + 1
    ws_ana.cell(91, 2, str(disc))
    ws_ana.cell(91, 7, "仅安全监测，不作因变量")

    # --- G 段：开放题编码 --------------------------------------------------------
    THEME_COLS = {
        "Stationary_Jitter": 4, "Viewpoint_Drift": 5, "Absolute_Offset": 6, "Motion_Lag": 7,
        "Motion_Sliding": 8, "Orientation_Mismatch": 9, "PostPlacement_Settling": 10,
        "Recovery_Jump": 11, "Wrong_Recovery": 12, "Predictability": 13,
        "Embedding_Blending": 14, "No_Noticeable_Difference": 15, "Other": 16,
    }
    CODING = {
        "P001": ["Viewpoint_Drift", "Stationary_Jitter", "Wrong_Recovery"],
        "P002": ["Stationary_Jitter"],
        "P003": ["No_Noticeable_Difference"],
        "P004": ["Motion_Lag", "Predictability"],
        "P005": ["PostPlacement_Settling"],
        "P006": ["Stationary_Jitter", "Recovery_Jump", "Predictability"],
        "P007": ["Absolute_Offset", "Embedding_Blending"],
        "P008": ["Stationary_Jitter", "Recovery_Jump", "Predictability"],
        "P009": ["Embedding_Blending", "Other"],
        "P010": ["Wrong_Recovery", "Recovery_Jump"],
        "P011": ["No_Noticeable_Difference"],
        "P012": ["Viewpoint_Drift"],
        "P013": ["Wrong_Recovery", "Predictability"],
        "P014": ["Motion_Lag", "Recovery_Jump"],
        "P015": ["Absolute_Offset", "Viewpoint_Drift"],
        "P016": ["Orientation_Mismatch", "Absolute_Offset", "Embedding_Blending"],
        "P017": ["Recovery_Jump"],
        "P018": ["Viewpoint_Drift", "Stationary_Jitter", "Motion_Lag"],
        "P019": ["Motion_Sliding", "Other"],
        "P020": ["Predictability", "Other"],
        "P021": ["Absolute_Offset", "Wrong_Recovery"],
        "P022": ["Absolute_Offset", "Other"],
        "P023": ["Stationary_Jitter", "Motion_Sliding"],
        "P024": ["Orientation_Mismatch", "Recovery_Jump"],
    }
    theme_counts = {name: 0 for name in THEME_COLS}
    for offset, pid in enumerate(participants):
        row_idx = 95 + offset
        ws_ana.cell(row_idx, 2, OPEN_DIFFERENCE[pid][:60])
        ws_ana.cell(row_idx, 3, OPEN_DISTRUST[pid][:60])
        for theme, col in THEME_COLS.items():
            hit = 1 if theme in CODING[pid] else 0
            ws_ana.cell(row_idx, col, hit)
            theme_counts[theme] += hit
        ws_ana.cell(row_idx, 17, "SIM-C1")
        ws_ana.cell(row_idx, 18, "SIM-C2")
        ws_ana.cell(row_idx, 19, "模拟数据由单一编码方案生成，无真实分歧")

    # --- 首屏合成声明与效应锚定表 -------------------------------------------------
    ws_warn = wb.create_sheet("00_SIM_WARNING", 0)
    warn_lines = [
        ["【AI 合成演练数据 · 严禁作为真实实验结果】"],
        [f"生成模型：{MODEL_NAME}；生成脚本：simulate_exp3_claude_opus_5_1m.py；随机种子：{SEED}（重复运行全部单元格取值一致；xlsx 哈希因创建时间元数据而变）"],
        ["本工作簿的每一个应答、时长与审计数值都是模拟生成的，不来自任何真实参与者。"],
        ["禁止：写入论文结果 / 效应量 / 功效或样本量推断；禁止与真实采集数据合并；禁止作为预期结果引用。"],
        ["允许：检验 v5.1 冻结设计的灵敏度、条目区分度、缺失与作废处理、分析链路是否跑得通。"],
        [""],
        ["结构来源：EgoAnchor_Experiment3_DataCollection_24P_v5_1_Beautified_Checked_VSCodeSafe.xlsx（原样复制后填写）"],
        ["未改动：README / Verification_Audit / Questionnaire 三表，以及 Participants 与 Records 的预填平衡设计列。"],
        ["已改动：Participants 背景与时间列、Records 全部应答与审计列、Analysis 全部结果列。"],
        ["已修复：Analysis!A78:S78（母版把“输出可用率”数据行误合并为节标题带）在本副本中取消合并。"],
        [""],
        ["本次演练与设计的关键差异（正式采集前需确认）"],
        ["Q10_OPT 默认停用，全部留空；TiA 的“无法回答”按缺失写空单元格，未写入任何占位符。"],
        ["TiA 反向项列（*_REV）填写的是原始作答分，反向计分 6−原始分只在分析中派生，未回写采集列。"],
        ["P010 第 3 区块按技术作废规则整块留空，三物体均值退化为两物体均值；另有 4 个区块记录了非作废技术事件。"],
        [""],
        ["工作表导览"],
        ["00_SIM_WARNING", "本页：合成声明与导览"],
        ["01_SIM_EFFECT_ANCHOR", "效应量锚定表：每个条目的 δ 从实验一/二哪项实测差异推出"],
        ["02_SIM_PERSONA", "24 名虚拟参与者的人格参数与预设技术事件"],
        ["Records", "A 段 144 区块 / B 段 48 方法级 / C 段 24 最终问卷，均已填满"],
        ["Analysis", "主证实家族、已发表量表家族、顺序诊断、逐物体、操纵检验、最终计数、开放题编码"],
    ]
    for offset, line in enumerate(warn_lines, start=1):
        for col, value in enumerate(line, start=1):
            ws_warn.cell(offset, col, value)

    ws_eff = wb.create_sheet("01_SIM_EFFECT_ANCHOR", 1)
    eff_header = ["条目", "简称", "两方法均值基线", "δ = EgoAnchor − One-Euro（潜在七点尺）",
                  "锚定的实验一/二实测差异", "预期"]
    anchor_notes = {
        "Q1": ("头动泄漏 P95 10.65→0.82 mm（13.0×）；静止帧间增量 0.85→0.06 mm（13.3×）", "强正向"),
        "Q2": ("平移对齐 RMSE 15.69→9.12 mm（正向）对 Start-transition 334→510 ms（反向）", "预期近零"),
        "Q9": ("旋转对齐 RMSE 5.55→4.81°（仅 1.15×）；静止旋转泄漏 3.29→0.33° 不在本条目问域内", "弱正向/临界"),
        "Q10_OPT": ("前向过冲 0.485→0.217 mm 对反向回动 0.497→0.695 mm", "默认停用"),
        "AQ_IQ2": ("Start-transition 334→510 ms；有效时延 380→360 ms", "预期轻微反向"),
        "AQ_IQ3": ("静止帧间增量 13.3×；Linear/SLERP 显示步长连续性", "强正向"),
        "Q3": ("遮挡平移 P95 10.41→4.85 mm（2.1×）；灾难性失效 >40 mm 1→0 次", "强正向"),
        "Q8": ("绝对注册 P95 14.00→6.60 mm（2.1×）", "中强正向"),
        "AQ_EQ1": ("综合配准质量", "中正向"),
        "AQ_EQ2": ("综合配准质量（接管原 Q4 合理可信构念）", "中正向"),
        "AQ_EQ3": ("尺度不随方法变化", "天花板、预期零"),
        "AQ_IQ1": ("交互方式两方法完全相同，仅受整体印象牵引", "弱正向"),
        "Q6": ("应用侧结局；灾难性失效 1→0，静止/遮挡通道全面占优", "强正向"),
        "Q7": ("稳定性收益与时延代价的整体权衡", "中正向"),
    }
    ws_eff.append(eff_header)
    for item_id, short, base, delta in BLOCK_ITEMS:
        note, expect = anchor_notes[item_id]
        ws_eff.append([item_id, short,
                       "—" if item_id == "Q10_OPT" else base,
                       "—" if item_id == "Q10_OPT" else delta, note, expect])
    ws_eff.append([])
    ws_eff.append(["方法级", "", "", "", "", ""])
    for item_id, sub, reverse, base, delta in TIA_ITEMS:
        ws_eff.append([item_id, f"TiA-{sub}" + ("（反向项）" if reverse else ""), base, delta,
                       "累积可靠性/可预测性判断，由三物体区块体验传导（传导系数 0.30）"
                       "＋同方法各条目共享的总体印象项（SD 0.46）", ""])
    for item_id, base, delta in STIAS_ITEMS:
        ws_eff.append([item_id, "S-TIAS", base, delta,
                       "总体信任，由区块体验传导（传导系数 0.38）"
                       "＋共享总体印象项（SD 0.46）与 S-TIAS 专属印象项（SD 0.42）", ""])
    ws_eff.append([])
    ws_eff.append(["物体调制", "增益（放大条件效应）", "共同质量偏移", "候选到达率 Hz", "VCD 中位"])
    for obj_key, tuning in OBJECT_TUNING.items():
        ws_eff.append([obj_key, tuning["gain"], tuning["shift"], tuning["cand"], tuning["vcd"]])
    ws_eff.append([])
    ws_eff.append(["方差成分（决定统计量是否真实，与 δ 同等重要）", "取值", "作用层级", "缺失后果", "", ""])
    for name, value, level, consequence in [
        ("条目层随机波动 ITEM_NOISE_SCALE", ITEM_NOISE_SCALE, "区块×条目",
         "过大则相邻配对与三物体求均值的设计优势被抹掉；过小则效应虚高"),
        ("区块整体印象 HALO_SCALE", HALO_SCALE, "区块（13 项共享）",
         "同区块条目相关性消失，不在配对差中抵消"),
        ("人×条目随机斜率 PERSON_ITEM_SLOPE_SD", PERSON_ITEM_SLOPE_SD, "人×条目",
         "缺失时三物体求均值使配对符号完全由恒正 δ 决定，r_rb 被结构性推到 1.0"),
        ("分量表共享因子 SUBSCALE_FACTOR_SD", SUBSCALE_FACTOR_SD, "区块×分量表",
         "缺失时 AQ-EQ/AQ-IQ 条目只靠 halo 相连，α 被人为压到 0.2–0.4"),
        ("人×条目稳定偏好 person_item_bias", 0.34, "人×条目（跨方法恒定）",
         "缺失时各人条目难度完全一致，条目间差异只剩噪声"),
        ("方法级总体印象 impression", 0.46, "人×方法（10 项共享）",
         "缺失时 6 项均值几乎无残差，TiA 配对差产出不真实的 r_rb≈1"),
        ("同物体内对比效应", 0.14, "同物体第二个区块",
         "相邻 A/B 的对比判断被低估"),
    ]:
        ws_eff.append([name, value, level, consequence, "", ""])

    ws_persona = wb.create_sheet("02_SIM_PERSONA", 2)
    ws_persona.append(["Participant_ID", "人格标签", "感知敏锐度", "宽严倾向", "量表风格",
                       "噪声", "疲劳", "判定阈值", "条目特异乘子", "物体特异乘子", "备注"])
    for pid in participants:
        persona = PERSONAS[pid]
        ws_persona.append([
            pid, persona.tag, persona.acuity, persona.leniency, persona.style_gain,
            persona.noise, persona.fatigue, persona.decisiveness,
            json.dumps(persona.item_mult, ensure_ascii=False) if persona.item_mult else "—",
            json.dumps(persona.object_mult, ensure_ascii=False) if persona.object_mult else "—",
            persona.note or "—",
        ])
    ws_persona.append([])
    ws_persona.append(["预设技术事件", "区块", "整块作废", "说明"])
    for (pid, blk_idx), (voided, desc, _pen) in TECH_EVENTS.items():
        ws_persona.append([pid, blk_idx, "是" if voided else "否", desc])
    ws_persona.append([])
    ws_persona.append(["TiA 无法回答（记缺失）", "条目", "范围"])
    for (pid, item_id), scope in TIA_MISSING.items():
        ws_persona.append([pid, item_id, "两次施测" if scope == "both" else "首次施测"])

    # README 首屏补合成标记与禁用声明
    ws_readme = wb["README"]
    ws_readme.cell(1, 1, "【AI 合成演练副本 · SYNTHETIC · 非正式采集表】" + str(ws_readme.cell(1, 1).value))
    ws_readme.cell(
        2, 1,
        "【本工作簿全部数据由 AI 模型合成，不得进入论文结果、效应量、功效或样本量推断，"
        "也不得与真实参与者数据合并】仅用于采集表与分析流程演练。"
        "唯一权威规格：2026-EgoAnchor/experiment_3_questionnaire_design_zh.md（v5.1）。"
        "本副本保留母版 6 表结构：问卷条目集中在 Questionnaire，实验记录集中在 Records（三段堆叠）；"
        "另加 00_SIM_WARNING / 01_SIM_EFFECT_ANCHOR / 02_SIM_PERSONA 三张演练说明表。"
        "为适配 VS Code Office Viewer，本副本已移除数据验证、条件格式与冻结窗格"
        "（这些仅是空表输入辅助，对已填数据无用），统计约束由生成与验证代码执行。"
        "采集表中不派生复合分；原始评分未被反向分覆盖，TiA 反向计分（6−原始分）仅在分析代码中生成。",
    )

    # 移除继承自母版的数据验证与条件格式：本副本已填满，输入辅助无意义，
    # 且 AGENTS.md 要求演练工作簿不依赖 Excel 的这些特性。
    for ws in wb.worksheets:
        ws.data_validations.dataValidation = []
        ws.conditional_formatting = ConditionalFormattingList()
        ws.freeze_panes = None

    wb.save(OUTPUT_WORKBOOK)

    # =======================================================================
    # 自检与摘要（写 JSON 供报告核对，不作为交付物）
    # =======================================================================
    summary = {
        "model": MODEL_NAME,
        "seed": SEED,
        "workbook": OUTPUT_WORKBOOK.name,
        "primary": [
            {
                "item": s["item"], "short": s["short"],
                "oe": median_iqr(s["oe"]), "ea": median_iqr(s["ea"]),
                "median_diff": float(np.median(s["ea"] - s["oe"])),
                "W": s["W"], "p": s["p"], "p_holm": s["p_holm"],
                "r": s["r"], "ci": s["ci"], "dz": s["dz"],
                "pos_neg_tie": [s["pos"], s["neg"], s["tie"]],
            }
            for s in primary_stats
        ],
        "scales": [
            {
                "name": s["name"], "oe": median_iqr(s["oe"]), "ea": median_iqr(s["ea"]),
                "median_diff": float(np.median(s["ea"] - s["oe"])),
                "p": s["p"], "p_holm": s["p_holm"], "r": s["r"], "dz": s["dz"],
                "alpha": s["alpha"], "omega": s["omega"],
            }
            for s in scale_stats
        ],
        "per_object": {f"{k[0]}|{k[1]}": round(v, 3) for k, v in per_object.items()},
        "manipulation": manip_results,
        "lifecycle": lifecycle,
        "final": {
            "method_choice": mc, "trust_choice": tc,
            "mismatch": mismatch,
            "strength_median": float(np.median(strengths)),
            "confidence_median": float(np.median(confid)),
            "discomfort": disc,
        },
        "themes": theme_counts,
        "durations": {
            "median": float(np.median(durations)),
            "p95": float(np.percentile(durations, 95)),
            "over150": over,
            "n": int(durations.size),
        },
        "diagnostics": diag_rows,
        "per_participant": {
            pid: {
                "persona": PERSONAS[pid].tag,
                "choice": final_records[pid]["method_choice"],
                "trust": final_records[pid]["trust_choice"],
                "strength": final_records[pid]["strength"],
                "confidence": final_records[pid]["confidence"],
                "pref_score": round(final_records[pid]["pref_score"], 3),
                "diff_Q1": round(obj_mean(pid, "EgoAnchor", "Q1") - obj_mean(pid, "One-Euro", "Q1"), 2),
                "diff_Q3": round(obj_mean(pid, "EgoAnchor", "Q3") - obj_mean(pid, "One-Euro", "Q3"), 2),
                "diff_Q6": round(obj_mean(pid, "EgoAnchor", "Q6") - obj_mean(pid, "One-Euro", "Q6"), 2),
                "diff_Q7": round(obj_mean(pid, "EgoAnchor", "Q7") - obj_mean(pid, "One-Euro", "Q7"), 2),
                "diff_AQIQ2": round(obj_mean(pid, "EgoAnchor", "AQ_IQ2") - obj_mean(pid, "One-Euro", "AQ_IQ2"), 2),
                "diff_AQEQ2": round(obj_mean(pid, "EgoAnchor", "AQ_EQ2") - obj_mean(pid, "One-Euro", "AQ_EQ2"), 2),
            }
            for pid in participants
        },
    }

    # N=18 稳健性：按 Participants 顺序抽掉最后 6 人（相当于提前止收整组）
    subset = participants[:18]
    robust = []
    for item_id, short in primary:
        oe = np.array([obj_mean(p, "One-Euro", item_id) for p in subset])
        ea = np.array([obj_mean(p, "EgoAnchor", item_id) for p in subset])
        _w, p = wilcoxon_pair(oe, ea)
        robust.append({"item": item_id, "p": p})
    for record, p_adj in zip(robust, holm([r["p"] for r in robust])):
        record["p_holm"] = p_adj
    summary["n18"] = robust

    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"已生成 {OUTPUT_WORKBOOK.name}")
    print("主证实家族（Holm 后）：")
    for s in primary_stats:
        mark = "*" if s["p_holm"] < 0.05 else " "
        print(f"  {mark} {s['item']:<8}{s['short']:<12} Δ={np.median(s['ea']-s['oe']):+.2f} "
              f"dz={s['dz']:+.2f} p_Holm={s['p_holm']:.4g} r={s['r']:+.2f}")
    print("已发表量表家族（Holm 后）：")
    for s in scale_stats:
        mark = "*" if s["p_holm"] < 0.05 else " "
        print(f"  {mark} {s['name']:<20} Δ={np.median(s['ea']-s['oe']):+.2f} "
              f"dz={s['dz']:+.2f} p_Holm={s['p_holm']:.4g} α={s['alpha'][0]:.2f}/{s['alpha'][1]:.2f}")
    print(f"最终选择：{mc}")
    print(f"信任选择：{tc}")
    print(f"问卷时长 Mdn={np.median(durations):.0f}s，>150s 的区块 {over}/{durations.size}")


if __name__ == "__main__":
    main()
