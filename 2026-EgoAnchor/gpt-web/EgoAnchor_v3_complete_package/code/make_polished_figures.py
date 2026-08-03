from openpyxl import load_workbook
from pathlib import Path
from collections import defaultdict
import math, re, json
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import patches
from matplotlib.lines import Line2D
from matplotlib.font_manager import FontProperties
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / 'data'
OUT = ROOT / 'figures'
OUT.mkdir(parents=True, exist_ok=True)

# ---------- visual system ----------
COLORS = {
    'Arrival-Hold':'#4C78A8',
    'Capture-Hold':'#F28E2B',
    'One-Euro Anchor':'#59A14F',
    'EgoAnchor':'#E15759',
    'grey':'#8A8F98',
    'light':'#E9EBEF',
    'dark':'#2A2F35',
    'teal':'#2A9D8F',
    'purple':'#8E63CE',
}
METHODS = ['Arrival-Hold','Capture-Hold','One-Euro Anchor','EgoAnchor']
METHOD_LABELS = ['Arrival','Capture','One-Euro','EgoAnchor']

plt.rcParams.update({
    'font.family': 'DejaVu Sans',
    'font.size': 7.5,
    'axes.titlesize': 8.5,
    'axes.labelsize': 7.5,
    'xtick.labelsize': 6.7,
    'ytick.labelsize': 6.7,
    'legend.fontsize': 6.7,
    'axes.linewidth': 0.7,
    'pdf.fonttype': 42,
    'ps.fonttype': 42,
    'savefig.transparent': False,
})

def clean_ax(ax, grid='y'):
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.tick_params(length=2.5, width=0.6, pad=1.5)
    if grid:
        ax.grid(axis=grid, color='#D9DDE3', linewidth=0.55, alpha=0.8, zorder=0)
    ax.set_axisbelow(True)

def q(vals, p):
    vals = sorted(float(v) for v in vals if v is not None)
    if not vals: return float('nan')
    pos = (len(vals)-1)*p
    lo, hi = int(math.floor(pos)), int(math.ceil(pos))
    if lo == hi: return vals[lo]
    return vals[lo] + (vals[hi]-vals[lo])*(pos-lo)

def save(fig, stem):
    fig.savefig(OUT/f'{stem}.pdf', bbox_inches='tight', pad_inches=0.02)
    fig.savefig(OUT/f'{stem}.png', dpi=320, bbox_inches='tight', pad_inches=0.02)
    plt.close(fig)

# ---------- data ----------
def sheet_values(path, sheet_name, min_row, max_row, min_col, max_col):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    values = [list(row) for row in ws.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col,
        values_only=True,
    )]
    wb.close()
    return values

f2 = sheet_values(DATA / 'figure_plot_data.xlsx', 'Figure2', 1, 417, 1, 11)
f3 = sheet_values(DATA / 'figure_plot_data.xlsx', 'Figure3', 1, 460, 1, 11)
h2, h3 = f2[0], f3[0]
rows2 = [dict(zip(h2,r)) for r in f2[1:] if any(x is not None for x in r)]
rows3 = [dict(zip(h3,r)) for r in f3[1:] if any(x is not None for x in r)]

# ---------- teaser crop ----------
src = Image.open(DATA / 'replay_grid_source.jpeg').convert('RGB')
# Source grid geometry from replay_grid metadata.
origin_x, origin_y = 251, 122
cell_w, cell_h, gx, gy = 320, 213, 4, 4
row_idx = [1,2,4,5]  # reference, arrival, one-euro, ours
row_labels = ['Quest Reference','Arrival','One-Euro','EgoAnchor']
row_colors = ['#2D2D2D', COLORS['Arrival-Hold'], COLORS['One-Euro Anchor'], COLORS['EgoAnchor']]
col_idx = [0,3,5]
col_labels = ['0.00 s','3.75 s','6.25 s']
scale = 1.0
label_w, top_h, pad = 285, 74, 12
out_w = label_w + len(col_idx)*cell_w + (len(col_idx)-1)*8 + pad*2
out_h = top_h + len(row_idx)*cell_h + (len(row_idx)-1)*5 + pad
canvas = Image.new('RGB',(out_w,out_h),'white')
d = ImageDraw.Draw(canvas)
font_candidates = [
    ('/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc', '/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc'),
    ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
]
font_path, font_bold_path = next(((a, b) for a, b in font_candidates if Path(a).exists() and Path(b).exists()), font_candidates[-1])
font_row = ImageFont.truetype(font_bold_path, 27)
font_time = ImageFont.truetype(font_path, 28)
font_tag = ImageFont.truetype(font_bold_path, 25)
# header
for ci,lab in enumerate(col_labels):
    x = label_w + pad + ci*(cell_w+8)
    bbox=d.textbbox((0,0), lab, font=font_time)
    d.text((x+(cell_w-(bbox[2]-bbox[0]))/2, 20), lab, fill='#2A2F35', font=font_time)
# row cells
for ri,(ridx,rlab,rcol) in enumerate(zip(row_idx,row_labels,row_colors)):
    y = top_h + ri*(cell_h+5)
    # label and color key
    d.rounded_rectangle((14,y+cell_h//2-19,42,y+cell_h//2+19), radius=7, fill=rcol)
    d.text((52,y+cell_h//2-20), rlab, fill='#222222', font=font_row)
    for ci,cidx in enumerate(col_idx):
        x0=origin_x+cidx*(cell_w+gx); y0=origin_y+ridx*(cell_h+gy)
        crop=src.crop((x0,y0,x0+cell_w,y0+cell_h))
        x=label_w+pad+ci*(cell_w+8)
        canvas.paste(crop,(x,y))
        # subtle colored border
        d.rectangle((x,y,x+cell_w-1,y+cell_h-1), outline=rcol, width=3)
# key result badges
badge_y = out_h-2
canvas.save(OUT/'teaser_sequence.png', quality=95)
# PDF via matplotlib to preserve exact dimensions
fig,ax=plt.subplots(figsize=(11.4,4.0))
ax.imshow(canvas)
ax.axis('off')
fig.subplots_adjust(0,0,1,1)
save(fig,'teaser_sequence')

# ---------- system overview diagram ----------
fig, ax = plt.subplots(figsize=(11.7,3.25))
ax.set_xlim(0, 100); ax.set_ylim(0, 30); ax.axis('off')

def stage_box(x,y,w,h,title,lines,face,edge):
    r=patches.FancyBboxPatch((x,y),w,h,boxstyle='round,pad=0.28,rounding_size=.8',facecolor=face,edgecolor=edge,linewidth=1.0)
    ax.add_patch(r)
    ax.text(x+w/2,y+h-1.3,title,ha='center',va='top',fontsize=8.0,fontweight='bold',color='#242A30')
    ax.text(x+1.0,y+h-3.5,'\n'.join(lines),ha='left',va='top',fontsize=6.25,color='#4E5660',linespacing=1.28)
    return r

def arrow(x1,y1,x2,y2,color='#707781',lw=1.1):
    ax.annotate('',xy=(x2,y2),xytext=(x1,y1),arrowprops=dict(arrowstyle='-|>',lw=lw,color=color,shrinkA=1,shrinkB=1))

ax.text(2,28.4,'EgoAnchor system overview',fontsize=10.6,fontweight='bold',color=COLORS['dark'])
ax.text(2,26.4,'A zero-shot perception backend emits asynchronous pose observations; an anchor runtime turns them into render-rate world anchors.',fontsize=7.0,color='#59616B')

# layer labels
ax.add_patch(patches.FancyBboxPatch((16.2,22.7),29.0,2.0,boxstyle='round,pad=.15,rounding_size=.6',facecolor='#E9F1FA',edgecolor='none'))
ax.text(30.7,23.7,'PERCEPTION BACKEND',ha='center',va='center',fontsize=7.0,fontweight='bold',color='#496D9B')
ax.add_patch(patches.FancyBboxPatch((47.7,22.7),48.0,2.0,boxstyle='round,pad=.15,rounding_size=.6',facecolor='#FDEDEA',edgecolor='none'))
ax.text(71.7,23.7,'ANCHOR RUNTIME',ha='center',va='center',fontsize=7.0,fontweight='bold',color='#B94B4F')

stage_box(2,12.4,11.2,9.1,'Inputs',['Quest stereo','HMD pose history','Text + 3D mesh'],'#F4F7FA','#8CA7C6')
stage_box(16.2,12.4,12.5,9.1,'Object perception',['open-vocabulary init','temporal mask','metric stereo depth'],'#F4F7FA','#8CA7C6')
stage_box(31.0,12.4,14.2,9.1,'6DoF + VCD',['model-based pose','visibility gate','color-depth check'],'#F7F2FB',COLORS['purple'])
stage_box(47.7,12.4,14.0,9.1,'Temporal reconcile',['frame-ID lookup','capture-time world','registration'],'#FFF5F3',COLORS['EgoAnchor'])
stage_box(64.0,12.4,14.0,9.1,'State-aware anchor',['moving: interpolate','stationary: lock','occluded: hold'],'#FFF5F3',COLORS['EgoAnchor'])
stage_box(80.3,12.4,15.4,9.1,'Application output',['world pose @ render rate','Tracked / Stationary','Frozen / Lost'],'#FFF5F3',COLORS['EgoAnchor'])
for a,b in [(13.2,16.2),(28.7,31.0),(45.2,47.7),(61.7,64.0),(78.0,80.3)]: arrow(a,16.8,b,16.8)
ax.text(46.4,15.0,'{frame ID, t$_{cap}$, T$_o^c$, R}',ha='center',fontsize=6.2,color='#555C66',family='monospace')
ax.text(79.2,15.0,'pose + lifecycle state',ha='center',fontsize=6.2,color='#555C66')

# timeline band
ax.add_patch(patches.FancyBboxPatch((2,2.0),93.7,7.2,boxstyle='round,pad=.30,rounding_size=.8',facecolor='#FAFAFB',edgecolor='#D1D5DB',linewidth=.8))
ax.text(4,8.0,'Capture-time semantics',fontsize=7.5,fontweight='bold',color=COLORS['dark'])
ax.plot([18,91],[5.1,5.1],color='#888F98',lw=1.1)
for x,lab,col in [(26,'capture $t_f$',COLORS['Capture-Hold']),(53,'arrival $t_a$',COLORS['Arrival-Hold']),(84,'render $t_r$',COLORS['EgoAnchor'])]:
    ax.plot([x,x],[4.35,5.85],color=col,lw=2)
    ax.text(x,3.25,lab,ha='center',fontsize=6.5,color=col,fontweight='bold')
ax.annotate('retrieve HMD pose at $t_f$',xy=(26,5.1),xytext=(36,7.05),ha='center',fontsize=6.3,color='#4B5159',arrowprops=dict(arrowstyle='->',color='#6F7680',lw=.75))
ax.annotate('avoid composition at $t_a$',xy=(53,5.1),xytext=(64,7.05),ha='center',fontsize=6.3,color='#4B5159',arrowprops=dict(arrowstyle='->',color='#6F7680',lw=.75))
ax.text(79,2.6,'low-rate observations  ->  60 Hz anchor output',fontsize=6.3,color='#4B5159',ha='center')
save(fig,'system_overview')

# ---------- Experiment 1: one-axis, shared visual grammar ----------
configs = [
    ('(a) Static translation','centered_p95_mm','frame_increment_p95_mm','mm','Static translation'),
    ('(b) Static rotation','centered_rotation_p95_deg','frame_rotation_increment_p95_deg','deg','Static rotation'),
    ('(c) Dynamic translation','aligned_rmse_mm','aligned_residual_increment_p95_mm','mm','Dynamic translation'),
    ('(d) Dynamic rotation','aligned_rmse_deg','aligned_residual_increment_p95_deg','deg','Dynamic rotation'),
]
fig, axes = plt.subplots(1,4,figsize=(11.6,3.25))
for ax,(panel,metric_e,metric_j,unit,title) in zip(axes,configs):
    for i,m in enumerate(METHODS):
        ev=[float(r['y_value']) for r in rows2 if r['panel']==panel and r['variant_id']==m and r['y_metric']==metric_e]
        jv=[float(r['y_value']) for r in rows2 if r['panel']==panel and r['variant_id']==m and r['y_metric']==metric_j]
        # deterministic spread
        for vals,offset,marker,fill in [(ev,-0.12,'o',True),(jv,0.12,'D',False)]:
            n=len(vals)
            offs=[0] if n<=1 else [(-0.07+0.14*k/(n-1)) for k in range(n)]
            xs=[i+offset+o for o in offs]
            if fill:
                ax.scatter(xs,vals,s=10,color=COLORS[m],alpha=.18,linewidths=0,zorder=2)
            else:
                ax.scatter(xs,vals,s=11,facecolors='none',edgecolors=COLORS[m],alpha=.20,linewidths=.7,zorder=2,marker=marker)
        for vals,offset,marker,fill in [(ev,-0.12,'o',True),(jv,0.12,'D',False)]:
            med=q(vals,.5); lo=q(vals,.25); hi=q(vals,.75)
            ax.errorbar([i+offset],[med],yerr=[[med-lo],[hi-med]],fmt=marker,ms=4.2,capthick=.8,capsize=2.5,elinewidth=1.2,
                        color=COLORS[m],mfc=COLORS[m] if fill else 'white',mec=COLORS[m],zorder=4)
    ax.set_title(title,pad=3)
    ax.set_yscale('log')
    ax.set_ylabel(unit)
    ax.set_xticks(range(4),METHOD_LABELS,rotation=24,ha='right')
    clean_ax(ax,'y')
    ax.tick_params(axis='x',length=0)
# shared legend
handles=[Line2D([0],[0],marker='o',color='#333',linestyle='none',markersize=4,label='Error / aligned RMSE'),
         Line2D([0],[0],marker='D',markerfacecolor='white',markeredgecolor='#333',color='#333',linestyle='none',markersize=4,label='Residual jitter')]
fig.legend(handles=handles,loc='upper center',bbox_to_anchor=(0.5,1.04),ncol=2,frameon=False)
fig.subplots_adjust(left=.055,right=.995,bottom=.25,top=.79,wspace=.32)
save(fig,'exp1_behavior')

# ---------- Experiment 2 ----------
fig, axes = plt.subplots(1,4,figsize=(11.6,3.15),gridspec_kw={'width_ratios':[1,1,1.15,1.55]})
# paired helper
for ax,panel,order,title,ylabel in [
    (axes[0],'(a) Capture-time alignment',['Capture time','Arrival time'],'Capture-time alignment','P95 error (mm)'),
    (axes[1],'(b) StaticLock',['EgoAnchor','EgoAnchor w/o StaticLock'],'StaticLock','Centered P95 (mm)')]:
    grouped=defaultdict(dict)
    for r in rows3:
        if r['panel']!=panel: continue
        key=(r['session_id'],r['trial_id'],r['segment_id'])
        grouped[key][r['series']]=float(r['y_value'])
    for vals in grouped.values():
        if all(s in vals for s in order):
            ax.plot([0,1],[vals[order[0]],vals[order[1]]],color='#B7BBC2',lw=.75,alpha=.75,zorder=1)
            ax.scatter([0,1],[vals[order[0]],vals[order[1]]],s=9,color='#B7BBC2',alpha=.8,zorder=2)
    meds=[]; los=[]; his=[]
    for s in order:
        vv=[float(r['y_value']) for r in rows3 if r['panel']==panel and r['series']==s]
        med=q(vv,.5); meds.append(med); los.append(med-q(vv,.25)); his.append(q(vv,.75)-med)
    ax.plot([0,1],meds,color=COLORS['EgoAnchor'],lw=1.8,zorder=3)
    ax.errorbar([0,1],meds,yerr=[los,his],fmt='s',ms=4.5,capsize=2.5,color=COLORS['EgoAnchor'],mfc='white',mec=COLORS['EgoAnchor'],zorder=4)
    ax.set_xticks([0,1],[x.replace('EgoAnchor w/o StaticLock','w/o lock').replace('EgoAnchor','On').replace('Capture time','Capture').replace('Arrival time','Arrival') for x in order])
    ax.set_title(title,pad=3); ax.set_ylabel(ylabel); clean_ax(ax,'y'); ax.tick_params(axis='x',length=0)
# VCD
ax=axes[2]
events=defaultdict(list)
for r in rows3:
    if r['panel']=='(c) VCD score risk-coverage' and r['series']=='Event curve':
        events[(r['session_id'],r['trial_id'],r['segment_id'])].append((float(r['x_value']),float(r['y_value'])))
for pts in events.values():
    pts=sorted(pts); ax.plot([x for x,_ in pts],[y for _,y in pts],color='#C4C8CE',alpha=.65,lw=.55)
med=sorted((float(r['x_value']),float(r['y_value'])) for r in rows3 if r['panel']=='(c) VCD score risk-coverage' and r['series']=='Median')
lo=sorted((float(r['x_value']),float(r['y_value'])) for r in rows3 if r['panel']=='(c) VCD score risk-coverage' and r['series']=='IQR lower')
hi=sorted((float(r['x_value']),float(r['y_value'])) for r in rows3 if r['panel']=='(c) VCD score risk-coverage' and r['series']=='IQR upper')
ax.fill_between([x*100 for x,_ in med],[y for _,y in lo],[y for _,y in hi],color=COLORS['purple'],alpha=.16,linewidth=0)
ax.plot([x*100 for x,_ in med],[y for _,y in med],color=COLORS['purple'],lw=1.8)
ax.set_title('VCD risk-coverage',pad=3); ax.set_xlabel('Retained observations (%)'); ax.set_ylabel('Selective risk (mm)'); ax.set_xlim(0,100); clean_ax(ax,'both')
# temporal
ax=axes[3]
strategies=[('EgoAnchor w/o StaticLock','Linear/SLERP',COLORS['EgoAnchor'],'o'),('Hermite Interpolation','Hermite',COLORS['Capture-Hold'],'^'),('Smoothed KF Extrapolation','Smoothed KF',COLORS['Arrival-Hold'],'s')]
for sid,label,col,mark in strategies:
    rr=[r for r in rows3 if r['panel']=='(d) Runtime temporal strategies' and r['series']==sid]
    xs=[float(r['x_value']) for r in rr]; ys=[float(r['y_value']) for r in rr]
    ax.scatter(xs,ys,s=18,color=col,alpha=.42,marker=mark,edgecolors='none',label=label)
    ax.scatter([q(xs,.5)],[q(ys,.5)],s=55,color=col,marker=mark,edgecolors='white',linewidth=.8,zorder=4)
ax.set_title('Temporal synthesis',pad=3); ax.set_xlabel('Effective lag (ms)'); ax.set_ylabel('Aligned RMSE (mm)'); ax.set_yscale('log'); clean_ax(ax,'both'); ax.legend(frameon=False,loc='upper right',handletextpad=.4,borderaxespad=.2)
fig.subplots_adjust(left=.055,right=.995,bottom=.24,top=.82,wspace=.42)
save(fig,'exp2_attribution')

# ---------- Experiment 3 forest + choices ----------
main = sheet_values(DATA / 'experiment3_analysis.xlsx', '主结果', 1, 16, 1, 12)
headers=main[3]; mainrows=[dict(zip(headers,r)) for r in main[4:]]

def parse_interval(s):
    m=re.match(r'\s*([-+]?\d+(?:\.\d+)?)\s*\[\s*([-+]?\d+(?:\.\d+)?)\s*,\s*([-+]?\d+(?:\.\d+)?)\s*\]\s*',str(s))
    return tuple(float(x) for x in m.groups())

fig = plt.figure(figsize=(11.6,4.35))
gs=fig.add_gridspec(1,3,width_ratios=[1.55,1.35,.9],wspace=.38)
for ax,family,title in [(fig.add_subplot(gs[0,0]),'主证实条目','Anchor-quality outcomes'),(fig.add_subplot(gs[0,1]),'已发表量表','Published scales')]:
    rr=[r for r in mainrows if r['检验家族']==family]
    labels=[]
    for r in rr:
        s=str(r['指标'])
        s=re.sub(r'^Q\d+\s*','',s)
        labels.append(s.replace('稳定—响应平衡','Stability-response').replace('静止稳定','Static stability').replace('位置正确','Position correctness').replace('运动附着','Motion attachment').replace('姿态一致','Orientation consistency').replace('恢复一致','Recovery consistency').replace('依赖意愿','Willingness to rely').replace('AQ 嵌入质量','AQ embedding').replace('AQ 交互质量','AQ interaction').replace('TiA 可靠性/能力','TiA reliability').replace('TiA 理解/可预测性','TiA predictability').replace('S-TIAS 信任','S-TIAS trust'))
    ys=list(range(len(rr)))[::-1]
    for y,r in zip(ys,rr):
        medv,loi,hii=parse_interval(r['配对差中位数 [Q1, Q3]'])
        p=float(r['Holm 校正 p']); sig=p<.05
        col=COLORS['EgoAnchor'] if sig else '#8A8F98'
        ax.errorbar([medv],[y],xerr=[[medv-loi],[hii-medv]],fmt='o' if sig else 'D',ms=4.7,mfc=col if sig else 'white',mec=col,color=col,capsize=2.5,elinewidth=1.25,zorder=3)
        ptxt='<.001' if p<.001 else f'{p:.3f}'
        ax.text(1.02,y,ptxt,transform=ax.get_yaxis_transform(),ha='left',va='center',fontsize=6.4,color=col,fontweight='bold' if sig else 'normal')
    ax.axvline(0,color='#555B63',lw=.8)
    ax.set_yticks(ys,labels)
    ax.set_xlabel('Median paired difference (EgoAnchor - One-Euro)')
    ax.set_title(title,pad=5)
    ax.set_xlim(-.5,1.6 if family=='主证实条目' else 1.1)
    clean_ax(ax,'x')
    ax.text(1.02,1.02,'p$_{Holm}$',transform=ax.transAxes,ha='left',va='bottom',fontsize=6.4,color='#555B63')
# choice bars
ax=fig.add_subplot(gs[0,2]); ax.set_title('Final choices',pad=5)
choices=[('Overall preference',[15,4,5]),('Trust choice',[18,1,5])]
cols=[COLORS['EgoAnchor'],COLORS['One-Euro Anchor'],'#C8CCD2']
for i,(lab,vals) in enumerate(choices):
    left=0
    for v,c in zip(vals,cols):
        ax.barh(i,v,left=left,color=c,height=.52,edgecolor='white',linewidth=.6)
        if v>=3: ax.text(left+v/2,i,str(v),ha='center',va='center',fontsize=7,color='white' if c!='#C8CCD2' else '#333',fontweight='bold')
        left+=v
ax.set_yticks([0,1],[x[0] for x in choices]); ax.set_xlim(0,24); ax.invert_yaxis(); ax.set_xlabel('Participants (N=24)'); clean_ax(ax,'x')
legend=[patches.Patch(color=cols[0],label='EgoAnchor'),patches.Patch(color=cols[1],label='One-Euro'),patches.Patch(color=cols[2],label='No preference')]
ax.legend(handles=legend,frameon=False,loc='lower center',bbox_to_anchor=(.5,-.34),ncol=1)
fig.subplots_adjust(left=.14,right=.98,bottom=.20,top=.86)
save(fig,'exp3_perception')

print('Generated polished figures in',OUT)

# ---------- Experiment 2 combined figure + compact attribution table ----------
from PIL import Image as PILImage
img = PILImage.open(OUT/'exp2_attribution.png').convert('RGB')
fig = plt.figure(figsize=(11.6,4.55))
gs = fig.add_gridspec(2,1,height_ratios=[3.25,1.25],hspace=.04)
ax0=fig.add_subplot(gs[0]); ax0.imshow(img); ax0.axis('off')
ax1=fig.add_subplot(gs[1]); ax1.axis('off')
cols=['Design','Enabled','Alternative','Paired effect','Consistency']
rows=[
 ['Capture-time alignment','15.38 mm','43.75 mm','2.69x','12/12'],
 ['StaticLock','0.82 mm','13.73 mm','17.06x','12/12'],
 ['VCD ranking','4.79 mm','7.25 mm','1.46x','12/12'],
 ['Historical synthesis','8.93 mm / 4.70 deg','45.64 mm / 12.21 deg','5.46x / 2.63x','16/16, 12/12'],
]
tab=ax1.table(cellText=rows,colLabels=cols,loc='center',cellLoc='center',colLoc='center',colWidths=[.19,.21,.24,.19,.13])
tab.auto_set_font_size(False); tab.set_fontsize(7.0); tab.scale(1,1.28)
for (r,c),cell in tab.get_celld().items():
    cell.set_linewidth(.45)
    cell.set_edgecolor('#D3D7DD')
    if r==0:
        cell.set_facecolor('#F0F2F5'); cell.set_text_props(weight='bold',color='#30363D')
    else:
        cell.set_facecolor('white')
        if c==3: cell.set_text_props(weight='bold',color=COLORS['EgoAnchor'])
fig.subplots_adjust(left=.02,right=.98,top=.99,bottom=.02)
save(fig,'exp2_combined')
